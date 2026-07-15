"""
╔══════════════════════════════════════════════════════════════════╗
║     ثبت تختال های فولاد سفید دشت                               ║
║     شرکت سازه پیشگام مدیسه                                      ║
║     نسخه ۳.۰ - سیستم مدیریت اسلب (تاریخ شمسی)               ║
╚══════════════════════════════════════════════════════════════════╝

نصب پیش‌نیازها:
    pip install customtkinter openpyxl pillow jdatetime pyodbc

اجرا:
    python slab_system_v4.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font as tkfont, colorchooser, simpledialog
import json, os, hashlib, random, string, datetime, re
import threading, time, sys, shutil, atexit

# کیوسک پیش‌فرض — عین فایل مرجع؛ فقط STF_ADMIN=1 معاف است
if os.environ.get("STF_ADMIN") != "1":
    os.environ["STF_KIOSK"] = "1"

try:
    from keyboard_layout import install_keyboard_switcher, attach_switcher_bar, register_refresh
except ImportError:
    def install_keyboard_switcher(_root):
        return None
    def attach_switcher_bar(parent, colors=None, compact=False):
        return None, lambda: None
    def register_refresh(_fn):
        pass

# ═══════════════════════════════════════════════════════════
#  DPI Awareness (ویندوز) — بدون این تنظیم، روی صفحه‌نمایش‌هایی که
#  Scale (مثلاً ۱۲۵٪/۱۵۰٪) روشن دارند، ویندوز کل پنجره را به‌صورت
#  مجازی/scaled رندر می‌کند و آن را در وسط صفحه کوچک‌تر نشان می‌دهد —
#  دقیقاً همان چیزی که باعث می‌شد با وجود تنظیم عرض/طول برابر با کل
#  صفحه، باز هم حاشیه‌ی خالی از چپ و راست دیده شود. با اعلام
#  DPI-Awareness به ویندوز، پیکسل‌های واقعی صفحه در اختیار برنامه قرار
#  می‌گیرد و دیگر هیچ scale/centering مجازی‌ای اتفاق نمی‌افتد.
# ═══════════════════════════════════════════════════════════
if os.name == "nt":
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)   # Per-Monitor DPI Aware
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()        # fallback برای نسخه‌های قدیمی‌تر ویندوز
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════
#  مدیریت اسکرول مرکزی — اسکرول ماوس روی ویجت زیر نشانگر
# ═══════════════════════════════════════════════════════════
_active_scroll_target = None   # canvas/tree فعال (پشتیبان)

def _find_scrollable_under(widget):
    """از ویجت زیر موس به بالا می‌رود تا Treeview/Canvas/Listbox/Text پیدا کند."""
    w = widget
    seen = set()
    while w is not None and id(w) not in seen:
        seen.add(id(w))
        try:
            cls = w.winfo_class()
        except Exception:
            break
        if cls in ("Treeview", "Canvas", "Listbox", "Text"):
            return w
        try:
            w = w.master
        except Exception:
            break
    return None


def _global_mousewheel(event):
    """اسکرول ماوس — بدون نیاز به کلیک/فوکوس قبلی روی نوار اسکرول."""
    global _active_scroll_target
    target = None
    try:
        under = event.widget.winfo_containing(event.x_root, event.y_root)
        if under is not None:
            target = _find_scrollable_under(under)
    except Exception:
        target = None
    if target is None:
        target = _active_scroll_target
    if target is None:
        return
    # صفحه لاگین: اگر اسکرول قفل است زمینه نباید بپرد
    try:
        root = target.winfo_toplevel()
        if getattr(root, "_login_scroll_enabled", None) is False and target is getattr(root, "_login_canvas", None):
            try:
                target.yview_moveto(0)
            except Exception:
                pass
            return "break"
    except Exception:
        pass
    try:
        delta = int(-1 * (event.delta / 120))
        if delta == 0:
            delta = -1 if event.delta > 0 else 1
        target.yview_scroll(delta, "units")
        # بعد از اسکرول لاگین، شیشه را فوری با زمینهٔ ثابت هم‌تراز کن
        try:
            root = target.winfo_toplevel()
            if hasattr(root, "_reapply_login_glass_layers") and target is getattr(root, "_login_canvas", None):
                root.after_idle(root._reapply_login_glass_layers)
        except Exception:
            pass
        return "break"
    except Exception:
        if target is _active_scroll_target:
            _active_scroll_target = None


def register_scroll_canvas(canvas, *extra_widgets):
    """canvas/tree را برای اسکرول ماوس ثبت می‌کند — وقتی موس داخل آن است"""
    global _active_scroll_target
    def _enter(e):
        global _active_scroll_target
        _active_scroll_target = canvas
    def _leave(e):
        global _active_scroll_target
        if _active_scroll_target is canvas:
            _active_scroll_target = None
    canvas.bind("<Enter>", _enter, add="+")
    canvas.bind("<Leave>", _leave, add="+")
    for w in extra_widgets:
        try:
            w.bind("<Enter>", _enter, add="+")
            w.bind("<Leave>", _leave, add="+")
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════
#  تاریخ شمسی
# ═══════════════════════════════════════════════════════════
try:
    import jdatetime
    def to_shamsi(dt: datetime.datetime) -> str:
        """تبدیل datetime میلادی به رشته شمسی  ۱۴۰۴/۰۳/۱۵  ۱۴:۳۰:۲۵"""
        jdt = jdatetime.datetime.fromgregorian(datetime=dt)
        return jdt.strftime("%Y/%m/%d  %H:%M:%S")
    SHAMSI_OK = True
except ImportError:
    # fallback خودنوشته بدون کتابخانه خارجی
    SHAMSI_OK = False
    def _g2j(gy, gm, gd):
        """تبدیل تاریخ میلادی به شمسی — الگوریتم صحیح"""
        g_days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        j_days_in_month = [31, 31, 31, 31, 31, 31, 30, 30, 30, 30, 30, 29]
        gy2 = gy - 1600
        gm -= 1
        gd -= 1
        g_day_no = 365*gy2 + (gy2+3)//4 - (gy2+99)//100 + (gy2+399)//400
        for i in range(gm):
            g_day_no += g_days_in_month[i]
        if gm > 1 and (gy%4==0 and (gy%100!=0 or gy%400==0)):
            g_day_no += 1
        g_day_no += gd
        j_day_no = g_day_no - 79
        j_np = j_day_no // 12053
        j_day_no %= 12053
        jy = 979 + 33*j_np + 4*(j_day_no//1461)
        j_day_no %= 1461
        if j_day_no >= 366:
            jy += (j_day_no-1)//365
            j_day_no = (j_day_no-1)%365
        jm = 12; jd = j_day_no + 1
        for i in range(11):
            if j_day_no >= j_days_in_month[i]:
                j_day_no -= j_days_in_month[i]
            else:
                jm = i+1; jd = j_day_no+1; break
        return jy, jm, jd

    def to_shamsi(dt: datetime.datetime) -> str:
        jy, jm, jd = _g2j(dt.year, dt.month, dt.day)
        return f"{jy}/{jm:02d}/{jd:02d}  {dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}"

try:
    import customtkinter as ctk
    CTK = True
except ImportError:
    CTK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLSX = True
except ImportError:
    XLSX = False

# ═══════════════════════════════════════════════════════════
#  پایگاه داده — SQL Server / API (SteelFactory2)
# ═══════════════════════════════════════════════════════════
from db_backend import (
    load_db as _sql_load_db,
    save_db as _sql_save_db,
    is_storage_alive,
    set_current_user as _set_db_user,
    export_current_db,
    restore_db_from_file,
    log_audit as _db_log_audit,
)

def _app_base_dir():
    """
    مسیر پوشه‌ی برنامه — چه به‌صورت اسکریپت پایتون اجرا شود چه به‌صورت exe
    (PyInstaller). این تابع مستقل از این است که برنامه از کجا اجرا شده
    (دابل‌کلیک، شورتکات با "Start in" متفاوت، Task Scheduler، یا ترمینال
    در مسیر دیگر) — همیشه پوشه‌ای را برمی‌گرداند که خودِ exe/اسکریپت در آن
    قرار دارد. قبلاً DB_FILE و مسیرهای بک‌آپ نسبی (وابسته به cwd) بودند که
    باعث می‌شد روی سیستم دیگر، یا حتی همین سیستم با یک شورتکات متفاوت،
    برنامه دیتابیس را پیدا نکند و انگار از نو خالی شروع شود.
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ── پوشه‌ی بک‌آپ — snapshot فایلی (علاوه بر SQL) کنار exe ──
# هرجا نرم‌افزار کپی/منتقل شود، پوشه‌ی بک‌آپ هم خودش همان‌جا کنارش ساخته می‌شود
# (به‌جای مسیر ثابت در Documents که به سیستم اول وابسته بود)
BACKUP_DIR = os.path.join(_app_base_dir(), "backups")
BACKUP_DIR2 = ""   # مسیر دوم اختیاری است؛ کاربر می‌تواند از تب بک‌آپ یک مسیر اضافه انتخاب کند

# رمز قفل Safe Mode و دسترسی به مدیریت بک‌آپ
STF_LOCK_PASSWORD = "Reza9063"

# سقف نمایش ردیف در جداول — فقط UI (داده در DB کامل می‌ماند)
# برای سرعت و جلوگیری از هنگ با داده‌های خیلی بزرگ
TREE_ROW_LIMIT = 800


def _backup_data_dir(path: str) -> str:
    """مسیر داده بک‌آپ — همان پوشه (ساده)."""
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass
    return path


def _ensure_backup_dirs_protected():
    """فقط ساخت پوشه‌های backups و archives — بدون قفل پراکنده."""
    paths = [BACKUP_DIR]
    if BACKUP_DIR2:
        paths.append(BACKUP_DIR2)
    try:
        db = load_db()
        if isinstance(db, dict):
            for p in (db.get("settings") or {}).get("backup_paths") or []:
                if p and str(p).strip():
                    paths.append(str(p).strip())
    except Exception:
        pass
    paths.append(os.path.join(_app_base_dir(), "archives"))
    paths.append(os.path.join(_app_base_dir(), "backups"))
    seen = set()
    for p in paths:
        if not p:
            continue
        ap = os.path.abspath(p)
        if ap in seen:
            continue
        seen.add(ap)
        try:
            os.makedirs(ap, exist_ok=True)
        except Exception:
            pass
    # پاکسازی پوشه‌های قدیمی backup / *.secure / *.gate اگر مانده
    try:
        if r"D:\SteelFactory2-v2" not in sys.path:
            sys.path.insert(0, r"D:\SteelFactory2-v2")
        from shared.backup_vault import cleanup_extra_backup_dirs
        cleanup_extra_backup_dirs(_app_base_dir())
    except Exception:
        pass


def _ask_lock_password(parent=None, title="قفل امنیتی", prompt=None) -> bool:
    """درخواست رمز Reza9063 — حداکثر ۳ بار (Safe Mode و موارد امنیتی)."""
    msg = prompt or "برای ادامه، رمز مدیریت را وارد کنید:"
    for _ in range(3):
        pw = simpledialog.askstring(title, msg, show="*", parent=parent)
        if pw is None:
            return False
        if pw == STF_LOCK_PASSWORD:
            return True
        messagebox.showerror("رمز اشتباه", "رمز وارد شده صحیح نیست.", parent=parent)
    return False

# پوشه‌ی دارایی‌های قابل‌حمل (تصویر پس‌زمینه و امثال آن) — کنار exe ساخته می‌شود
# تا وقتی برنامه روی سیستم دیگری نصب/کپی می‌شود، فایل‌ها هم همراهش بیایند
# (برخلاف ذخیره‌ی مسیر مطلق فایل روی سیستم اول که فقط همان‌جا کار می‌کند)
ASSETS_DIR = os.path.join(_app_base_dir(), "app_assets")


def _sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _clear_local_background_files():
    try:
        os.makedirs(ASSETS_DIR, exist_ok=True)
    except Exception:
        return
    for e in (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"):
        p = os.path.join(ASSETS_DIR, "login_background" + e)
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass


def _resolve_server_base_url():
    """آدرس API ادمین برای همگام‌سازی تصویر زمینه."""
    if os.environ.get("STF_ADMIN") == "1":
        return "http://127.0.0.1:8080"
    try:
        from client.db_bridge import reload_server_url
        return (reload_server_url(discover=False) or "").rstrip("/")
    except Exception:
        pass
    try:
        from shared.config import load_config
        return (load_config().get("server_url") or "http://192.168.0.101:8080").rstrip("/")
    except Exception:
        return "http://192.168.0.101:8080"


def _ensure_stf_shared_on_path():
    """مسیر فریم‌ورک SteelFactory2 را برای import shared.* اضافه می‌کند."""
    candidates = [
        os.environ.get("STF_ROOT"),
        r"D:\SteelFactory2-v2",
        os.path.join(_app_base_dir(), "..", "SteelFactory2-v2"),
        os.path.join(_app_base_dir(), "..", "..", "SteelFactory2-v2"),
    ]
    for c in candidates:
        if not c:
            continue
        root = os.path.abspath(c)
        if os.path.isdir(os.path.join(root, "shared")) and root not in sys.path:
            sys.path.insert(0, root)
            return root
    return None


def _publish_background_to_server(local_path):
    """آپلود فایل پس‌زمینه به API تا همه کلاینت‌ها دانلود کنند."""
    import urllib.request
    import urllib.parse
    ext = os.path.splitext(local_path)[1].lower() or ".jpg"
    base = _resolve_server_base_url()
    url = f"{base}/api/v1/assets/background?ext={urllib.parse.quote(ext)}"
    with open(local_path, "rb") as f:
        body = f.read()
    req = urllib.request.Request(
        url, data=body, method="PUT",
        headers={"Content-Type": "application/octet-stream"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _delete_background_on_server():
    import urllib.request
    base = _resolve_server_base_url()
    req = urllib.request.Request(f"{base}/api/v1/assets/background", method="DELETE")
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode("utf-8"))


def ensure_login_background_local(db=None):
    """
    تصویر پس‌زمینه را محلی می‌کند؛ در کلاینت در صورت نیاز از سرور دانلود می‌شود.
    مسیر فایل محلی را برمی‌گرداند یا "".
    """
    if db is None:
        try:
            db = load_db()
        except Exception:
            db = {}
    settings = (db or {}).get("settings") or {}
    # اول مسیر محلی معتبر
    for e in (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"):
        cand = os.path.join(ASSETS_DIR, "login_background" + e)
        if os.path.isfile(cand):
            want = (settings.get("background_image_sha256") or "").strip()
            if want:
                try:
                    if _sha256_file(cand) == want:
                        return cand
                except Exception:
                    pass
            elif settings.get("background_image"):
                return cand
    # دانلود از سرور (کلاینت یا ادمین از API)
    try:
        from client.db_bridge import sync_background_asset
        path = sync_background_asset(db)
        if path and os.path.isfile(path):
            return path
    except Exception:
        pass
    try:
        import urllib.request
        want_sha = (settings.get("background_image_sha256") or "").strip()
        if not want_sha and not (settings.get("background_image") or "").strip():
            return ""
        base = _resolve_server_base_url()
        req = urllib.request.Request(f"{base}/api/v1/assets/background", method="GET")
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = resp.read()
            ext = (resp.headers.get("X-Background-Ext") or settings.get("background_image_ext") or ".jpg").lower()
            if not ext.startswith("."):
                ext = "." + ext
        _clear_local_background_files()
        os.makedirs(ASSETS_DIR, exist_ok=True)
        dest = os.path.join(ASSETS_DIR, "login_background" + ext)
        with open(dest, "wb") as f:
            f.write(data)
        return dest
    except Exception:
        return ""


# ═══════════════════════════════════════════════════════════
#  پوشه گزارش‌ها — یک پوشه اصلی "report" کنار برنامه، با ۵ زیرپوشه
#  اختصاصی برای هر نوع گزارش، تا هر گزارش به‌صورت پیش‌فرض دقیقاً
#  در پوشه‌ی مخصوص خودش ذخیره شود (کاربر همچنان می‌تواند مسیر را
#  در پنجره‌ی ذخیره تغییر دهد، ولی مسیر پیش‌فرض همیشه همین‌جاست)
# ═══════════════════════════════════════════════════════════
REPORT_DIR = os.path.join(_app_base_dir(), "report")

REPORT_SUBDIRS = {
    "overview":   "01- گزارش کلی (نمای کلی)",
    "qc_excel":   "02- گزارش اکسل کنترل کیفی",
    "melt_excel": "03- گزارش اکسل ثبت ذوب",
    "lab_pdf":    "04- گزارش PDF آزمایشگاه",
    "qc_pdf":     "05- گزارش PDF اسلب‌های کنترل کیفی شده",
}

def get_report_dir(key: str) -> str:
    """پوشه‌ی اختصاصی گزارش را برمی‌گرداند و در صورت نبود، آن را می‌سازد."""
    sub = REPORT_SUBDIRS.get(key, key)
    path = os.path.join(REPORT_DIR, sub)
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass
    return path

def ensure_report_dirs():
    """در ابتدای اجرای برنامه، پوشه اصلی report و هر ۵ زیرپوشه را می‌سازد."""
    for key in REPORT_SUBDIRS:
        get_report_dir(key)


# ═══════════════════════════════════════════════════════════
#  نام‌گذاری تمیز گزارش‌ها — نام گزارش + تاریخ شمسی (بدون ساعت)
#  و در صورت تکراری بودن نام، به‌جای جایگزینی فایل قبلی، پسوند
#  -2 و -3 و... اضافه می‌شود تا هیچ گزارشی روی گزارش قبلی ذخیره نشود.
# ═══════════════════════════════════════════════════════════
def shamsi_date_for_filename() -> str:
    """فقط بخش تاریخ (بدون ساعت) به‌صورت مناسب برای نام فایل — مثلاً 1405-04-13"""
    return to_shamsi(datetime.datetime.now()).split("  ")[0].replace("/", "-")

def make_unique_filename(folder: str, filename: str) -> str:
    """اگر فایلی با همین نام از قبل در پوشه وجود داشته باشد، به‌جای
    جایگزین‌کردن آن، با افزودن -2، -3، ... یک نام جدید و غیرتکراری
    می‌سازد. مثال: «گزارش کنترل کیفی1405-04-13.pdf» ← دومین بار همان روز
    می‌شود «گزارش کنترل کیفی1405-04-13-2.pdf»."""
    try:
        if not os.path.exists(os.path.join(folder, filename)):
            return filename
        base, ext = os.path.splitext(filename)
        n = 2
        while os.path.exists(os.path.join(folder, f"{base}-{n}{ext}")):
            n += 1
        return f"{base}-{n}{ext}"
    except Exception:
        return filename

def open_folder_in_explorer(path: str):
    """پوشه‌ی داده‌شده را در File Explorer (یا معادل آن) باز می‌کند."""
    try:
        os.makedirs(path, exist_ok=True)
        if os.name == "nt":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as ex:
        messagebox.showerror("خطا", f"امکان باز کردن پوشه وجود ندارد:\n{ex}")


# ═══════════════════════════════════════════════════════════
#  کنترل نوار تسک‌بار ویندوز — «قفل نرم» صفحه: تا وقتی برنامه باز
#  و روی صفحه است تسک‌بار/استارت مخفی و برنامه جلو می‌ماند.
# ═══════════════════════════════════════════════════════════
_SHELL_CLASSES = (
    "Shell_TrayWnd",
    "Shell_SecondaryTrayWnd",
    "Windows.UI.Core.CoreWindow",
    "XamlExplorerDockHost",
    "TopLevelWindowForOverflowXamlIsland",
    "MultitaskingViewFrame",
)


def _toggle_taskbars(show: bool):
    if os.name != "nt":
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        SW_SHOW, SW_HIDE = 1, 0
        for cls in ("Shell_TrayWnd", "Shell_SecondaryTrayWnd"):
            hwnd = user32.FindWindowW(cls, None)
            while hwnd:
                user32.ShowWindow(hwnd, SW_SHOW if show else SW_HIDE)
                hwnd = user32.FindWindowExW(None, hwnd, cls, None)
    except Exception:
        pass


def hide_windows_shell():
    """تسک‌بار، منوی استارت و پنجره‌های Win11 را مخفی می‌کند."""
    if os.name != "nt":
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        SW_HIDE = 0
        for cls in _SHELL_CLASSES:
            hwnd = user32.FindWindowW(cls, None)
            while hwnd:
                user32.ShowWindow(hwnd, SW_HIDE)
                hwnd = user32.FindWindowExW(None, hwnd, cls, None)
    except Exception:
        pass


def show_windows_shell(force_broadcast=False):
    """بازگرداندن تسک‌بار و shell — معکوس hide_windows_shell."""
    if os.name != "nt":
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        SW_SHOW = 5
        for cls in _SHELL_CLASSES:
            hwnd = user32.FindWindowW(cls, None)
            while hwnd:
                user32.ShowWindow(hwnd, SW_SHOW)
                hwnd = user32.FindWindowExW(None, hwnd, cls, None)
        # TaskbarCreated فقط گاه‌به‌گاه — هر ۳۵۰ms سیستم را قفل می‌کرد
        now = time.time()
        last = getattr(show_windows_shell, "_last_broadcast", 0.0)
        if force_broadcast or (now - last) > 8.0:
            show_windows_shell._last_broadcast = now
            try:
                msg = user32.RegisterWindowMessageW("TaskbarCreated")
                if msg:
                    user32.SendNotifyMessageW(0xFFFF, msg, 0, 0)
            except Exception:
                pass
    except Exception:
        pass


def _ui_heavy_ok(win):
    """حلقه‌های سنگین UI فقط در تمام‌صفحهٔ فعال — نه مینیمایز/کوچک."""
    try:
        if win is None or not win.winfo_exists():
            return False
    except Exception:
        return False
    if getattr(win, "_app_closing", False):
        return False
    if getattr(win, "_minimizing_authorized", False):
        return False
    if getattr(win, "_was_in_taskbar", False):
        return False
    try:
        if _hwnd_is_iconic(win) or win.state() == "iconic":
            return False
    except Exception:
        return False
    return _app_shell_lock_active(win)


def _ui_idle_ms(win, active_ms=500, idle_ms=2500, minimized_ms=8000):
    """فاصلهٔ پولینگ — وقتی مینیمایز/کوچک است خیلی کمتر کار کن."""
    try:
        if _hwnd_is_iconic(win) or getattr(win, "_minimizing_authorized", False):
            return minimized_ms
    except Exception:
        pass
    if not _ui_heavy_ok(win):
        return idle_ms
    return active_ms


def hide_all_taskbars(win=None):
    """تسک‌بار را مخفی می‌کند — نه هنگام مینیمایز."""
    if win is not None and not _app_shell_lock_active(win):
        return
    hide_windows_shell()


def show_all_taskbars():
    show_windows_shell()
    _toggle_taskbars(True)


def _restore_taskbar_on_exit():
    """Safety net — taskbar must work after app closes (admin + client)."""
    try:
        show_all_taskbars()
    except Exception:
        pass


atexit.register(_restore_taskbar_on_exit)


def _is_kiosk_station() -> bool:
    """فقط پنل ادمین از قفل کیوسک معاف است — مثل فایل مرجع، پیش‌فرض همیشه کیوسک."""
    return os.environ.get("STF_ADMIN") != "1"


def _kiosk_key_block_active(win) -> bool:
    """مسدودسازی کلیدهای سیستمی — حتی قبل از تکمیل UI."""
    if not _is_kiosk_station():
        return False
    if getattr(win, "_app_closing", False):
        return False
    if getattr(win, "_minimizing_authorized", False):
        return False
    return True


def _app_shell_lock_active(win):
    """آیا برنامه در حالت تمام‌صفحه/قفل است؟ (نه حالت کوچک یا مینیمایز)"""
    if getattr(win, "_app_closing", False):
        return False
    if getattr(win, "_was_full", True) is False:
        return False
    if getattr(win, "_minimizing_authorized", False):
        return False
    if getattr(win, "_was_in_taskbar", False):
        return False
    try:
        if not win.winfo_exists():
            return False
        if _hwnd_is_iconic(win):
            return False
        if win.state() == "iconic":
            return False
    except Exception:
        return False
    return True


def _modal_dialog_active(win):
    """آیا دیالوگ/پاپ‌آپ باز است؟"""
    if getattr(win, "_modal_depth", 0) > 0:
        return True
    if getattr(win, "_lock_dialog_open", False):
        return True
    if getattr(win, "_minimize_dialog_open", False):
        return True
    if getattr(win, "_restore_dialog_open", False):
        return True
    try:
        for ch in win.winfo_children():
            if isinstance(ch, tk.Toplevel):
                try:
                    if ch.winfo_exists() and ch.winfo_viewable():
                        return True
                except Exception:
                    pass
    except Exception:
        pass
    return False


def _modal_enter(win):
    win._modal_depth = getattr(win, "_modal_depth", 0) + 1


def _modal_leave(win):
    win._modal_depth = max(0, getattr(win, "_modal_depth", 0) - 1)


def _register_popup(root, popup):
    if root is None or popup is None:
        return
    popups = getattr(root, "_active_popups", None)
    if popups is None:
        root._active_popups = []
        popups = root._active_popups
    if popup not in popups:
        popups.append(popup)
    _start_popup_zorder_guard(root)


def _unregister_popup(root, popup):
    if root is None:
        return
    popups = getattr(root, "_active_popups", [])
    try:
        popups.remove(popup)
    except ValueError:
        pass


def _get_top_popup(root):
    for popup in reversed(getattr(root, "_active_popups", [])):
        try:
            if popup.winfo_exists() and popup.winfo_viewable():
                return popup
        except Exception:
            pass
    return None


def _win32_set_topmost(win, topmost=True):
    if os.name != "nt":
        return
    try:
        import ctypes
        hwnd = _get_toplevel_hwnd(win)
        insert_after = -1 if topmost else -2
        flags = 0x0002 | 0x0001 | 0x0040  # NOMOVE | NOSIZE | SHOWWINDOW
        ctypes.windll.user32.SetWindowPos(
            hwnd, insert_after, 0, 0, 0, 0, flags)
    except Exception:
        pass


def _win32_stack_popup_above_main(root, popup):
    """Place popup HWND directly above main window in Z-order."""
    if os.name != "nt" or root is None or popup is None:
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        main_hwnd = _get_toplevel_hwnd(root)
        pop_hwnd = _get_toplevel_hwnd(popup)
        if not main_hwnd or not pop_hwnd:
            return
        flags = 0x0002 | 0x0001 | 0x0040
        user32.SetWindowPos(pop_hwnd, main_hwnd, 0, 0, 0, 0, flags)
        user32.SetWindowPos(pop_hwnd, -1, 0, 0, 0, 0, flags)
    except Exception:
        pass


def _bring_popup_to_front(popup, focus=True):
    """پاپ‌آپ/دیالوگ همیشه جلوتر از پنجرهٔ اصلی کیوسک."""
    if popup is None:
        return
    try:
        if not popup.winfo_exists():
            return
    except Exception:
        return
    root = _root_window(popup)
    try:
        popup.deiconify()
        popup.lift()
        popup.attributes("-topmost", True)
    except Exception:
        pass
    if root is not None:
        _win32_stack_popup_above_main(root, popup)
    _win32_set_topmost(popup, True)
    try:
        popup.update_idletasks()
    except Exception:
        pass
    if os.name == "nt" and focus:
        # فقط وقتی واقعاً می‌خواهیم فوکوس بگیریم — وگرنه Entry رمز تایپ نمی‌پذیرد
        try:
            import ctypes
            pop_hwnd = _get_toplevel_hwnd(popup)
            if root is not None:
                _win32_stack_popup_above_main(root, popup)
            ctypes.windll.user32.SetForegroundWindow(pop_hwnd)
        except Exception:
            pass
    if focus:
        try:
            popup.focus_force()
        except Exception:
            pass
        # اگر دیالوگ فیلد ورودی دارد، فوکوس را به همان برگردان
        try:
            ent = getattr(popup, "_focus_entry", None)
            if ent is not None and ent.winfo_exists():
                ent.focus_set()
        except Exception:
            pass


def _bind_popup_kiosk_keys(popup, parent=None):
    """Block system shortcuts while a popup is open."""
    if not _is_kiosk_station():
        return
    root = _root_window(parent or popup)

    def _block_sys(event=None):
        if root is not None:
            _react_to_blocked_system_key(root)
        return "break"

    for seq in (
        "<Alt-Escape>", "<Alt-Tab>", "<Control-Escape>",
        "<Control-Shift-Escape>", "<Meta-Key>", "<Meta-Tab>",
        "<Control-Shift-Tab>", "<F11>", "<F12>",
    ):
        try:
            popup.bind(seq, _block_sys, add="+")
        except Exception:
            pass


def _focus_existing_edit_popup(owner, key):
    """اگر پنجره ویرایش همین رکورد باز است، فقط آن را جلو بیاور — از باز شدن چندباره جلوگیری می‌کند."""
    store = getattr(owner, "_open_edit_popups", None)
    if not store:
        return None
    pop = store.get(key)
    if pop is None or pop is True:
        # True = در حال باز شدن (قفل موقت) — پنجره جدید نساز
        return True if pop is True else None
    try:
        if pop.winfo_exists():
            _bring_popup_to_front(pop, focus=True)
            try:
                pop.lift()
                pop.focus_force()
            except Exception:
                pass
            return pop
    except Exception:
        pass
    store.pop(key, None)
    return None


def _acquire_edit_popup(owner, key):
    """قفل اتمی: اگر پنجره باز/در حال باز شدن است True برمی‌گرداند؛ وگرنه اسلات را رزرو می‌کند و None."""
    existing = _focus_existing_edit_popup(owner, key)
    if existing:
        return existing
    if not hasattr(owner, "_open_edit_popups") or owner._open_edit_popups is None:
        owner._open_edit_popups = {}
    # رزرو فوری تا کلیک‌های پشت‌سرهم قبل از Toplevel پنجره دوم نسازند
    if owner._open_edit_popups.get(key) is True:
        return True
    owner._open_edit_popups[key] = True
    return None


def _release_edit_popup_claim(owner, key):
    """اگر ساخت پنجره شکست خورد، قفل موقت را آزاد کن."""
    store = getattr(owner, "_open_edit_popups", None)
    if store is not None and store.get(key) is True:
        store.pop(key, None)


def _register_edit_popup(owner, key, popup):
    """ثبت پنجره ویرایش تا کلیک‌های بعدی همان پنجره را باز کنند نه پنجره جدید."""
    if not hasattr(owner, "_open_edit_popups") or owner._open_edit_popups is None:
        owner._open_edit_popups = {}
    owner._open_edit_popups[key] = popup

    def _clear(_event=None, k=key, o=owner, p=popup):
        store = getattr(o, "_open_edit_popups", None)
        if store is not None and store.get(k) is p:
            store.pop(k, None)

    popup.bind("<Destroy>", _clear, add="+")


def finalize_popup_window(popup, parent=None):
    """Call after popup UI is built — keeps edit/dialog windows on top."""
    root = _root_window(parent or popup)
    if root is not None and not getattr(popup, "_popup_prepared", False):
        prepare_popup_window(popup, parent)
    _bind_popup_kiosk_keys(popup, parent)
    _bring_popup_to_front(popup)

    def _refocus(focus=False):
        try:
            if popup.winfo_exists():
                _bring_popup_to_front(popup, focus=focus)
        except Exception:
            pass

    def _on_map(_event=None):
        _refocus(focus=False)

    popup.bind("<Map>", _on_map, add="+")
    popup.bind("<FocusIn>", _on_map, add="+")
    for delay in (20, 80, 200, 500, 1000):
        popup.after(delay, lambda f=False: _refocus(f))


def _start_popup_zorder_guard(root):
    if getattr(root, "_popup_zorder_guard", False):
        return
    root._popup_zorder_guard = True

    def _tick():
        try:
            if not root.winfo_exists():
                root._popup_zorder_guard = False
                return
            # هنگام دیالوگ رمز، SetForeground را تکرار نکن — تایپ قطع می‌شود
            if _modal_dialog_active(root):
                root.after(800, _tick)
                return
            top = _get_top_popup(root)
            if top is None:
                root._popup_zorder_guard = False
                return
            try:
                if _hwnd_is_iconic(root):
                    root.after(2000, _tick)
                    return
            except Exception:
                pass
            # فقط z-order (topmost/lift) بدون دزدیدن فوکوس کیبورد
            try:
                top.lift()
                top.attributes("-topmost", True)
            except Exception:
                pass
            root.after(600, _tick)
        except Exception:
            root._popup_zorder_guard = False

    root.after(120, _tick)


def _react_to_blocked_system_key(win):
    """Win/Alt+Tab — تسک‌بار مخفی + برگرداندن فوکوس به برنامه."""
    try:
        if getattr(win, "_app_closing", False):
            return
        hide_windows_shell()
        if _is_kiosk_station() and _kiosk_key_block_active(win):
            top = _get_top_popup(win)
            if top is not None:
                _bring_popup_to_front(top)
            else:
                _raise_window_front(win)
    except Exception:
        pass


def _get_app_root_hwnd(win):
    if os.name != "nt":
        return None
    try:
        return _get_toplevel_hwnd(win)
    except Exception:
        return None


def _hwnd_belongs_to_app(hwnd, app_root):
    """آیا hwnd متعلق به همین برنامه (یا دیالوگ آن) است؟"""
    if not hwnd or not app_root:
        return False
    try:
        import ctypes
        user32 = ctypes.windll.user32
        GA_ROOT = 2
        root = user32.GetAncestor(hwnd, GA_ROOT)
        if root == app_root:
            return True
        pid_app = ctypes.c_ulong()
        pid_fg = ctypes.c_ulong()
        user32.GetWindowThreadProcessId(app_root, ctypes.byref(pid_app))
        user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid_fg))
        return pid_app.value == pid_fg.value and pid_app.value != 0
    except Exception:
        return False


def _suppress_foreign_windows(win):
    """بستن Task Manager و پنجره‌های shell که روی کیوسک ظاهر می‌شوند."""
    if os.name != "nt" or not _is_kiosk_station():
        return
    # حتی وقتی دیالوگ رمز باز است Task Manager را ببند
    if getattr(win, "_app_closing", False):
        return
    if getattr(win, "_minimizing_authorized", False):
        return
    try:
        import ctypes
        from ctypes import wintypes
        user32 = ctypes.windll.user32
        WM_CLOSE = 0x0010
        SW_HIDE = 0
        app_root = _get_app_root_hwnd(win)
        block_classes = {
            "TaskManagerWindow",
            "MultitaskingViewFrame",
            "XamlExplorerDockHost",
            "TopLevelWindowForOverflowXamlIsland",
        }

        def _each(hwnd, _):
            try:
                if not user32.IsWindowVisible(hwnd):
                    return True
                if _hwnd_belongs_to_app(hwnd, app_root):
                    return True
                buf = ctypes.create_unicode_buffer(256)
                user32.GetClassNameW(hwnd, buf, 256)
                cls = buf.value
                if cls in block_classes:
                    user32.PostMessageW(hwnd, WM_CLOSE, 0, 0)
                    user32.ShowWindow(hwnd, SW_HIDE)
                elif cls in _SHELL_CLASSES:
                    user32.ShowWindow(hwnd, SW_HIDE)
            except Exception:
                pass
            return True

        WNDENUMPROC = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
        user32.EnumWindows(WNDENUMPROC(_each), 0)
    except Exception:
        pass


def _start_foreground_guard(win):
    """برنامه همیشه جلو بماند — کاربر فقط با همین برنامه کار کند."""
    if not _is_kiosk_station() or getattr(win, "_fg_guard_started", False):
        return
    win._fg_guard_started = True

    def _tick():
        try:
            if not win.winfo_exists() or getattr(win, "_app_closing", False):
                return
            # مینیمایز/کوچک: هیچ EnumWindows / hide تسک‌بار — جلوگیری از هنگ
            if not _ui_heavy_ok(win):
                win.after(_ui_idle_ms(win, 2000, 5000, 12000), _tick)
                return
            if _kiosk_key_block_active(win):
                if _modal_dialog_active(win):
                    pass
                else:
                    top = _get_top_popup(win)
                    if top is not None:
                        try:
                            top.lift()
                            top.attributes("-topmost", True)
                        except Exception:
                            pass
                    elif os.name == "nt":
                        import ctypes
                        fg = ctypes.windll.user32.GetForegroundWindow()
                        app_root = _get_app_root_hwnd(win)
                        if fg and app_root and not _hwnd_belongs_to_app(fg, app_root):
                            _raise_window_front(win)
        except Exception:
            pass
        try:
            win.after(_ui_idle_ms(win, 2000, 5000, 12000), _tick)
        except Exception:
            pass

    win.after(800, _tick)


def _reinforce_kiosk_security(win):
    """بازاعمال فول‌اسکرین، تسک‌بار مخفی، و hook کیبورد."""
    if os.environ.get("STF_ADMIN") == "1":
        return
    try:
        if not win.winfo_exists():
            return
        if _modal_dialog_active(win):
            return
        hide_all_taskbars(win)
        force_full_screen(win)
        lock_window_chrome(win)
        _suppress_foreign_windows(win)
        if not getattr(win, "_kb_hook_id", None):
            start_keyboard_lock(win)
    except Exception:
        pass


def _start_kiosk_reinforce_loop(win):
    """هر چند ثانیه قفل کیوسک را بازاعمال می‌کند."""
    if os.environ.get("STF_ADMIN") == "1":
        return
    if getattr(win, "_kiosk_reinforce_started", False):
        return
    win._kiosk_reinforce_started = True

    def _tick():
        try:
            if not win.winfo_exists() or getattr(win, "_app_closing", False):
                return
            if _ui_heavy_ok(win) and not _modal_dialog_active(win):
                _reinforce_kiosk_security(win)
        except Exception:
            pass
        try:
            win.after(_ui_idle_ms(win, 6000, 15000, 30000), _tick)
        except Exception:
            pass

    win.after(800, lambda w=win: _reinforce_kiosk_security(w) if _ui_heavy_ok(w) else None)
    win.after(6000, _tick)


def start_taskbar_hide_loop(win):
    """تسک‌بار مخفی — بدون دست زدن به فوکوس یا دیالوگ‌ها."""
    if getattr(win, "_taskbar_loop_started", False):
        return
    win._taskbar_loop_started = True

    def _loop():
        try:
            if not win.winfo_exists() or getattr(win, "_app_closing", False):
                return
            # هنگام مینیمایز/کوچک، تسک‌بار را نشان بده و دست نزن — عامل هنگ شدید سیستم
            if not _ui_heavy_ok(win):
                try:
                    show_windows_shell(force_broadcast=False)
                except Exception:
                    pass
                win.after(_ui_idle_ms(win, 2000, 5000, 15000), _loop)
                return
            hide_windows_shell()
            win.after(_ui_idle_ms(win, 2000, 5000, 15000), _loop)
        except Exception:
            pass

    win.after(800, _loop)


def is_lock_password_required(db=None):
    """آیا برای بستن/مینیمایز/کوچک‌کردن رمز قفل لازم است؟"""
    if db is None:
        db = load_db()
    settings = db.get("settings", {})
    if settings.get("minimize_password_disabled"):
        return False
    pw = settings.get("minimize_password")
    if pw is None:
        return True
    return bool(str(pw).strip())


def get_lock_password(db=None):
    if db is None:
        db = load_db()
    pw = db.get("settings", {}).get("minimize_password", "1234")
    if pw is None or not str(pw).strip():
        return "1234"
    return str(pw).strip()


def is_backup_vault_password_required(db=None):
    """آیا بک‌آپ/آرشیو باید با رمز قفل شود؟"""
    if db is None:
        db = load_db()
    settings = db.get("settings", {}) or {}
    if settings.get("backup_vault_password_disabled"):
        return False
    return True


def get_backup_vault_password(db=None):
    """رمز فعلی vault بک‌آپ — پیش‌فرض Reza9063."""
    if db is None:
        db = load_db()
    settings = db.get("settings", {}) or {}
    if settings.get("backup_vault_password_disabled"):
        return ""
    pw = settings.get("backup_vault_password")
    if pw is None or not str(pw).strip():
        return "Reza9063"
    return str(pw).strip()


def _clear_lock_bypass_flags(win):
    """هر بار قبل/بعد از عملیات — پرچم‌های عبور موقت رمز را صفر می‌کند."""
    win._allow_iconify = False
    win._allow_restore = False
    win._lock_minimize_ok = False
    win._lock_restore_ok = False


def _maximize_without_password(win):
    """کوچک → بزرگ (تمام‌صفحه) — بدون رمز."""
    win._was_full = True
    win._drag_locked = True
    win._allow_restore = False
    hide_all_taskbars(win)
    force_full_screen(win)
    update_window_drag_state(win)
    try:
        win.update_idletasks()
        win.update()
    except Exception:
        pass
    _raise_window_front(win)


def _schedule_lock_action(win, cb):
    """اجرای امن callback قفل — بدون انباشت."""
    if cb is None:
        return
    if getattr(win, "_app_closing", False):
        return
    try:
        win.after(0, cb)
    except Exception:
        pass


def do_window_restore_down(win):
    """خروج از تمام‌صفحه — فقط کوچک کردن پنجره، بدون بستن برنامه."""
    if getattr(win, "_window_action_in_progress", False):
        return
    win._window_action_in_progress = True
    _clear_lock_bypass_flags(win)
    try:
        win._allow_restore = True
        win._was_full = False
        win._drag_locked = False
        show_all_taskbars()
        win.update_idletasks()
        try:
            if win.state() in ("zoomed", "iconic"):
                win.state("normal")
                win.update_idletasks()
        except Exception:
            pass
        win.geometry("1320x840")
        win.update_idletasks()
        try:
            win.eval('tk::PlaceWindow . center')
        except Exception:
            pass
        win._drag_locked = False
        win._was_full = False
        update_window_drag_state(win)
        show_all_taskbars()
        _raise_window_front(win)
    finally:
        win._allow_restore = False
        _clear_lock_bypass_flags(win)
        try:
            win.after(400, lambda w=win: setattr(w, "_window_action_in_progress", False))
        except Exception:
            win._window_action_in_progress = False


def _hwnd_is_iconic(win):
    """آیا پنجره واقعاً در تسک‌بار مینیمایز شده؟ (سطح ویندوز)"""
    if os.name == "nt":
        try:
            import ctypes
            return bool(ctypes.windll.user32.IsIconic(_get_toplevel_hwnd(win)))
        except Exception:
            pass
    try:
        return win.state() == "iconic"
    except Exception:
        return False


def _window_is_minimized(win):
    """آیا پنجره در تسک‌بار مینیمایز شده؟"""
    if getattr(win, "_minimizing_authorized", False):
        return True
    if _hwnd_is_iconic(win):
        return True
    try:
        return win.state() == "iconic"
    except Exception:
        return False


def _suspend_wndproc_hook(win):
    """hook را موقت برمی‌دارد تا مینیمایز ویندوز درست به تسک‌بار برود."""
    if os.name != "nt" or not getattr(win, "_drag_lock_installed", False):
        return False
    if getattr(win, "_wndproc_suspended", False):
        return True
    try:
        import ctypes
        user32 = ctypes.windll.user32
        hwnd = _get_toplevel_hwnd(win)
        GWLP_WNDPROC = -4
        orig = getattr(win, "_drag_lock_orig_proc", None)
        if not orig:
            return False
        set_wndproc = getattr(user32, "SetWindowLongPtrW", user32.SetWindowLongW)
        set_wndproc(hwnd, GWLP_WNDPROC, orig)
        win._wndproc_suspended = True
        return True
    except Exception:
        return False


def _resume_wndproc_hook(win):
    """hook را بعد از بازگشت از تسک‌بار برمی‌گرداند."""
    if os.name != "nt" or not getattr(win, "_wndproc_suspended", False):
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        hwnd = _get_toplevel_hwnd(win)
        GWLP_WNDPROC = -4
        proc = getattr(win, "_drag_lock_proc_ref", None)
        if proc:
            set_wndproc = getattr(user32, "SetWindowLongPtrW", user32.SetWindowLongW)
            set_wndproc(hwnd, GWLP_WNDPROC, proc)
        win._wndproc_suspended = False
    except Exception:
        win._wndproc_suspended = False


def _restore_hwnd_styles(win):
    """بازگرداندن استایل ذخیره‌شده قبل از مینیمایز."""
    if not getattr(win, "_pre_minimize_style_saved", False):
        return
    if os.name != "nt":
        win._pre_minimize_style_saved = False
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        hwnd = _get_toplevel_hwnd(win)
        GWL_STYLE, GWL_EXSTYLE = -16, -20
        get_long = getattr(user32, "GetWindowLongPtrW", user32.GetWindowLongW)
        set_long = getattr(user32, "SetWindowLongPtrW", user32.SetWindowLongW)
        set_long(hwnd, GWL_STYLE, win._saved_minimize_style)
        set_long(hwnd, GWL_EXSTYLE, win._saved_minimize_exstyle)
        user32.SetWindowPos(hwnd, 0, 0, 0, 0, 0, 0x2 | 0x1 | 0x4 | 0x20)
    except Exception:
        pass
    win._pre_minimize_style_saved = False


def _abort_failed_minimize(win):
    """مینیمایز نشد — وضعیت را برگردان (نه بستن برنامه)."""
    win._minimizing_authorized = False
    win._was_in_taskbar = False
    win._allow_iconify = False
    if os.name == "nt":
        try:
            import ctypes
            hwnd = _get_toplevel_hwnd(win)
            ctypes.windll.user32.ShowWindow(hwnd, 9)
        except Exception:
            pass
    try:
        win.wm_state("normal")
    except Exception:
        pass
    if getattr(win, "_frameless_done", False) and _app_shell_lock_active(win):
        try:
            lock_window_chrome(win)
        except Exception:
            pass


def _verify_minimize_state(win, attempt=0):
    """بررسی رفتن به تسک‌بار — بدون update که باعث بستن می‌شود."""
    try:
        if not win.winfo_exists():
            return
        if not getattr(win, "_minimizing_authorized", False):
            return
        if _hwnd_is_iconic(win):
            win._was_in_taskbar = True
            _sync_tk_iconic_state(win)
            _start_taskbar_keepalive(win)
            return
        if attempt < 10:
            win.after(80 + attempt * 40, lambda w=win, n=attempt + 1: _verify_minimize_state(w, n))
            return
        _abort_failed_minimize(win)
        show_all_taskbars()
        _raise_window_front(win)
        start_keyboard_lock(win)
    except Exception:
        pass


def do_window_minimize_safe(win):
    """مینیمایز — فقط به تسک‌بار، هرگز بستن برنامه."""
    if getattr(win, "_window_action_in_progress", False):
        return
    win._window_action_in_progress = True
    win._allow_iconify = True
    win._minimizing_authorized = True
    win._was_in_taskbar = False
    try:
        show_all_taskbars()
        stop_keyboard_lock(win)
        win.update_idletasks()
        win_minimize(win)
        win.after(60, lambda w=win: _verify_minimize_state(w, 0))
    finally:
        def _done():
            try:
                if getattr(win, "_minimizing_authorized", False):
                    if _hwnd_is_iconic(win):
                        return
            except Exception:
                if getattr(win, "_minimizing_authorized", False):
                    return
            win._window_action_in_progress = False
        try:
            win.after(2000, _done)
        except Exception:
            _done()


def _minimize_failure_watchdog(win):
    """سازگاری — همان verify."""
    _verify_minimize_state(win, 8)


def _finish_authorized_minimize(win):
    """پایان مینیمایز مجاز — frameless دوباره (فقط بعد از restore از تسک‌بار)."""
    if _hwnd_is_iconic(win):
        return
    try:
        if win.state() == "iconic":
            return
    except Exception:
        pass
    if not getattr(win, "_minimizing_authorized", False):
        return
    win._minimizing_authorized = False
    win._was_in_taskbar = False
    win._allow_iconify = False
    win._lock_minimize_ok = False
    win._window_action_in_progress = False
    try:
        win.wm_state("normal")
    except Exception:
        pass
    if getattr(win, "_frameless_done", False) and _app_shell_lock_active(win):
        try:
            lock_window_chrome(win)
            force_full_screen(win)
        except Exception:
            pass
    if _app_shell_lock_active(win):
        start_keyboard_lock(win)
    # بعد از restore از تسک‌بار — رفرش فوری شیشه (مثل Configure)، نه تأخیر چندمرحله‌ای
    stabilize = getattr(win, "_stabilize_login_after_geometry_change", None)
    if callable(stabilize):
        try:
            stabilize()
        except Exception:
            pass
        try:
            win.after_idle(stabilize)
        except Exception:
            pass


def _mark_minimize_in_taskbar(win):
    """تأیید رفتن واقعی به تسک‌بار — قبل از آن frameless برنگردد."""
    if not getattr(win, "_minimizing_authorized", False):
        return
    if _hwnd_is_iconic(win):
        win._was_in_taskbar = True
        return
    try:
        if win.state() == "iconic":
            win._was_in_taskbar = True
            return
    except Exception:
        pass
    try:
        win.after(120, lambda w=win: _mark_minimize_in_taskbar(w))
    except Exception:
        pass


def _try_finish_authorized_minimize_on_map(win):
    """فقط restore واقعی از تسک‌بار — نه Map وسط مینیمایز."""
    if not getattr(win, "_minimizing_authorized", False):
        return
    if _hwnd_is_iconic(win):
        return
    try:
        if win.state() == "iconic":
            return
    except Exception:
        pass
    if not getattr(win, "_was_in_taskbar", False):
        return
    _finish_authorized_minimize(win)


def _start_taskbar_keepalive(win):
    """تا وقتی در تسک‌بار است — تسک‌بار مخفی نشود (کم‌فرکانس تا هنگ نشود)."""
    def _loop():
        try:
            if not win.winfo_exists():
                return
            if not getattr(win, "_minimizing_authorized", False) and not _hwnd_is_iconic(win):
                return
            show_all_taskbars()
            win.after(2500, _loop)
        except Exception:
            pass
    try:
        win.after(80, _loop)
    except Exception:
        pass


def start_keyboard_lock(win):
    """کلید ویندوز، استارت، Alt+Tab و میانبرهای خروج را مسدود می‌کند."""
    if os.name != "nt" or getattr(win, "_kb_hook_handler", None):
        return
    try:
        import ctypes
        from ctypes import wintypes

        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32

        WH_KEYBOARD_LL = 13
        WM_KEYDOWN, WM_SYSKEYDOWN = 0x0100, 0x0104
        WM_KEYUP, WM_SYSKEYUP = 0x0101, 0x0105
        VK_LWIN, VK_RWIN, VK_APPS, VK_TAB, VK_ESCAPE = 0x5B, 0x5C, 0x5D, 0x09, 0x1B
        VK_CONTROL, VK_MENU, VK_SHIFT, VK_F4 = 0x11, 0x12, 0x10, 0x73
        VK_F11, VK_F12, VK_SNAPSHOT, VK_SPACE = 0x7A, 0x7B, 0x2C, 0x20
        VK_F1, VK_PAUSE, VK_DELETE = 0x70, 0x13, 0x2E

        class KBDLLHOOKSTRUCT(ctypes.Structure):
            _fields_ = [
                ("vkCode", wintypes.DWORD),
                ("scanCode", wintypes.DWORD),
                ("flags", wintypes.DWORD),
                ("time", wintypes.DWORD),
                ("dwExtraInfo", ctypes.c_ulonglong),
            ]

        LRESULT = ctypes.c_longlong if ctypes.sizeof(ctypes.c_void_p) == 8 else ctypes.c_long
        HOOKPROC = ctypes.WINFUNCTYPE(LRESULT, ctypes.c_int, wintypes.WPARAM, wintypes.LPARAM)
        hook_ref = [None]

        def _win_down():
            return (user32.GetAsyncKeyState(VK_LWIN) & 0x8000) or (
                user32.GetAsyncKeyState(VK_RWIN) & 0x8000)

        def _alt_down():
            return user32.GetAsyncKeyState(VK_MENU) & 0x8000

        def _ctrl_down():
            return user32.GetAsyncKeyState(VK_CONTROL) & 0x8000

        def _shift_down():
            return user32.GetAsyncKeyState(VK_SHIFT) & 0x8000

        def _blocked(vk):
            # همیشه Task Manager shortcut را ببند (حتی در دیالوگ رمز)
            if vk == VK_ESCAPE and _ctrl_down() and _shift_down():
                return True
            # تلاش برای مسدود کردن ترکیب‌های نزدیک به CAD (خود CAD از user-mode کامل بسته نمی‌شود)
            if vk == VK_DELETE and _ctrl_down() and _alt_down():
                return True
            if not _kiosk_key_block_active(win):
                return False
            modal = _modal_dialog_active(win)
            # هرگز خود Ctrl / Shift را مسدود نکن — تعویض زبان (Ctrl+Shift)
            if vk in (VK_CONTROL, VK_SHIFT):
                return False
            if vk in (VK_LWIN, VK_RWIN, VK_APPS):
                return True
            if _win_down():
                return True
            if vk == VK_TAB and (_alt_down() or _ctrl_down()):
                return True
            if vk == VK_F4 and _alt_down():
                return True
            if vk == VK_ESCAPE and _alt_down():
                return True
            if vk == VK_ESCAPE and _ctrl_down():
                return True
            if vk in (VK_F11, VK_F12, VK_SNAPSHOT, VK_F1, VK_PAUSE):
                return True
            if vk == VK_SPACE and _alt_down():
                return True
            if modal:
                return False
            return False

        def _handler(nCode, wParam, lParam):
            if nCode >= 0 and wParam in (
                    WM_KEYDOWN, WM_SYSKEYDOWN, WM_KEYUP, WM_SYSKEYUP):
                kb = ctypes.cast(lParam, ctypes.POINTER(KBDLLHOOKSTRUCT)).contents
                if _blocked(kb.vkCode):
                    if wParam in (WM_KEYDOWN, WM_SYSKEYDOWN):
                        _react_to_blocked_system_key(win)
                    return 1
            return user32.CallNextHookEx(hook_ref[0], nCode, wParam, lParam)

        proc = HOOKPROC(_handler)
        hook_id = user32.SetWindowsHookExW(
            WH_KEYBOARD_LL, proc, kernel32.GetModuleHandleW(None), 0)
        if hook_id:
            hook_ref[0] = hook_id
            win._kb_hook_handler = proc
            win._kb_hook_id = hook_id

            def _unhook(event=None):
                hid = getattr(win, "_kb_hook_id", None)
                if hid:
                    try:
                        user32.UnhookWindowsHookEx(hid)
                    except Exception:
                        pass
                    win._kb_hook_id = None
            win.bind("<Destroy>", _unhook, add="+")
    except Exception:
        pass


def stop_keyboard_lock(win):
    hid = getattr(win, "_kb_hook_id", None)
    if hid and os.name == "nt":
        try:
            import ctypes
            ctypes.windll.user32.UnhookWindowsHookEx(hid)
        except Exception:
            pass
        win._kb_hook_id = None


# ═══════════════════════════════════════════════════════════
#  تمام‌صفحه‌ی واقعی — بدون فاصله خالی در پایین صفحه
#  state("zoomed") فقط پنجره را تا لبه‌ی «work area» ویندوز بزرگ می‌کند؛
#  این work area حتی وقتی تسک‌بار مخفی می‌شود بلافاصله به‌روزرسانی
#  نمی‌شود و همان فضای خالیِ نوار تسک‌بار در پایین صفحه باقی می‌ماند.
#  برای رفع این باگ، به‌جای state("zoomed")، مستقیماً اندازه‌ی پنجره را
#  برابر با کل عرض/ارتفاع صفحه‌نمایش قرار می‌دهیم.
# ═══════════════════════════════════════════════════════════
FULLSCREEN_WIDTH_MARGIN = 0
FULLSCREEN_EDGE_BLEED = 0
FULLSCREEN_TOP_BLEED = 0


def _fit_window_rect(win):
    """مستطیل پنجره — دقیقاً داخل مانیتور (بدون بیرون‌زدگی لبه)."""
    mx, my, mw, mh = _get_monitor_rect(win)
    return mx, my, mw, mh


def _get_toplevel_hwnd(win):
    """هندل واقعی پنجره روی ویندوز (نه فقط فریم داخلی Tk)."""
    if os.name != "nt":
        return win.winfo_id()
    try:
        import ctypes
        user32 = ctypes.windll.user32
        hwnd = win.winfo_id()
        GA_ROOT = 2
        root = user32.GetAncestor(hwnd, GA_ROOT)
        if root:
            return root
        parent = user32.GetParent(hwnd)
        return parent if parent else hwnd
    except Exception:
        return win.winfo_id()


def _get_monitor_rect(win):
    """مستطیل کامل مانیتور (rcMonitor) — نه work area که فضای تسک‌بار را کم می‌کند."""
    try:
        win.update_idletasks()
    except Exception:
        pass
    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes
            user32 = ctypes.windll.user32
            hwnd = _get_toplevel_hwnd(win)

            class MONITORINFO(ctypes.Structure):
                _fields_ = [
                    ("cbSize", wintypes.DWORD),
                    ("rcMonitor", wintypes.RECT),
                    ("rcWork", wintypes.RECT),
                    ("dwFlags", wintypes.DWORD),
                ]

            MONITOR_DEFAULTTONEAREST = 2
            hmon = user32.MonitorFromWindow(hwnd, MONITOR_DEFAULTTONEAREST)
            mi = MONITORINFO()
            mi.cbSize = ctypes.sizeof(MONITORINFO)
            if user32.GetMonitorInfoW(hmon, ctypes.byref(mi)):
                r = mi.rcMonitor
                return r.left, r.top, r.right - r.left, r.bottom - r.top
        except Exception:
            pass
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    return 0, 0, sw, sh


def force_full_screen(win):
    """پنجره را داخل مانیتور فیت می‌کند — حاشیه جزئی DWM بدون بیرون‌زدن دکمه‌ها."""
    try:
        root = win if isinstance(win, tk.Tk) else _root_window(win)
        if root is not None and (_modal_dialog_active(root) or _get_top_popup(root)):
            return
    except Exception:
        pass
    if getattr(win, "_fullscreen_guard", False):
        return
    if getattr(win, "_allow_restore", False):
        return
    if getattr(win, "_was_full", True) is False:
        return
    if _window_is_minimized(win):
        return
    win._fullscreen_guard = True
    try:
        win.update_idletasks()
        try:
            st = win.state()
            if st in ("zoomed", "iconic"):
                win.state("normal")
                win.update_idletasks()
        except Exception:
            pass

        ox, oy, ow, oh = _fit_window_rect(win)

        if os.name == "nt":
            try:
                import ctypes
                hwnd = _get_toplevel_hwnd(win)
                SWP_NOZORDER, SWP_NOACTIVATE, SWP_SHOWWINDOW = 0x4, 0x10, 0x40
                ctypes.windll.user32.SetWindowPos(
                    hwnd, 0, ox, oy, ow, oh,
                    SWP_NOZORDER | SWP_NOACTIVATE | SWP_SHOWWINDOW)
                win.update_idletasks()
                return
            except Exception:
                pass
        win.geometry(f"{ow}x{oh}+{ox}+{oy}")
        win.update_idletasks()
    except Exception:
        try:
            ox, oy, ow, oh = _fit_window_rect(win)
            win.geometry(f"{ow}x{oh}+{ox}+{oy}")
        except Exception:
            pass
    finally:
        try:
            win.after(250, lambda w=win: setattr(w, "_fullscreen_guard", False))
        except Exception:
            win._fullscreen_guard = False


def _begin_window_transition(win):
    """لایهٔ تمام‌صفحه هنگام تعویض محتوا — دسکتاپ لحظه‌ای دیده نشود."""
    win._window_action_in_progress = True
    win._app_closing = False
    win._was_full = True
    win._allow_restore = False
    win._fullscreen_guard = False
    hide_all_taskbars(win)
    try:
        win.configure(bg=C["bg"])
    except Exception:
        pass
    force_full_screen(win)
    cover = tk.Frame(win, bg=C["bg"], bd=0, highlightthickness=0)
    cover.place(x=0, y=0, relwidth=1, relheight=1)
    try:
        cover.lift()
    except Exception:
        pass
    win.update_idletasks()
    win.update()
    if os.name == "nt":
        try:
            import ctypes
            hwnd = _get_toplevel_hwnd(win)
            ctypes.windll.user32.SetForegroundWindow(hwnd)
        except Exception:
            pass
    win._transition_cover = cover
    return cover


def _end_window_transition(win):
    cover = getattr(win, "_transition_cover", None)
    if cover is not None:
        try:
            if cover.winfo_exists():
                cover.destroy()
        except Exception:
            pass
        win._transition_cover = None
    win._window_action_in_progress = False
    force_full_screen(win)
    win.update_idletasks()
    win.update()


def enforce_full_screen_robust(win):
    """نسخه‌ی مطمئن‌تر از force_full_screen برای پنجره‌هایی مثل صفحه‌ی لاگین."""
    def _apply(_event=None):
        try:
            root = win if isinstance(win, tk.Tk) else _root_window(win)
            if root is not None and (_modal_dialog_active(root) or _get_top_popup(root)):
                return
            if _window_is_minimized(win):
                return
            force_full_screen(win)
        except Exception:
            pass

    # اجرای فوری + روی اولین Map واقعیِ پنجره
    _apply()
    try:
        win.bind("<Map>", _apply, add="+")
    except Exception:
        pass
    # چند تلاشِ اضافه با فاصله‌ی کوتاه، برای پوشش دادنِ سیستم‌های کندتر
    # یا مانیتورهایی با DPI Scale غیرِ ۱۰۰٪
    for delay in (0, 150, 500, 1000):
        try:
            win.after(delay, _apply)
        except Exception:
            pass


def _finalize_fullscreen_window(win):
    """قفل لبه + حذف نوار ویندوز + hook + فیت تمام‌صفحه."""
    try:
        lock_window_chrome(win)
        lock_window_drag(win, force=True)
        force_full_screen(win)
        lock_window_chrome(win)
    except Exception:
        pass


def enforce_frameless_window(win):
    """نوار عنوان ویندوز — یک‌بار اعمال می‌شود (بدون رفرش پشت‌سرهم)."""
    if getattr(win, "_frameless_done", False):
        return
    win._frameless_done = True
    try:
        lock_window_chrome(win)
    except Exception:
        pass
    try:
        win.after(300, lambda w=win: lock_window_chrome(w)
                              if w.winfo_exists() and not _window_is_minimized(w) else None)
    except Exception:
        pass


def bind_window_lock_keys(win):
    """Alt+F4 و سایر میانبرهای خروج — فقط با رمز (کلاینت: همه مسیرها)."""
    def _block_close(event=None):
        cb = getattr(win, "_on_close_attempt", None)
        if cb:
            cb()
        return "break"

    def _block_any(event=None):
        _react_to_blocked_system_key(win)
        return "break"

    try:
        win.bind("<Alt-F4>", _block_close, add="+")
    except Exception:
        pass
    if not _is_kiosk_station():
        return
    for seq in (
        "<F11>", "<F12>", "<F1>",
        "<Control-Escape>",
        "<Control-Shift-Escape>",
        "<Alt-Tab>", "<Alt-Escape>", "<Alt-Shift-Tab>",
        "<Meta-Key>", "<Meta-L>", "<Meta-Tab>",
        "<Control-Shift-Tab>",
    ):
        try:
            win.bind_all(seq, _block_any, add="+")
        except Exception:
            pass


def is_window_full_screen(win) -> bool:
    """آیا پنجره تقریباً کل مانیتور را پوشانده (بدون بیرون‌زدن دکمه‌ها)؟"""
    try:
        mx, my, mw, mh = _get_monitor_rect(win)
        target_w = mw - FULLSCREEN_WIDTH_MARGIN + FULLSCREEN_EDGE_BLEED
        max_right = mx + mw + 1
        max_bottom = my + mh + FULLSCREEN_EDGE_BLEED + 1
        if os.name == "nt":
            try:
                import ctypes
                from ctypes import wintypes
                hwnd = _get_toplevel_hwnd(win)
                rect = wintypes.RECT()
                if ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect)):
                    if rect.left < mx - FULLSCREEN_EDGE_BLEED - 1:
                        return False
                    if rect.top < my - FULLSCREEN_TOP_BLEED - 1:
                        return False
                    if rect.right > max_right or rect.bottom > max_bottom:
                        return False
                    w = rect.right - rect.left
                    h = rect.bottom - rect.top
                    return w >= target_w - 4 and h >= mh - 4
            except Exception:
                pass
        w, h = win.winfo_width(), win.winfo_height()
        return w >= (target_w - 8) and h >= mh - 8
    except Exception:
        return False


# ═══════════════════════════════════════════════════════════
#  حذف نوار عنوان ویندوز — دکمه‌های ✕/▢/– فقط داخل برنامه
#  (با رمز). دکمه‌های پیش‌فرض ویندوز غیرفعال می‌شوند.
# ═══════════════════════════════════════════════════════════
def lock_window_chrome(win):
    """حذف نوار عنوان ویندوز — بدون overrideredirect تا مینیمایز در تسک‌بار بماند."""
    if _window_is_minimized(win):
        return
    try:
        win.update_idletasks()
        win.overrideredirect(False)
    except Exception:
        pass
    if os.name != "nt":
        return
    try:
        import ctypes
        win.update_idletasks()
        hwnd = _get_toplevel_hwnd(win)
        GWL_STYLE = -16
        WS_CAPTION     = 0x00C00000
        WS_THICKFRAME  = 0x00040000
        WS_MAXIMIZEBOX = 0x00010000
        WS_MINIMIZEBOX = 0x00020000
        WS_SYSMENU     = 0x00080000
        get_style = getattr(ctypes.windll.user32, "GetWindowLongPtrW",
                             ctypes.windll.user32.GetWindowLongW)
        set_style = getattr(ctypes.windll.user32, "SetWindowLongPtrW",
                             ctypes.windll.user32.SetWindowLongW)
        style = get_style(hwnd, GWL_STYLE)
        style &= ~(WS_CAPTION | WS_THICKFRAME | WS_MAXIMIZEBOX | WS_MINIMIZEBOX | WS_SYSMENU)
        set_style(hwnd, GWL_STYLE, style)
        SWP_NOMOVE, SWP_NOSIZE, SWP_NOZORDER, SWP_FRAMECHANGED = 0x2, 0x1, 0x4, 0x20
        ctypes.windll.user32.SetWindowPos(
            hwnd, 0, 0, 0, 0, 0,
            SWP_NOMOVE | SWP_NOSIZE | SWP_NOZORDER | SWP_FRAMECHANGED)
        _ensure_taskbar_button(win)
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════
#  قفل جابه‌جاییِ پنجره با موس (Windows) — با «زیرکلاس‌بندی» (subclassing)
#  پنجره در سطح ویندوز، پیام‌های مربوط به گرفتن-و-کشیدنِ نوار عنوان
#  (WM_NCLBUTTONDOWN روی ناحیه‌ی caption) و دستور جابه‌جایی از منوی
#  سیستم (WM_SYSCOMMAND / SC_MOVE، مثلاً از طریق Alt+Space) خنثی
#  می‌شوند — بدون هیچ تأثیری روی خودِ سه دکمه‌ی ✕/▢/‗ که هر کدام
#  hit-test جداگانه‌ی خودشان را دارند و این‌جا دست‌نخورده می‌مانند.
# ═══════════════════════════════════════════════════════════
def lock_window_drag(win, force=False):
    """قفل جابه‌جایی + رمزگذاری دکمه‌های ✕/▢/– نوار عنوان ویندوز."""
    if os.name != "nt":
        return
    if getattr(win, "_drag_lock_installed", False) and not force:
        return
    try:
        import ctypes
        from ctypes import wintypes
        win.update_idletasks()
        hwnd = _get_toplevel_hwnd(win)

        GWLP_WNDPROC = -4
        WM_CLOSE           = 0x0010
        WM_DESTROY         = 0x0002
        WM_SIZE            = 0x0005
        SIZE_MINIMIZED     = 1
        WM_NCLBUTTONDOWN   = 0x00A1
        WM_NCLBUTTONDBLCLK = 0x00A3
        WM_SYSCOMMAND      = 0x0112
        HTCAPTION   = 2
        HTMINBUTTON = 8
        HTMAXBUTTON = 9
        HTREDUCE    = 8
        HTZOOM      = 9
        HTCLOSE     = 20
        SC_MOVE     = 0xF010
        SC_MINIMIZE = 0xF020
        SC_MAXIMIZE = 0xF030
        SC_RESTORE  = 0xF120
        SC_CLOSE    = 0xF060
        SC_MASK     = 0xFFF0

        user32 = ctypes.windll.user32
        get_wndproc = getattr(user32, "GetWindowLongPtrW", user32.GetWindowLongW)
        set_wndproc = getattr(user32, "SetWindowLongPtrW", user32.SetWindowLongW)

        is64 = ctypes.sizeof(ctypes.c_void_p) == 8
        LRESULT = ctypes.c_longlong if is64 else ctypes.c_long
        get_wndproc.restype  = ctypes.c_void_p
        get_wndproc.argtypes = [wintypes.HWND, ctypes.c_int]
        set_wndproc.restype  = ctypes.c_void_p
        set_wndproc.argtypes = [wintypes.HWND, ctypes.c_int, ctypes.c_void_p]
        user32.CallWindowProcW.restype  = LRESULT
        user32.CallWindowProcW.argtypes = [
            ctypes.c_void_p, wintypes.HWND, ctypes.c_uint,
            wintypes.WPARAM, wintypes.LPARAM]

        WNDPROCTYPE = ctypes.WINFUNCTYPE(
            LRESULT, wintypes.HWND, ctypes.c_uint, wintypes.WPARAM, wintypes.LPARAM)

        if not getattr(win, "_drag_lock_orig_proc", None):
            win._drag_lock_orig_proc = get_wndproc(hwnd, GWLP_WNDPROC)
        old_proc = win._drag_lock_orig_proc

        def _handle_title_button(hit):
            """کلیک روی ✕/▢/– نوار عنوان — همیشه از مسیر رمز."""
            closing = getattr(win, "_app_closing", False)
            if closing:
                return False
            if hit == HTCLOSE:
                _schedule_lock_action(
                    win, getattr(win, "_on_close_attempt", None))
                return True
            if hit in (HTMINBUTTON, HTREDUCE):
                if getattr(win, "_allow_iconify", False):
                    win._allow_iconify = False
                    return False
                if not getattr(win, "_minimize_dialog_open", False):
                    _schedule_lock_action(
                        win, getattr(win, "_on_minimize_attempt", None))
                return True
            if hit in (HTMAXBUTTON, HTZOOM):
                is_full = getattr(win, "_was_full", True)
                if not is_full:
                    _schedule_lock_action(
                        win, lambda: _maximize_without_password(win))
                    return True
                if getattr(win, "_allow_restore", False):
                    win._allow_restore = False
                    return False
                if not getattr(win, "_restore_dialog_open", False):
                    _schedule_lock_action(
                        win, getattr(win, "_on_restore_attempt", None))
                return True
            return False

        def _wnd_proc(hWnd, msg, wParam, lParam):
            try:
                closing = getattr(win, "_app_closing", False)

                if msg == WM_SIZE and wParam == SIZE_MINIMIZED:
                    if getattr(win, "_minimizing_authorized", False):
                        win._was_in_taskbar = True
                    return user32.CallWindowProcW(
                        old_proc, hWnd, msg, wParam, lParam)

                if msg == WM_DESTROY and not closing:
                    if getattr(win, "_minimizing_authorized", False) or _hwnd_is_iconic(win):
                        return 0

                # WM_CLOSE — بعضی ویندوزها مستقیم می‌فرستند
                if msg == WM_CLOSE and not closing:
                    if getattr(win, "_minimizing_authorized", False) or _hwnd_is_iconic(win):
                        return 0
                    _schedule_lock_action(
                        win, getattr(win, "_on_close_attempt", None))
                    return 0

                # کلیک مستقیم روی دکمه‌های نوار عنوان (✕ ▢ –)
                if msg in (WM_NCLBUTTONDOWN, WM_NCLBUTTONDBLCLK) and not closing:
                    if _handle_title_button(wParam):
                        return 0
                    if getattr(win, "_drag_locked", True) and wParam == HTCAPTION:
                        return 0

                if msg == WM_SYSCOMMAND and not closing:
                    cmd = wParam & SC_MASK

                    if cmd == SC_CLOSE:
                        if getattr(win, "_minimizing_authorized", False) or _hwnd_is_iconic(win):
                            return 0
                        _schedule_lock_action(
                            win, getattr(win, "_on_close_attempt", None))
                        return 0

                    if cmd == SC_MINIMIZE:
                        if (getattr(win, "_allow_iconify", False)
                                or getattr(win, "_minimizing_authorized", False)):
                            win._allow_iconify = False
                            return user32.CallWindowProcW(
                                old_proc, hWnd, msg, wParam, lParam)
                        if not getattr(win, "_minimize_dialog_open", False):
                            _schedule_lock_action(
                                win, getattr(win, "_on_minimize_attempt", None))
                        return 0

                    if cmd in (SC_MAXIMIZE, SC_RESTORE):
                        is_full = getattr(win, "_was_full", True)
                        if not is_full and cmd == SC_MAXIMIZE:
                            _schedule_lock_action(
                                win, lambda: _maximize_without_password(win))
                            return 0
                        if is_full:
                            if getattr(win, "_allow_restore", False):
                                win._allow_restore = False
                                return user32.CallWindowProcW(
                                    old_proc, hWnd, msg, wParam, lParam)
                            if not getattr(win, "_restore_dialog_open", False):
                                _schedule_lock_action(
                                    win, getattr(win, "_on_restore_attempt", None))
                            return 0

                if getattr(win, "_drag_locked", True):
                    if msg == WM_SYSCOMMAND and (wParam & SC_MASK) == SC_MOVE:
                        return 0
            except Exception:
                pass
            try:
                return user32.CallWindowProcW(old_proc, hWnd, msg, wParam, lParam)
            except Exception:
                return 0

        new_proc = WNDPROCTYPE(_wnd_proc)
        win._drag_lock_proc_ref = new_proc
        win._drag_lock_installed = True
        set_wndproc(hwnd, GWLP_WNDPROC, new_proc)
    except Exception:
        pass


def _ensure_taskbar_button(win):
    """آیکون برنامه در نوار وظیفه — حتی بعد از borderless."""
    if os.name != "nt":
        return
    try:
        import ctypes
        hwnd = _get_toplevel_hwnd(win)
        GWL_EXSTYLE = -20
        GWL_HWNDPARENT = -8
        WS_EX_APPWINDOW = 0x00040000
        WS_EX_TOOLWINDOW = 0x00000080
        get_long = getattr(ctypes.windll.user32, "GetWindowLongPtrW",
                           ctypes.windll.user32.GetWindowLongW)
        set_long = getattr(ctypes.windll.user32, "SetWindowLongPtrW",
                           ctypes.windll.user32.SetWindowLongW)
        ex = get_long(hwnd, GWL_EXSTYLE)
        ex = (ex | WS_EX_APPWINDOW) & ~WS_EX_TOOLWINDOW
        set_long(hwnd, GWL_EXSTYLE, ex)
        set_long(hwnd, GWL_HWNDPARENT, 0)
        SWP = 0x2 | 0x1 | 0x4 | 0x20 | 0x40
        ctypes.windll.user32.SetWindowPos(hwnd, 0, 0, 0, 0, 0, SWP)
        win.update_idletasks()
    except Exception:
        pass


def _sync_tk_iconic_state(win):
    """همگام‌سازی Tk با وضعیت مینیمایز ویندوز — جلوگیری از بستن."""
    if not _hwnd_is_iconic(win):
        return
    try:
        win.tk.call("wm", "iconstate", win._w, "iconic")
    except Exception:
        try:
            win.wm_state("iconic")
        except Exception:
            pass


def _prepare_hwnd_for_minimize(win):
    """استایل Win32 استاندارد — فراخوانی فقط وقتی hook موقتاً برداشته شده."""
    if os.name != "nt":
        return
    try:
        import ctypes
        user32 = ctypes.windll.user32
        hwnd = _get_toplevel_hwnd(win)
        GWL_STYLE = -16
        GWL_EXSTYLE = -20
        GWL_HWNDPARENT = -8
        WS_OVERLAPPEDWINDOW = 0x00CF0000
        WS_EX_APPWINDOW = 0x00040000
        WS_EX_TOOLWINDOW = 0x00000080
        get_long = getattr(user32, "GetWindowLongPtrW", user32.GetWindowLongW)
        set_long = getattr(user32, "SetWindowLongPtrW", user32.SetWindowLongW)
        if not getattr(win, "_pre_minimize_style_saved", False):
            win._saved_minimize_style = get_long(hwnd, GWL_STYLE)
            win._saved_minimize_exstyle = get_long(hwnd, GWL_EXSTYLE)
            win._pre_minimize_style_saved = True
        set_long(hwnd, GWL_STYLE, WS_OVERLAPPEDWINDOW)
        ex = get_long(hwnd, GWL_EXSTYLE)
        ex = (ex | WS_EX_APPWINDOW) & ~WS_EX_TOOLWINDOW
        set_long(hwnd, GWL_EXSTYLE, ex)
        set_long(hwnd, GWL_HWNDPARENT, 0)
        SWP = 0x2 | 0x1 | 0x4 | 0x20
        user32.SetWindowPos(hwnd, 0, 0, 0, 0, 0, SWP)
    except Exception:
        pass


def win_minimize(win):
    """مینیمایز به تسک‌بار — PostMessage (بدون reentrant crash)."""
    if os.name != "nt":
        try:
            win._allow_iconify = True
            iconify = getattr(win, "_real_iconify", win.iconify)
            iconify()
            return True
        except Exception:
            return False
    try:
        import ctypes
        user32 = ctypes.windll.user32
        hwnd = _get_toplevel_hwnd(win)
        WM_SYSCOMMAND = 0x0112
        SC_MINIMIZE = 0xF020
        win._allow_iconify = True
        user32.PostMessageW(hwnd, WM_SYSCOMMAND, SC_MINIMIZE, 0)
        return True
    except Exception:
        return False


def win_restore(win):
    """بازگرداندن پنجره از حالت مینیمایز و اطمینان از تمام‌صفحه و بالاترین لایه بودن آن."""
    if os.name == "nt":
        try:
            import ctypes
            win.update_idletasks()
            hwnd = _get_toplevel_hwnd(win)
            SW_RESTORE = 9
            ctypes.windll.user32.ShowWindow(hwnd, SW_RESTORE)
        except Exception:
            pass
    try:
        win.deiconify()
        win.lift()
    except Exception:
        pass


def _root_window(widget):
    """پنجرهٔ اصلی Tk (نه Toplevel فرزند)."""
    w = widget
    while w is not None:
        if isinstance(w, tk.Tk):
            return w
        try:
            w = w.master
        except Exception:
            break
    try:
        return widget.winfo_toplevel() if widget else None
    except Exception:
        return None


def _ensure_app_not_topmost(win):
    try:
        win.attributes("-topmost", False)
    except Exception:
        pass


def _is_topmost_window(win):
    try:
        return bool(win and win.winfo_exists() and win.attributes("-topmost"))
    except Exception:
        return False


def _lower_root_topmost(parent):
    """topmost والد را موقتاً برمی‌دارد تا دیالوگ جلوتر دیده شود."""
    root = _root_window(parent)
    lowered = False
    if root and _is_topmost_window(root):
        try:
            root.attributes("-topmost", False)
            root.update_idletasks()
            lowered = True
        except Exception:
            pass
    return root, lowered


def _restore_root_topmost(root, lowered):
    if lowered and root and root.winfo_exists():
        try:
            root.attributes("-topmost", True)
            root.lift()
            root.update_idletasks()
        except Exception:
            pass


def _release_grab_safe(widget):
    """grab گیر کرده را آزاد می‌کند — وگرنه کل برنامه قفل می‌شود."""
    w = widget
    for _ in range(8):
        if w is None:
            break
        try:
            if w.winfo_exists():
                w.grab_release()
        except Exception:
            pass
        try:
            w = w.master
        except Exception:
            break


def _raise_window_front(win, skip_win_focus=False):
    """پنجره را جلو می‌آورد — اگر پاپ‌آپ باز است همان جلو می‌ماند."""
    try:
        if isinstance(win, tk.Tk):
            top = _get_top_popup(win)
            if top is not None:
                _bring_popup_to_front(top, focus=not skip_win_focus)
                return
            if _modal_dialog_active(win):
                return
    except Exception:
        pass
    try:
        win.deiconify()
        win.lift()
        win.update_idletasks()
    except Exception:
        pass
    if os.name == "nt":
        try:
            import ctypes
            hwnd = _get_toplevel_hwnd(win)
            ctypes.windll.user32.SetForegroundWindow(hwnd)
        except Exception:
            pass
    if not skip_win_focus:
        try:
            win.focus_force()
        except Exception:
            pass


def _focus_entry_widget(entry, dlg):
    """فوکوس اولیه روی فیلد رمز — بدون قفل کردن فوکوس (دکمه‌ها قابل کلیک باشند)."""
    def _do():
        try:
            if entry.winfo_exists() and dlg.winfo_exists():
                entry.focus_set()
                try:
                    entry.icursor("end")
                except Exception:
                    pass
        except Exception:
            pass

    def _on_focus_out(event=None):
        # فقط اگر فوکوس از کل دیالوگ خارج شد، برگردان — نه وقتی روی دکمه Confirm/Cancel است
        try:
            if not dlg.winfo_exists():
                return
            focused = dlg.focus_get()
            if focused is None:
                dlg.after(30, _do)
                return
            # اگر فوکوس هنوز داخل همین دیالوگ است (مثلاً روی دکمه)، دست نزن
            w = focused
            while w is not None:
                if w == dlg:
                    return
                try:
                    w = w.master
                except Exception:
                    break
            dlg.after(30, _do)
        except Exception:
            pass

    _do()
    try:
        dlg.after(40, _do)
        dlg.after(150, _do)
        dlg.bind("<Map>", lambda e: _do(), add="+")
        entry.bind("<FocusOut>", _on_focus_out, add="+")
    except Exception:
        pass


def _present_modal_dialog(dlg, focus_widget=None):
    """نمایش دیالوگ جلو — بدون grab_set (grab باعث قفل کل برنامه می‌شود)."""
    parent = getattr(dlg, "master", None)
    root = _root_window(parent or dlg)
    if root is not None:
        _modal_enter(root)
        _register_popup(root, dlg)
    _release_grab_safe(parent or dlg)
    try:
        _bring_popup_to_front(dlg, focus=False)
        if focus_widget is not None:
            _focus_entry_widget(focus_widget, dlg)
        dlg.wait_window()
    finally:
        if root is not None:
            _unregister_popup(root, dlg)
            _modal_leave(root)
        _release_grab_safe(dlg)
        _release_grab_safe(parent)


def _ask_string_topmost(parent, title, prompt, show=None):
    """دیالوگ رمز/ورودی — همیشه وسط صفحه و جلوی برنامه."""
    result = {"value": None}
    dlg = tk.Toplevel(parent)
    dlg.withdraw()
    dlg.title(title)
    dlg.resizable(False, False)
    dlg.configure(bg=C["bg"])
    try:
        dlg.transient(parent)
    except Exception:
        pass

    frm = tk.Frame(dlg, bg=C["bg"], padx=24, pady=18,
                   highlightthickness=2, highlightbackground=C["accent"])
    frm.pack(fill="both", expand=True)
    tk.Label(frm, text=title, bg=C["bg"], fg=C["accent"],
             font=(_MAIN_FONT, 12, "bold")).pack(pady=(0, 6))
    tk.Label(frm, text=prompt, bg=C["bg"], fg=C["text_bright"],
             font=(_MAIN_FONT, 10, "bold"), wraplength=320, justify="right").pack(pady=(0, 12))
    var = tk.StringVar()
    ent = tk.Entry(frm, textvariable=var, show=(show or ""), justify="center",
                   font=("B Nazanin", 14, "bold"), bd=0, relief="flat",
                   bg=C["entry_bg"], fg=C["accent"], insertbackground=C["accent"])
    # رمز: از سوییچر فارسی/انگلیسی رد شو تا کیبورد گیر نکند
    ent._kb_skip = True
    dlg._focus_entry = ent
    ent.pack(fill="x", ipady=8)

    # نوار سوییچ EN/FA برای خروج از حالت نوشتاری فارسی
    try:
        attach_switcher_bar(frm, compact=True)
    except Exception:
        pass

    def ok(event=None):
        result["value"] = var.get()
        dlg.destroy()

    def cancel(event=None):
        result["value"] = None
        dlg.destroy()

    btn_f = tk.Frame(frm, bg=C["bg"])
    btn_f.pack(fill="x", pady=(16, 0))
    ok_btn = tk.Button(btn_f, text=tr("confirm"), command=ok, width=10)
    ok_btn.pack(side="right", padx=4)
    cancel_btn = tk.Button(btn_f, text=tr("cancel"), command=cancel, width=10)
    cancel_btn.pack(side="right", padx=4)
    # دکمه‌ها باید فوکوس بگیرند (بدون دزدیده شدن توسط Entry)
    for b in (ok_btn, cancel_btn):
        b.configure(takefocus=1)

    ent.bind("<Return>", ok)
    dlg.bind("<Escape>", cancel)
    dlg.protocol("WM_DELETE_WINDOW", cancel)

    dlg.update_idletasks()
    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    w, h = max(360, dlg.winfo_reqwidth()), dlg.winfo_reqheight()
    dlg.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
    try:
        dlg.attributes("-topmost", True)
    except Exception:
        pass
    dlg.deiconify()
    dlg.lift()
    # قبل از باز شدن دیالوگ رمز، کیبورد را انگلیسی کن
    try:
        from keyboard_layout import get_manager, LANG_EN
        get_manager().switch_to(LANG_EN)
    except Exception:
        pass
    _present_modal_dialog(dlg, ent)
    # بعد از بستن دیالوگ هم انگلیسی بماند
    try:
        from keyboard_layout import get_manager, LANG_EN
        get_manager().switch_to(LANG_EN)
    except Exception:
        pass
    return result["value"]


def _show_message_topmost(parent, title, message, msg_type="info"):
    """پیام ساده — جایگزین messagebox"""
    dlg = tk.Toplevel(parent)
    dlg.withdraw()
    dlg.title(title)
    dlg.resizable(False, False)
    dlg.configure(bg=C["bg"])
    try:
        dlg.transient(parent)
    except Exception:
        pass
    colors = {"info": C["accent"], "error": C["danger"], "warning": C["warning"]}
    fg = colors.get(msg_type, C["accent"])
    frm = tk.Frame(dlg, bg=C["bg"], padx=24, pady=18,
                   highlightthickness=2, highlightbackground=fg)
    frm.pack(fill="both", expand=True)
    tk.Label(frm, text=title, bg=C["bg"], fg=fg,
             font=(_MAIN_FONT, 12, "bold")).pack(pady=(0, 8))
    tk.Label(frm, text=message, bg=C["bg"], fg=C["text_bright"],
             font=(_MAIN_FONT, 10, "bold"), wraplength=360, justify="right").pack(pady=(0, 14))

    def close(event=None):
        dlg.destroy()

    tk.Button(frm, text=tr("ok"), command=close, width=12).pack()
    dlg.bind("<Return>", close)
    dlg.bind("<Escape>", close)
    dlg.protocol("WM_DELETE_WINDOW", close)
    dlg.update_idletasks()
    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    w, h = max(380, dlg.winfo_reqwidth()), dlg.winfo_reqheight()
    dlg.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
    _present_modal_dialog(dlg)


def _ask_okcancel_topmost(parent, title, message):
    """تایید/انصراف"""
    result = {"value": False}
    dlg = tk.Toplevel(parent)
    dlg.withdraw()
    dlg.title(title)
    dlg.resizable(False, False)
    dlg.configure(bg=C["bg"])
    try:
        dlg.transient(parent)
    except Exception:
        pass
    frm = tk.Frame(dlg, bg=C["bg"], padx=24, pady=18,
                   highlightthickness=2, highlightbackground=C["accent"])
    frm.pack(fill="both", expand=True)
    tk.Label(frm, text=title, bg=C["bg"], fg=C["accent"],
             font=(_MAIN_FONT, 12, "bold")).pack(pady=(0, 6))
    tk.Label(frm, text=message, bg=C["bg"], fg=C["text_bright"],
             font=(_MAIN_FONT, 10, "bold"), wraplength=360, justify="right").pack(pady=(0, 14))

    def ok(event=None):
        result["value"] = True
        dlg.destroy()

    def cancel(event=None):
        result["value"] = False
        dlg.destroy()

    btn_f = tk.Frame(frm, bg=C["bg"])
    btn_f.pack(fill="x")
    tk.Button(btn_f, text=tr("confirm"), command=ok, width=10).pack(side="right", padx=4)
    tk.Button(btn_f, text=tr("cancel"), command=cancel, width=10).pack(side="right", padx=4)
    dlg.bind("<Return>", ok)
    dlg.bind("<Escape>", cancel)
    dlg.protocol("WM_DELETE_WINDOW", cancel)
    dlg.update_idletasks()
    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    w, h = max(380, dlg.winfo_reqwidth()), dlg.winfo_reqheight()
    dlg.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
    _present_modal_dialog(dlg)
    return result["value"]


def _ask_yesno_topmost(parent, title, message):
    """بله/خیر"""
    result = {"value": False}
    dlg = tk.Toplevel(parent)
    dlg.withdraw()
    dlg.title(title)
    dlg.resizable(False, False)
    dlg.configure(bg=C["bg"])
    try:
        dlg.transient(parent)
    except Exception:
        pass
    frm = tk.Frame(dlg, bg=C["bg"], padx=24, pady=18,
                   highlightthickness=2, highlightbackground=C["accent2"])
    frm.pack(fill="both", expand=True)
    tk.Label(frm, text=title, bg=C["bg"], fg=C["accent2"],
             font=(_MAIN_FONT, 12, "bold")).pack(pady=(0, 6))
    tk.Label(frm, text=message, bg=C["bg"], fg=C["text_bright"],
             font=(_MAIN_FONT, 10, "bold"), wraplength=360, justify="right").pack(pady=(0, 14))

    def yes(event=None):
        result["value"] = True
        dlg.destroy()

    def no(event=None):
        result["value"] = False
        dlg.destroy()

    btn_f = tk.Frame(frm, bg=C["bg"])
    btn_f.pack(fill="x")
    tk.Button(btn_f, text=tr("yes"), command=yes, width=10).pack(side="right", padx=4)
    tk.Button(btn_f, text=tr("no"), command=no, width=10).pack(side="right", padx=4)
    dlg.bind("<Return>", yes)
    dlg.bind("<Escape>", no)
    dlg.protocol("WM_DELETE_WINDOW", no)
    dlg.update_idletasks()
    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    w, h = max(380, dlg.winfo_reqwidth()), dlg.winfo_reqheight()
    dlg.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")
    _present_modal_dialog(dlg)
    return result["value"]


def verify_lock_password(win, title, prompt):
    """رمز قفل نرم‌افزار — همیشه پرسیده می‌شود (بستن / مینیمایز / کوچک‌کردن)."""
    if getattr(win, "_lock_dialog_open", False):
        return False
    _release_grab_safe(win)
    win._lock_dialog_open = True
    try:
        db = load_db()
        real_pw = get_lock_password(db)
        pw = _ask_string_topmost(win, title, prompt, show="●")
        if pw is None:
            return False
        if normalize_digits(pw) == normalize_digits(real_pw):
            return True
        _show_message_topmost(win, tr("wrong_pw_title"), tr("wrong_pw"), "error")
        return False
    finally:
        win._lock_dialog_open = False


def request_window_minimize(win):
    """مینیمایز — فقط با رمز قفل."""
    if getattr(win, "_app_closing", False):
        return
    if getattr(win, "_minimize_dialog_open", False):
        return
    win._minimize_dialog_open = True
    try:
        if verify_lock_password(win, tr("minimize_pw_title"), tr("minimize_pw_msg")):
            do_window_minimize_safe(win)
    finally:
        win._minimize_dialog_open = False
        if not getattr(win, "_minimizing_authorized", False):
            _clear_lock_bypass_flags(win)


def request_window_restore_down(win):
    """کوچک‌کردن از تمام‌صفحه — فقط با رمز قفل."""
    if getattr(win, "_app_closing", False):
        return
    _clear_lock_bypass_flags(win)
    if getattr(win, "_restore_dialog_open", False):
        return
    win._restore_dialog_open = True
    try:
        if verify_lock_password(win, tr("restore_pw_title"), tr("restore_pw_msg")):
            do_window_restore_down(win)
        else:
            win._drag_locked = True
            win._was_full = True
            force_full_screen(win)
    finally:
        win._restore_dialog_open = False
        _clear_lock_bypass_flags(win)


def request_window_restore_or_maximize(win):
    """▢ — اگر تمام‌صفحه: کوچک‌کردن با رمز. اگر کوچک: بزرگ بدون رمز."""
    if getattr(win, "_was_full", True):
        request_window_restore_down(win)
    else:
        _maximize_without_password(win)


def request_window_close(win, after_password=None):
    """بستن — فقط با رمز قفل. after_password تابع بستن واقعی است."""
    if getattr(win, "_app_closing", False):
        return
    _clear_lock_bypass_flags(win)
    if not verify_lock_password(win, tr("close_pw_title"), tr("close_pw_msg")):
        return
    if after_password:
        after_password()


# ═══════════════════════════════════════════════════════════
#  اولویت نمایش — دیالوگ‌ها و پنجره‌های درون‌برنامه‌ای همیشه
#  روی پنجره‌ی اصلی (topmost) دیده شوند، نه پشت آن
# ═══════════════════════════════════════════════════════════
def _popup_context_menu(parent, menu, x_root, y_root):
    """منوی راست‌کلیک."""
    try:
        menu.tk_popup(x_root, y_root)
    finally:
        try:
            menu.grab_release()
        except Exception:
            pass


def prepare_popup_window(win, parent=None):
    """پاپ‌آپ ورود/پیام — همیشه جلوی پنجرهٔ اصلی."""
    if parent is None:
        parent = getattr(win, "master", None)
    root = _root_window(parent or win)
    if root is not None and not getattr(win, "_popup_prepared", False):
        _modal_enter(root)
        win._popup_prepared = True
        win._popup_modal_root = root
        _register_popup(root, win)
    try:
        if root:
            win.transient(root)
    except Exception:
        pass
    try:
        win.attributes("-topmost", True)
    except Exception:
        pass

    if not getattr(win, "_popup_destroy_bound", False):
        win._popup_destroy_bound = True

        def _on_popup_destroy(event=None):
            r = getattr(win, "_popup_modal_root", None)
            if r is not None and getattr(win, "_popup_prepared", False):
                _modal_leave(r)
                win._popup_prepared = False
                win._popup_modal_root = None
            _unregister_popup(r, win)

        win.bind("<Destroy>", _on_popup_destroy, add="+")

    _bring_popup_to_front(win)
    return win


def _dialog_parent(kwargs):
    p = kwargs.get("parent")
    if p is not None:
        return p
    try:
        import tkinter as _tk
        return _tk._default_root
    except Exception:
        return None


def _patch_modal_dialogs():
    """messagebox همیشه جلوی پنجره — نه پشت آن (که باعث قفل کل برنامه می‌شود)."""
    def _show(title=None, message=None, msg_type="info", **kwargs):
        parent = _dialog_parent(kwargs)
        _release_grab_safe(parent)
        _show_message_topmost(parent, title or "", message or "", msg_type)

    def _yesno(title=None, message=None, **kwargs):
        parent = _dialog_parent(kwargs)
        _release_grab_safe(parent)
        return _ask_yesno_topmost(parent, title or "", message or "")

    def _okcancel(title=None, message=None, **kwargs):
        parent = _dialog_parent(kwargs)
        _release_grab_safe(parent)
        return _ask_okcancel_topmost(parent, title or "", message or "")

    def _question(title=None, message=None, **kwargs):
        return "yes" if _yesno(title=title, message=message, **kwargs) else "no"

    def _retrycancel(title=None, message=None, **kwargs):
        return _okcancel(title=title, message=message, **kwargs)

    def _yesnocancel(title=None, message=None, **kwargs):
        if _yesno(title=title, message=message, **kwargs):
            return True
        return None

    _replacements = {
        "showinfo": lambda t=None, m=None, **kw: _show(t, m, "info", **kw),
        "showwarning": lambda t=None, m=None, **kw: _show(t, m, "warning", **kw),
        "showerror": lambda t=None, m=None, **kw: _show(t, m, "error", **kw),
        "askyesno": _yesno,
        "askokcancel": _okcancel,
        "askquestion": _question,
        "askretrycancel": _retrycancel,
        "askyesnocancel": _yesnocancel,
    }
    for _name, _fn in _replacements.items():
        setattr(messagebox, _name, _fn)


_patch_modal_dialogs()


# ═══════════════════════════════════════════════════════════
#  ردیف آیکون‌های کنترل پنجره — ثابت و یکسان در کل نرم‌افزار
#  (هم صفحه‌ی لاگین، هم صفحه‌ی اصلی)، همیشه سمت راست، به همان
#  اندازه، و به همین ترتیب از راست به چپ:
#      ✕  خروج    │   ▢  کوچک‌کردن   │   –  مینیمایز
# ═══════════════════════════════════════════════════════════
def build_window_control_icons(parent, win, size=40):
    """دکمه‌های ✕/▢/– — بدون کادر، استایل تمیز."""
    bg = C["header_bg"]
    row = tk.Frame(parent, bg=bg, bd=0, highlightthickness=0)

    def _run(action_attr):
        def _click(e=None):
            cb = getattr(win, action_attr, None)
            if cb:
                cb()
            return "break"
        return _click

    icon_font = ("Segoe UI Symbol", max(14, int(size * 0.42)))

    def _make(glyph, hover_bg, fg, hover_fg, click_handler):
        box = tk.Frame(row, bg=bg, width=size, height=size,
                       bd=0, highlightthickness=0, cursor="hand2")
        box.pack(side="right", padx=1)
        box.pack_propagate(False)
        lbl = tk.Label(box, text=glyph, bg=bg, fg=fg, font=icon_font)
        lbl.place(relx=.5, rely=.5, anchor="center")

        def _enter(e=None):
            box.config(bg=hover_bg)
            lbl.config(bg=hover_bg, fg=hover_fg)
        def _leave(e=None):
            box.config(bg=bg)
            lbl.config(bg=bg, fg=fg)

        for w in (box, lbl):
            w.bind("<Enter>", _enter)
            w.bind("<Leave>", _leave)
            w.bind("<Button-1>", click_handler)
        return box

    _make("✕", "#e81123", "#b0b8c4", "#ffffff", _run("_on_close_attempt"))
    _make("□", "#3d4a5c", "#b0b8c4", "#ffffff", _run("_on_restore_attempt"))
    _make("−", "#3d4a5c", "#b0b8c4", "#ffffff", _run("_on_minimize_attempt"))
    row.pack()
    return row


def mount_window_controls(header, win, size=40, margin_right=14):
    """دکمه‌های ✕/▢/– — گوشهٔ راست هدر، بدون قاب."""
    bar = tk.Frame(header, bg=C["header_bg"], bd=0, highlightthickness=0)
    bar._is_win_controls = True
    build_window_control_icons(bar, win, size=size)
    bar.place(relx=1.0, rely=0.5, anchor="e", x=-margin_right)
    try:
        bar.lift()
    except Exception:
        pass
    return bar


def _drag_excluded_widget(widget):
    """دکمه‌ها و کنترل‌های ✕/▢/– — نه کل هدر."""
    hdr = None
    w = widget
    while w is not None:
        if getattr(w, "_is_win_controls", False):
            return True
        if isinstance(w, tk.Button):
            return True
        try:
            if str(w.cget("cursor")) == "hand2":
                return True
        except Exception:
            pass
        if getattr(w, "_is_drag_header", False):
            hdr = w
            break
        w = getattr(w, "master", None)
    return False


def _window_drag_allowed(win):
    return not getattr(win, "_drag_locked", True) and not getattr(win, "_was_full", True)


def _point_in_header_drag_zone(win, x_root, y_root):
    """آیا نقطه داخل ناحیهٔ هدر (قابل کشیدن) است؟"""
    zones = []
    h = getattr(win, "_drag_header", None)
    if h is not None:
        zones.append(h)
    zones.extend(getattr(win, "_header_drag_zones", []))
    for zone in zones:
        try:
            if zone is None or not zone.winfo_exists():
                continue
            zx, zy = zone.winfo_rootx(), zone.winfo_rooty()
            zw, zh = zone.winfo_width(), zone.winfo_height()
            if zx <= x_root < zx + zw and zy <= y_root < zy + zh:
                return True
        except Exception:
            pass
    return False


def register_header_drag_zone(win, *widgets):
    """نوارهای رنگی بالای هدر — در ناحیهٔ کشیدن."""
    zones = list(getattr(win, "_header_drag_zones", []))
    for w in widgets:
        if w is not None and w not in zones:
            zones.append(w)
    win._header_drag_zones = zones
    update_window_drag_state(win)


def bind_window_header_drag(win, header):
    """در حالت کوچک — هر نقطه از هدر (چپ/راست/وسط) قابل کشیدن."""
    header._is_drag_header = True
    win._drag_header = header

    def _begin_drag(event):
        if not _window_drag_allowed(win):
            return
        if not _point_in_header_drag_zone(win, event.x_root, event.y_root):
            return
        if _drag_excluded_widget(event.widget):
            return
        try:
            if os.name == "nt":
                import ctypes
                from ctypes import wintypes
                rect = wintypes.RECT()
                ctypes.windll.user32.GetWindowRect(
                    _get_toplevel_hwnd(win), ctypes.byref(rect))
                win._drag_offset = (event.x_root - rect.left, event.y_root - rect.top)
            else:
                win._drag_offset = (event.x_root - win.winfo_x(), event.y_root - win.winfo_y())
            win._header_dragging = True
        except Exception:
            win._drag_offset = (0, 0)
            win._header_dragging = True

    def _motion(event):
        if not getattr(win, "_header_dragging", False):
            return
        if not _window_drag_allowed(win):
            win._header_dragging = False
            return
        if not hasattr(win, "_drag_offset"):
            return
        ox, oy = win._drag_offset
        nx, ny = int(event.x_root - ox), int(event.y_root - oy)
        try:
            if os.name == "nt":
                import ctypes
                from ctypes import wintypes
                user32 = ctypes.windll.user32
                hwnd = _get_toplevel_hwnd(win)
                rect = wintypes.RECT()
                user32.GetWindowRect(hwnd, ctypes.byref(rect))
                w, h = rect.right - rect.left, rect.bottom - rect.top
                user32.MoveWindow(hwnd, nx, ny, w, h, True)
            else:
                ww = win.winfo_width() or 1320
                wh = win.winfo_height() or 840
                win.geometry(f"{ww}x{wh}+{nx}+{ny}")
        except Exception:
            pass

    def _stop(event=None):
        win._header_dragging = False

    if not getattr(win, "_header_drag_global_bound", False):
        win._header_drag_global_bound = True
        win.bind("<B1-Motion>", _motion, add="+")
        win.bind("<ButtonRelease-1>", _stop, add="+")
        win.bind_all("<ButtonPress-1>", _begin_drag, add="+")

    def _refresh_cursor(_event=None):
        if not _window_drag_allowed(win):
            cur = "arrow"
        else:
            cur = "sizeall"

        def _set(w):
            try:
                if getattr(w, "_is_win_controls", False):
                    return
                if isinstance(w, tk.Button):
                    return
                try:
                    if str(w.cget("cursor")) == "hand2":
                        return
                except Exception:
                    pass
                w.configure(cursor=cur)
            except Exception:
                pass
            for ch in w.winfo_children():
                _set(ch)

        _set(header)
        for zone in getattr(win, "_header_drag_zones", []):
            try:
                if zone.winfo_exists():
                    _set(zone)
            except Exception:
                pass

    win._refresh_header_drag_cursor = _refresh_cursor
    _refresh_cursor()
    try:
        header.bind("<Enter>", _refresh_cursor, add="+")
        header.bind("<Leave>", _refresh_cursor, add="+")
    except Exception:
        pass


def refresh_header_drag_bindings(win):
    """به‌روزرسانی نشانگر موس هدر."""
    update_window_drag_state(win)


def update_window_drag_state(win):
    """بعد از کوچک/بزرگ — نشانگر موس هدر را به‌روز کند."""
    fn = getattr(win, "_refresh_header_drag_cursor", None)
    if fn:
        try:
            fn()
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════
#  پروسه‌ی نگهبان مستقل (safety net) — اگر برنامه‌ی اصلی به هر
#  دلیل غیرمنتظره (کرش، قطع برق، Kill از Task Manager) از بین
#  برود، این پروسه‌ی جدا و مستقل، نبودِ پروسه‌ی اصلی را تشخیص
#  می‌دهد و نوار تسک‌بار مخفی‌شده را فوراً برمی‌گرداند تا کاربر
#  پشت سیستم بدون دسترسی به ویندوز گیر نکند.
# ═══════════════════════════════════════════════════════════
_TASKBAR_WATCHDOG_FLAG = "--taskbar-watchdog"

def _run_taskbar_watchdog(parent_pid: int):
    if os.name != "nt":
        return
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
        SYNCHRONIZE = 0x00100000
        handle = kernel32.OpenProcess(
            PROCESS_QUERY_LIMITED_INFORMATION | SYNCHRONIZE, False, parent_pid)
        if not handle:
            show_all_taskbars()
            return
        try:
            WAIT_OBJECT_0 = 0x0
            while True:
                ret = kernel32.WaitForSingleObject(handle, 1500)  # هر ۱.۵ ثانیه چک می‌کند
                if ret == WAIT_OBJECT_0:   # پروسه‌ی اصلی دیگر زنده نیست
                    break
        finally:
            kernel32.CloseHandle(handle)
    except Exception:
        pass
    show_all_taskbars()


def start_taskbar_watchdog():
    """پروسه‌ی نگهبان مستقل تسک‌بار — عین فایل مرجع."""
    if os.environ.get("STF_ADMIN") == "1":
        return None
    if os.name != "nt":
        return None
    try:
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        if getattr(sys, "frozen", False):
            args = [sys.executable, _TASKBAR_WATCHDOG_FLAG, str(os.getpid())]
        else:
            wd_script = os.environ.get("STF_WATCHDOG_SCRIPT", "").strip()
            if wd_script and os.path.isfile(wd_script):
                args = [sys.executable, wd_script, _TASKBAR_WATCHDOG_FLAG, str(os.getpid())]
            else:
                args = [sys.executable, os.path.abspath(__file__), _TASKBAR_WATCHDOG_FLAG, str(os.getpid())]
        return subprocess.Popen(args, creationflags=creationflags)
    except Exception:
        return None


def _is_watchdog_argv() -> bool:
    return _TASKBAR_WATCHDOG_FLAG in sys.argv


def _watchdog_parent_pid() -> int | None:
    try:
        idx = sys.argv.index(_TASKBAR_WATCHDOG_FLAG)
        return int(sys.argv[idx + 1])
    except (ValueError, IndexError, TypeError):
        return None



# ── checksum برای تشخیص دستکاری ──
def _db_checksum(data: str) -> str:
    return hashlib.sha256(data.encode("utf-8")).hexdigest()

def _write_checksum(path: str, checksum: str):
    try:
        with open(path + ".sha256", "w", encoding="utf-8") as f:
            f.write(checksum)
    except Exception:
        pass

def _verify_checksum(path: str) -> bool:
    """True اگر فایل سالم باشد یا checksum وجود نداشته باشد"""
    chk_path = path + ".sha256"
    if not os.path.exists(chk_path):
        return True
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = f.read()
        with open(chk_path, "r", encoding="utf-8") as f:
            expected = f.read().strip()
        return _db_checksum(data) == expected
    except Exception:
        return False

def _set_readonly(path: str):
    """فایل را فقط‌خواندنی می‌کند (Windows & Linux)"""
    try:
        import stat
        cur = os.stat(path).st_mode
        os.chmod(path, cur & ~(stat.S_IWRITE | stat.S_IWGRP | stat.S_IWOTH))
    except Exception:
        pass

def _clear_readonly(path: str):
    """قبل از نوشتن، حالت فقط‌خواندنی را برمی‌دارد"""
    try:
        import stat
        os.chmod(path, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    except Exception:
        pass

def _load_json_safe(path: str):
    """بارگذاری JSON با بررسی checksum — None اگر خراب باشد"""
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read()
        chk_path = path + ".sha256"
        if os.path.exists(chk_path):
            with open(chk_path, "r", encoding="utf-8") as f:
                expected = f.read().strip()
            if _db_checksum(raw) != expected:
                return None          # دستکاری شده — رد می‌شود
        return json.loads(raw)
    except Exception:
        return None

def _find_best_backup() -> dict | None:
    """جدیدترین بک‌آپ سالم را برمی‌گرداند"""
    for bdir in [BACKUP_DIR, BACKUP_DIR2]:
        if not bdir:
            continue
        data_dir = _backup_data_dir(bdir)
        if not os.path.exists(data_dir):
            continue
        try:
            files = sorted([
                x for x in os.listdir(data_dir)
                if x.startswith("slab_db_") and (
                    x.endswith(".json") or x.endswith(".zip")
                )
            ], reverse=True)
        except Exception:
            continue
        for fname in files:
            fp = os.path.join(data_dir, fname)
            if fname.endswith(".zip"):
                try:
                    if r"D:\SteelFactory2-v2" not in sys.path:
                        sys.path.insert(0, r"D:\SteelFactory2-v2")
                    from shared.backup_vault import decrypt_zip_to_bytes
                    from pathlib import Path as _P
                    raw = decrypt_zip_to_bytes(_P(fp)).decode("utf-8")
                    db = json.loads(raw)
                    if db is not None:
                        return db
                except Exception:
                    continue
            else:
                db = _load_json_safe(fp)
                if db is not None:
                    return db
    return None


# ═══════════════════════════════════════════════════════════
#  سیستم همگام‌سازی شبکه (LAN Sync) — حداکثر ۵ سیستم
#  بدون نیاز به تنظیم اضافه — خودکار سیستم‌ها را پیدا می‌کند
# ═══════════════════════════════════════════════════════════
import socket, struct, select

SYNC_PORT       = 57321          # پورت sync — باید روی همه سیستم‌ها یکسان باشد
SYNC_MAGIC      = b"SLAB_SYNC_V1"
SYNC_TIMEOUT    = 1.5            # ثانیه — منتظر پاسخ سیستم‌های دیگر
MAX_PEERS       = 5              # حداکثر تعداد سیستم‌های همزمان
LOCK_FILE       = "slab_db.lock" # فایل قفل برای جلوگیری از نوشتن همزمان
LOCK_TIMEOUT    = 8              # ثانیه — اگر lock بیشتر از این نگه داشته شد، آزاد می‌شود

_sync_lock      = threading.Lock()   # قفل داخلی thread-safe

# ── شناسه یکتا برای این سیستم ──
def _get_machine_id() -> str:
    id_file = os.path.join(_app_base_dir(), "slab_machine_id.txt")
    if os.path.exists(id_file):
        try:
            with open(id_file, "r") as f:
                mid = f.read().strip()
            if mid: return mid
        except Exception:
            pass
    import uuid
    mid = str(uuid.uuid4())
    try:
        with open(id_file, "w") as f:
            f.write(mid)
    except Exception:
        pass
    return mid

_MACHINE_ID = _get_machine_id()

# ── پیدا کردن IP محلی این سیستم ──
def _get_local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


# ═══════════════════════════════════════════════════════════
#  بررسی سلامت شبکه‌ی محلی (LAN) — مستقل از اینترنت بیرونی
#  هدف: تشخیص قطعی کابل LAN / مودم محلی، نه قطعی اینترنت بیرونی
# ═══════════════════════════════════════════════════════════
import subprocess

_NET_OK_CACHE = {"ok": True, "ts": 0.0}
_NET_CACHE_TTL = 3.0   # ثانیه — از پینگ مکرر در ذخیره‌های پشت‌سرهم جلوگیری می‌کند

def _get_default_gateway() -> str | None:
    """آی‌پی مودم/گیت‌وی پیش‌فرض شبکه محلی را برمی‌گرداند (ویندوز و لینوکس)"""
    try:
        if os.name == "nt":
            out = subprocess.run(
                ["ipconfig"], capture_output=True, text=True,
                timeout=4, creationflags=subprocess.CREATE_NO_WINDOW
            ).stdout
            for line in out.splitlines():
                if "Default Gateway" in line or "دروازه" in line or "Gateway" in line:
                    parts = line.split(":")
                    if len(parts) >= 2:
                        ip = parts[-1].strip()
                        if re.match(r"^\d+\.\d+\.\d+\.\d+$", ip):
                            return ip
        else:
            out = subprocess.run(
                ["ip", "route"], capture_output=True, text=True, timeout=4
            ).stdout
            m = re.search(r"default via (\d+\.\d+\.\d+\.\d+)", out)
            if m:
                return m.group(1)
    except Exception:
        pass
    return None

def _ping_host(ip: str, timeout_ms: int = 800) -> bool:
    """پینگ یک‌باره — بدون نیاز به دسترسی ادمین"""
    try:
        if os.name == "nt":
            cmd = ["ping", "-n", "1", "-w", str(timeout_ms), ip]
            result = subprocess.run(
                cmd, capture_output=True, timeout=(timeout_ms / 1000) + 1.5,
                creationflags=subprocess.CREATE_NO_WINDOW
            )
        else:
            cmd = ["ping", "-c", "1", "-W", str(max(1, timeout_ms // 1000)), ip]
            result = subprocess.run(cmd, capture_output=True, timeout=(timeout_ms / 1000) + 1.5)
        return result.returncode == 0
    except Exception:
        return False

def is_lan_alive(force: bool = False) -> bool:
    """در معماری SQL Server: اتصال به SQL (سرور) یا API (کلاینت) را بررسی می‌کند."""
    if os.environ.get("STF_ENABLE_SYNC") == "1" and not os.environ.get("STF_DISABLE_SYNC"):
        now = time.time()
        if not force and (now - _NET_OK_CACHE["ts"]) < _NET_CACHE_TTL:
            return _NET_OK_CACHE["ok"]
        ok = False
        try:
            local_ip = _get_local_ip()
            if local_ip and local_ip != "127.0.0.1":
                gw = _get_default_gateway()
                if gw and _ping_host(gw):
                    ok = True
                elif _discover_peers():
                    ok = True
        except Exception:
            ok = False
        _NET_OK_CACHE["ok"] = ok
        _NET_OK_CACHE["ts"] = now
        return ok
    return is_storage_alive(force)

# ── broadcast برای پیدا کردن سیستم‌های دیگر ──
def _discover_peers() -> list[str]:
    """سیستم‌های دیگر در LAN را پیدا می‌کند — آدرس IP آن‌ها را برمی‌گرداند"""
    peers = []
    try:
        local_ip = _get_local_ip()
        # ساخت پیام discovery
        msg = SYNC_MAGIC + b"|DISCOVER|" + _MACHINE_ID.encode() + b"|" + local_ip.encode()

        # ارسال broadcast
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        sock.settimeout(SYNC_TIMEOUT)
        sock.bind(("", 0))

        # broadcast به کل شبکه
        broadcast_ip = ".".join(local_ip.split(".")[:3]) + ".255"
        sock.sendto(msg, (broadcast_ip, SYNC_PORT))

        # منتظر پاسخ
        deadline = time.time() + SYNC_TIMEOUT
        while time.time() < deadline:
            try:
                ready = select.select([sock], [], [], max(0, deadline - time.time()))
                if ready[0]:
                    data, addr = sock.recvfrom(4096)
                    if data.startswith(SYNC_MAGIC + b"|ALIVE|"):
                        parts = data.split(b"|")
                        if len(parts) >= 4:
                            peer_id = parts[2].decode(errors="ignore")
                            peer_ip = parts[3].decode(errors="ignore")
                            if peer_id != _MACHINE_ID and peer_ip not in peers:
                                peers.append(peer_ip)
                                if len(peers) >= MAX_PEERS:
                                    break
            except Exception:
                break
        sock.close()
    except Exception:
        pass
    return peers

# ── سرور listener در background ──
_listener_started = False

def _start_sync_listener():
    """یک thread در background که به درخواست‌های سیستم‌های دیگر پاسخ می‌دهد"""
    global _listener_started
    if _listener_started:
        return
    _listener_started = True

    def _listener():
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            sock.bind(("", SYNC_PORT))
            sock.settimeout(2.0)
            local_ip = _get_local_ip()
            while True:
                try:
                    data, addr = sock.recvfrom(65535)
                    if not data.startswith(SYNC_MAGIC):
                        continue
                    parts = data.split(b"|")
                    if len(parts) < 3:
                        continue
                    cmd = parts[1].decode(errors="ignore")

                    if cmd == "DISCOVER":
                        # پاسخ alive
                        reply = SYNC_MAGIC + b"|ALIVE|" + _MACHINE_ID.encode() + b"|" + local_ip.encode()
                        sock.sendto(reply, addr)

                    elif cmd == "GET_DB":
                        # ارسال فایل DB به سیستم درخواست‌دهنده
                        sender_id = parts[2].decode(errors="ignore") if len(parts) > 2 else ""
                        if sender_id == _MACHINE_ID:
                            continue
                        try:
                            raw = json.dumps(load_db(), ensure_ascii=False, indent=2)
                            data_bytes = raw.encode("utf-8")
                            chunk_size = 60000
                            total = (len(data_bytes) + chunk_size - 1) // chunk_size
                            req_tag = parts[3] if len(parts) > 3 else b"0"
                            for i in range(total):
                                chunk = data_bytes[i*chunk_size:(i+1)*chunk_size]
                                reply = (SYNC_MAGIC + b"|DB_CHUNK|" +
                                         req_tag + b"|" +
                                         str(i).encode() + b"|" +
                                         str(total).encode() + b"|" +
                                         chunk)
                                sock.sendto(reply, addr)
                                time.sleep(0.01)
                        except Exception:
                            pass

                    elif cmd == "SAVE_DB":
                        # دریافت و ذخیره DB از سیستم دیگر
                        sender_id = parts[2].decode(errors="ignore") if len(parts) > 2 else ""
                        if sender_id == _MACHINE_ID:
                            continue
                        if len(parts) >= 4:
                            try:
                                remote_raw = parts[3].decode("utf-8")
                                remote_db  = json.loads(remote_raw)
                                _apply_remote_save(remote_db, remote_raw)
                            except Exception:
                                pass

                except socket.timeout:
                    continue
                except Exception:
                    time.sleep(0.5)
        except Exception:
            pass

    t = threading.Thread(target=_listener, daemon=True)
    t.start()


# ── همگام‌سازی فعال (pull): به‌جای فقط منتظر broadcast ماندن،
#    خودش از سیستم‌های روشن، آخرین نسخه‌ی کامل دیتا را می‌خواهد و merge می‌کند.
#    این کار باعث می‌شود وقتی سیستمی خاموش بوده و دوباره روشن می‌شود،
#    بدون نیاز به اینکه کسی تغییر تازه‌ای ثبت کند، خودش را کامل به‌روز کند.
def _pull_sync_once():
    try:
        peers = _discover_peers()
        if not peers:
            return
        req_tag = str(random.randint(1, 999999)).encode()
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        sock.settimeout(2.5)
        sock.bind(("", 0))
        msg = SYNC_MAGIC + b"|GET_DB|" + _MACHINE_ID.encode() + b"|" + req_tag
        for peer_ip in peers:
            try:
                sock.sendto(msg, (peer_ip, SYNC_PORT))
            except Exception:
                pass

        chunks = {}
        expected_total = None
        deadline = time.time() + 2.5
        while time.time() < deadline:
            try:
                remaining = max(0.1, deadline - time.time())
                ready = select.select([sock], [], [], remaining)
                if not ready[0]:
                    continue
                data, addr = sock.recvfrom(65535)
                if not data.startswith(SYNC_MAGIC + b"|DB_CHUNK|"):
                    continue
                rest = data[len(SYNC_MAGIC + b"|DB_CHUNK|"):]
                p = rest.split(b"|", 3)
                if len(p) < 4 or p[0] != req_tag:
                    continue
                idx, total = int(p[1]), int(p[2])
                chunks[idx] = p[3]
                expected_total = total
                if len(chunks) >= total:
                    break
            except Exception:
                continue
        sock.close()

        if expected_total and len(chunks) >= expected_total:
            full = b"".join(chunks[i] for i in range(expected_total))
            remote_raw = full.decode("utf-8")
            remote_db = json.loads(remote_raw)
            _apply_remote_save(remote_db, remote_raw)
    except Exception:
        pass

def _start_pull_sync_loop():
    """در بدو اجرا و سپس هر ۳۰ ثانیه یک‌بار، فعالانه از سیستم‌های دیگر sync می‌کشد"""
    def _loop():
        while True:
            try:
                if is_lan_alive():
                    _pull_sync_once()
            except Exception:
                pass
            time.sleep(30)
    t = threading.Thread(target=_loop, daemon=True)
    t.start()


# ── اعمال تغییرات دریافتی از سیستم دیگر ──
def _apply_remote_save(remote_db: dict, remote_raw: str):
    """داده دریافتی از سیستم دیگر را با first-write-wins ادغام می‌کند"""
    with _sync_lock:
        try:
            local_db = load_db()
            if local_db is None:
                save_db(remote_db)
                return

            # ── merge: first-write-wins بر اساس timestamp ──
            changed = False

            # melts
            local_ids = {r["slab_id"]: i for i, r in enumerate(local_db.get("melts", []))}
            for rec in remote_db.get("melts", []):
                sid = rec.get("slab_id")
                if not sid:
                    continue
                if sid not in local_ids:
                    # اسلب جدید — اضافه کن
                    local_db.setdefault("melts", []).append(rec)
                    changed = True
                else:
                    # اگر محلی هنوز «ثبت شده» است ولی remote تغییر کرده
                    local_rec = local_db["melts"][local_ids[sid]]
                    for field in ("qc_status","qc_by","qc_at",
                                  "exit_status","exit_by","exit_at",
                                  "registered_by","registered_at"):
                        local_val  = local_rec.get(field, "")
                        remote_val = rec.get(field, "")
                        # first-write-wins: اگر محلی خالی است ولی remote دارد
                        if not local_val and remote_val:
                            local_rec[field] = remote_val
                            changed = True

            # لیست‌های append-only — فقط رکوردهای جدید اضافه می‌شوند
            for key in ("movement_log","return_log","login_log","qc_history",
                        "tickets","file_exit_log","file_qc_log",
                        "file_melt_log","file_warehouse_log",
                        "lab_deliveries","scarf_cut","scrap","transfers_out"):
                local_list  = local_db.get(key, [])
                remote_list = remote_db.get(key, [])
                # شناسه‌های موجود
                local_batch_ids = set()
                for item in local_list:
                    for id_field in ("batch_id","id","slab_id","at","sent_at","registered_at"):
                        v = item.get(id_field)
                        if v:
                            local_batch_ids.add(f"{id_field}:{v}")
                            break
                for item in remote_list:
                    item_key = None
                    for id_field in ("batch_id","id","slab_id","at","sent_at","registered_at"):
                        v = item.get(id_field)
                        if v:
                            item_key = f"{id_field}:{v}"
                            break
                    if item_key and item_key not in local_batch_ids:
                        local_db.setdefault(key, []).append(item)
                        local_batch_ids.add(item_key)
                        changed = True

            if changed:
                save_db(local_db)
        except Exception:
            pass

# ── ارسال DB به همه سیستم‌های دیگر ──
def _broadcast_save(raw: str):
    """بعد از هر save، تغییرات را به بقیه سیستم‌ها می‌فرستد"""
    def _send():
        try:
            peers = _discover_peers()
            if not peers:
                return
            data_bytes = (SYNC_MAGIC + b"|SAVE_DB|" +
                          _MACHINE_ID.encode() + b"|" +
                          raw.encode("utf-8"))
            # اگر حجم بزرگ است از TCP استفاده می‌کند
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            for peer_ip in peers:
                try:
                    # UDP برای فایل‌های کوچک‌تر از 60KB
                    if len(data_bytes) < 60000:
                        sock.sendto(data_bytes, (peer_ip, SYNC_PORT))
                    else:
                        # TCP برای فایل‌های بزرگ‌تر
                        _send_tcp(peer_ip, raw)
                except Exception:
                    pass
            sock.close()
        except Exception:
            pass
    threading.Thread(target=_send, daemon=True).start()

def _send_tcp(peer_ip: str, raw: str):
    """ارسال از طریق TCP برای فایل‌های بزرگ"""
    TCP_PORT = SYNC_PORT + 1
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(5)
        s.connect((peer_ip, TCP_PORT))
        data = raw.encode("utf-8")
        # ارسال سایز و بعد داده
        s.sendall(struct.pack("!I", len(data)))
        s.sendall(data)
        s.close()
    except Exception:
        pass

# ── TCP listener برای فایل‌های بزرگ ──
def _start_tcp_listener():
    TCP_PORT = SYNC_PORT + 1
    def _tcp_server():
        try:
            srv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            srv.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            srv.bind(("", TCP_PORT))
            srv.listen(5)
            srv.settimeout(2.0)
            while True:
                try:
                    conn, addr = srv.accept()
                    conn.settimeout(10)
                    # خواندن سایز
                    size_data = b""
                    while len(size_data) < 4:
                        chunk = conn.recv(4 - len(size_data))
                        if not chunk: break
                        size_data += chunk
                    if len(size_data) < 4:
                        conn.close(); continue
                    size = struct.unpack("!I", size_data)[0]
                    if size > 50 * 1024 * 1024:  # حداکثر 50MB
                        conn.close(); continue
                    # خواندن داده
                    raw_bytes = b""
                    while len(raw_bytes) < size:
                        chunk = conn.recv(min(65536, size - len(raw_bytes)))
                        if not chunk: break
                        raw_bytes += chunk
                    conn.close()
                    if len(raw_bytes) == size:
                        remote_raw = raw_bytes.decode("utf-8")
                        remote_db  = json.loads(remote_raw)
                        _apply_remote_save(remote_db, remote_raw)
                except socket.timeout:
                    continue
                except Exception:
                    pass
        except Exception:
            pass
    threading.Thread(target=_tcp_server, daemon=True).start()

# ── شروع listener‌ها — فقط در حالت legacy P2P (STF_ENABLE_SYNC=1)
# در معماری SQL Server / client-server با STF_DISABLE_SYNC=1 غیرفعال می‌شود
if os.environ.get("STF_ENABLE_SYNC") == "1" and not os.environ.get("STF_DISABLE_SYNC"):
    _start_sync_listener()
    _start_tcp_listener()
    _start_pull_sync_loop()

def load_db(force=False):
    """بارگذاری از SQL Server (سرور) یا API (کلاینت) — با کش حافظه"""
    return _sql_load_db(force=force)

def save_db(db, action="save"):
    """ذخیره در SQL Server (سرور) یا API (کلاینت) + بک‌آپ خودکار"""
    ok = _sql_save_db(db, action=action)
    if ok and os.environ.get("STF_ENABLE_SYNC") == "1" and not os.environ.get("STF_DISABLE_SYNC"):
        try:
            raw = json.dumps(db, ensure_ascii=False, indent=2)
            threading.Thread(target=_broadcast_save, args=(raw,), daemon=True).start()
        except Exception:
            pass
    return ok

def now_str():
    """تاریخ و ساعت فعلی به شمسی"""
    return to_shamsi(datetime.datetime.now())

def get_first_report_date_sh():
    """تاریخ شمسی (فقط تاریخ) اولین ذوبی که در کل نرم‌افزار ثبت شده —
    برای استفاده به عنوان مقدار پیش‌فرض «از تاریخ» در همه گزارش‌گیری‌ها.
    اگر هیچ ذوبی ثبت نشده باشد، تاریخ امروز برگردانده می‌شود."""
    try:
        db = load_db()
        melts = db.get("melts", []) if db else []
        dates = []
        for m in melts:
            ts = m.get("registered_at") or m.get("created_at") or m.get("date")
            if ts:
                d = str(ts).strip().replace("  ", " ").split(" ", 1)[0]
                if d:
                    dates.append(d)
        if dates:
            # مرتب‌سازی رشته‌ای تاریخ شمسی YYYY/MM/DD به‌درستی کار می‌کند
            return sorted(dates)[0]
    except Exception:
        pass
    return to_shamsi(datetime.datetime.now()).split("  ")[0]

def split_dt(val):
    """تاریخ شمسی را به (تاریخ, ساعت) تفکیک می‌کند
    فرمت‌های پشتیبانی‌شده:
      '1404/03/15  14:30:25'  →  ('1404/03/15', '14:30:25')
      '14:30:25  1404/03/15'  →  ('1404/03/15', '14:30:25')
      '1404/03/15 14:30:25'   →  ('1404/03/15', '14:30:25')
    """
    if not val or val in ("—", "", None):
        return "—", "—"
    v = str(val).strip()
    # نرمال‌سازی فاصله‌های دوتایی به یکی
    v = " ".join(v.split())
    parts = v.split(" ", 1)
    if len(parts) < 2:
        return v, "—"
    p0, p1 = parts[0].strip(), parts[1].strip()
    # تشخیص کدام بخش تاریخه (شامل /) و کدام ساعته (شامل :)
    if "/" in p0:
        return p0, p1   # p0 = تاریخ، p1 = ساعت
    elif "/" in p1:
        return p1, p0   # p1 = تاریخ، p0 = ساعت (فرمت برعکس)
    else:
        return p0, p1   # پیش‌فرض
def _make_dt_filter(parent, bg, today_str, to_str=None):
    """ساخت فیلتر از/تا با تاریخ و ساعت جداگانه — برگشت (from_v, to_v)"""
    if to_str is None:
        to_str = to_shamsi(datetime.datetime.now()).split("  ")[0]

    def mk_ent(parent, var, w):
        return tk.Entry(parent, textvariable=var,
                        bg=C["entry_bg"], fg=C["accent"],
                        insertbackground=C["accent"], font=("B Nazanin", 11, "bold"),
                        bd=0, relief="flat", highlightthickness=1,
                        highlightbackground=C["border"],
                        highlightcolor=C["accent"],
                        justify="center", width=w)

    def mk_pair(parent, label_text, var, w):
        """ساخت جفت لیبل+ورودی — لیبل سمت راست، ورودی سمت چپ"""
        f = tk.Frame(parent, bg=bg)
        f.pack(side="right", padx=(0, 8))
        tk.Label(f, text=label_text, bg=bg, fg=C["text_dim"],
                 font=FONT_NORM).pack(side="right")
        mk_ent(f, var, w).pack(side="right", padx=(4, 0))
        return f

    dt_ctrl = tk.Frame(parent, bg=bg)
    dt_ctrl.pack(fill="x", padx=14, pady=6)

    tk.Label(dt_ctrl, text="(فرمت: ۱۴۰۴/۰۳/۱۵  —  ساعت: ۱۴:۳۰:۰۰)", bg=bg,
             fg=C["text_dim"], font=(_MAIN_FONT, 9, "bold")).pack(anchor="e", pady=(0,2))

    row_dates = tk.Frame(dt_ctrl, bg=bg)
    row_dates.pack(fill="x", pady=(0,3))
    from_date = tk.StringVar(value=today_str)
    to_date   = tk.StringVar(value=to_str)
    mk_pair(row_dates, "از تاریخ", from_date, 13)
    mk_pair(row_dates, "تا تاریخ", to_date,   13)

    row_times = tk.Frame(dt_ctrl, bg=bg)
    row_times.pack(fill="x", pady=(0,4))
    from_time = tk.StringVar(value="00:00:00")
    to_time   = tk.StringVar(value="23:59:59")
    mk_pair(row_times, "از ساعت", from_time, 13)
    mk_pair(row_times, "تا ساعت", to_time,   13)

    from_v = tk.StringVar()
    to_v   = tk.StringVar()
    def _sync(*_):
        from_v.set(f"{from_date.get().strip()}  {from_time.get().strip()}")
        to_v.set(f"{to_date.get().strip()}  {to_time.get().strip()}")
    for v in (from_date, from_time, to_date, to_time):
        v.trace_add("write", _sync)
    _sync()
    return from_v, to_v


def fmt_time(t):
    """فرمت زمان شمسی با صفر پیشرو — ۱۴:۰۴:۰۲ نه ۱۴:۴:۲"""
    if not t: return "—"
    t = t.strip()
    parts = t.replace("  ", " ").split(" ")
    if len(parts) < 2:
        return t
    date_p = parts[0]
    time_p = parts[1]
    tp = time_p.split(":")
    if len(tp) >= 2:
        hh = tp[0].zfill(2)
        mm = tp[1].zfill(2)
        ss = tp[2].zfill(2) if len(tp) > 2 else "00"
        time_p = f"{hh}:{mm}:{ss}"
    return f"{date_p}  {time_p}"

def normalize_digits(s):
    """اعداد فارسی/عربی را به انگلیسی تبدیل می‌کند — رمز، شماره اسلب، تاریخ و ..."""
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    return s.translate(str.maketrans(
        "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
        "01234567890123456789",
    ))

def hash_pw(pw):
    return hashlib.sha256(normalize_digits(pw).encode()).hexdigest()

# ── کش نام‌های نمایشی ──
_display_cache = {}

def get_display_name(username, db=None):
    """username → نام نمایشی کاربر (با کش)"""
    if not username or username in ("—", "شخص دیگر", "شیفت", "برش‌کار", ""):
        return username
    global _display_cache
    if username in _display_cache:
        return _display_cache[username]
    if db is None:
        db = load_db()
    disp = db.get("users", {}).get(username, {}).get("display", username)
    _display_cache[username] = disp
    return disp

def invalidate_display_cache():
    """بعد از تغییر کاربران کش رو پاک کن"""
    global _display_cache
    _display_cache = {}

# ═══════════════════════════════════════════════════════════
#  رنگ‌بندی و استایل
# ═══════════════════════════════════════════════════════════
C = {
    # ─── زمینه: طوسی گرم فولادی ───
    "bg":          "#3a3f45",   # طوسی تیره گرم
    "panel":       "#40464d",
    "card":        "#484e56",   # کارت طوسی روشن‌تر
    "card2":       "#444a52",
    "card_hover":  "#525860",

    # ─── رنگ اصلی: طلایی برنزی گرم ───
    "accent":      "#d4a043",   # برنز طلایی
    "accent2":     "#b8882e",
    "accent_glow": "#f0c060",

    # ─── رنگ‌های وضعیت ───
    "gold":        "#f0c060",   # طلایی درخشان
    "success":     "#4caf80",   # سبز روشن
    "warning":     "#f0c060",
    "danger":      "#e05050",

    # ─── متن ───
    "text":        "#f0f2f4",   # سفید گرم — خوانا روی طوسی
    "text_dim":    "#a0a8b0",   # طوسی روشن‌تر
    "text_bright": "#ffffff",

    # ─── حاشیه‌ها ───
    "border":      "#555c64",
    "border2":     "#606870",
    "border_hot":  "#d4a043",

    # ─── هدر: طوسی تیره‌تر ───
    "header_bg":   "#2e3338",
    "header_line": "#d4a043",

    # ─── تب‌ها ───
    "tab_active":  "#d4a043",
    "tab_inactive":"#484e56",

    # ─── ورودی‌ها ───
    "entry_bg":    "#353b41",

    # ─── دکمه‌ها ───
    "btn_primary": "#2a6090",
    "btn_hover":   "#3a70a0",
    "btn_danger":  "#a03030",
    "btn_success": "#2a7850",
    "btn_ghost":   "#484e56",

    # ─── هایلایت جستجو ───
    "highlight":   "#f0c060",
}

# ── تشخیص فونت نازنین ──
def _get_main_font():
    try:
        import tkinter.font as tkf_
        import tkinter as tk_
        _r = tk_.Tk(); _r.withdraw()
        avail = tkf_.families(_r)
        _r.destroy()
        # اول دقیقاً همین اسم که روی سیستم هست
        if "B Nazanin" in avail:
            return "B Nazanin"
        # بقیه حالت‌ها
        for fn in ("BNazanin","B_Nazanin","Nazanin","B NAZANIN"):
            if fn in avail:
                return fn
        # جستجوی فازی
        for fn in avail:
            if "azanin" in fn.lower():
                return fn
    except:
        pass
    return _MAIN_FONT
_MAIN_FONT = _get_main_font()
FONT_TITLE  = (_MAIN_FONT, 26, "bold")
FONT_HEAD   = (_MAIN_FONT, 14, "bold")
FONT_NORM   = (_MAIN_FONT, 12, "bold")
FONT_SMALL  = (_MAIN_FONT, 10, "bold")
FONT_MONO   = ("B Nazanin", 12, "bold")
FONT_ICON   = ("Segoe UI Symbol", 20, "bold")
FONT_LARGE  = (_MAIN_FONT, 14, "bold")

# ═══════════════════════════════════════════════════════════
#  متن فارسی — tr() برای یکدستی پیام‌ها
# ═══════════════════════════════════════════════════════════
_CURRENT_LANG = "fa"

_I18N = {
    "fa": {
        "online": "● آنلاین",
        "select_user": "⬇  کاربر خود را انتخاب کنید",
        "logout": "🚪  خروج از حساب",
        "logout_title": "خروج از حساب",
        "logout_msg": "آیا می‌خواهید از حساب کاربری خارج شوید؟",
        "close_pw_title": "🔒  رمز بستن برنامه",
        "close_pw_msg": "برای بستن برنامه، رمز عبور را وارد کنید:",
        "restore_pw_title": "🔒  رمز خروج از حالت تمام‌صفحه",
        "restore_pw_msg": "برای کوچک کردن پنجره، رمز عبور را وارد کنید:",
        "minimize_pw_title": "🔒  رمز مینیمایز",
        "minimize_pw_msg": "برای مینیمایز کردن برنامه، رمز عبور را وارد کنید:",
        "wrong_pw_title": "رمز اشتباه",
        "wrong_pw": "رمز واردشده صحیح نیست.",
        "save_exit_title": "💾  ذخیره و خروج",
        "save_exit_msg": "قبل از خروج، اطلاعات سیستم ذخیره و بک‌آپ‌گیری می‌شود.\n\nآیا ادامه می‌دهید؟",
        "save_err_title": "⚠️  خطا در ذخیره",
        "save_err_msg": "ذخیره‌سازی اطلاعات با خطا مواجه شد:\n{err}\n\nآیا همچنان می‌خواهید بدون ذخیره خارج شوید؟",
        "refresh": "  🔄  بروزرسانی  ",
        "save_btn": "  💾  ذخیره‌سازی  ",
        "confirm": "تایید",
        "cancel": "انصراف",
        "yes": "بله",
        "no": "خیر",
        "ok": "تایید",
        "app_title": "سامانه مدیریت تختال",
        "app_subtitle": "سازه پیشگام مدیسه  ·  فولاد سفید دشت",
        "login_subtitle": "شرکت سازه پیشگام مدیسه  ·  فولاد سفید دشت",
        "tab_home": "نمای کلی", "tab_melts": "ثبت ذوب جدید", "tab_qc": "کنترل کیفی",
        "tab_rejected": "تایید نشده", "tab_transfer": "موجودی انبار",
        "tab_lab": "تحویل باومن به آزمایشگاه", "tab_scrap": "قراضه",
        "tab_pdf": "گزارش‌گیری", "tab_scarf": "اسکارف", "tab_cut": "برش",
        "tab_nobat": "خروج اسلب", "tab_ticket": "صندوق نامه", "tab_admin": "مدیریت سیستم",
    },
}

_TAB_ICONS = {
    "home": "🏠", "melts": "🔥", "qc": "✅", "rejected": "⛔", "transfer": "🏭",
    "lab": "🧪", "scrap": "♻️", "pdf": "📊", "scarf": "⚙", "cut": "✂",
    "nobat": "🔄", "ticket": "✉", "admin": "👑",
}


def get_lang():
    return "fa"


def tr(key, **fmt):
    text = _I18N["fa"].get(key, key)
    if fmt:
        try:
            return text.format(**fmt)
        except Exception:
            pass
    return text


def _refresh_client_connection_label(win, *, force_check: bool = False) -> None:
    """Refresh online/offline badge — client reads cache only; admin uses short local health."""
    lbl = getattr(win, "_online_lbl", None)
    if lbl is None:
        return
    try:
        offline = False
        if os.environ.get("STF_CLIENT") == "1":
            # Never call is_server_alive on UI thread (force_check ignored for network).
            from client.db_bridge import is_offline_mode
            offline = bool(is_offline_mode())
        else:
            # Admin: cached is_lan_alive, or short 127.0.0.1 health (0.4s)
            try:
                offline = not bool(is_lan_alive(False))
            except Exception:
                api_ok = False
                try:
                    import urllib.request
                    with urllib.request.urlopen("http://127.0.0.1:8080/api/v1/health", timeout=0.4) as resp:
                        import json as _json
                        data = _json.loads(resp.read().decode("utf-8"))
                        api_ok = data.get("status") in ("ok", "degraded")
                except Exception:
                    api_ok = False
                offline = not api_ok
        if offline:
            text, fg = "● آفلاین", "#e07b39"
        else:
            text, fg = tr("online"), C["success"]
        # شیشه‌ای (Canvas) یا Label معمولی
        if hasattr(lbl, "_glass_text"):
            lbl._glass_fg = fg
            setter = getattr(win, "_glass_set_text", None)
            if callable(setter):
                setter(lbl, text)
            else:
                lbl._glass_text = text
        else:
            lbl.config(text=text, fg=fg)
    except Exception:
        pass


def tab_label(key):
    ic = _TAB_ICONS.get(key, "")
    name = tr(f"tab_{key}")
    return f"{ic}  {name}" if ic else name


def _cancel_pending_window_timers(win):
    for attr in ("_restore_after_id", "_cfg_debounce_id", "_configure_after_id"):
        tid = getattr(win, attr, None)
        if tid is not None:
            try:
                win.after_cancel(tid)
            except Exception:
                pass
            setattr(win, attr, None)


def _prepare_window_transition(win):
    """قبل از تعویض لاگین↔اصلی — بدون علامت بستن برنامه."""
    _cancel_pending_window_timers(win)
    win._restore_dialog_open = False
    win._minimize_dialog_open = False
    win._window_action_in_progress = True


def _prepare_app_shutdown(win):
    """قبل از بستن واقعی برنامه."""
    win._app_closing = True
    win._minimizing_authorized = False
    win._was_in_taskbar = False
    win._allow_iconify = False
    _cancel_pending_window_timers(win)
    win._restore_dialog_open = False
    win._minimize_dialog_open = False
    try:
        win.attributes("-topmost", False)
    except Exception:
        pass
    show_all_taskbars()


def _install_minimize_lifecycle_guards(win):
    """جلوگیری از بستن/withdraw/iconify تصادفی هنگام مینیمایز."""
    if getattr(win, "_minimize_lifecycle_guards", False):
        return
    win._minimize_lifecycle_guards = True
    win._real_destroy = win.destroy
    win._real_withdraw = win.withdraw
    win._real_iconify = win.iconify

    def _guarded_destroy():
        if getattr(win, "_app_closing", False):
            win._real_destroy()
            return
        if getattr(win, "_minimizing_authorized", False) or _hwnd_is_iconic(win):
            return
        win._real_destroy()

    def _guarded_withdraw():
        if getattr(win, "_app_closing", False):
            win._real_withdraw()
            return
        if getattr(win, "_minimizing_authorized", False) or _hwnd_is_iconic(win):
            return
        win._real_withdraw()

    def _guarded_iconify():
        if getattr(win, "_app_closing", False):
            win._real_iconify()
            return
        if (getattr(win, "_minimizing_authorized", False)
                or getattr(win, "_allow_iconify", False)):
            win._real_iconify()
            return

    win.destroy = _guarded_destroy
    win.withdraw = _guarded_withdraw
    win.iconify = _guarded_iconify


def _install_taskbar_restore_guard(win):
    """بازگرداندن تسک‌بار هنگام destroy — برای admin و client."""
    if getattr(win, "_taskbar_restore_bound", False):
        return
    win._taskbar_restore_bound = True

    def _on_final_destroy(event=None):
        if event is not None and event.widget is not win:
            return
        show_all_taskbars()

    win.bind("<Destroy>", _on_final_destroy, add="+")


def _install_destroy_guard(win):
    """سازگاری — همان محافظ چرخهٔ عمر."""
    _install_minimize_lifecycle_guards(win)
    _install_taskbar_restore_guard(win)


def _do_window_minimize(win):
    if getattr(win, "_app_closing", False):
        return
    do_window_minimize_safe(win)


def _do_window_restore_down(win):
    if getattr(win, "_app_closing", False):
        return
    do_window_restore_down(win)

# ═══════════════════════════════════════════════════════════
#  ابزارهای UI
# ═══════════════════════════════════════════════════════════

def styled_btn(parent, text, command, color=None, width=140, height=36, **kw):
    bg = color or C["btn_primary"]
    fg = C["text_bright"]
    hover_bg = _lighten(bg, 18)

    btn = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=fg,
        activebackground=hover_bg,
        activeforeground=C["text_bright"],
        font=(_MAIN_FONT, 11, "bold"), bd=0, relief="flat",
        cursor="hand2",
        padx=18, pady=10,
        **kw
    )
    def on_enter(e):
        btn.config(bg=hover_bg)
    def on_leave(e):
        btn.config(bg=bg)
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    return btn

def _lighten(hex_color, amt=20):
    h = hex_color.lstrip('#')
    try:
        r,g,b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
        r,g,b = min(255,r+amt), min(255,g+amt), min(255,b+amt)
        return f"#{r:02x}{g:02x}{b:02x}"
    except:
        return hex_color

def make_entry(parent, width=30, show=None):
    e = tk.Entry(
        parent, width=width,
        bg=C["entry_bg"], fg=C["text"],
        insertbackground=C["accent"],
        font=FONT_NORM, justify="right", bd=0, relief="flat",
        highlightthickness=1,
        highlightbackground=C["border"],
        highlightcolor=C["accent"],
        show=show or ""
    )
    return e

def make_combo(parent, values, width=28):
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("Dark.TCombobox",
        fieldbackground=C["entry_bg"],
        background=C["entry_bg"],
        foreground=C["text"],
        selectbackground=C["accent"],
        selectforeground=C["text_bright"],
        arrowcolor=C["accent"],
        bordercolor=C["border"],
        lightcolor=C["border"],
        darkcolor=C["border"],
        insertcolor=C["text"],
        padding=(8, 6),
    )
    style.map("Dark.TCombobox",
        fieldbackground=[
            ("readonly", C["entry_bg"]),
            ("focus",    C["entry_bg"]),
            ("!disabled",C["entry_bg"]),
        ],
        foreground=[
            ("readonly",  C["text"]),
            ("focus",     C["text"]),
            ("!disabled", C["text"]),
        ],
        selectbackground=[("readonly", C["accent"])],
        selectforeground=[("readonly", C["text_bright"])],
        background=[
            ("readonly", C["card"]),
            ("active",   C["card_hover"]),
        ],
        arrowcolor=[
            ("readonly", C["accent"]),
            ("active",   C["accent_glow"]),
        ],
    )
    cb = ttk.Combobox(parent, values=values, width=width,
                      font=(_MAIN_FONT, 11, "bold"), style="Dark.TCombobox",
                      state="readonly")
    def fix_fg(e=None):
        cb.configure(foreground=C["text"])
    cb.bind("<<ComboboxSelected>>", fix_fg)
    return cb

def make_label(parent, text, font=None, fg=None, **kw):
    return tk.Label(parent, text=text,
                    bg=kw.pop("bg", C["card"]),
                    fg=fg or C["text"],
                    font=font or FONT_NORM, **kw)

def separator(parent, color=None):
    f = tk.Frame(parent, bg=color or C["border"], height=1)
    f.pack(fill="x", padx=10, pady=4)
    return f

def card_frame(parent, **kw):
    f = tk.Frame(parent, bg=C["card"],
                 highlightthickness=1,
                 highlightbackground=C["border"], **kw)
    return f

_TREE_STYLE_CONFIGURED = False  # جلوگیری از configure مکرر


def _tree_slab_col(tree):
    """ستون شماره اسلب در Treeview (در صورت وجود)."""
    try:
        cols = list(tree["columns"])
    except Exception:
        return None
    for c in ("slab_id", "sid", "id"):
        if c in cols:
            return c
    return cols[0] if cols else None


def _tree_row_sid(tree, iid):
    """شماره اسلب یک ردیف — مستقل از بقیهٔ مقادیر (که با ویرایش عوض می‌شوند)."""
    if not iid:
        return None
    try:
        col = _tree_slab_col(tree)
        if col:
            raw = str(tree.set(iid, col) or "")
        else:
            vals = tree.item(iid, "values")
            raw = str(vals[0] if vals else "")
        sid = raw.replace("↳", "").split()[0].strip()
        if sid and sid not in ("—", "-", "–", "○", "☑", "🚚"):
            return sid
    except Exception:
        pass
    return None


def _tree_capture_view(tree, force=False):
    """فقط کسر اسکرول را ذخیره کن — قبل از هر بازسازی."""
    if getattr(tree, "_stf_scroll_freeze", False) and not force:
        return getattr(tree, "_stf_locked_yview", None), getattr(tree, "_stf_locked_sid", None)
    yv = None
    try:
        yv = float(tree.yview()[0])
    except Exception:
        yv = getattr(tree, "_stf_user_yview", None)
    sid = None
    try:
        sel = tree.selection()
        if sel:
            sid = _tree_row_sid(tree, sel[0])
    except Exception:
        pass
    tree._stf_saved_yview = yv
    tree._stf_saved_sid = sid
    tree._stf_locked_yview = yv
    tree._stf_locked_sid = sid
    if yv is not None:
        tree._stf_user_yview = yv
    return yv, sid


def _tree_begin_rebuild(tree):
    """قفل اسکرول — مقدار را یک‌بار بگیر و دیگر عوض نکن."""
    if getattr(tree, "_stf_scroll_freeze", False):
        return
    yv, sid = _tree_capture_view(tree, force=True)
    tree._stf_scroll_freeze = True
    tree._stf_locked_yview = yv
    tree._stf_locked_sid = sid


def _tree_force_yview(tree):
    """برگرداندن اسکرول به همان کسر — بدون see و بدون انتخاب."""
    try:
        if not tree.winfo_exists():
            return
    except Exception:
        return
    yv = getattr(tree, "_stf_locked_yview", None)
    if yv is None:
        yv = getattr(tree, "_stf_saved_yview", None)
    if yv is None:
        return
    try:
        tree.yview_moveto(float(yv))
    except Exception:
        pass


def _tree_end_rebuild(tree, sid=None):
    """یک‌بار اسکرول را برگردان و قفل را باز کن — بدون afterهای تکراری (عامل پرش)."""
    yv = getattr(tree, "_stf_locked_yview", None)
    if yv is None:
        yv = getattr(tree, "_stf_saved_yview", None)
    tree._stf_saved_yview = yv
    if yv is not None:
        try:
            tree.yview_moveto(float(yv))
        except Exception:
            pass
        tree._stf_user_yview = float(yv)
    tree._stf_scroll_freeze = False
    # فقط یک‌بار بعد از رسم — نه چند بار پشت‌سرهم
    try:
        tree.after_idle(lambda t=tree: _tree_force_yview(t))
    except Exception:
        pass


def _tree_restore_view(tree, yv=None, sid=None, attempt=0):
    """سازگاری — فقط yview، بدون see و بدون retry پشت‌سرهم."""
    if yv is not None:
        tree._stf_locked_yview = yv
        tree._stf_saved_yview = yv
    _tree_force_yview(tree)


def scrolled_tree(parent, columns, headings, height=14):
    """
    جدول اسکرول‌دار — ستون‌ها از راست به چپ نمایش داده می‌شوند (RTL)
    کلیک روی هر سرستون → مرتب‌سازی صعودی/نزولی
    """
    global _TREE_STYLE_CONFIGURED
    frame = tk.Frame(parent, bg=C["card"])
    style = ttk.Style()
    if not _TREE_STYLE_CONFIGURED:
        style.theme_use('clam')
        _TREE_STYLE_CONFIGURED = True
    style.configure("Dark.Treeview",
        background=C["card2"],
        foreground=C["text"],
        fieldbackground=C["card2"],
        rowheight=32,
        font=(_MAIN_FONT, 11, "bold"),
        bordercolor=C["border"],
        relief="flat"
    )
    style.configure("Dark.Treeview.Heading",
        background="#2e3338",
        foreground=C["accent"],
        font=(_MAIN_FONT, 11, "bold"),
        relief="groove",
        bordercolor=C["border"],
    )
    style.map("Dark.Treeview.Heading",
        background=[("active", C["accent2"]), ("pressed", C["accent2"]), ("!active", "#1a2d40")],
        foreground=[("active", "#ffffff"),    ("pressed", "#ffffff"),    ("!active", C["accent"])],
        relief=[("active","groove"),("pressed","groove")]
    )
    style.map("Dark.Treeview",
        background=[("selected", "#d4a043")],
        foreground=[("selected", "#2e3338")]
    )
    # ── ساخت درخت با ستون‌های اصلی ──
    tree = ttk.Treeview(frame, columns=columns, show="headings",
                        height=height, style="Dark.Treeview")

    # ── مرتب‌سازی با کلیک روی سرستون ──
    _sort_state = {}  # col → bool (ascending)

    def _num_key(v):
        cleaned = normalize_digits(v).replace("/", "").replace(":", "").replace(" ", "")
        return float(cleaned) if cleaned else 0.0

    def _strip_arrow(text):
        return text.rstrip(" ▲▼")

    def _apply_sort(col, asc):
        """مرتب‌سازی واقعی ستون — بدون تغییر وضعیت صعودی/نزولی (برای استفاده‌ی مجدد بعد از رفرش).
        روی لیست زندهٔ ستون‌های tree کار می‌کند، نه تاپل اولیه — چون بعضی تب‌ها
        (مثل کنترل کیفی) تعداد ستون‌ها را به‌صورت داینامیک در هر رفرش تغییر می‌دهند."""
        if col not in tree["columns"]:
            return
        items = [(tree.set(iid, col), iid) for iid in tree.get_children("")]
        try:
            # مرتب‌سازی عددی: صعودی = کم→زیاد
            items.sort(key=lambda x: _num_key(x[0]), reverse=not asc)
        except Exception:
            # مرتب‌سازی الفبایی: صعودی = A→Z (بدون توجه به بزرگی/کوچکی حروف)
            items.sort(key=lambda x: x[0].casefold(), reverse=not asc)
        for i, (_, iid) in enumerate(items):
            tree.move(iid, "", i)
        # نمایش فلش فقط روی ستونی که مرتب‌سازی شده — بر اساس متن فعلی هر سرستون
        for c in tree["columns"]:
            base = _strip_arrow(tree.heading(c, "text"))
            if c == col:
                tree.heading(c, text=base + (" ▲" if asc else " ▼"))
            else:
                tree.heading(c, text=base)

    def _sort_by_col(col):
        # کلیک اول → صعودی (▲)، کلیک دوم → نزولی (▼) و الی آخر
        asc = not _sort_state.get(col, False)
        _sort_state[col] = asc
        _apply_sort(col, asc)
        # ثبت آخرین مرتب‌سازی روی خود tree تا بعد از هر رفرش/بازسازی ردیف‌ها
        # (مثلاً در تب‌هایی که هر چند ثانیه خودکار به‌روزرسانی می‌شوند) دوباره اعمال شود
        tree._last_sort_action = lambda c=col, a=asc: _apply_sort(c, a)

    def _bind_sort_commands():
        """به همه‌ی ستون‌های فعلیِ tree، دستور مرتب‌سازی با کلیک را وصل می‌کند"""
        for c in tree["columns"]:
            tree.heading(c, command=lambda col=c: _sort_by_col(col))

    for col, head in zip(columns, headings):
        tree.heading(col, text=head, anchor="center",
                     command=lambda c=col: _sort_by_col(c))
        tree.column(col, width=120, anchor="center")

    # ── حفظ کلیک‌پذیریِ سرستون‌ها وقتی ستون‌ها به‌صورت داینامیک بازسازی می‌شوند ──
    # برخی تب‌ها (مثل کنترل کیفی) هر بار رفرش، tree.configure(columns=...) را با
    # تعداد ستون متفاوت صدا می‌زنند. در ttk، این کار دستور (command) کلیک روی همه‌ی
    # سرستون‌ها را پاک می‌کند. این پچ، بعد از هر چنین تغییری، دوباره دستور مرتب‌سازی
    # را به همه‌ی سرستون‌های فعلی وصل می‌کند.
    _orig_configure = tree.configure

    def _configure_keep_sort(cnf=None, **kw):
        res = _orig_configure(cnf, **kw)
        changed_cols = bool(cnf and "columns" in cnf) or ("columns" in kw)
        if changed_cols:
            _bind_sort_commands()
        return res

    tree.configure = _configure_keep_sort
    tree.config = _configure_keep_sort

    # ── حفظ مرتب‌سازی و موقعیت اسکرول پس از رفرش/بازسازی خودکار ردیف‌ها ──
    # بسیاری از تب‌ها (مثل کنترل کیفی، رد شده، باومن و ...) هر چند ثانیه یک‌بار
    # tree.delete(*tree.get_children()) را صدا می‌زنند و دوباره ردیف‌ها را insert می‌کنند؛
    # این کار قبلاً هم مرتب‌سازی دستی کاربر و هم موقعیت اسکرول را از بین می‌برد و باعث
    # می‌شد جدول هر چند ثانیه «بپرد» (برود بالا و دوباره به جای قبلی برگردد).
    # با این پچ، وقتی این بازسازیِ کامل اتفاق می‌افتد، موقعیت اسکرول و آخرین
    # مرتب‌سازی ذخیره شده و بعد از پر شدن دوباره‌ی جدول، فوراً بازگردانده می‌شود.
    tree._last_sort_action = None
    tree._default_sort = None  # بعداً با apply_date_time_sort پر می‌شود
    _orig_delete = tree.delete

    def _delete_and_keep_sort(*iids):
        children_before = tree.get_children("")
        is_full_clear = (not iids) or (set(iids) >= set(children_before))
        if is_full_clear:
            _tree_begin_rebuild(tree)
        _orig_delete(*iids)
        if is_full_clear:
            def _reapply():
                try:
                    if not tree.winfo_exists():
                        return
                    # سورت فقط اگر هنوز قفل است (وسط رفرش) — یک‌بار، بعد اسکرول
                    action = getattr(tree, "_last_sort_action", None) or getattr(tree, "_default_sort", None)
                    if action and getattr(tree, "_stf_scroll_freeze", False):
                        action()
                    _tree_end_rebuild(tree)
                except Exception:
                    try:
                        tree._stf_scroll_freeze = False
                    except Exception:
                        pass
            try:
                tree.after_idle(_reapply)
            except Exception:
                pass

    tree.delete = _delete_and_keep_sort

    # ── RTL: معکوس کردن ترتیب نمایش ستون‌ها ──
    tree["displaycolumns"] = list(reversed(columns))

    vsb = tk.Scrollbar(frame, orient="vertical", command=tree.yview,
                       bg="#707070", troughcolor="#1a1a1a", width=16,
                       activebackground=C["accent"])
    hsb = tk.Scrollbar(frame, orient="horizontal", command=tree.xview,
                       bg="#707070", troughcolor="#1a1a1a", width=16,
                       activebackground=C["accent"])

    def _yscroll_set(*args):
        try:
            vsb.set(*args)
        except Exception:
            pass
        # وسط rebuild اسکرول موقتاً ته می‌رود — مقدار کاربر را خراب نکن
        if getattr(tree, "_stf_scroll_freeze", False):
            return
        try:
            tree._stf_user_yview = float(tree.yview()[0])
            top = tree.identify_row(8) or tree.identify_row(2)
            if top:
                sid = _tree_row_sid(tree, top)
                if sid:
                    tree._stf_user_sid = sid
        except Exception:
            pass

    tree.configure(yscrollcommand=_yscroll_set, xscrollcommand=hsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    # اسکرول ماوس روی جدول بدون نیاز به کلیک قبلی
    register_scroll_canvas(tree, frame, vsb)

    # مرتب‌سازی پیش‌فرض: تاریخ سپس ساعت — جدیدترین اول
    # (اگر sort_toolbar وصل شود، همان را به‌عنوان _last_sort_action تنظیم می‌کند)
    def _default_sort_fn():
        apply_date_time_sort(tree)

    tree._default_sort = _default_sort_fn

    def _boot_default_sort(attempt=0):
        try:
            if not tree.winfo_exists():
                return
            # اگر نوار مرتب‌سازی قبلاً اکشن گذاشته، همان اعمال می‌شود
            action = getattr(tree, "_last_sort_action", None) or tree._default_sort
            if tree.get_children(""):
                if getattr(tree, "_last_sort_action", None) is None:
                    tree._last_sort_action = tree._default_sort
                action()
            elif attempt < 30:
                tree.after(100, lambda: _boot_default_sort(attempt + 1))
        except Exception:
            pass

    tree.after(150, lambda: _boot_default_sort(0))
    return frame, tree


def _find_date_time_cols(tree, date_col=None):
    """یافتن ستون تاریخ و ستون ساعت در یک Treeview"""
    cols = list(tree["columns"])
    if date_col and date_col in cols:
        dc = date_col
        tc = None
        for c in cols:
            cl = c.lower()
            if any(k in cl for k in ("time", "ساعت", "reg_time", "qc_time", "del_time", "log_time", "ret_time")):
                tc = c
                break
        return dc, tc
    date_c = time_c = None
    date_keywords = ("date", "تاریخ", "reg_date", "qc_date", "del_date", "log_date", "ret_date")
    time_keywords = ("time", "ساعت", "reg_time", "qc_time", "del_time", "log_time", "ret_time")
    for c in cols:
        cl = c.lower()
        if date_c is None and any(k in cl for k in date_keywords):
            date_c = c
        if time_c is None and any(k in cl for k in time_keywords):
            time_c = c
    if date_c is None:
        for c in cols:
            cl = c.lower()
            if any(k in cl for k in ("at", "registered", "delivered", "transferred", "returned")):
                date_c = c
                break
    return date_c, time_c


def apply_date_time_sort(tree, date_col=None, reverse=True):
    """مرتب‌سازی پیش‌فرض کل نرم‌افزار: تاریخ (اول) + ساعت (دوم) — جدیدترین در سطر اول."""
    date_c, time_c = _find_date_time_cols(tree, date_col=date_col)
    if date_c is None:
        return False
    try:
        # فقط از مقدار قفل/ذخیره‌شده — هرگز yview زنده وسط rebuild (که ته صفحه است)
        saved_y = getattr(tree, "_stf_locked_yview", None)
        if saved_y is None:
            saved_y = getattr(tree, "_stf_saved_yview", None)
        if saved_y is None and not getattr(tree, "_stf_scroll_freeze", False):
            try:
                saved_y = tree.yview()[0]
            except Exception:
                saved_y = None
        saved_sid = (
            getattr(tree, "_stf_locked_sid", None)
            or getattr(tree, "_stf_saved_sid", None)
        )
        def _key(iid):
            d = normalize_digits(str(tree.set(iid, date_c)).strip())
            t = normalize_digits(str(tree.set(iid, time_c)).strip()) if time_c else ""
            if not d or d in ("—", "-", "–"):
                return ""
            if t and t not in ("—", "-", "–", ""):
                return f"{d}  {t}"
            return d
        items = [(iid, _key(iid)) for iid in tree.get_children("")]
        items.sort(key=lambda x: x[1], reverse=reverse)
        for i, (iid, _) in enumerate(items):
            tree.move(iid, "", i)
        kids = list(tree.get_children(""))
        if len(kids) > TREE_ROW_LIMIT:
            for iid in kids[TREE_ROW_LIMIT:]:
                try:
                    tree.detach(iid)
                    tree.delete(iid)
                except Exception:
                    pass
        if saved_y is not None:
            try:
                tree.yview_moveto(float(saved_y))
            except Exception:
                pass
        if saved_sid:
            try:
                for iid in tree.get_children(""):
                    if _tree_row_sid(tree, iid) == saved_sid:
                        tree.selection_set(iid)
                        tree.focus(iid)
                        break
            except Exception:
                pass
        if saved_y is not None:
            try:
                tree.yview_moveto(float(saved_y))
            except Exception:
                pass
        return True
    except Exception:
        return False


def search_bar(parent, tree, col_indices=None):
    frm = tk.Frame(parent, bg=C["card"])
    tk.Label(frm, text="🔍  جستجو:", bg=C["card"], fg=C["accent"],
             font=FONT_NORM).pack(side="right", padx=(0,4))
    sv = tk.StringVar()
    ent = tk.Entry(frm, textvariable=sv, bg=C["entry_bg"], fg=C["text"],
                   insertbackground=C["accent"], font=FONT_NORM, bd=0,
                   relief="flat", highlightthickness=1,
                   highlightbackground=C["border"],
                   highlightcolor=C["accent"], width=22)
    ent.pack(side="right", padx=2)

    def do_search(*_):
        q = sv.get().strip().lower()
        tree.tag_configure("found", background="#1a3a00", foreground=C["highlight"])
        tree.tag_configure("normal", background=C["card2"], foreground=C["text"])
        found_first = None
        for iid in tree.get_children():
            vals = [str(v).lower() for v in tree.item(iid, "values")]
            cols = col_indices if col_indices else range(len(vals))
            hit = q and any(q in vals[i] for i in cols if i < len(vals))
            tree.item(iid, tags=("found" if hit else "normal",))
            if hit and not found_first:
                found_first = iid
        if found_first:
            tree.see(found_first)
            tree.selection_set(found_first)

    sv.trace_add("write", do_search)
    return frm


def sort_toolbar(parent, tree,
                 slab_col="slab_id",
                 date_col=None,
                 bg=None):
    """
    نوار مرتب‌سازی — یک دکمه:
      مرتب‌سازی بر اساس تاریخ و ساعت (جدید→قدیم)
    پیش‌فرض همه جداول همین است: جدیدترین ثبت در سطر اول.
    """
    BG = bg or C["panel"]
    frm = tk.Frame(parent, bg=BG)

    def sort_by_date():
        apply_date_time_sort(tree, date_col=date_col, reverse=True)
        tree._last_sort_action = sort_by_date

    # پیش‌فرض کل نرم‌افزار: تاریخ+ساعت، جدیدترین اول
    tree._default_sort = sort_by_date
    if getattr(tree, "_last_sort_action", None) is None:
        tree._last_sort_action = sort_by_date

    def _auto_default_sort(attempt=0):
        try:
            if not tree.winfo_exists():
                return
            if tree.get_children(""):
                sort_by_date()
            elif attempt < 25:
                tree.after(120, lambda: _auto_default_sort(attempt + 1))
        except Exception:
            pass

    tree.after(80, lambda: _auto_default_sort(0))

    _btn_style = dict(
        bg=C["card"], fg=C["accent"],
        font=(_MAIN_FONT, 9, "bold"),
        bd=0, relief="flat", cursor="hand2",
        padx=10, pady=4,
        activebackground=C["card_hover"],
        activeforeground=C["accent_glow"],
    )

    tk.Label(frm, text="مرتب‌سازی:", bg=BG,
             fg=C["text_dim"], font=FONT_SMALL).pack(side="right", padx=(0,4))

    btn_date = tk.Button(frm, text="مرتب‌سازی بر اساس تاریخ و ساعت",
                          command=sort_by_date, **_btn_style)
    btn_date.pack(side="right", padx=2)

    for btn in (btn_date,):
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg=C["card_hover"]))
        btn.bind("<Leave>", lambda e, b=btn: b.config(bg=C["card"]))

    return frm

# ═══════════════════════════════════════════════════════════
#  اعتبارسنجی شماره اسلب
# ═══════════════════════════════════════════════════════════
def validate_slab_id(sid):
    sid = normalize_digits(str(sid).strip())
    if not re.fullmatch(r'\d{11}', sid):
        return False, "شماره اسلب باید دقیقاً ۱۱ رقم عددی باشد.", ""
    return True, "", sid


# اسکارف/برش فقط در مرحلهٔ ثبت ذوب («ثبت شده») مجاز است — نه بعد از تأیید QC
SCARF_CUT_QC_BLOCK_MSG = (
    "این اسلب قبلاً به تأیید کنترل کیفی رسیده و نمی‌توانید اصلاح اسکارف و برش ثبت کنید. "
    "این موضوع را با سرپرست در جریان بگذارید."
)
_MELT_STAGE_QC_STATUSES = frozenset({"ثبت شده", ""})


def melt_qc_status(melt) -> str:
    if not isinstance(melt, dict):
        return ""
    return str(melt.get("qc_status") or "").strip()


def melt_allows_scarf_cut(melt) -> bool:
    """فقط ذوب‌هایی که هنوز در مرحلهٔ ثبت ذوب‌اند (تایید QC نشده)."""
    st = melt_qc_status(melt)
    return st in _MELT_STAGE_QC_STATUSES


def find_melt_by_slab(db, slab_id):
    sid = str(slab_id or "").strip()
    for m in (db or {}).get("melts") or []:
        if str(m.get("slab_id") or "").strip() == sid:
            return m
    return None


def assert_scarf_cut_allowed(db, slab_id, *, parent=None) -> bool:
    """اگر اسلب از مرحلهٔ ثبت ذوب گذشته باشد، خطا نشان بده و False برگردان."""
    melt = find_melt_by_slab(db, slab_id)
    if melt is None:
        return True  # نبودن ذوب جداگانه در مسیر ثبت چک می‌شود
    if melt_allows_scarf_cut(melt):
        return True
    try:
        messagebox.showerror("مجاز نیست", SCARF_CUT_QC_BLOCK_MSG, parent=parent)
    except Exception:
        messagebox.showerror("مجاز نیست", SCARF_CUT_QC_BLOCK_MSG)
    return False


def _flatten(widget):
    """همه فرزندان یک ویجت را به صورت تخت برمی‌گرداند"""
    result = []
    for ch in widget.winfo_children():
        result.append(ch)
        result.extend(_flatten(ch))
    return result

def check_duplicate(db, section_key, slab_id):
    for rec in db.get(section_key, []):
        if rec.get("slab_id") == slab_id:
            return True
    return False

# ─── تعداد برگشت‌های یک اسلب ───
def count_returns_for_slab(db, slab_id):
    """تعداد دفعاتی که این اسلب به انبار داخلی برگشته"""
    return sum(1 for r in db.get("return_log", []) if r.get("slab_id") == slab_id)

def return_ordinal_fa(n):
    """اول، دوم، سوم، ..."""
    words = ["اول", "دوم", "سوم", "چهارم", "پنجم", "ششم", "هفتم", "هشتم", "نهم", "دهم"]
    if 1 <= n <= len(words):
        return words[n-1]
    return f"{n}ام"

# ─── سه مکان نگهداری اسلب و مکان فعلی هر اسلب ───
WAREHOUSE_LOCATIONS = ["انبار داخلی", "انبار روباز ۱", "انبار روباز ۲"]

def _normalize_warehouse_name(loc):
    """نام انبار را به یکی از WAREHOUSE_LOCATIONS نرمال کن (رقم فارسی/عربی/لاتین)."""
    if not loc:
        return "انبار داخلی"
    s = str(loc).strip()
    for wh in WAREHOUSE_LOCATIONS:
        if wh == s:
            return wh
    # ارقام لاتین/عربی → فارسی برای تطبیق
    try:
        s2 = normalize_digits(s)
    except Exception:
        s2 = s
    for wh in WAREHOUSE_LOCATIONS:
        if wh == s2 or normalize_digits(wh) == s2:
            return wh
    if "داخلی" in s:
        return "انبار داخلی"
    # رقم ۲ (فارسی/عربی/لاتین)
    if any(ch in s for ch in ("۲", "٢", "2")):
        return "انبار روباز ۲"
    if any(ch in s for ch in ("۱", "١", "1")):
        return "انبار روباز ۱"
    return "انبار داخلی"


def get_current_location(db, sid):
    """مکان فعلی اسلب — movement_log جدیدتر از transfers_out اولویت دارد."""
    # ۱) آخرین انتقال واقعی در لاگ حرکت (منبع حقیقت پس از انتقال)
    moves = [
        m for m in (db.get("movement_log") or [])
        if isinstance(m, dict)
        and m.get("slab_id") == sid
        and m.get("operation", "انتقال") == "انتقال"
        and m.get("to")
    ]
    if moves:
        latest = max(moves, key=lambda m: str(m.get("at") or ""))
        return _normalize_warehouse_name(latest.get("to"))

    # ۲) رکورد transfers_out (ترجیح جدیدترین transferred_at در صورت تکراری)
    trs = [
        t for t in (db.get("transfers_out") or [])
        if isinstance(t, dict) and t.get("slab_id") == sid
    ]
    if not trs:
        melt = next(
            (r for r in (db.get("melts") or [])
             if isinstance(r, dict) and r.get("slab_id") == sid),
            None,
        )
        if melt:
            return _normalize_warehouse_name(
                melt.get("location") or melt.get("rej_location") or "انبار داخلی"
            )
        return "انبار داخلی"
    tr = max(
        trs,
        key=lambda t: str(
            t.get("updated_at") or t.get("transferred_at") or t.get("at") or ""
        ),
    )
    loc = tr.get("current_location") or tr.get("destination") or tr.get("to") or tr.get("location")
    return _normalize_warehouse_name(loc)

def other_locations(cur_loc):
    """دو مکان دیگر غیر از مکان فعلی (برای پر کردن کشویی‌ها)"""
    return [l for l in WAREHOUSE_LOCATIONS if l != cur_loc]


def open_slab_transfer_popup(app, sid, cur, *, status_lbl=None, on_done=None, source="تب انتقال"):
    """پنجره انتقال اسلب — همان منطق تب کنترل کیفی (قابل استفاده در ردشده/قراضه)."""
    edit_key = f"transfer:{sid}"
    if _acquire_edit_popup(app, edit_key):
        return

    opts = [w for w in WAREHOUSE_LOCATIONS if w != cur]
    if not opts:
        _release_edit_popup_claim(app, edit_key)
        messagebox.showinfo("انتقال", "مقصد دیگری موجود نیست.", parent=app)
        return

    pop = tk.Toplevel(app)
    prepare_popup_window(pop, app)
    _register_edit_popup(app, edit_key, pop)
    pop.title(f"انتقال اسلب  {sid}")
    pop.configure(bg=C["card"])
    pop.resizable(False, False)
    pop.focus_force()

    is_from_outside = cur != "انبار داخلی"
    pw = 460
    ph = 360 if is_from_outside else 260
    app.update_idletasks()
    sx = app.winfo_screenwidth(); sy = app.winfo_screenheight()
    pop.geometry(f"{pw}x{ph}+{(sx-pw)//2}+{(sy-ph)//2}")

    tk.Frame(pop, bg=C["accent"], height=3).pack(fill="x")
    hf = tk.Frame(pop, bg=C["header_bg"])
    hf.pack(fill="x")
    tk.Label(hf, text="🚚  انتقال اسلب", bg=C["header_bg"],
             fg=C["accent"], font=FONT_HEAD).pack(side="right", padx=16, pady=10)
    tk.Label(hf, text=f"شماره: {sid}", bg=C["header_bg"],
             fg=C["text_dim"], font=FONT_SMALL).pack(side="left", padx=16)

    body = tk.Frame(pop, bg=C["card"])
    body.pack(fill="both", expand=True, padx=20, pady=12)

    r0 = tk.Frame(body, bg=C["card"]); r0.pack(fill="x", pady=4)
    tk.Label(r0, text="مکان فعلی:", bg=C["card"], fg=C["text_dim"],
             font=FONT_NORM, width=14, anchor="e").pack(side="right")
    tk.Label(r0, text=cur, bg=C["card"], fg=C["warning"],
             font=(_MAIN_FONT,12,"bold")).pack(side="right", padx=10)

    tk.Frame(body, bg=C["border"], height=1).pack(fill="x", pady=6)

    r1 = tk.Frame(body, bg=C["card"]); r1.pack(fill="x", pady=4)
    tk.Label(r1, text="انتقال به:", bg=C["card"], fg=C["text"],
             font=FONT_NORM, width=14, anchor="e").pack(side="right")
    dest_cb = make_combo(r1, opts, width=20)
    dest_cb.set(opts[0])
    dest_cb.pack(side="right", padx=10)

    reason_frame = tk.Frame(body, bg=C["card"])
    reason_var = tk.StringVar(value="")
    detail_var = tk.StringVar()
    detail_frame = tk.Frame(body, bg=C["card"])
    REASON_OPTS = ["مسائل مربوط به کنترل کیفی", "موارد دیگر"]

    def on_dest_change(e=None):
        dest = dest_cb.get()
        if dest == "انبار داخلی":
            reason_frame.pack(fill="x", pady=4)
            if reason_var.get() == "موارد دیگر":
                detail_frame.pack(fill="x", pady=4)
            pop.geometry(f"{pw}x{ph}+{(sx-pw)//2}+{(sy-ph)//2}")
        else:
            reason_frame.pack_forget()
            detail_frame.pack_forget()
            pop.geometry(f"{pw}x260+{(sx-pw)//2}+{(sy-ph)//2}")

    def on_reason_change(e=None):
        if reason_var.get() == "موارد دیگر":
            detail_frame.pack(fill="x", pady=4)
        else:
            detail_frame.pack_forget()
            detail_var.set("")

    dest_cb.bind("<<ComboboxSelected>>", on_dest_change)

    tk.Label(reason_frame, text="دلیل انتقال به داخلی:", bg=C["card"],
             fg=C["danger"], font=FONT_NORM, width=14, anchor="e").pack(side="right")
    reason_cb = make_combo(reason_frame, REASON_OPTS, width=24)
    reason_cb.set(REASON_OPTS[0])
    reason_var.set(REASON_OPTS[0])
    reason_cb.pack(side="right", padx=10)
    reason_cb.bind("<<ComboboxSelected>>",
                    lambda e: (reason_var.set(reason_cb.get()), on_reason_change()))

    tk.Label(detail_frame, text="توضیحات:", bg=C["card"],
             fg=C["danger"], font=FONT_NORM, width=14, anchor="e").pack(side="right")
    detail_ent = tk.Entry(detail_frame, textvariable=detail_var,
                           bg=C["entry_bg"], fg=C["text"],
                           insertbackground=C["accent"],
                           font=FONT_NORM, bd=0, relief="flat",
                           highlightthickness=1,
                           highlightbackground=C["danger"],
                           highlightcolor=C["danger"], width=24)
    detail_ent.pack(side="right", padx=10, ipady=4)
    on_dest_change()

    tk.Frame(body, bg=C["border"], height=1).pack(fill="x", pady=8)
    btn_row = tk.Frame(body, bg=C["card"])
    btn_row.pack(fill="x")
    tk.Button(btn_row, text="انصراف", command=pop.destroy,
              bg=C["card2"], fg=C["text"], font=(_MAIN_FONT,11, "bold"),
              bd=0, relief="flat", cursor="hand2", padx=16, pady=8
              ).pack(side="left")

    def do_transfer():
        new_dest = dest_cb.get()
        reason = reason_cb.get() if new_dest == "انبار داخلی" else ""
        reason_detail = detail_var.get()
        if not new_dest or new_dest == "—":
            messagebox.showwarning("خطا", "مقصد انتخاب نشده", parent=pop); return
        if cur == new_dest:
            messagebox.showwarning("خطا", f"اسلب هم‌اکنون در {new_dest} است", parent=pop); return
        if new_dest == "انبار داخلی" and not reason:
            messagebox.showwarning("خطا", "لطفاً دلیل انتقال به انبار داخلی را انتخاب کنید.", parent=pop)
            return
        if new_dest == "انبار داخلی" and reason == "موارد دیگر" and not reason_detail.strip():
            messagebox.showwarning("خطا", "لطفاً توضیحات «موارد دیگر» را بنویسید.", parent=pop)
            return
        final_reason = ""
        if new_dest == "انبار داخلی":
            final_reason = reason_detail.strip() if reason == "موارد دیگر" else reason
        if not messagebox.askyesno("تأیید انتقال",
                f"اسلب: {sid}\nاز: {cur}\nبه: {new_dest}"
                + (f"\nدلیل: {final_reason}" if final_reason else "")
                + "\n\nآیا مطمئن هستید؟",
                parent=pop):
            return
        db = load_db(); ts = now_str()
        prev = [m for m in db.get("movement_log",[]) if m.get("slab_id")==sid]
        db.setdefault("movement_log",[]).append({
            "slab_id":     sid,
            "move_number": len(prev)+1,
            "operation":   "انتقال",
            "from":        cur,
            "to":          new_dest,
            "reason":      final_reason,
            "by":          app.username,
            "at":          ts,
        })
        tr = next((r for r in db.get("transfers_out",[]) if r["slab_id"]==sid), None)
        if tr:
            tr["current_location"] = new_dest
            tr["destination"]      = new_dest
            tr["location"]         = new_dest
            tr["to"]               = new_dest
            tr["transferred_by"]   = app.username
            tr["transferred_at"]   = ts
            tr["updated_at"]       = ts
            tr["at"]               = ts
        else:
            db.setdefault("transfers_out",[]).append({
                "slab_id":          sid,
                "destination":      new_dest,
                "current_location": new_dest,
                "location":         new_dest,
                "to":               new_dest,
                "transferred_by":   app.username,
                "transferred_at":   ts,
                "updated_at":       ts,
                "at":               ts,
                "source":           source,
            })
        # همگام‌سازی محل در رکورد ذوب (برای تب رد شده)
        melt = next((r for r in db.get("melts",[]) if r.get("slab_id")==sid), None)
        if melt is not None:
            melt["location"] = new_dest
            melt["rej_location"] = new_dest
            melt["updated_at"] = ts
        save_db(db)
        reason_txt = f"  |  دلیل: {final_reason}" if final_reason else ""
        if status_lbl is not None:
            try:
                status_lbl.config(
                    text=f"✔  انتقال اسلب {sid}  از {cur}  به {new_dest}{reason_txt}  |  {ts}",
                    fg=C["success"])
            except Exception:
                pass
        pop.destroy()
        if on_done:
            try: on_done()
            except Exception: pass

    styled_btn(btn_row, "✔  ثبت انتقال", do_transfer,
               color=C["btn_success"]).pack(side="right")
    pop.bind("<Escape>", lambda e: pop.destroy())


# ═══════════════════════════════════════════════════════════
#  پنجره اصلی – لاگین
# ═══════════════════════════════════════════════════════════

def get_slab_full_info(db, sid):
    """
    اطلاعات کامل یک اسلب را از همه بخش‌های دیتابیس جمع‌آوری می‌کند.
    برای گزارش Excel ادمین و PDF شیفت.
    """
    melt = next((r for r in db["melts"] if r["slab_id"]==sid), {})
    sc_recs = [r for r in db["scarf_cut"] if r["slab_id"]==sid]
    scarf_r = next((r for r in sc_recs if r.get("operation")=="اسکارفی"), None)
    cut_r   = next((r for r in sc_recs if r.get("operation")=="برشی"), None)
    bauman_r = cut_r if (cut_r and cut_r.get("bauman_done")) else None
    lab_r    = next((r for r in db["lab_deliveries"] if r["slab_id"]==sid), None)
    transfer_r = next((r for r in db["transfers_out"] if r["slab_id"]==sid), None)
    scrap_r  = next((r for r in db["scrap"] if r["slab_id"]==sid), None)
    returns  = [r for r in db.get("return_log",[]) if r.get("slab_id")==sid]
    nobat_r  = melt  # exit_status در همان رکورد ذوب

    # محل فعلی — همیشه نام دقیق انبار (داخلی / روباز ۱ / روباز ۲)، صرف‌نظر از وضعیت خروج
    # وضعیت خروج به‌صورت جدا در exit_status / exit_by / exit_at مشخص می‌شود
    cur_loc = get_current_location(db, sid)

    # تعداد برگشت‌ها
    ret_to_internal = len(returns)
    ret_to_outside  = sum(1 for r in db.get("movement_log",[])
                          if r.get("slab_id")==sid and "انتقال مجدد" in r.get("operation",""))

    # وضعیت خروج: اگر خروج نزده، نام دقیق انباری که اسلب الان در آن است ذکر می‌شود
    _exit_raw = melt.get("exit_status", "در انبار")
    exit_status_txt = _exit_raw if _exit_raw == "خروج زده شده" else f"در {cur_loc}"

    return {
        "slab_id":          sid,
        "registered_by":    get_display_name(melt.get("registered_by","—"), db),
        "registered_at":    melt.get("registered_at","—"),
        "qc_status":        melt.get("qc_status","—"),
        "qc_by":            get_display_name(melt.get("qc_by","—"), db),
        "qc_at":            melt.get("qc_at","—"),
        "location":         melt.get("location","انبار داخلی"),
        "scarf":            "دارد: " + scarf_r.get("reason","") if scarf_r else "ندارد",
        "cut":              "دارد: " + cut_r.get("reason","") if cut_r else "ندارد",
        "bauman_done":      "دارد" if bauman_r else "ندارد",
        "bauman_at":        bauman_r.get("registered_at","—") if bauman_r else "—",
        "bauman_by":        get_display_name(bauman_r.get("registered_by","—"), db) if bauman_r else "—",
        "lab_delivered":    "تحویل داده شده" if lab_r else ("آماده تحویل" if bauman_r else "ندارد"),
        "lab_delivered_at": lab_r.get("delivered_at","—") if lab_r else "—",
        "lab_delivered_by": get_display_name(lab_r.get("delivered_by","—"), db) if lab_r else "—",
        "transfer_dest":    transfer_r.get("destination","—") if transfer_r else "—",
        "transfer_at":      transfer_r.get("transferred_at","—") if transfer_r else "—",
        "transfer_by":      get_display_name(transfer_r.get("transferred_by","—"), db) if transfer_r else "—",
        "ret_to_internal":  ret_to_internal,
        "ret_to_outside":   ret_to_outside,
        "returns_detail":   returns,
        "scrap":            "قراضه: " + scrap_r.get("reason","") if scrap_r else "ندارد",
        "exit_status":      exit_status_txt,
        "exit_by":          get_display_name(melt.get("exit_by","—"), db),
        "exit_at":          melt.get("exit_at","—"),
        "current_location": cur_loc,
        "note":             melt.get("note",""),
    }

# ── helper برای تغییر رنگ پس‌زمینه recursive ──
def _set_bg_recursive(widget, bg):
    try:
        widget.configure(bg=bg)
    except Exception:
        pass
    for ch in widget.winfo_children():
        _set_bg_recursive(ch, bg)

# ── reference به پنجره لاگین برای اعمال پس‌زمینه ──
_ACTIVE_LOGIN_WIN = None
_ACTIVE_MAIN_APP = None

class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        global _ACTIVE_LOGIN_WIN
        _ACTIVE_LOGIN_WIN = self
        self._init_login_state()
        self._finish_login_setup()

    def _init_login_state(self, db=None):
        """مقداردهی اولیهٔ وضعیت صفحهٔ لاگین."""
        self.db = db if db is not None else load_db()
        try:
            from client.db_bridge import normalize_db
            self.db = normalize_db(self.db)
        except Exception:
            if not isinstance(self.db.get("users"), dict) or not self.db.get("users"):
                self.db["users"] = {}
        self.title("سیستم مدیریت تختال فولاد")
        self.resizable(True, True)
        self._app_closing = False
        self._on_restore_attempt = self._ask_login_restore_password
        self._on_minimize_attempt = self._ask_login_minimize_password
        self._on_close_attempt = self._on_login_close
        self._window_action_in_progress = False
        self.configure(bg=C["bg"])
        self.selected_user = None
        self._was_minimized = False
        self._was_full = True
        self._allow_restore = False
        self._restore_after_id = None
        self._restore_dialog_open = False
        self._minimize_dialog_open = False
        self._lock_dialog_open = False
        self._configure_after_id = None
        self._allow_iconify = False
        self._minimizing_authorized = False
        self._was_in_taskbar = False
        self._lock_minimize_ok = False
        self._lock_restore_ok = False
        self._login_restore_cover = None
        self._suppress_configure_bg_refresh = False
        self._watchdog_proc = None
        self._drag_locked = True

    def _finish_login_setup(self, in_place=False):
        if os.environ.get("STF_ADMIN") == "1":
            self.protocol("WM_DELETE_WINDOW", self._on_login_close)
            _install_taskbar_restore_guard(self)
            self._build()
            self.geometry("1320x840")
            self.minsize(1040, 680)
            self._was_full = False
            # Admin هم Map/Unmap می‌خواهد: مینیمایز→ریستور PhotoImageها را خراب
            # می‌کند ولی اندازه عوض نمی‌شود؛ مسیر double-click (Configure/سایز)
            # شیشه را تازه می‌کند — اینجا همان رفرش را بعد از Map اجرا می‌کنیم.
            self.bind("<Unmap>", self._on_login_unmap)
            self.bind("<Map>", self._on_login_map)
            self.bind("<Configure>", self._on_login_configure)
            return
        hide_all_taskbars(self)
        if in_place:
            self._window_action_in_progress = True
            self._was_full = True
            self._allow_restore = False
            self._fullscreen_guard = False
            force_full_screen(self)
            self.update_idletasks()
            self.update()
        if self._watchdog_proc is None:
            self._watchdog_proc = start_taskbar_watchdog()
        self.protocol("WM_DELETE_WINDOW", self._on_login_close)
        force_full_screen(self)
        lock_window_chrome(self)
        lock_window_drag(self)
        enforce_frameless_window(self)
        bind_window_lock_keys(self)
        _install_destroy_guard(self)
        start_keyboard_lock(self)
        start_taskbar_hide_loop(self)
        _start_kiosk_reinforce_loop(self)
        _start_foreground_guard(self)
        try:
            install_keyboard_switcher(self)
        except Exception:
            pass
        self._build()
        self.update_idletasks()
        hide_all_taskbars(self)
        force_full_screen(self)
        if in_place:
            _finalize_fullscreen_window(self)
        else:
            enforce_full_screen_robust(self)
            self.after(400, lambda: _finalize_fullscreen_window(self))
        self.bind("<Unmap>", self._on_login_unmap)
        self.bind("<Map>", self._on_login_map)
        self.bind("<Configure>", self._on_login_configure)
        self.after(300, lambda: _release_grab_safe(self))
        _ensure_app_not_topmost(self)
        self.update_idletasks()
        _raise_window_front(self)
        if in_place:
            force_full_screen(self)
            self.update_idletasks()
            self.update()
            _end_window_transition(self)
        else:
            self._window_action_in_progress = False

    @classmethod
    def from_main_app(cls, app):
        """تبدیل MainApp به صفحهٔ لاگین روی همان پنجره — بدون بستن و بدون مکث."""
        _prepare_window_transition(app)
        stop_keyboard_lock(app)
        old_watchdog = getattr(app, "_watchdog_proc", None)
        app._watchdog_proc = None

        _begin_window_transition(app)
        cover = getattr(app, "_transition_cover", None)

        for ev in ("<Unmap>", "<Map>", "<Configure>"):
            try:
                app.unbind(ev)
            except Exception:
                pass
        for seq in ("<F5>", "<MouseWheel>"):
            try:
                app.unbind_all(seq)
            except Exception:
                pass

        app._drag_locked = True
        for w in list(app.winfo_children()):
            if w is not cover:
                w.destroy()
        app.update_idletasks()
        app.update()

        app._datetime_lbl = None
        app._ticket_btns = None
        app._tab_buttons = None
        app._drag_lock_installed = False
        try:
            app.pack_propagate(True)
        except Exception:
            pass

        app.__class__ = cls
        global _ACTIVE_LOGIN_WIN
        _ACTIVE_LOGIN_WIN = app

        app._init_login_state(db=getattr(app, "db", None))
        app._finish_login_setup(in_place=True)

        if old_watchdog is not None:
            try:
                old_watchdog.terminate()
            except Exception:
                pass

    def _on_login_close(self):
        """بستن برنامه از صفحهٔ لاگین — فقط با رمز قفل."""
        def _finish():
            _prepare_app_shutdown(self)
            stop_keyboard_lock(self)
            show_all_taskbars()
            self.destroy()
        request_window_close(self, _finish)

    def _ask_login_minimize_password(self):
        request_window_minimize(self)

    def _ask_login_restore_password(self):
        request_window_restore_or_maximize(self)

    def _build(self):
        self.configure(bg=C["bg"])

        # ════════════════════════════════════════════
        # هدر صفحه لاگین — فولادی صنعتی
        # ════════════════════════════════════════════
        # نوار رنگی سه‌گانه
        login_stripe1 = tk.Frame(self, bg="#d4a043", height=2)
        login_stripe1.pack(fill="x")
        login_stripe2 = tk.Frame(self, bg="#b8882e", height=2)
        login_stripe2.pack(fill="x")
        login_stripe3 = tk.Frame(self, bg="#8a6020", height=1)
        login_stripe3.pack(fill="x")

        # هدر اصلی — شیشه‌ای روی تصویر زمینه
        hdr = tk.Frame(self, bg=C["header_bg"], height=80)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        self._hdr_frame = hdr

        # آیکون + عنوان
        logo_f = tk.Frame(hdr, bg=C["header_bg"])
        logo_f.place(relx=.5, rely=.5, anchor="center")
        self._hdr_logo_frame = logo_f

        # لوگوی مربع فولادی
        logo_box = tk.Frame(logo_f, bg=C["accent"], width=48, height=48)
        logo_box.pack(side="right", padx=(14, 0))
        logo_box.pack_propagate(False)
        tk.Label(logo_box, text="Fe", bg=C["accent"],
                 fg="#000", font=("B Nazanin", 18, "bold")).place(relx=.5, rely=.5, anchor="center")

        title_f = tk.Frame(logo_f, bg=C["header_bg"])
        title_f.pack(side="right")
        self._hdr_title_frame = title_f
        # عنوان بدون کادر مشکی — روی تصویر زمینه
        self._login_title_lbl = self._make_glass_label(
            title_f, text=tr("app_title"), font=(_MAIN_FONT, 22, "bold"),
            fg="#ffffff", padx=6, pady=2, anchor="e",
            tint_hex=C["header_bg"], tint_alpha=0.10,
        )
        self._login_title_lbl.pack(anchor="e")
        self._login_subtitle_lbl = self._make_glass_label(
            title_f, text=tr("login_subtitle"), font=(_MAIN_FONT, 11, "bold"),
            fg="#d4a043", padx=6, pady=1, anchor="e",
            tint_hex=C["header_bg"], tint_alpha=0.10,
        )
        self._login_subtitle_lbl.pack(anchor="e", pady=(2, 0))

        # تاریخ و ساعت + وضعیت اتصال — شیشه‌ای
        time_f = tk.Frame(hdr, bg=C["header_bg"])
        time_f.pack(side="left", padx=(40, 20), fill="y")
        self._hdr_time_frame = time_f
        self._login_datetime = self._make_glass_label(
            time_f, text="۱۴۰۵/۰۴/۰۹  ۰۰:۰۰:۰۰", font=("B Nazanin", 11, "bold"),
            fg="#d4a043", padx=8, pady=2, anchor="w",
            tint_hex=C["header_bg"], tint_alpha=0.10,
        )
        self._login_datetime.pack(anchor="w", pady=(8, 0), padx=(8, 0))
        self._online_lbl = self._make_glass_label(
            time_f, text=tr("online"), font=(_MAIN_FONT, 10, "bold"),
            fg=C["success"], padx=8, pady=2, anchor="w",
            tint_hex=C["header_bg"], tint_alpha=0.10,
        )
        self._online_lbl.pack(anchor="w", padx=(8, 0), pady=(0, 4))

        win_ctrl = mount_window_controls(hdr, self, size=40, margin_right=12)
        self._win_ctrl = win_ctrl
        bind_window_header_drag(self, hdr)
        register_header_drag_zone(self, login_stripe1, login_stripe2, login_stripe3)

        self._tick_login()

        # خط جداکننده نازک
        tk.Frame(self, bg=C["accent"], height=1).pack(fill="x")

        # راهنما — شیشه‌ای
        guide = tk.Frame(self, bg=C["header_bg"])
        guide.pack(fill="x")
        self._guide_frame = guide
        self._guide_lbl = self._make_glass_label(
            guide, text=tr("select_user"), font=(_MAIN_FONT, 11, "bold"),
            fg="#d4a043", padx=8, pady=2,
            tint_hex=C["header_bg"], tint_alpha=0.10,
        )
        self._guide_lbl.pack(pady=4)
        tk.Frame(self, bg="#d4a043", height=1).pack(fill="x")

        # ════════════════════════════════════════════
        # گرید کاربران با اسکرول
        # ════════════════════════════════════════════
        # ── پس‌زمینه لاگین — با place پشت همه ویجت‌ها ──
        self._login_bg_img = None
        self._login_bg_lbl = None
        self._login_bg_pil = None   # تصویر کامل PIL برای برش لایه‌های شفاف‌نما

        scroll_container = tk.Frame(self, bg=C["bg"])
        scroll_container.pack(fill="both", expand=True)
        self._scroll_container = scroll_container

        canvas = tk.Canvas(scroll_container, bg=C["bg"],
                           highlightthickness=0)
        self._login_canvas = canvas
        scrollbar = tk.Scrollbar(scroll_container, orient="vertical",
                                  bg="#707070", troughcolor="#1a1a1a",
                                  activebackground=C["accent"], width=14)
        self._login_scrollbar = scrollbar

        def _on_scrollbar(*args):
            if not getattr(self, "_login_scroll_enabled", False):
                canvas.yview_moveto(0)
                return
            canvas.yview(*args)
            # هم‌ترازی فوری شیشه با تصویر ثابت — بدون تأخیر تا زمینه «نپرد»
            try:
                self._reapply_login_glass_layers()
            except Exception:
                pass

        def _on_yscroll_set(first, last):
            scrollbar.set(first, last)
            try:
                self._update_login_scroll_visibility()
            except Exception:
                pass

        scrollbar.configure(command=_on_scrollbar)
        canvas.configure(yscrollcommand=_on_yscroll_set)

        self._users_frame = tk.Frame(canvas, bg=C["bg"])
        self._win_id = canvas.create_window((0, 0), window=self._users_frame, anchor="nw")
        self._login_scroll_enabled = False

        def on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            self._update_login_scroll_visibility()
            self._schedule_login_glass_refresh(80)

        def on_canvas_resize(e):
            canvas.itemconfig(self._win_id, width=e.width)
            self._update_login_scroll_visibility()
            self._schedule_login_glass_refresh(80)

        self._users_frame.bind("<Configure>", on_frame_configure)
        canvas.bind("<Configure>", on_canvas_resize)
        register_scroll_canvas(canvas, self._users_frame)

        def _login_wheel(event):
            # اگر محتوا در یک صفحه جا شده، اسکرول را قفل کن تا زمینه نپرد
            if not getattr(self, "_login_scroll_enabled", False):
                try:
                    canvas.yview_moveto(0)
                except Exception:
                    pass
                return "break"
            try:
                self.after_idle(self._reapply_login_glass_layers)
            except Exception:
                pass
        canvas.bind("<MouseWheel>", _login_wheel, add="+")
        self._users_frame.bind("<MouseWheel>", _login_wheel, add="+")
        self.bind_all("<MouseWheel>", _global_mousewheel)
        canvas.pack(side="left", fill="both", expand=True)
        self._build_user_grid()
        self.after(80, self._refresh_login_icons)
        self.after(200, self._update_login_scroll_visibility)

        # تصویر پس‌زمینه — چند پاس تا کل ارتفاع پنجره (با فوتر) پوشش داده شود
        self.after(300, self._refresh_login_background)
        self.after(700, self._refresh_login_background)
        self.after(1200, self._refresh_login_background)

        # کردیت برنامه‌نویس — شیشه‌ای روی تصویر (بدون نوار مشکی)
        credit_bar = tk.Frame(self, bg=C["header_bg"])
        credit_bar.pack(fill="x", side="bottom")
        self._credit_bar = credit_bar
        tk.Frame(credit_bar, bg=C["accent"], height=1).pack(fill="x")
        credit_in = tk.Frame(credit_bar, bg=C["header_bg"])
        credit_in.pack(fill="x", pady=6)
        self._credit_in_frame = credit_in
        self._credit_lbl = self._make_glass_label(
            credit_in,
            text="برنامه‌نویس:   Reza Borzoei  09107606159    |    Ali Jazayeri  09930245658",
            font=(_MAIN_FONT, 11, "bold"), fg=C["gold"], padx=10, pady=2,
            tint_hex=C["header_bg"], tint_alpha=0.10,
        )
        self._credit_lbl.pack()

    def _on_login_configure(self, event):
        """
        بعد از هر تغییر اندازه/جابه‌جایی پنجره (شامل maximize شدن خودکار اول
        برنامه)، با کمی تأخیر (debounce) پس‌زمینه را دوباره اعمال می‌کند تا
        مطمئن شویم هیچ‌جا رنگ توپر/سیاه باقی نمی‌ماند. اگر اندازه‌ی پنجره
        از تمام‌صفحه خارج شود، بی‌سروصدا (بدون رمز) به حالت تمام‌صفحه
        برگردانده می‌شود — رمز فقط برای مینیمایز و بستن پرسیده می‌شود.
        """
        if event.widget is not self:
            return
        if getattr(self, "_app_closing", False):
            return
        if getattr(self, "_fullscreen_guard", False):
            return
        if getattr(self, "_cfg_debounce_id", None):
            try:
                self.after_cancel(self._cfg_debounce_id)
            except Exception:
                pass
        self._cfg_debounce_id = self.after(80, self._login_check_fullscreen)

    def _login_check_fullscreen(self):
        self._cfg_debounce_id = None
        if getattr(self, "_app_closing", False):
            return
        if getattr(self, "_fullscreen_guard", False):
            return
        if getattr(self, "_window_action_in_progress", False):
            return
        if _modal_dialog_active(self):
            return
        if _window_is_minimized(self):
            return
        is_full = is_window_full_screen(self)
        if not is_full and self._was_full:
            if self._allow_restore:
                self._allow_restore = False
                self._was_full = False
            else:
                force_full_screen(self)
                self._was_full = True
                self._drag_locked = True
                if self._restore_after_id is not None:
                    try:
                        self.after_cancel(self._restore_after_id)
                    except Exception:
                        pass
                    self._restore_after_id = None
                if not self._restore_dialog_open:
                    self._restore_after_id = self.after(
                        60, self._ask_login_restore_password)
                if getattr(self, "_configure_after_id", None):
                    try:
                        self.after_cancel(self._configure_after_id)
                    except Exception:
                        pass
                self._configure_after_id = self.after(
                    400, self._refresh_login_background)
                return
        self._was_full = is_full
        if getattr(self, "_suppress_configure_bg_refresh", False):
            # restore از تسک‌بار همین الان شیشه را تازه کرده — debounce ۴۰۰ms فلش می‌سازد
            if getattr(self, "_configure_after_id", None):
                try:
                    self.after_cancel(self._configure_after_id)
                except Exception:
                    pass
                self._configure_after_id = None
            return
        if getattr(self, "_configure_after_id", None):
            try:
                self.after_cancel(self._configure_after_id)
            except Exception:
                pass
        self._configure_after_id = self.after(400, self._refresh_login_background)

    def _arm_login_restore_cover(self):
        """
        قبل از مینیمایز یک پوشش هم‌رنگ/هم‌تصویر زمینه می‌گذارد تا هنگام
        restore، شیشه‌ی کهنه/ناهم‌تراز دیده نشود و بعد از رفرش فوری برداشته شود.
        """
        if getattr(self, "_app_closing", False):
            return
        if getattr(self, "_login_restore_cover", None) is not None:
            try:
                if self._login_restore_cover.winfo_exists():
                    return
            except Exception:
                self._login_restore_cover = None
        try:
            cover = tk.Label(self, bd=0, highlightthickness=0)
            img = getattr(self, "_login_bg_img", None)
            if img is not None:
                try:
                    cover.configure(image=img)
                    cover.image = img
                except Exception:
                    cover.configure(bg=C.get("bg", "#1a1f28"))
            else:
                cover.configure(bg=C.get("bg", "#1a1f28"))
            bleed = int(getattr(self, "_login_bg_bleed", 0) or 0)
            sw = max(1, int(self.winfo_width() or 1))
            sh = max(1, int(self.winfo_height() or 1))
            cover.place(x=-bleed, y=-bleed, width=sw + bleed * 2, height=sh + bleed * 2)
            try:
                cover.lift()
            except Exception:
                pass
            self._login_restore_cover = cover
        except Exception:
            self._login_restore_cover = None

    def _disarm_login_restore_cover(self):
        cover = getattr(self, "_login_restore_cover", None)
        self._login_restore_cover = None
        if cover is None:
            return
        try:
            if cover.winfo_exists():
                cover.place_forget()
                cover.destroy()
        except Exception:
            pass

    def _on_login_unmap(self, event):
        """اگر بدون رمز مینیمایز شد — برگردان (فقط کیوسک). Admin فقط فلگ ریستور می‌زند."""
        if event.widget is not self:
            return
        if getattr(self, "_app_closing", False):
            return
        self._was_minimized = True
        # پوشش تا رفرش فوریِ Map — جلوگیری از فلش شیشهٔ غلط
        self._arm_login_restore_cover()
        if self._allow_iconify or getattr(self, "_minimizing_authorized", False):
            self._allow_iconify = False
            if getattr(self, "_minimizing_authorized", False):
                self.after(80, lambda: _mark_minimize_in_taskbar(self))
            return
        # Admin / بدون قفل شل: مینیمایز native مجاز است — فقط بعد از Map شیشه را تازه کن
        if not _app_shell_lock_active(self):
            return
        try:
            if self.state() == "iconic" or _hwnd_is_iconic(self):
                self.after(1, self._recover_login_from_minimize)
        except Exception:
            pass

    def _recover_login_from_minimize(self):
        try:
            win_restore(self)
            hide_all_taskbars(self)
            force_full_screen(self)
            if not getattr(self, "_minimize_dialog_open", False):
                self.after(80, self._ask_login_minimize_password)
        except Exception:
            pass

    def _on_login_map(self, event):
        """
        بعد از مینیمایز→ریستور، اندازه پنجره مثل double-click عوض نمی‌شود؛
        پس Configure/شیشه خودکار تازه نمی‌شود و PhotoImageها می‌پرند/ناپدید
        می‌شوند. همان مسیر رفرشِ تغییر اندازه را اینجا فوری اجرا می‌کنیم — بدون
        تخریب و ساخت دوبارهٔ گرید و بدون تأخیر چندمرحله‌ای که فلش می‌سازد.
        """
        if event.widget is not self:
            return
        if getattr(self, "_minimizing_authorized", False):
            _try_finish_authorized_minimize_on_map(self)
        # Map وسطِ مینیمایز (قبل از تسک‌بار) — هنوز iconic؛ فلگ را نگه دار
        try:
            if _hwnd_is_iconic(self) or self.state() == "iconic":
                return
        except Exception:
            pass
        if getattr(self, "_was_full", True) and _app_shell_lock_active(self):
            hide_all_taskbars(self)
        if self._was_minimized:
            self._was_minimized = False
            self._suppress_configure_bg_refresh = True
            # Configure بعد از restore هم debounce ۴۰۰ms می‌گذارد — همان را کنسل کن
            # تا رفرش دیرهنگام دوباره layout را «بپراند».
            aid = getattr(self, "_configure_after_id", None)
            if aid is not None:
                try:
                    self.after_cancel(aid)
                except Exception:
                    pass
                self._configure_after_id = None
            # فوری زیر پوشش — سپس یک پاس idle و برداشتن پوشش
            try:
                self._stabilize_login_after_geometry_change()
            except Exception:
                pass

            def _finish_restore_paint():
                try:
                    self._stabilize_login_after_geometry_change()
                except Exception:
                    pass
                self._disarm_login_restore_cover()

            try:
                self.after_idle(_finish_restore_paint)
            except Exception:
                self._disarm_login_restore_cover()
            # ایمنی: پوشش نماند اگر idle اجرا نشد
            try:
                self.after(250, self._disarm_login_restore_cover)
            except Exception:
                pass
            try:
                self.after(500, lambda: setattr(self, "_suppress_configure_bg_refresh", False))
            except Exception:
                self._suppress_configure_bg_refresh = False
        else:
            self.after(200, self._reapply_login_glass_layers)

    def _stabilize_login_after_geometry_change(self):
        """
        همان کاری که مسیر double-click / Configure می‌کند:
        عرض canvas + پس‌زمینه + لایه‌های شیشه/آیکون — بدون _build_user_grid.
        """
        if getattr(self, "_app_closing", False):
            return
        if type(self).__name__ != "LoginWindow":
            return
        try:
            if _hwnd_is_iconic(self) or self.state() == "iconic":
                return
        except Exception:
            pass
        if _modal_dialog_active(self) or getattr(self, "_minimize_dialog_open", False) or getattr(self, "_restore_dialog_open", False):
            try:
                self.after(250, self._stabilize_login_after_geometry_change)
            except Exception:
                pass
            return
        if getattr(self, "_stabilize_login_busy", False):
            return
        self._stabilize_login_busy = True
        try:
            # کیوسک: فقط اگر هنوز تمام‌صفحه است، هندسه را محکم کن (قبل از شیشه)
            if getattr(self, "_was_full", True) and _app_shell_lock_active(self):
                try:
                    force_full_screen(self)
                except Exception:
                    pass
                try:
                    hide_all_taskbars(self)
                except Exception:
                    pass
            self.update_idletasks()
            try:
                self.update()
            except Exception:
                pass

            canvas = getattr(self, "_login_canvas", None)
            if canvas is not None and canvas.winfo_exists():
                uf = getattr(self, "_users_frame", None)
                if uf is None or not uf.winfo_exists():
                    self._users_frame = tk.Frame(canvas, bg=C["bg"])
                    self._win_id = canvas.create_window(
                        (0, 0), window=self._users_frame, anchor="nw")
                    self._build_user_grid()
                self.update_idletasks()
                cw = int(canvas.winfo_width())
                if cw < 50:
                    cw = max(int(self.winfo_width()) - 20, 400)
                if hasattr(self, "_win_id"):
                    canvas.itemconfig(self._win_id, width=cw)
                try:
                    self._users_frame.update_idletasks()
                except Exception:
                    pass
                try:
                    canvas.configure(scrollregion=canvas.bbox("all"))
                except Exception:
                    pass

            # عین مسیر Configure بعد از کوچک شدن با double-click
            self._refresh_login_background(force=True)
            self._update_login_scroll_visibility()
            self._reapply_login_glass_layers()
        except Exception:
            try:
                self.after(400, self._stabilize_login_after_geometry_change)
            except Exception:
                pass
        finally:
            # فوری آزاد کن تا after_idle بتواند یک پاس settle بزند
            self._stabilize_login_busy = False

    def _recover_after_restore(self):
        """سازگاری با نام قدیمی — همان stabilize مسیر Configure."""
        self._stabilize_login_after_geometry_change()

    def _schedule_login_glass_refresh(self, delay_ms=200):
        """
        بعد از بزرگ/کوچک شدن پنجره یا اسکرول، لایه‌های شیشه‌ای را با تأخیر کوتاه
        دوباره می‌کشد تا با تصویر ثابت پنجره هم‌تراز بمانند (زمینه با اسکرول نپرد).
        """
        if type(self).__name__ != "LoginWindow":
            return
        if getattr(self, "_app_closing", False):
            return
        aid = getattr(self, "_glass_refresh_after_id", None)
        if aid is not None:
            try:
                self.after_cancel(aid)
            except Exception:
                pass
        self._glass_refresh_after_id = self.after(
            delay_ms, self._run_scheduled_login_glass_refresh)

    def _run_scheduled_login_glass_refresh(self):
        self._glass_refresh_after_id = None
        if getattr(self, "_app_closing", False):
            return
        if _modal_dialog_active(self):
            return
        try:
            self._reapply_login_glass_layers()
        except Exception:
            pass

    def _update_login_scroll_visibility(self):
        """اگر همهٔ کاربران در یک صفحه جا شوند، اسکرول کاملاً قفل شود تا زمینه نپرد."""
        try:
            canvas = getattr(self, "_login_canvas", None)
            scrollbar = getattr(self, "_login_scrollbar", None)
            if canvas is None or scrollbar is None:
                return
            if not canvas.winfo_exists():
                return
            canvas.update_idletasks()
            bbox = canvas.bbox("all")
            if not bbox:
                self._login_scroll_enabled = False
                return
            content_h = max(0, int(bbox[3] - bbox[1]))
            view_h = max(1, int(canvas.winfo_height()))
            if content_h <= view_h + 8:
                self._login_scroll_enabled = False
                try:
                    canvas.yview_moveto(0)
                except Exception:
                    pass
                if scrollbar.winfo_ismapped():
                    scrollbar.pack_forget()
            else:
                self._login_scroll_enabled = True
                if not scrollbar.winfo_ismapped():
                    scrollbar.pack(side="right", fill="y")
        except Exception:
            self._login_scroll_enabled = False

    def _reapply_login_glass_layers(self):
        """
        لایه‌های شیشه‌ای را با اندازه/مکان فعلی ویجت‌ها دوباره می‌کشد.
        برش همیشه از تصویر ثابت کل پنجره است — پس زمینه با اسکرول جابه‌جا نمی‌شود.
        """
        if type(self).__name__ != "LoginWindow":
            return
        if getattr(self, "_app_closing", False):
            return
        try:
            self.update_idletasks()
            sw = int(self.winfo_width())
            sh = int(self.winfo_height())
            pil = getattr(self, "_login_bg_pil", None)
            if pil is None:
                return
            bleed = int(getattr(self, "_login_bg_bleed", 0) or 0)
            if sw > 10 and sh > 10 and (
                abs(pil.width - (sw + bleed * 2)) > 6 or abs(pil.height - (sh + bleed * 2)) > 6
            ):
                self._refresh_login_background()
                return

            for w in (getattr(self, "_scroll_container", None),
                      getattr(self, "_login_canvas", None),
                      getattr(self, "_users_frame", None)):
                if w is not None and w.winfo_exists():
                    self._apply_transparent_layer(w, tint_hex=C["bg"], tint_alpha=0.04)

            for w in (getattr(self, "_hdr_frame", None),
                      getattr(self, "_guide_frame", None),
                      getattr(self, "_credit_bar", None),
                      getattr(self, "_hdr_title_frame", None),
                      getattr(self, "_hdr_time_frame", None),
                      getattr(self, "_hdr_logo_frame", None),
                      getattr(self, "_credit_in_frame", None)):
                if w is not None and w.winfo_exists():
                    self._apply_transparent_layer(w, tint_hex=C["header_bg"], tint_alpha=0.08)

            for w in (getattr(self, "_admin_wrapper_frames", []) or []):
                if w is not None and w.winfo_exists():
                    self._apply_transparent_layer(w, tint_hex=C["bg"], tint_alpha=0.04)

            for inner_w, bg_color, tint_a in getattr(self, "_user_card_widgets", []):
                if inner_w is not None and inner_w.winfo_exists():
                    self._apply_transparent_layer(inner_w, tint_hex=bg_color, tint_alpha=tint_a)

            self._apply_all_glass_texts()
            self._refresh_login_icons()
        except Exception:
            pass

    def _clear_bg_layer(self, widget):
        """لایهٔ تصویر قدیمی را برمی‌دارد تا باکس سفید/ناهم‌اندازه نماند."""
        try:
            lbl = getattr(widget, "_bg_layer_lbl", None)
            if lbl is not None:
                try:
                    if lbl.winfo_exists():
                        lbl.place_forget()
                        lbl.destroy()
                except Exception:
                    pass
            widget._bg_layer_lbl = None
            widget._bg_layer_img = None
        except Exception:
            pass

    def _refresh_login_background(self, bg_path=None, force=False):
        """
        تصویر پس‌زمینه را روی کل پنجره لاگین می‌گذارد و سپس همان تصویر را
        به‌صورت بریده‌شده (crop) پشت کانتینرها/کارت‌های شیشه‌ای می‌گذارد.
        فقط وقتی اندازه عوض شده یا force — نه هر Configure (علت فلاشر).
        """
        if type(self).__name__ != "LoginWindow":
            return
        if (not force) and _modal_dialog_active(self):
            return
        try:
            self.update_idletasks()
            sw = max(64, int(self.winfo_width() or 0))
            sh = max(64, int(self.winfo_height() or 0))
            if sw < 100 or sh < 100:
                sw = int(self.winfo_screenwidth())
                sh = int(self.winfo_screenheight())
            prev_sz = getattr(self, "_login_bg_win_size", None)
            if (
                (not force)
                and prev_sz == (sw, sh)
                and getattr(self, "_login_bg_img", None) is not None
                and getattr(self, "_login_bg_lbl", None) is not None
            ):
                try:
                    if self._login_bg_lbl.winfo_exists():
                        return  # اندازه یکی است — رفرش کامل نکن (فلاشر نمی‌سازد)
                except Exception:
                    pass

            if bg_path is None:
                try:
                    self.db = load_db()
                except Exception:
                    pass
                bg_path = ensure_login_background_local(getattr(self, "db", None))
                if not bg_path:
                    bg_path = (getattr(self, "db", {}) or {}).get("settings", {}).get("background_image", "")
            if not bg_path or not os.path.exists(bg_path):
                for _ext in (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"):
                    _cand = os.path.join(ASSETS_DIR, "login_background" + _ext)
                    if os.path.exists(_cand):
                        bg_path = _cand
                        break
            has_bg = bool(bg_path) and os.path.exists(bg_path)
            from PIL import Image, ImageTk, ImageEnhance

            if has_bg:
                # دقیقاً هم‌اندازه پنجره + bleed تا لبه سیاه/خالی نماند
                bleed = 8
                pil_img = Image.open(bg_path).convert("RGB").resize(
                    (sw + bleed * 2, sh + bleed * 2), Image.LANCZOS)
                pil_img = ImageEnhance.Color(pil_img).enhance(0.72)
                pil_img = ImageEnhance.Brightness(pil_img).enhance(0.92)
                pil_img = ImageEnhance.Contrast(pil_img).enhance(0.96)
                self._login_bg_pil = pil_img
                self._login_bg_bleed = bleed
                self._login_bg_win_size = (sw, sh)

                ph = ImageTk.PhotoImage(pil_img)
                if self._login_bg_lbl is None or not self._login_bg_lbl.winfo_exists():
                    self._login_bg_lbl = tk.Label(self, image=ph, bd=0, highlightthickness=0, bg="#000000")
                else:
                    self._login_bg_lbl.configure(image=ph, bg="#000000")
                # place با پیکسل — نه فقط rel — تا لبه چپ/راست سیاه نماند
                self._login_bg_lbl.place(x=-bleed, y=-bleed, width=sw + bleed * 2, height=sh + bleed * 2)
                self._login_bg_lbl.lower()
                self._login_bg_lbl.image = ph
                self._login_bg_img = ph
            else:
                self._login_bg_pil = None
                self._login_bg_bleed = 0
                self._login_bg_win_size = (sw, sh)
                if self._login_bg_lbl is not None and self._login_bg_lbl.winfo_exists():
                    self._login_bg_lbl.place_forget()

            # همه لایه‌های viewport + خودِ users_frame تا دور کادرها سیاه نماند
            for w in (getattr(self, "_scroll_container", None),
                      getattr(self, "_login_canvas", None),
                      getattr(self, "_users_frame", None)):
                if w is not None and w.winfo_exists():
                    self._apply_transparent_layer(w, tint_hex=C["bg"], tint_alpha=0.04)

            for w in (getattr(self, "_admin_wrapper_frames", []) or []):
                if w is not None and w.winfo_exists():
                    self._apply_transparent_layer(w, tint_hex=C["bg"], tint_alpha=0.04)

            for w in (getattr(self, "_hdr_frame", None),
                      getattr(self, "_guide_frame", None),
                      getattr(self, "_credit_bar", None),
                      getattr(self, "_hdr_title_frame", None),
                      getattr(self, "_hdr_time_frame", None),
                      getattr(self, "_hdr_logo_frame", None),
                      getattr(self, "_credit_in_frame", None)):
                if w is not None and w.winfo_exists():
                    self._apply_transparent_layer(w, tint_hex=C["header_bg"], tint_alpha=0.08)

            for inner_w, bg_color, tint_a in getattr(self, "_user_card_widgets", []):
                if inner_w is not None and inner_w.winfo_exists():
                    if has_bg:
                        self._apply_transparent_layer(inner_w, tint_hex=bg_color, tint_alpha=tint_a)
                    else:
                        self._clear_bg_layer(inner_w)
                        try:
                            inner_w.configure(bg=bg_color)
                        except Exception:
                            pass

            self._apply_all_glass_texts()
            self._refresh_login_icons()
            self._update_login_scroll_visibility()
        except Exception:
            pass

    def _make_glass_label(self, parent, text, font, fg, padx=4, pady=2,
                           justify="center", anchor="center", wraplength=None,
                           tint_hex=None, tint_alpha=0.0):
        """
        به‌جای Label معمولی (که همیشه یک باکس تو‌پر پشت متنش دارد)، یک
        Canvas کوچک می‌سازد که متن مستقیماً روی برشِ تصویر پس‌زمینه رسم
        می‌شود — یعنی واقعاً هیچ باکس سیاه/تو‌پری پشت نوشته نیست، متن
        مثل این‌که شناور روی خود عکس است.
        tint_hex/tint_alpha باید دقیقاً همان تینتی باشد که کارت/فریم اطراف
        این متن استفاده می‌کند، وگرنه این Canvas چون تصویر «خام» را نشان
        می‌دهد، یک مستطیل کمی روشن‌تر از اطرافش (که تیره‌تر شده) دیده می‌شود.
        """
        import tkinter.font as tkfont
        f = tkfont.Font(font=font)
        if wraplength:
            words = text.split()
            line, lines = "", 1
            for wd in words:
                test = (line + " " + wd).strip()
                if f.measure(test) > wraplength and line:
                    lines += 1
                    line = wd
                else:
                    line = test
            tw = wraplength
            th = f.metrics("linespace") * lines
        else:
            tw = f.measure(text) if text else f.measure(" ")
            th = f.metrics("linespace")
        w = max(4, tw + padx * 2)
        h = max(4, th + pady * 2)
        cv = tk.Canvas(parent, width=w, height=h, highlightthickness=0, bd=0,
                       bg=(tint_hex or C.get("bg", "#1a1f28")))
        cv._glass_text = text
        cv._glass_font = font
        cv._glass_fg = fg
        cv._glass_justify = justify
        cv._glass_anchor = anchor
        cv._glass_wraplength = wraplength
        cv._glass_tint_hex = tint_hex
        cv._glass_tint_alpha = tint_alpha
        if not hasattr(self, "_glass_text_widgets"):
            self._glass_text_widgets = []
        self._glass_text_widgets.append(cv)
        try:
            cv.after_idle(lambda c=cv: self._apply_glass_text(c))
        except Exception:
            pass
        return cv

    def _glass_set_text(self, cv, text):
        """متن یک glass-label را عوض می‌کند (مثلاً برای ساعت زنده)"""
        try:
            cv._glass_text = text
            self._apply_glass_text(cv)
        except Exception:
            pass

    def _apply_glass_text(self, cv):
        """تصویر پشت متن را برش می‌زند و متن را مستقیماً رویش رسم می‌کند"""
        try:
            if not cv.winfo_exists():
                return
            from PIL import Image, ImageTk
            cv.update_idletasks()
            w = cv.winfo_width()
            h = cv.winfo_height()
            if w < 2 or h < 2:
                try:
                    cv.after(80, lambda c=cv: self._apply_glass_text(c))
                except Exception:
                    pass
                return
            cv.delete("all")
            if self._login_bg_pil is not None:
                bleed = int(getattr(self, "_login_bg_bleed", 0) or 0)
                ox = cv.winfo_rootx() - self.winfo_rootx() + bleed
                oy = cv.winfo_rooty() - self.winfo_rooty() + bleed
                full = self._login_bg_pil
                box = (max(0, ox), max(0, oy),
                       min(full.width, ox + w), min(full.height, oy + h))
                if box[2] > box[0] and box[3] > box[1]:
                    crop = full.crop(box)
                    if crop.size != (w, h):
                        crop = crop.resize((w, h))
                    tint_hex = getattr(cv, "_glass_tint_hex", None)
                    tint_alpha = getattr(cv, "_glass_tint_alpha", 0.0)
                    if tint_hex and tint_alpha > 0:
                        tint_layer = Image.new("RGB", crop.size, tint_hex)
                        crop = Image.blend(crop.convert("RGB"), tint_layer, tint_alpha)
                    ph = ImageTk.PhotoImage(crop)
                    cv.create_image(0, 0, anchor="nw", image=ph)
                    cv._glass_img_ref = ph
            else:
                bg = getattr(cv, "_glass_tint_hex", None) or C["bg"]
                try:
                    cv.configure(bg=bg)
                except Exception:
                    pass
            anchor = cv._glass_anchor
            pos = {"center": (w / 2, h / 2), "e": (w - 2, h / 2), "w": (2, h / 2)}.get(
                anchor, (w / 2, h / 2))
            cv.create_text(pos[0], pos[1], text=cv._glass_text, font=cv._glass_font,
                            fill=cv._glass_fg, anchor=anchor, justify=cv._glass_justify,
                            width=(cv._glass_wraplength or 0))
        except Exception:
            pass

    def _apply_all_glass_texts(self):
        """رسم همهٔ متن‌های شیشه‌ای — با یا بدون تصویر زمینه."""
        for cv in getattr(self, "_glass_text_widgets", []):
            try:
                if cv is not None and cv.winfo_exists():
                    self._apply_glass_text(cv)
            except Exception:
                pass

    def _apply_transparent_layer(self, widget, tint_hex=None, tint_alpha=0.0):
        """
        پشت widget یک Label با برش دقیق تصویر پس‌زمینه (هم‌راستا با مکان واقعی
        widget روی صفحه) قرار می‌دهد تا انگار آن widget شفاف است و تصویر
        پس‌زمینه از پشتش دیده می‌شود.

        اگر tint_hex داده شود، تصویر برش‌خورده با آن رنگ به نسبت tint_alpha
        ترکیب (blend) می‌شود — یعنی به‌جای پس‌زمینه‌ی کاملاً کدر، یک حالت
        «شیشه‌ی مات رنگی» ساخته می‌شود: هم تصویر زمینه از پشت کارت‌های
        کاربران دیده می‌شود و هم کارت‌ها (آیکون/اسم) همچنان مشخص و خوانا
        باقی می‌مانند. tint_alpha بین 0 (کاملاً شفاف، فقط تصویر) تا 1
        (کاملاً همان رنگ کارت، بدون تصویر) است.

        مهم: رنگ bg خودِ Label همیشه همان tint است — وگرنه وقتی پنجره
        بزرگ می‌شود و تصویر هنوز هم‌اندازه نیست، حاشیهٔ سفید سیستم دیده می‌شود.
        """
        try:
            fill_bg = tint_hex or C.get("bg", "#1a1f28")
            if self._login_bg_pil is None:
                self._clear_bg_layer(widget)
                try:
                    widget.configure(bg=fill_bg)
                except Exception:
                    pass
                return
            from PIL import Image, ImageTk
            widget.update_idletasks()
            w = int(widget.winfo_width())
            h = int(widget.winfo_height())
            if w < 2 or h < 2:
                return
            bleed = int(getattr(self, "_login_bg_bleed", 0) or 0)
            ox = int(widget.winfo_rootx() - self.winfo_rootx()) + bleed
            oy = int(widget.winfo_rooty() - self.winfo_rooty()) + bleed
            full = self._login_bg_pil
            # اگر هنوز خارج از تصویر هستیم (وسط انیمیشن resize)، فقط رنگ یکدست بگذار
            if ox + 2 >= full.width or oy + 2 >= full.height or ox + w <= 0 or oy + h <= 0:
                self._clear_bg_layer(widget)
                try:
                    widget.configure(bg=fill_bg)
                except Exception:
                    pass
                return
            box = (max(0, ox), max(0, oy),
                   min(full.width, ox + w), min(full.height, oy + h))
            if box[2] <= box[0] or box[3] <= box[1]:
                return
            crop = full.crop(box)
            # همیشه دقیقاً به اندازهٔ ویجت — حتی اگر crop ناقص بود
            if crop.size != (w, h):
                crop = crop.resize((w, h), Image.LANCZOS)

            if tint_hex and tint_alpha > 0:
                tint_layer = Image.new("RGB", crop.size, tint_hex)
                crop = Image.blend(crop.convert("RGB"), tint_layer, tint_alpha)

            ph = ImageTk.PhotoImage(crop)

            lbl = getattr(widget, "_bg_layer_lbl", None)
            if lbl is None or not lbl.winfo_exists():
                lbl = tk.Label(widget, image=ph, bd=0, highlightthickness=0, bg=fill_bg)
                lbl.place(x=0, y=0, relwidth=1, relheight=1)
                lbl.lower()
                lbl._is_bg_layer = True
                widget._bg_layer_lbl = lbl
            else:
                # تصویر قبلی را عوض کن و bg را حتماً تیره نگه دار تا حاشیه سفید نماند
                lbl.configure(image=ph, bg=fill_bg)
                lbl.place(x=0, y=0, relwidth=1, relheight=1)
                lbl.lower()
            lbl.image = ph
            widget._bg_layer_img = ph
            widget._bg_layer_size = (w, h)
        except Exception:
            pass

    def _tick_login(self):
        if not getattr(self, "_login_datetime", None):
            return
        try:
            if not self._login_datetime.winfo_exists():
                return
        except Exception:
            return
        now = datetime.datetime.now()
        sh = to_shamsi(now)
        parts = sh.split("  ")
        d = parts[0] if parts else ""
        t = parts[1] if len(parts) > 1 else ""
        try:
            txt = f"{d}  {t}"
            if hasattr(self._login_datetime, "_glass_text"):
                self._glass_set_text(self._login_datetime, txt)
            else:
                self._login_datetime.config(text=txt)
            if not hasattr(self, "_conn_check_tick"):
                self._conn_check_tick = 0
            self._conn_check_tick += 1
            force = self._conn_check_tick >= 5
            if force:
                self._conn_check_tick = 0
            _refresh_client_connection_label(self, force_check=force)
            self.after(1000, self._tick_login)
        except Exception:
            pass

    def _build_user_grid(self):
        for w in self._users_frame.winfo_children():
            w.destroy()
        self._users_frame._bg_layer_lbl = None
        self._user_card_widgets = []
        self._admin_wrapper_frames = []
        self._grid_frame = None
        self._glass_text_widgets = [cv for cv in getattr(self, "_glass_text_widgets", [])
                                     if cv.winfo_exists()]
        self._icon_widgets = [t for t in getattr(self, "_icon_widgets", [])
                               if t[0].winfo_exists()]
        try:
            from client.db_bridge import normalize_db
            self.db = normalize_db(self.db)
        except Exception:
            pass
        users = self.db.get("users") if isinstance(self.db.get("users"), dict) else {}
        sorted_users = sorted(users.items(),
            key=lambda x: (0 if x[1].get("role") == "admin" else 1, x[0]))
        # کاربرانی که «عدم نمایش» برایشان فعال شده، اصلاً در صفحه ورود نمایش داده نمی‌شوند
        sorted_users = [(u, d) for u, d in sorted_users if not d.get("hidden", False)]

        role_labels = {
            "admin":   ("مدیریت سیستم", "#d4a043", "▣"),   # طلایی
            "shift":   ("اپراتور شیفت", "#4a9fd4", "◆"),   # آبی فولادی
            "scarf":   ("اسکارف‌کار و برش‌کار", "#3db880", "◉"),
            "shift_n": ("نفرات شیفت فروش", "#9a70c8", "◍"),
        }

        bg0 = self._users_frame.cget("bg")
        admins = [(u, d) for u, d in sorted_users if d["role"] == "admin"]
        scarfs = [(u, d) for u, d in sorted_users if d["role"] == "scarf"]
        shifts = [(u, d) for u, d in sorted_users if d["role"] == "shift"]
        sales  = [(u, d) for u, d in sorted_users if d["role"] == "shift_n"]
        known = {"admin", "scarf", "shift", "shift_n"}
        others = [(u, d) for u, d in sorted_users if d["role"] not in known]

        def _make_role_section(title, accent, users, cols):
            if not users:
                return None
            wrap = tk.Frame(self._users_frame, bg=bg0)
            wrap.pack(fill="x", padx=18, pady=(2, 4))
            self._admin_wrapper_frames.append(wrap)

            box = tk.Frame(
                wrap, bg="#243040",
                highlightthickness=2,
                highlightbackground=accent,
            )
            box.pack(fill="x")
            # شیشه‌ای‌تر — تصویر زمینه بیشتر دیده شود
            self._user_card_widgets.append((box, "#243040", 0.06))

            hdr = tk.Frame(box, bg="#1a2838")
            hdr.pack(fill="x")
            self._user_card_widgets.append((hdr, "#1a2838", 0.06))
            tk.Frame(hdr, bg=accent, height=2).pack(fill="x")
            title_cv = self._make_glass_label(
                hdr, text=title, font=("B Nazanin", 12, "bold"),
                fg=accent, padx=6, pady=2,
                tint_hex="#1a2838", tint_alpha=0.06,
            )
            title_cv.pack(anchor="center", pady=2)

            grid = tk.Frame(box, bg="#243040")
            grid.pack(fill="x", padx=8, pady=(2, 6))
            self._user_card_widgets.append((grid, "#243040", 0.06))
            for i in range(cols):
                grid.grid_columnconfigure(i, weight=1, uniform="rolecol")
            for i, (uname, udata) in enumerate(users):
                self._make_user_card(
                    grid, uname, udata, role_labels,
                    row=i // cols, col=i % cols,
                )
            return grid

        # ── مدیریت ──
        for uname, udata in admins:
            admin_wrapper = tk.Frame(self._users_frame, bg=bg0)
            admin_wrapper.pack(fill="x", padx=30, pady=(4, 4))
            self._admin_wrapper_frames.append(admin_wrapper)
            admin_card_frame = tk.Frame(admin_wrapper, bg=bg0)
            admin_card_frame.pack(anchor="center")
            self._admin_wrapper_frames.append(admin_card_frame)
            self._make_user_card(admin_card_frame, uname, udata, role_labels, full_width=True)

        # ── اسکارف‌کار و برش‌کار (روبروی هم) ──
        _make_role_section("اسکارف‌کار و برش‌کار", "#3db880", scarfs, cols=2)

        # ── نفرات شیفت ──
        g_shift = _make_role_section("نفرات شیفت", "#4a9fd4", shifts, cols=4)
        if g_shift is not None:
            self._grid_frame = g_shift

        # ── نفرات شیفت فروش ──
        _make_role_section("نفرات شیفت فروش", "#9a70c8", sales, cols=2)

        # نقش‌های ناشناخته (اگر باشد)
        if others:
            _make_role_section("سایر کاربران", "#8a9098", others, cols=4)

    def _build_avatar_rgba(self, role_color, size, crown):
        """تصویر RGBA دایره‌ای آواتار را می‌سازد (با گوشه‌های کاملاً شفاف)"""
        if not hasattr(self, "_avatar_rgba_cache"):
            self._avatar_rgba_cache = {}
        key = (role_color, size, crown)
        if key in self._avatar_rgba_cache:
            return self._avatar_rgba_cache[key]
        from PIL import Image, ImageDraw
        SS = 4  # سوپرسمپل برای لبه‌های نرم
        S = size * SS
        img = Image.new("RGBA", (S, S), (0, 0, 0, 0))
        d = ImageDraw.Draw(img)

        pad = S * 0.04
        d.ellipse([pad, pad, S - pad, S - pad], fill=role_color)
        inner_pad = S * 0.07
        d.ellipse([inner_pad, inner_pad, S - inner_pad, S - inner_pad],
                  fill=self._shade_hex(role_color, 0.86))

        icon_color = "#ffffff"
        head_r = S * 0.155
        head_cx, head_cy = S / 2, S * 0.40
        d.ellipse([head_cx - head_r, head_cy - head_r,
                   head_cx + head_r, head_cy + head_r], fill=icon_color)
        body_w = S * 0.46
        body_top = S * 0.57
        d.pieslice([S / 2 - body_w / 2, body_top,
                    S / 2 + body_w / 2, body_top + body_w],
                   180, 360, fill=icon_color)

        if crown:
            cw = S * 0.46
            cx, cy = S / 2, S * 0.03
            base_y = cy + S * 0.13
            pts = [
                (cx - cw / 2, base_y), (cx - cw / 2, cy + S * 0.05),
                (cx - cw / 4, cy + S * 0.10), (cx, cy),
                (cx + cw / 4, cy + S * 0.10), (cx + cw / 2, cy + S * 0.05),
                (cx + cw / 2, base_y),
            ]
            d.polygon(pts, fill="#f0c84a")
            d.ellipse([cx - S * 0.018, cy - S * 0.018, cx + S * 0.018, cy + S * 0.018],
                      fill="#fff2c0")

        img = img.resize((size, size), Image.LANCZOS)
        self._avatar_rgba_cache[key] = img
        return img

    def _get_avatar_icon(self, role_color, size=30, crown=False):
        """
        نسخه‌ی PhotoImage ساده (روی پس‌زمینه‌ی تخت) برای جاهایی که پشت‌زمینه
        ثابت دارند (مثل پاپ‌آپ رمز عبور). برای آیکون‌های روی تصویر زمینه از
        _register_icon_widget استفاده کنید که گوشه‌هایش هم شفاف می‌شوند.
        """
        if not hasattr(self, "_avatar_icon_cache"):
            self._avatar_icon_cache = {}
        key = (role_color, size, crown)
        if key in self._avatar_icon_cache:
            return self._avatar_icon_cache[key]
        try:
            from PIL import ImageTk
            img = self._build_avatar_rgba(role_color, size, crown)
            ph = ImageTk.PhotoImage(img)
            self._avatar_icon_cache[key] = ph
            return ph
        except Exception:
            return None

    def _register_icon_widget(self, label, role_color, size, crown, tint_hex=None, tint_alpha=0.0):
        """
        این Label را برای ترکیب (alpha composite) با تصویر پس‌زمینه ثبت می‌کند
        — یعنی به‌جای گوشه‌های سیاه دور دایره‌ی آیکون، خودِ عکس پس‌زمینه را
        دقیقاً پشت آیکون می‌چسباند تا گوشه‌ها هم شفاف به‌نظر برسند.
        tint_hex/tint_alpha باید همان مقداری باشد که کارت اطراف آیکون با آن
        رنگ‌آمیزی شده، وگرنه گوشه‌های آیکون روشن‌تر از اطرافش به‌نظر می‌رسند
        و انگار یک «قاب مربعی» دور آیکون افتاده است.
        """
        if not hasattr(self, "_icon_widgets"):
            self._icon_widgets = []
        self._icon_widgets.append((label, role_color, size, crown, tint_hex, tint_alpha))

    def _refresh_login_icons(self):
        """آیکون‌های کاربران — با یا بدون تصویر زمینه."""
        for lbl, role_color, size, crown, tint_hex, tint_alpha in getattr(
                self, "_icon_widgets", []):
            if lbl is not None and lbl.winfo_exists():
                self._apply_icon_composite(
                    lbl, role_color, size, crown, tint_hex, tint_alpha)

    def _icon_label_bg(self, label):
        """رنگ پس‌زمینهٔ کارت والد — برای جلوگیری از قاب سفید دور آیکون."""
        w = label
        for _ in range(6):
            if w is None:
                break
            try:
                bg = w.cget("bg")
                if bg and str(bg).lower() not in ("", "systemwindow", "systembuttonface"):
                    return bg
            except Exception:
                pass
            w = getattr(w, "master", None)
        return C["bg"]

    def _apply_icon_composite(self, label, role_color, size, crown, tint_hex=None, tint_alpha=0.0):
        try:
            if not label.winfo_exists():
                return
            card_bg = tint_hex or self._icon_label_bg(label)
            try:
                label.configure(bg=card_bg, highlightthickness=0, bd=0)
            except Exception:
                pass
            if self._login_bg_pil is None:
                ph = self._get_avatar_icon(role_color, size, crown)
                if ph:
                    label.configure(image=ph, bg=card_bg)
                    label.image = ph
                return
            from PIL import Image, ImageTk
            label.update_idletasks()
            avatar_rgba = self._build_avatar_rgba(role_color, size, crown)
            if self._login_bg_pil is not None:
                bleed = int(getattr(self, "_login_bg_bleed", 0) or 0)
                ox = label.winfo_rootx() - self.winfo_rootx() + bleed
                oy = label.winfo_rooty() - self.winfo_rooty() + bleed
                full = self._login_bg_pil
                box = (max(0, ox), max(0, oy),
                       min(full.width, ox + size), min(full.height, oy + size))
                if box[2] > box[0] and box[3] > box[1]:
                    crop = full.crop(box).convert("RGB")
                    if crop.size != (size, size):
                        crop = crop.resize((size, size))
                    if tint_hex and tint_alpha > 0:
                        tint_layer = Image.new("RGB", crop.size, tint_hex)
                        crop = Image.blend(crop, tint_layer, tint_alpha)
                    crop = crop.convert("RGBA")
                else:
                    crop = Image.new("RGBA", (size, size), (40, 50, 65, 255))
            else:
                crop = Image.new("RGBA", (size, size), (40, 50, 65, 255))
            composed = Image.alpha_composite(crop, avatar_rgba)
            ph = ImageTk.PhotoImage(composed)
            label.configure(image=ph, bg=card_bg)
            label.image = ph
        except Exception:
            try:
                ph = self._get_avatar_icon(role_color, size, crown)
                if ph:
                    label.configure(image=ph, bg=self._icon_label_bg(label))
                    label.image = ph
            except Exception:
                pass

    @staticmethod
    def _shade_hex(hex_color, factor):
        """رنگ هگز را factor (0..1) برابر تیره‌تر/روشن‌تر می‌کند"""
        try:
            hex_color = hex_color.lstrip("#")
            r, g, b = (int(hex_color[i:i+2], 16) for i in (0, 2, 4))
            r, g, b = (max(0, min(255, int(c * factor))) for c in (r, g, b))
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return hex_color

    def _make_user_card(self, parent, uname, udata, role_labels,
                         full_width=False, row=0, col=0):
        role_lbl, role_color, role_icon = role_labels.get(
            udata["role"], ("کاربر", C["text_dim"], "◎"))
        suspended = udata.get("suspended", False)

        BG     = "#3a4050" if suspended else "#2e3a4a"
        FG     = "#8aaaC0" if suspended else "#e8edf5"
        BORDER = "#4a5060" if suspended else role_color
        # شیشه‌ای خیلی شفاف — دور کادر رنگی سیاه دیده نشود
        TINT = 0.08

        # فقط یک لایه کارت (بدون سایه‌های تیره که سیاهی اضافه می‌کنند)
        card = tk.Frame(parent, bg=BG,
                        highlightthickness=2,
                        highlightbackground=BORDER,
                        cursor="hand2" if not suspended else "arrow")
        if full_width:
            card.pack(pady=(0, 6), padx=60, anchor="center")
        else:
            card.grid(row=row, column=col, padx=6, pady=4, sticky="nsew")
        self._user_card_widgets.append((card, BG, TINT))

        tk.Frame(card, bg=BORDER if not suspended else "#555", height=2).pack(fill="x")

        if full_width:
            inner = tk.Frame(card, bg=BG)
            inner.pack(padx=18, pady=8)
            self._user_card_widgets.append((inner, BG, TINT))

            admin_icon_ph = self._get_avatar_icon(role_color, size=36, crown=True)
            admin_icon_lbl = tk.Label(inner, image=admin_icon_ph, bg=BG, bd=0, highlightthickness=0)
            admin_icon_lbl.pack(side="left", padx=(0, 12))
            admin_icon_lbl.image = admin_icon_ph
            self._register_icon_widget(admin_icon_lbl, role_color, 36, True, tint_hex=BG, tint_alpha=TINT)

            mid = tk.Frame(inner, bg=BG)
            mid.pack(side="left")
            self._user_card_widgets.append((mid, BG, TINT))
            name_cv = self._make_glass_label(
                mid, text=udata["display"], font=("B Nazanin", 15, "bold"),
                fg="#f0c040", padx=4, pady=2, anchor="w",
                tint_hex=BG, tint_alpha=TINT,
            )
            name_cv.pack(anchor="w")

        else:
            inner = tk.Frame(card, bg=BG)
            inner.pack(fill="both", expand=True, padx=8, pady=6)
            self._user_card_widgets.append((inner, BG, TINT))

            user_role_color = role_color if not suspended else "#6a7280"
            user_icon_ph = self._get_avatar_icon(user_role_color, size=30)
            user_icon_lbl = tk.Label(inner, image=user_icon_ph, bg=BG, bd=0, highlightthickness=0)
            user_icon_lbl.pack(anchor="center", pady=(2, 3))
            user_icon_lbl.image = user_icon_ph
            self._register_icon_widget(user_icon_lbl, user_role_color, 30, False, tint_hex=BG, tint_alpha=TINT)

            name_cv = self._make_glass_label(
                inner, text=udata["display"], font=("B Nazanin", 11, "bold"),
                fg=FG, padx=4, pady=2, wraplength=140, justify="center",
                tint_hex=BG, tint_alpha=TINT,
            )
            name_cv.pack(anchor="center", pady=(0, 2))

            if suspended:
                sus_cv = self._make_glass_label(
                    inner, text="⊘  تعلیق شده", font=("B Nazanin", 8, "bold"),
                    fg="#ff6060", padx=6, pady=2,
                    tint_hex="#3a1010", tint_alpha=0.25,
                )
                sus_cv.pack(pady=(2, 0))

        if not suspended:
            def bind_all(widget, u=uname):
                widget.bind("<Button-1>", lambda e, uu=u: self._open_pw_popup(uu))
                for ch in widget.winfo_children():
                    try:
                        bind_all(ch, u)
                    except Exception:
                        pass
            card.after(60, lambda c=card, u=uname: bind_all(c, u))

            def on_enter(e, c=card, rc=role_color):
                try:
                    c.config(highlightbackground=rc, highlightthickness=2)
                except Exception:
                    pass

            def on_leave(e, c=card, rc=BORDER):
                try:
                    c.config(highlightbackground=rc, highlightthickness=2)
                except Exception:
                    pass

            card.bind("<Enter>", on_enter)
            card.bind("<Leave>", on_leave)


    def _finish_user_login(self, udata):
        self._was_full = True
        self._allow_restore = False
        self._fullscreen_guard = False
        force_full_screen(self)
        self.update_idletasks()
        self.update()
        self._login(udata)

    def _open_pw_popup(self, uname):
        self.selected_user = uname
        udata = self.db["users"][uname]
        role_labels = {
            "admin":   ("مدیریت سیستم", "#d4a043"),
            "shift":   ("اپراتور شیفت", "#4a9fd4"),
            "scarf":   ("برش‌کار",       "#3db880"),
            "shift_n": ("نوبت‌کار",      "#9a70c8"),
        }
        role_icons = {
            "admin": "▣", "shift": "◆", "scarf": "◉", "shift_n": "◍"
        }
        role_name, role_color = role_labels.get(udata["role"], ("کاربر", C["text_dim"]))
        role_icon = role_icons.get(udata["role"], "◎")

        popup = tk.Toplevel(self)
        popup.title("")
        popup.configure(bg=C["bg"])
        popup.resizable(False, False)
        popup.overrideredirect(True)
        prepare_popup_window(popup, self)
        popup.lift()
        popup.focus_force()

        def _free_popup_grab(event=None):
            _release_grab_safe(popup)
            _release_grab_safe(self)

        popup.bind("<Destroy>", _free_popup_grab, add="+")

        W, H = 440, 400
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        popup.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

        # ── کانتینر با حاشیه رنگی ۳بعدی ──
        outer = tk.Frame(popup, bg=role_color, padx=2, pady=2)
        outer.pack(fill="both", expand=True)
        main = tk.Frame(outer, bg=C["bg"])
        main.pack(fill="both", expand=True)

        # ── سایه داخلی ──
        tk.Frame(main, bg=C["header_bg"], height=1).pack(fill="x")

        # ── نوار بالا ──
        top_bar = tk.Frame(main, bg=C["header_bg"], height=40)
        top_bar.pack(fill="x")
        top_bar.pack_propagate(False)

        # drag
        self._drag_x = self._drag_y = 0
        def start_drag(e): self._drag_x=e.x_root; self._drag_y=e.y_root
        def do_drag(e):
            dx=e.x_root-self._drag_x; dy=e.y_root-self._drag_y
            x=popup.winfo_x()+dx; y=popup.winfo_y()+dy
            popup.geometry(f"+{x}+{y}")
            self._drag_x=e.x_root; self._drag_y=e.y_root
        top_bar.bind("<ButtonPress-1>", start_drag)
        top_bar.bind("<B1-Motion>", do_drag)

        tk.Label(top_bar, text="  احراز هویت", bg=C["header_bg"],
                 fg=C["text_dim"], font=(_MAIN_FONT, 10, "bold")).pack(side="left", padx=8, pady=8)
        close_btn = tk.Label(top_bar, text="  ✕  ", bg=C["header_bg"],
                              fg=C["text_dim"], font=(_MAIN_FONT, 11, "bold"), cursor="hand2")
        close_btn.pack(side="right", padx=4)
        def _close_popup(e=None):
            try:
                popup.grab_release()
            except Exception:
                pass
            popup.destroy()
            return "break"
        close_btn.bind("<Button-1>", _close_popup)
        close_btn.bind("<Enter>", lambda e: close_btn.config(fg=C["danger"]))
        close_btn.bind("<Leave>", lambda e: close_btn.config(fg=C["text_dim"]))

        # ── محتوای اصلی ──
        body = tk.Frame(main, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=30, pady=20)

        # آیکون آدمک
        icon_f = tk.Frame(body, bg=C["bg"])
        icon_f.pack(pady=(0,12))
        popup_icon_ph = self._get_avatar_icon(role_color, size=64, crown=(udata.get("role") == "admin"))
        tk.Label(icon_f, image=popup_icon_ph, bg=C["bg"]).pack()
        icon_f._avatar_ref = popup_icon_ph  # جلوگیری از garbage collect

        # نام
        tk.Label(body, text=udata["display"], bg=C["bg"], fg=C["text_bright"],
                 font=(_MAIN_FONT, 16, "bold")).pack(pady=(0,20))

        # ── ورودی رمز ──
        pw_label = tk.Label(body, text="رمز عبور:", bg=C["bg"],
                             fg=C["text_dim"], font=(_MAIN_FONT, 10, "bold"))
        pw_label.pack(anchor="e", pady=(0,4))

        pw_wrap = tk.Frame(body, bg=role_color, padx=2, pady=2)
        pw_wrap.pack(fill="x")
        pw_inner = tk.Frame(pw_wrap, bg=C["entry_bg"])
        pw_inner.pack(fill="x")

        pw_var = tk.StringVar()
        pw_entry = tk.Entry(pw_inner, textvariable=pw_var,
                            show="●", bg=C["entry_bg"], fg=role_color,
                            insertbackground=role_color,
                            font=("B Nazanin", 14, "bold"),
                            bd=0, relief="flat", justify="center")
        pw_entry.pack(fill="x", ipady=8, padx=8)

        # نمایش/پنهان رمز
        show_var = tk.BooleanVar(value=False)
        def toggle_show():
            show_var.set(not show_var.get())
            pw_entry.config(show="" if show_var.get() else "●")
            eye_lbl.config(text="🙈" if show_var.get() else "👁")
        eye_lbl = tk.Label(pw_inner, text="👁", bg=C["entry_bg"],
                            fg=C["text_dim"], font=("Segoe UI Emoji", 13, "bold"),
                            cursor="hand2")
        eye_lbl.place(relx=1.0, rely=0.5, anchor="e", x=-8)
        eye_lbl.bind("<Button-1>", lambda e: toggle_show())

        err_lbl = tk.Label(body, text="", bg=C["bg"], fg=C["danger"],
                            font=(_MAIN_FONT, 10, "bold"))
        err_lbl.pack(pady=(6,0))

        def shake():
            orig = popup.winfo_x()
            for dx in [10,-10,8,-8,5,-5,2,-2,0]:
                popup.geometry(f"+{orig+dx}+{popup.winfo_y()}")
                popup.update()
                time.sleep(0.03)

        def do_login():
            pw = pw_var.get()
            h = hash_pw(pw)
            if h == udata["password"]:
                popup.destroy()
                self.update_idletasks()
                self._finish_user_login(udata)
            else:
                err_lbl.config(text="رمز عبور اشتباه است")
                pw_var.set("")
                pw_entry.focus_set()
                shake()

        # ── دکمه ورود — همیشه سبز ──
        _login_btn_color = "#2a7850"
        _login_btn_hover = "#1e6040"
        btn_frame = tk.Frame(body, bg=_login_btn_color, cursor="hand2")
        btn_frame.pack(fill="x", pady=(16,0))
        login_lbl = tk.Label(btn_frame, text="ورود به سیستم  ▶",
                              bg=_login_btn_color, fg="#ffffff",
                              font=(_MAIN_FONT, 16, "bold"), pady=16, cursor="hand2")
        login_lbl.pack(fill="x")
        btn_frame.bind("<Button-1>", lambda e: do_login())
        login_lbl.bind("<Button-1>", lambda e: do_login())
        btn_frame.bind("<Enter>", lambda e: btn_frame.config(bg=_login_btn_hover) or login_lbl.config(bg=_login_btn_hover))
        btn_frame.bind("<Leave>", lambda e: btn_frame.config(bg=_login_btn_color) or login_lbl.config(bg=_login_btn_color))

        pw_entry.focus_set()
        pw_entry.bind("<Return>", lambda e: do_login())
        popup.bind("<Escape>", lambda e: popup.destroy())

        # ── تضمین نمایش کامل دکمه «ورود به سیستم» ──
        # برای کاربر «مدیریت سیستم» به‌خاطر آیکون تاج اضافه، محتوا کمی بلندتر از
        # ارتفاع ثابت پاپ‌آپ می‌شود و دکمه ورود از دید خارج می‌شد و فقط Enter کار می‌کرد.
        # اینجا بعد از چیدمان کامل ویجت‌ها، ارتفاع واقعی لازم را محاسبه و در صورت نیاز
        # پنجره را بزرگ‌تر و دوباره وسط صفحه قرار می‌دهیم تا دکمه برای همه کاربران دیده شود.
        popup.update_idletasks()
        req_h = outer.winfo_reqheight() + 4
        if req_h > H:
            popup.geometry(f"{W}x{req_h}+{(sw-W)//2}+{(sh-req_h)//2}")
        try:
            popup.attributes("-topmost", True)
            popup.lift()
            pw_entry.focus_set()
        except Exception:
            pass

    def _login(self, udata):
        _set_db_user(udata.get("display") or self.selected_user)
        # ثبت لاگ ورود
        db = load_db()
        db.setdefault("login_log", []).append({
            "username": self.selected_user,
            "display":  udata.get("display",""),
            "role":     udata.get("role",""),
            "at":       now_str(),
            "logout_at": "",
        })
        log_idx = len(db["login_log"]) - 1
        save_db(db)
        MainApp.from_login_window(self, self.selected_user, udata, log_idx)

# ═══════════════════════════════════════════════════════════
#  برنامه اصلی
# ═══════════════════════════════════════════════════════════
class MainApp(tk.Tk):
    def __init__(self, username, udata, login_log_index=None):
        super().__init__()
        self._init_main_state(username, udata, login_log_index)
        self._finish_main_setup()

    def _init_main_state(self, username, udata, login_log_index=None, in_place=False, db=None):
        self.username = username
        self.udata = udata
        _set_db_user(udata.get("display") or username)
        self.db = db if db is not None else load_db()
        self.role = udata["role"]
        self._login_log_index = login_log_index
        _ensure_backup_dirs_protected()
        if os.environ.get("STF_CLIENT") == "1":
            try:
                from client.db_bridge import set_current_user
                set_current_user(username)
            except Exception:
                pass
        self.title(f"فولاد سفید دشت  |  {udata['display']}")
        if not in_place:
            self.geometry("1320x840")
        self.minsize(1040, 680)
        self.resizable(True, True)
        self._on_restore_attempt = self._ask_restore_password
        self._on_minimize_attempt = self._ask_minimize_password
        self._on_close_attempt = self._on_close
        self._window_action_in_progress = False
        self.configure(bg=C["bg"])
        self._drag_locked = True
        self._app_closing = False
        self._allow_iconify = False
        self._minimizing_authorized = False
        self._was_in_taskbar = False
        self._allow_restore = False
        self._lock_minimize_ok = False
        self._lock_restore_ok = False
        self._was_full = True
        self._restore_after_id = None
        self._restore_dialog_open = False
        self._minimize_dialog_open = False
        self._lock_dialog_open = False
        self._cfg_debounce_id = None
        self._watchdog_proc = None

    def _finish_main_setup(self, in_place=False):
        if os.environ.get("STF_ADMIN") == "1":
            self._apply_bg()
            self._build_ui()
            self.protocol("WM_DELETE_WINDOW", self._on_close)
            _install_taskbar_restore_guard(self)
            self.geometry("1320x840")
            self.minsize(1040, 680)
            self._was_full = False
            return
        hide_all_taskbars(self)
        if in_place:
            self._window_action_in_progress = True
            self._was_full = True
            self._allow_restore = False
            self._fullscreen_guard = False
            force_full_screen(self)
            self.update_idletasks()
            self.update()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        force_full_screen(self)
        lock_window_chrome(self)
        lock_window_drag(self)
        enforce_frameless_window(self)
        bind_window_lock_keys(self)
        _install_destroy_guard(self)
        start_keyboard_lock(self)
        start_taskbar_hide_loop(self)
        _start_kiosk_reinforce_loop(self)
        _start_foreground_guard(self)
        try:
            install_keyboard_switcher(self)
        except Exception:
            pass
        self._apply_bg()
        self._build_ui()
        self.pack_propagate(False)
        self.update_idletasks()
        hide_all_taskbars(self)
        force_full_screen(self)
        if in_place:
            _finalize_fullscreen_window(self)
        else:
            enforce_full_screen_robust(self)
            self.after(400, lambda: _finalize_fullscreen_window(self))
        self.bind("<Unmap>", self._on_unmap)
        self.bind("<Map>", self._on_map)
        self.bind("<Configure>", self._on_configure)
        self._start_taskbar_watchdog()
        self.after(300, lambda: _release_grab_safe(self))
        _ensure_app_not_topmost(self)
        _ensure_app_not_topmost(self)
        self.update_idletasks()
        _raise_window_front(self)
        if in_place:
            force_full_screen(self)
            self.update_idletasks()
            self.update()
            _end_window_transition(self)
        else:
            self._window_action_in_progress = False

    @classmethod
    def from_login_window(cls, win, username, udata, login_log_index=None):
        """تبدیل LoginWindow به MainApp روی همان پنجره — بدون بستن و بدون مکث."""
        _prepare_window_transition(win)
        stop_keyboard_lock(win)
        old_watchdog = getattr(win, "_watchdog_proc", None)
        win._watchdog_proc = None

        _begin_window_transition(win)
        cover = getattr(win, "_transition_cover", None)

        for ev in ("<Unmap>", "<Map>", "<Configure>"):
            try:
                win.unbind(ev)
            except Exception:
                pass

        win._drag_locked = True
        for w in list(win.winfo_children()):
            if w is not cover:
                w.destroy()
        win.update_idletasks()
        win.update()

        win._login_datetime = None
        win._drag_lock_installed = False
        global _ACTIVE_LOGIN_WIN
        _ACTIVE_LOGIN_WIN = None

        win.__class__ = cls
        win._init_main_state(username, udata, login_log_index, in_place=True, db=getattr(win, "db", None))
        win._finish_main_setup(in_place=True)
        try:
            import main as _main_mod
            _main_mod._ACTIVE_MAIN_APP = win
        except Exception:
            pass

        if old_watchdog is not None:
            try:
                old_watchdog.terminate()
            except Exception:
                pass

    def _on_configure(self, event):
        """جلوگیری از Restore Down آزاد — debounce برای جلوگیری از حلقه و کندی"""
        if event.widget is not self:
            return
        if getattr(self, "_app_closing", False):
            return
        if getattr(self, "_fullscreen_guard", False):
            return
        if getattr(self, "_cfg_debounce_id", None):
            try:
                self.after_cancel(self._cfg_debounce_id)
            except Exception:
                pass
        self._cfg_debounce_id = self.after(80, self._check_fullscreen_state)

    def _check_fullscreen_state(self):
        self._cfg_debounce_id = None
        if getattr(self, "_app_closing", False):
            return
        if getattr(self, "_fullscreen_guard", False):
            return
        if getattr(self, "_window_action_in_progress", False):
            return
        if _modal_dialog_active(self):
            return
        if _window_is_minimized(self):
            return
        is_full = is_window_full_screen(self)
        if not is_full and self._was_full:
            if self._allow_restore:
                self._allow_restore = False
                self._was_full = False
                return
            force_full_screen(self)
            self._was_full = True
            self._drag_locked = True
            if self._restore_after_id is not None:
                try:
                    self.after_cancel(self._restore_after_id)
                except Exception:
                    pass
                self._restore_after_id = None
            if not self._restore_dialog_open:
                self._restore_after_id = self.after(
                    60, self._ask_restore_password)
            return
        self._was_full = is_full
        if is_full:
            hide_all_taskbars(self)

    def _ask_restore_password(self):
        self._restore_after_id = None
        request_window_restore_or_maximize(self)

    def _ask_minimize_password(self):
        request_window_minimize(self)

    def _start_taskbar_watchdog(self):
        """پروسه‌ی نگهبان مستقل تسک‌بار را اجرا می‌کند (safety net)."""
        self._watchdog_proc = start_taskbar_watchdog()

    def _on_map(self, event):
        """بازگشت از مینیمایز — در حالت تمام‌صفحه تسک‌بار مخفی می‌ماند."""
        if event.widget is not self:
            return
        if getattr(self, "_minimizing_authorized", False):
            _try_finish_authorized_minimize_on_map(self)
        if getattr(self, "_was_full", True) and _app_shell_lock_active(self):
            hide_all_taskbars(self)

    def _on_unmap(self, event):
        """اگر بدون رمز مینیمایز شد — برگردان."""
        if event.widget is not self:
            return
        if getattr(self, "_app_closing", False):
            return
        if self._allow_iconify or getattr(self, "_minimizing_authorized", False):
            self._allow_iconify = False
            if getattr(self, "_minimizing_authorized", False):
                self.after(80, lambda: _mark_minimize_in_taskbar(self))
            return
        try:
            if self.state() == "iconic" or _hwnd_is_iconic(self):
                self.after(1, self._recover_from_minimize)
        except Exception:
            pass

    def _recover_from_minimize(self):
        try:
            win_restore(self)
            hide_all_taskbars(self)
            force_full_screen(self)
            if not getattr(self, "_minimize_dialog_open", False):
                self.after(80, self._ask_minimize_password)
        except Exception:
            pass

    def _verify_lock_password(self, title, prompt):
        """رمز قفل نرم‌افزار را می‌پرسد (همان رمزی که برای مینیمایز و بستن
        هر دو استفاده می‌شود) و True/False برمی‌گرداند."""
        return verify_lock_password(self, title, prompt)

    def _apply_bg(self):
        s = self.db.get("settings", {})
        # بارگذاری تم ذخیره شده
        saved_theme = s.get("theme", {})
        if saved_theme:
            C.update(saved_theme)
        else:
            bg = s.get("background_color", C["bg"])
            C["bg"] = bg
        self.configure(bg=C["bg"])

    def _apply_saved_icons(self):
        """بارگذاری آیکون‌های ذخیره شده روی دکمه‌های تب"""
        saved_icons = self.db.get("settings", {}).get("tab_icons", {})
        for key, ic in saved_icons.items():
            if key in self._tab_buttons:
                lbl = tr(f"tab_{key}")
                self._tab_buttons[key].config(text=f"{ic}  {lbl}")

    def _record_logout(self):
        """ثبت تاریخ و ساعت خروج در همان رکورد لاگ ورود این نشست"""
        try:
            if self._login_log_index is None:
                return
            idx = self._login_log_index

            def _bg_logout_save():
                try:
                    db = load_db()
                    logs = db.get("login_log", [])
                    if 0 <= idx < len(logs):
                        logs[idx]["logout_at"] = now_str()
                        save_db(db, action="logout")
                except Exception:
                    pass

            # Client: هرگز save سنگین روی UI thread — آفلاین/دیسک هنگ می‌کند
            if os.environ.get("STF_CLIENT") == "1":
                import threading
                threading.Thread(target=_bg_logout_save, daemon=True, name="stf-logout-save").start()
            else:
                _bg_logout_save()
        except Exception:
            pass

    def _on_close(self):
        def _finish():
            if _ask_yesno_topmost(
                    self,
                    tr("save_exit_title"),
                    tr("save_exit_msg")):
                try:
                    # Client: ذخیره روی پس‌زمینه تا UI قفل نشود
                    if os.environ.get("STF_CLIENT") == "1":
                        import threading
                        def _bg_save():
                            try:
                                save_db(load_db())
                            except Exception:
                                pass
                        threading.Thread(target=_bg_save, daemon=True, name="stf-exit-save").start()
                        time.sleep(0.15)
                    else:
                        save_db(load_db())
                except Exception as ex:
                    if not _ask_yesno_topmost(
                            self,
                            tr("save_err_title"),
                            tr("save_err_msg", err=ex)):
                        return
            self._record_logout()
            _prepare_app_shutdown(self)
            stop_keyboard_lock(self)
            show_all_taskbars()
            self.destroy()
        request_window_close(self, _finish)

    def _build_ui(self):
        # ════════════════════════════════════════════
        # هدر فولادی صنعتی
        # ════════════════════════════════════════════
        tk.Frame(self, bg="#d4a043", height=3).pack(fill="x")
        tk.Frame(self, bg="#8a6020", height=1).pack(fill="x")

        header = tk.Frame(self, bg=C["header_bg"], height=82)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Frame(self, bg="#d4a043", height=1).pack(fill="x")
        tk.Frame(self, bg=C["border"],  height=1).pack(fill="x")

        # ══════════════════════════════════
        # چپ: تاریخ و ساعت
        # ══════════════════════════════════
        left_area = tk.Frame(header, bg=C["header_bg"])
        left_area.pack(side="left", padx=(28, 0), fill="y")

        # تاریخ و ساعت + نام کاربر
        clock_f = tk.Frame(left_area, bg=C["header_bg"])
        clock_f.pack(side="left", fill="y")
        role_icons_m  = {"admin":"▣","shift":"◆","scarf":"◉","shift_n":"◍"}
        role_colors_m = {"admin":"#d4a043","shift":"#4a9fd4","scarf":"#3db880","shift_n":"#9a70c8"}
        u_icon  = role_icons_m.get(self.role, "◎")
        u_color = role_colors_m.get(self.role, C["text_dim"])
        tk.Label(clock_f, text=f"{u_icon}  {self.udata['display']}",
                 bg=C["header_bg"], fg=u_color,
                 font=(_MAIN_FONT, 10, "bold")).pack(anchor="w", pady=(4, 0))
        self._datetime_lbl = tk.Label(clock_f, text="",
                                   bg=C["header_bg"], fg="#d4a043",
                                   font=("B Nazanin", 11, "bold"))
        self._datetime_lbl.pack(anchor="w", padx=(8, 0), pady=(0, 0))
        # وضعیت آنلاین/آفلاین — برای ادمین و کلاینت (مثل صفحه لاگین)
        self._online_lbl = tk.Label(
            clock_f, text=tr("online"), bg=C["header_bg"], fg=C["success"],
            font=(_MAIN_FONT, 10, "bold"), anchor="w")
        self._online_lbl.pack(anchor="w", padx=(8, 0), pady=(0, 4))
        self._tick()

        # ══════════════════════════════════
        # راست: ذخیره‌سازی + بروزرسانی
        # ══════════════════════════════════
        right_area = tk.Frame(header, bg=C["header_bg"])
        right_area.pack(side="right", padx=(0, 168), fill="y")

        mount_window_controls(header, self, size=40, margin_right=12)

        # بروزرسانی F5
        rf_f = tk.Frame(right_area, bg=C["card"], cursor="hand2",
                         highlightthickness=1, highlightbackground=C["border"])
        rf_f.pack(side="right", padx=(6,0), pady=14)
        self._rf_lbl = tk.Label(rf_f, text=tr("refresh"),
                           bg=C["card"], fg=C["text_dim"],
                           font=(_MAIN_FONT, 10, "bold"),
                           padx=10, pady=8)
        self._rf_lbl.pack()
        rf_f.bind("<Button-1>", lambda e: self._refresh_current_tab())
        self._rf_lbl.bind("<Button-1>", lambda e: self._refresh_current_tab())
        rf_f.bind("<Enter>", lambda e: rf_f.config(bg=C["card_hover"]) or self._rf_lbl.config(bg=C["card_hover"]))
        rf_f.bind("<Leave>", lambda e: rf_f.config(bg=C["card"]) or self._rf_lbl.config(bg=C["card"]))
        self.bind_all("<F5>", lambda e: self._refresh_current_tab())
        # اسکرول مرکزی — یک bind_all برای کل برنامه
        self.bind_all("<MouseWheel>", _global_mousewheel)

        # ذخیره‌سازی — همه کاربران
        bk_f = tk.Frame(right_area, bg="#2a4838", cursor="hand2",
                        highlightthickness=1, highlightbackground="#3a5848")
        bk_f.pack(side="right", padx=(6,0), pady=14)
        bk_lbl = tk.Label(bk_f, text=tr("save_btn"),
                           bg="#2a4838", fg="#3db880",
                           font=(_MAIN_FONT, 10, "bold"),
                           padx=10, pady=8)
        bk_lbl.pack()
        self._bk_status = tk.Label(right_area, text="", bg=C["header_bg"],
                                    fg=C["success"], font=(_MAIN_FONT, 8, "bold"))
        self._bk_status.pack(side="right", padx=2)

        def do_manual_backup():
            bd = BACKUP_DIR
            try:
                if r"D:\SteelFactory2-v2" not in sys.path:
                    sys.path.insert(0, r"D:\SteelFactory2-v2")
                from shared.backup_vault import encrypt_bytes_to_zip, set_vault_password
                if is_backup_vault_password_required():
                    set_vault_password(get_backup_vault_password())
                else:
                    set_vault_password("")
                import json as _json
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                raw = _json.dumps(load_db(), ensure_ascii=False, indent=2).encode("utf-8")
                data_dir = _backup_data_dir(bd)
                os.makedirs(data_dir, exist_ok=True)
                dst = os.path.join(data_dir, f"slab_db_MANUAL_{ts}.json")
                encrypt_bytes_to_zip(raw, __import__("pathlib").Path(dst),
                                    arcname=f"slab_db_MANUAL_{ts}.json")
                self._bk_status.config(text="✔ ذخیره شد")
                bk_f.config(bg="#1a3828", highlightbackground=C["success"])
                bk_lbl.config(bg="#1a3828")
                self.after(3000, lambda: [
                    bk_f.config(bg="#2a4838", highlightbackground="#3a5848"),
                    bk_lbl.config(bg="#2a4838"),
                    self._bk_status.config(text="")
                ])
            except Exception as ex:
                messagebox.showerror("خطا", f"ذخیره‌سازی ناموفق:\n{ex}", parent=self)

        bk_f.bind("<Button-1>", lambda e: do_manual_backup())
        bk_lbl.bind("<Button-1>", lambda e: do_manual_backup())
        bk_f.bind("<Enter>", lambda e: bk_f.config(bg="#1a3820") or bk_lbl.config(bg="#1a3820"))
        bk_f.bind("<Leave>", lambda e: bk_f.config(bg="#2a4838") or bk_lbl.config(bg="#2a4838"))

        # ══════════════════════════════════
        # وسط: Fe + عنوان سیستم (دقیقاً مرکز)
        # ══════════════════════════════════
        center_f = tk.Frame(header, bg=C["header_bg"])
        center_f.place(relx=.5, rely=.5, anchor="center")

        # آیکون Fe
        icon_box = tk.Frame(center_f, bg="#d4a043", width=46, height=46)
        icon_box.pack(side="right", padx=(0,12))
        icon_box.pack_propagate(False)
        tk.Label(icon_box, text="Fe", bg="#d4a043",
                 fg="#2e3338", font=("B Nazanin", 15, "bold")).place(relx=.5,rely=.5,anchor="center")

        # عناوین
        txt_f = tk.Frame(center_f, bg=C["header_bg"])
        txt_f.pack(side="right", fill="y", pady=8)
        tk.Label(txt_f, text=tr("app_title"),
                 bg=C["header_bg"], fg="#ffffff",
                 font=(_MAIN_FONT, 15, "bold")).pack(anchor="e")
        self._hdr_title_lbl = txt_f.winfo_children()[-1]
        tk.Label(txt_f, text=tr("app_subtitle"),
                 bg=C["header_bg"], fg="#d4a043",
                 font=(_MAIN_FONT, 9, "bold")).pack(anchor="e", pady=(2,0))
        self._hdr_subtitle_lbl = txt_f.winfo_children()[-1]

        bind_window_header_drag(self, header)
        refresh_header_drag_bindings(self)

        # ── Layout: sidebar راست + محتوا ──
        self._main_frame = tk.Frame(self, bg=C["bg"])
        self._main_frame.pack(fill="both", expand=True)

        # Sidebar راست
        self._sidebar = tk.Frame(self._main_frame, bg=C["panel"], width=200)
        self._sidebar.pack(side="right", fill="y")
        self._sidebar.pack_propagate(False)
        tk.Frame(self._main_frame, bg=C["border"], width=1).pack(side="right", fill="y")

        # محتوا
        self._content_area = tk.Frame(self._main_frame, bg=C["bg"])
        self._content_area.pack(side="right", fill="both", expand=True)

        self._tab_frames = {}    # key -> Frame
        self._tab_buttons = {}   # key -> Button
        self._current_tab = None

        self._build_tabs()
        self._apply_saved_icons()

        # انتخاب اولین تب
        if self._tab_frames:
            first_key = next(iter(self._tab_frames))
            self._switch_tab(first_key)

    def _add_tab(self, key, label, builder_fn):
        """اضافه کردن یک تب به sidebar"""
        btn = tk.Button(
            self._sidebar, text=label,
            bg=C["panel"], fg=C["text"],
            font=(_MAIN_FONT, 11, "bold"), bd=0, relief="flat",
            anchor="e", padx=16, pady=10,
            cursor="hand2", wraplength=170, justify="right",
            command=lambda k=key: self._switch_tab(k)
        )
        btn.pack(fill="x")
        tk.Frame(self._sidebar, bg=C["border"], height=1).pack(fill="x")

        def on_enter(e, b=btn): 
            if self._current_tab != key:
                b.config(bg=C["card_hover"])
        def on_leave(e, b=btn):
            if self._current_tab != key:
                b.config(bg=C["panel"])
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

        self._tab_buttons[key] = btn

        # محتوا رو lazy می‌سازیم
        frame = tk.Frame(self._content_area, bg=C["bg"])
        self._tab_frames[key] = (frame, builder_fn, False)  # False=not built yet

    def _switch_tab(self, key):
        if self._current_tab == key:
            return
        # پنهان کردن تب فعلی
        if self._current_tab and self._current_tab in self._tab_frames:
            old_frame = self._tab_frames[self._current_tab][0]
            old_frame.pack_forget()
            old_btn = self._tab_buttons.get(self._current_tab)
            if old_btn:
                old_btn.config(bg=C["panel"], fg=C["text"])

        self._current_tab = key
        frame, builder_fn, built = self._tab_frames[key]

        # ساختن محتوا اگه هنوز ساخته نشده
        if not built:
            try:
                builder_fn(frame)
            except Exception as _build_ex:
                import traceback
                err_msg = traceback.format_exc()
                tk.Label(frame, text=f"⚠️ خطا در بارگذاری تب:\n{_build_ex}",
                         bg=C["bg"], fg=C["danger"],
                         font=(_MAIN_FONT, 10), justify="right",
                         wraplength=600).pack(expand=True)
                print(f"[TAB BUILD ERROR] {key}:\n{err_msg}")
            self._tab_frames[key] = (frame, builder_fn, True)

        frame.pack(fill="both", expand=True)

        # هایلایت دکمه فعال
        btn = self._tab_buttons.get(key)
        if btn:
            btn.config(bg=C["accent"], fg="#ffffff")

        # وقتی تب نامه باز می‌شود، badge را فوری به‌روز کن
        if key == "ticket":
            self.after(300, self._ticket_badge_loop)

    def _tick(self):
        if not getattr(self, "_datetime_lbl", None):
            return
        try:
            if not self._datetime_lbl.winfo_exists():
                return
        except Exception:
            return
        delay = _ui_idle_ms(self, 1000, 3000, 10000)
        if _ui_heavy_ok(self) or delay <= 3000:
            now   = datetime.datetime.now()
            sh    = to_shamsi(now)
            parts = sh.split("  ")
            date_str = parts[0] if parts else ""
            time_str = parts[1] if len(parts) > 1 else ""
            if hasattr(self, "_datetime_lbl"):
                self._datetime_lbl.config(text=f"{date_str}  {time_str}")
            if not hasattr(self, "_conn_check_tick"):
                self._conn_check_tick = 0
            self._conn_check_tick += 1
            force = self._conn_check_tick >= 5
            if force:
                self._conn_check_tick = 0
            if _ui_heavy_ok(self):
                _refresh_client_connection_label(self, force_check=force)
        self.after(delay, self._tick)

    def _logout(self):
        if not _ask_okcancel_topmost(self, tr("logout_title"), tr("logout_msg")):
            return
        self._record_logout()
        self._was_full = True
        self._allow_restore = False
        self._fullscreen_guard = False
        force_full_screen(self)
        self.update_idletasks()
        # بدون self.update() اضافه — باعث هنگ می‌شد
        LoginWindow.from_main_app(self)

    def _soft_refresh_current_tab(self):
        """رفرش سبک تب فعلی — بدون destroy کل UI (اسکرول و سرعت).
        True = رفرش انجام شد؛ False = انجام نشد (debounce/بدون ویجت) تا caller دوباره تلاش کند.
        """
        now = time.time()
        last = float(getattr(self, "_soft_refresh_at", 0) or 0)
        if now - last < 0.55:
            return False  # debounce — نه موفقیت قلابی؛ وگرنه دادهٔ ادمین در UI می‌ماند
        self._soft_refresh_at = now
        key = self._current_tab
        if not key or key not in self._tab_frames:
            return False
        frame = self._tab_frames[key][0]
        # فقط خود تب + فرزندان مستقیم — عمیق‌تر (نوه) هنگ سنگین می‌آورد
        candidates = [frame]
        try:
            candidates.extend(list(frame.winfo_children()))
        except Exception:
            pass
        for w in candidates:
            for name in (
                "refresh_melts", "refresh_qc", "refresh_rejected",
                "refresh_warehouse", "refresh_all", "refresh",
            ):
                fn = getattr(w, name, None)
                if callable(fn):
                    try:
                        fn()
                        return True
                    except Exception:
                        pass
        return False

    def _refresh_current_tab(self):
        """بروزرسانی کلی — کش را در پس‌زمینه نو می‌کند؛ UI را بلاک نمی‌کند."""
        key = self._current_tab
        if not key or key not in self._tab_frames:
            return
        # جلوگیری از رفرش هم‌پوشان / حلقه
        if getattr(self, "_refresh_busy", False):
            # اگر بیش از ۸ ثانیه گیر کرده، آزاد کن
            started = float(getattr(self, "_refresh_busy_since", 0) or 0)
            if started and (time.time() - started) < 8:
                return
            self._refresh_busy = False
        self._refresh_busy = True
        self._refresh_busy_since = time.time()
        try:
            self._rf_lbl.config(text="  …  ")
        except Exception:
            pass

        def _finish_ui():
            self._refresh_busy = False
            try:
                self._rf_lbl.config(text=tr("refresh"))
            except Exception:
                pass

        def _apply_ui():
            try:
                # اسکرول را قبل از هر رفرشی قفل کن
                try:
                    frame = self._tab_frames[key][0]
                    for w in frame.winfo_children():
                        try:
                            if hasattr(w, "yview"):
                                self._stf_tab_yview = float(w.yview()[0])
                                break
                        except Exception:
                            pass
                except Exception:
                    pass
                if not self._soft_refresh_current_tab():
                    # فقط soft — destroy کامل هنگام ثبت اسلب هنگ می‌آورد
                    pass
            finally:
                _finish_ui()

        def _bg_reload():
            try:
                db = load_db(force=True)
                try:
                    self.db = db
                except Exception:
                    pass
            except Exception:
                pass
            try:
                self.after(0, _apply_ui)
            except Exception:
                self._refresh_busy = False

        # ایمنی: اگر thread گیر کرد، UI را آزاد کن
        try:
            self.after(10000, lambda: _finish_ui() if getattr(self, "_refresh_busy", False) else None)
        except Exception:
            pass

        try:
            import threading
            threading.Thread(target=_bg_reload, daemon=True, name="stf-ui-refresh").start()
        except Exception:
            try:
                # هرگز force روی UI thread نزن
                pass
            except Exception:
                pass
            _apply_ui()

    def _is_typing_now(self, frame):
        """بررسی می‌کند آیا کاربر هم‌اکنون در حال تایپ داخل یک فیلد ورودی، در همین تب است"""
        try:
            focused = self.focus_get()
        except Exception:
            return False
        if focused is None:
            return False
        if isinstance(focused, (tk.Entry, tk.Text, tk.Spinbox)):
            w = focused
            while w is not None:
                if w is frame:
                    return True
                try:
                    w = w.master
                except Exception:
                    break
            return False
        return False

    def _build_tabs(self):
        role = self.role
        if role == "admin":
            self._add_tab("home",    tab_label("home"),        self._build_home_tab)
        if role in ("shift", "admin"):
            self._add_tab("melts",   tab_label("melts"),   self._build_melts_tab)
            self._add_tab("qc",      tab_label("qc"),      self._build_qc_tab)
            self._add_tab("rejected",tab_label("rejected"),       self._build_rejected_tab)
            self._add_tab("transfer",tab_label("transfer"),     self._build_transfer_tab)
            self._add_tab("lab",     tab_label("lab"), self._build_lab_tab)
            self._add_tab("scrap",   tab_label("scrap"),           self._build_scrap_tab)
            self._add_tab("pdf",     tab_label("pdf"),      self._build_pdf_tab)
        if role == "shift_n":
            self._add_tab("qc",      tab_label("qc"),      self._build_qc_tab)
            self._add_tab("transfer",tab_label("transfer"),     self._build_transfer_tab)
        if role in ("scarf", "admin"):
            self._add_tab("scarf",   tab_label("scarf"),          self._build_scarf_tab)
            self._add_tab("cut",     tab_label("cut"),              self._build_cut_tab)
        if role == "scarf":
            pass  # اسکارف تب آزمایشگاه ندارد
        if role in ("shift_n", "admin"):
            self._add_tab("nobat",   tab_label("nobat"),       self._build_nobat_tab)
        # ── تب تیکت/نامه — برای همه کاربران ──
        if role != "admin":
            self._add_tab("ticket",  tab_label("ticket"),       self._build_ticket_tab)
        if role == "admin":
            self._add_tab("admin",   tab_label("admin"),    self._build_admin_tab)
            self._add_tab("ticket",  tab_label("ticket"),       self._build_ticket_tab)
        # ── شروع به‌روزرسانی badge نامه‌های خوانده‌نشده ──
        self.after(2000, self._ticket_badge_loop)

        # ── فضای خالی + دکمه خروج از حساب — پایین sidebar، زیر صندوق نامه ──
        tk.Frame(self._sidebar, bg=C["panel"]).pack(fill="both", expand=True)
        tk.Frame(self._sidebar, bg=C["border"], height=1).pack(fill="x")
        self._logout_btn = tk.Button(
            self._sidebar, text=tr("logout"),
            bg=C["panel"], fg=C["danger"],
            activebackground="#4a2020", activeforeground="#ff8080",
            font=(_MAIN_FONT, 11, "bold"), bd=0, relief="flat",
            anchor="e", padx=16, pady=12,
            cursor="hand2", wraplength=170, justify="right",
            command=self._logout,
        )
        self._logout_btn.pack(side="bottom", fill="x")
        self._logout_btn.bind("<Enter>", lambda e: self._logout_btn.config(bg="#3a1818"))
        self._logout_btn.bind("<Leave>", lambda e: self._logout_btn.config(bg=C["panel"]))

    # ═══════════════════════
    #  تب ۱: ثبت ذوب جدید
    # ═══════════════════════

    def _ticket_badge_loop(self):
        """هر ۸ ثانیه یک‌بار تعداد نامه‌های خوانده‌نشده را بررسی و badge دکمه را به‌روز می‌کند"""
        if not getattr(self, "_tab_buttons", None):
            return
        if getattr(self, "_app_closing", False):
            return
        try:
            if not self.winfo_exists():
                return
            db = load_db()
            unread = sum(
                1 for t in db.get("tickets", [])
                if t.get("to") == self.username and not t.get("read", False)
            )
            btn = self._tab_buttons.get("ticket")
            if btn:
                if unread > 0:
                    badge = f"+{unread}" if unread <= 99 else "+99"
                    btn.config(text=f"✉  صندوق نامه  🔴{badge}")
                else:
                    btn.config(text="✉  صندوق نامه")
            self.after(8000, self._ticket_badge_loop)
        except Exception:
            pass
    def _admin_popup(self, event, tree, db_key, get_rec_by_sid, on_refresh):
        """سازگاری با کدهای قدیمی — به inline editor هدایت می‌کند"""
        if self.role != "admin": return
        sel = tree.selection()
        if not sel: return
        vals = tree.item(sel[0], "values")
        sid_raw = vals[0] if vals else ""
        sid = sid_raw.replace("↳","").replace(" ","").replace("  ","")
        result = get_rec_by_sid(sid)
        if result is None: return
        self._admin_popup_direct(tree, db_key, result, on_refresh, sid)

    def _bind_admin_popup(self, tree, db_key, on_refresh, extra_key=None, extra_col_idx=None, id_col_idx=0):
        """
        دابل‌کلیک روی هر ردیف → پنجره ویرایش کامل با همه فیلدهای رکورد.
        فقط برای ادمین. فیلدها بر اساس نوع داده: کشویی یا متنی.

        id_col_idx: اندیس ستونی که شماره اسلب (slab_id) در آن قرار دارد.
        پیش‌فرض ۰ است، اما در جداولی که یک ستون اضافه مثل چک‌باکس (☑) در
        ابتدای جدول دارند (مثلاً تب «ثبت ذوب») باید ۱ داده شود.
        """
        if self.role != "admin": return

        _tooltip = [None]

        # ── تعریف فیلدهای هر جدول ──
        # (field_name, label_fa, widget_type, options_or_None)
        # widget_type: "text" | "combo" | "user" | "datetime"
        # نکته: نوع "datetime" یعنی این فیلد در دیتابیس یک رشتهٔ ترکیبی
        # «تاریخ  ساعت» است، اما در پنجره ویرایش به‌صورت دو باکس کاملاً
        # جدا (یکی فقط تاریخ، یکی فقط ساعت) نمایش و ویرایش می‌شود.
        FIELD_DEFS = {
            "melts": [
                ("slab_id",       "شماره اسلب",       "text",     None),
                ("qc_status",     "وضعیت QC",          "combo",    ["ثبت شده","کنترل کیفی شده","عدم تایید کنترل کیفی","قراضه"]),
                ("registered_by", "ثبت‌کننده",         "user",     None),
                ("registered_at", "تاریخ ثبت",         "datetime", None),
                ("note",          "توضیحات",           "text",     None),
                ("qc_by",         "تأییدکننده QC",     "user",     None),
                ("qc_at",         "تاریخ/ساعت QC",     "datetime", None),
                ("re_approved",   "تأیید مجدد؟",       "combo",    ["True","False"]),
                ("re_approved_by","تأییدکننده مجدد",   "user",     None),
                ("re_approved_at","تاریخ تأیید مجدد",  "datetime", None),
                ("exit_status",   "وضعیت خروج",        "combo",    ["—","خروج زده شده"]),
                ("exit_by",       "خروج‌دهنده",        "user",     None),
                ("exit_at",       "تاریخ/ساعت خروج",   "datetime", None),
                ("location",      "محل",               "combo",    ["انبار داخلی","انبار روباز ۱","انبار روباز ۲"]),
                ("reason",        "دلیل آخرین انتقال به انبار داخلی", "text", None),
            ],
            "scarf_cut": [
                ("slab_id",       "شماره اسلب",        "text",     None),
                ("operation",     "نوع عملیات",        "combo",    ["اسکارفی","برشی"]),
                ("reason",        "دلایل",             "text",     None),
                ("note",          "توضیحات",           "text",     None),
                ("bauman_done",   "باومن انجام شده؟",  "combo",    ["True","False"]),
                ("cut_count",     "تعداد برش",         "text",     None),
                ("registered_by", "ثبت‌کننده",         "user",     None),
                ("registered_at", "تاریخ/ساعت ثبت",   "datetime", None),
            ],
            "scrap": [
                ("slab_id",       "شماره اسلب",        "text",     None),
                ("reason",        "دلیل قراضه",        "text",     None),
                ("registered_by", "ثبت‌کننده",         "user",     None),
                ("registered_at", "تاریخ/ساعت ثبت",   "datetime", None),
            ],
            "lab_deliveries": [
                ("slab_id",       "شماره اسلب",        "text",     None),
                ("delivered_by",  "تحویل‌دهنده",       "user",     None),
                ("delivered_at",  "تاریخ/ساعت تحویل", "datetime", None),
            ],
            "bauman": [
                ("slab_id",       "شماره اسلب",        "text",     None),
                ("cut_by",        "برش‌کار",           "user",     None),
                ("cut_at",        "تاریخ/ساعت برش",   "datetime", None),
                ("lab_status",    "وضعیت آزمایشگاه",  "combo",    ["در انتظار","تحویل داده شده"]),
                ("delivered_by",  "تحویل‌دهنده",       "user",     None),
                ("delivered_at",  "تاریخ/ساعت تحویل", "datetime", None),
            ],
            "transfers_out": [
                ("slab_id",       "شماره اسلب",        "text",     None),
                ("destination",   "مقصد",              "combo",    ["انبار داخلی","انبار روباز ۱","انبار روباز ۲"]),
                ("current_location","محل فعلی",        "combo",    ["انبار داخلی","انبار روباز ۱","انبار روباز ۲"]),
                ("transferred_by","انتقال‌دهنده",      "user",     None),
                ("transferred_at","تاریخ/ساعت انتقال","datetime", None),
                ("reason",        "دلیل انتقال",       "text",     None),
            ],
            "returns": [
                ("slab_id",       "شماره اسلب",        "text",     None),
                ("reason",        "دلیل برگشت",        "text",     None),
                ("returned_by",   "ثبت‌کننده",         "user",     None),
                ("returned_at",   "تاریخ/ساعت برگشت", "datetime", None),
            ],
        }

        def _dt_labels(label):
            """از روی یک برچسب ترکیبی، برچسب جدای «تاریخ» و «ساعت» می‌سازد."""
            if "تاریخ/ساعت" in label:
                return label.replace("تاریخ/ساعت", "تاریخ"), label.replace("تاریخ/ساعت", "ساعت")
            if label.startswith("تاریخ"):
                rest = label[len("تاریخ"):]
                return label, "ساعت" + rest
            return label, label

        def _melt_last_transfer_move(db, sid):
            """آخرین رکورد «انتقال» واقعی (جابجایی فیزیکی بین انبارها) این اسلب
            در movement_log را برمی‌گرداند — همان دلیلی که در تب‌های انتقال/انبار
            به کاربر نمایش داده می‌شود."""
            moves = [m for m in db.get("movement_log", [])
                     if m.get("slab_id") == sid and m.get("operation", "انتقال") == "انتقال"]
            if not moves:
                return None
            moves.sort(key=lambda m: m.get("at", ""))
            return moves[-1]

        def _close_tooltip():
            if _tooltip[0]:
                try: _tooltip[0].destroy()
                except: pass
                _tooltip[0] = None

        def get_rec(sid, extra_val=None):
            db = load_db()
            for i, r in enumerate(db.get(db_key, [])):
                if r.get("slab_id") != sid: continue
                if extra_key and extra_val:
                    if r.get(extra_key) != extra_val: continue
                return (i, r)
            return None

        def get_all_users():
            db = load_db()
            return [f"{u}  ({d.get('display',u)})"
                    for u, d in db.get("users",{}).items()]

        def extract_username(val):
            if "  (" in val:
                return val.split("  (")[0].strip()
            return val.strip()

        # ── tooltip روی سلول ──
        def _show_tooltip(event):
            _close_tooltip()
            region = tree.identify_region(event.x, event.y)
            if region != "cell": return
            col_id = tree.identify_column(event.x)
            row_id = tree.identify_row(event.y)
            if not row_id: return
            try:
                disp = list(tree["displaycolumns"])
                di   = int(col_id.replace("#","")) - 1
                if di < 0 or di >= len(disp): return
                cn   = disp[di]
                ac   = list(tree["columns"])
                av   = list(tree.item(row_id,"values"))
                idx  = ac.index(cn)
                txt  = str(av[idx])
            except Exception:
                return
            if not txt or txt in ("—","","🚚") or len(txt) < 14: return
            tip = tk.Toplevel(tree)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{event.x_root+14}+{event.y_root+10}")
            tip.configure(bg=C["header_bg"])
            tk.Frame(tip, bg=C["gold"], height=2).pack(fill="x")
            tk.Label(tip, text=txt, bg=C["header_bg"], fg=C["text"],
                     font=(_MAIN_FONT,10,"bold"), justify="right",
                     wraplength=420, padx=10, pady=6).pack()
            tk.Frame(tip, bg=C["gold"], height=2).pack(fill="x")
            _tooltip[0] = tip
            tip.after(5000, _close_tooltip)

        tree.bind("<Motion>", _show_tooltip)
        tree.bind("<Leave>",  lambda e: _close_tooltip())

        # ── باز کردن پنجره ویرایش کامل برای یک ردیف مشخص ──
        def open_edit(row_id):
            all_vals = list(tree.item(row_id,"values"))
            if not all_vals or id_col_idx >= len(all_vals): return
            sid = str(all_vals[id_col_idx]).replace("↳","").strip()
            if not sid or sid.startswith("—"): return

            extra_val = (str(all_vals[extra_col_idx]).strip()
                         if extra_col_idx is not None and extra_col_idx < len(all_vals)
                         else None)
            result = get_rec(sid, extra_val) or get_rec(sid, None)
            if result is None: return
            rec_idx, rec = result

            edit_key = f"admin:{db_key}:{sid}:{rec.get(extra_key,'') if extra_key else ''}:{extra_val or ''}"
            if _acquire_edit_popup(self, edit_key):
                return

            # ── پنجره ویرایش ──
            win = tk.Toplevel(self)
            prepare_popup_window(win, self)
            _register_edit_popup(self, edit_key, win)
            win.title(f"✏️  ویرایش رکورد — {sid}")
            win.configure(bg=C["card"])
            win.resizable(True, True)
            self._center(win, 520, 560)

            # هدر
            hdr = tk.Frame(win, bg=C["header_bg"]); hdr.pack(fill="x")
            tk.Frame(hdr, bg=C["gold"], height=3).pack(fill="x")
            hi = tk.Frame(hdr, bg=C["header_bg"]); hi.pack(fill="x", padx=16, pady=10)
            tk.Label(hi, text="✏️  ویرایش رکورد", bg=C["header_bg"],
                     fg=C["gold"], font=FONT_HEAD).pack(side="right")
            tk.Label(hi, text=f"اسلب: {sid}  |  جدول: {db_key}",
                     bg=C["header_bg"], fg=C["text_dim"], font=FONT_SMALL).pack(side="left")

            # اسکرول
            outer = tk.Frame(win, bg=C["card"]); outer.pack(fill="both", expand=True, padx=12, pady=8)
            cv = tk.Canvas(outer, bg=C["card"], highlightthickness=0)
            vsb = tk.Scrollbar(outer, orient="vertical", command=cv.yview,
                               bg="#707070", troughcolor="#1a1a1a", width=16)
            cv.configure(yscrollcommand=vsb.set)
            vsb.pack(side="left", fill="y")
            cv.pack(side="right", fill="both", expand=True)
            ff = tk.Frame(cv, bg=C["card"])
            cv.create_window((0,0), window=ff, anchor="nw")
            ff.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
            register_scroll_canvas(cv, ff)

            users = get_all_users()
            field_defs = FIELD_DEFS.get(db_key, [])
            # اگه تعریف نداره، از کلیدهای رکورد بساز
            if not field_defs:
                field_defs = [(k, k, "text", None) for k in rec.keys()]

            db_snapshot = load_db()
            vars_ = {}

            for field, label, wtype, opts in field_defs:
                # ── دلیل آخرین انتقال یک اسلب در movement_log ذخیره می‌شود، نه در خود رکورد ذوب ──
                if db_key == "melts" and field == "reason":
                    _mv = _melt_last_transfer_move(db_snapshot, sid)
                    val = (_mv.get("reason", "") if _mv else "")
                else:
                    val = str(rec.get(field, ""))

                if wtype == "datetime":
                    # ── تاریخ و ساعت همیشه در دو ردیف/باکس کاملاً جدا از هم — هرگز توی یک سلول با هم نیستند ──
                    date_label, time_label = _dt_labels(label)
                    d_part, t_part = split_dt(val)
                    if d_part == "—": d_part = ""
                    if t_part == "—": t_part = ""

                    row = tk.Frame(ff, bg=C["card"])
                    row.pack(fill="x", padx=8, pady=(4,1))
                    tk.Label(row, text=f"{date_label}:", bg=C["card"], fg=C["gold"],
                             font=(_MAIN_FONT,9,"bold"), width=20, anchor="e").pack(side="right")
                    date_var = tk.StringVar(value=d_part)
                    tk.Entry(row, textvariable=date_var,
                             bg=C["entry_bg"], fg=C["text"],
                             insertbackground=C["accent"],
                             font=(_MAIN_FONT,10,"bold"),
                             justify="center", bd=0, relief="flat",
                             highlightthickness=1,
                             highlightbackground=C["border"],
                             highlightcolor=C["gold"], width=16
                             ).pack(side="right", padx=6)

                    row2 = tk.Frame(ff, bg=C["card"])
                    row2.pack(fill="x", padx=8, pady=(0,4))
                    tk.Label(row2, text=f"{time_label}:", bg=C["card"], fg=C["gold"],
                             font=(_MAIN_FONT,9,"bold"), width=20, anchor="e").pack(side="right")
                    time_var = tk.StringVar(value=t_part)
                    tk.Entry(row2, textvariable=time_var,
                             bg=C["entry_bg"], fg=C["text"],
                             insertbackground=C["accent"],
                             font=(_MAIN_FONT,10,"bold"),
                             justify="center", bd=0, relief="flat",
                             highlightthickness=1,
                             highlightbackground=C["border"],
                             highlightcolor=C["gold"], width=16
                             ).pack(side="right", padx=6)
                    vars_[field] = ("datetime", date_var, time_var)
                    continue

                row = tk.Frame(ff, bg=C["card"])
                row.pack(fill="x", padx=8, pady=4)

                tk.Label(row, text=f"{label}:", bg=C["card"], fg=C["gold"],
                         font=(_MAIN_FONT,9,"bold"), width=20, anchor="e").pack(side="right")

                if wtype == "user":
                    var = tk.StringVar()
                    cb = ttk.Combobox(row, textvariable=var, values=users,
                                      font=(_MAIN_FONT,10,"bold"),
                                      style="Dark.TCombobox", state="readonly", width=26)
                    # انتخاب مقدار فعلی
                    matched = next((u for u in users
                                   if u == val or u.startswith(val+"  (")), None)
                    cb.set(matched if matched else (users[0] if users else ""))
                    cb.pack(side="right", padx=6)
                    vars_[field] = ("user", var)

                elif wtype == "combo":
                    var = tk.StringVar(value=val)
                    cb = ttk.Combobox(row, textvariable=var, values=opts or [],
                                      font=(_MAIN_FONT,10,"bold"),
                                      style="Dark.TCombobox", state="readonly", width=26)
                    cb.set(val if val in (opts or []) else (opts[0] if opts else val))
                    cb.pack(side="right", padx=6)
                    vars_[field] = ("combo", var)

                else:  # text
                    var = tk.StringVar(value=val)
                    ent = tk.Entry(row, textvariable=var,
                                   bg=C["entry_bg"], fg=C["text"],
                                   insertbackground=C["accent"],
                                   font=(_MAIN_FONT,10,"bold"),
                                   justify="right", bd=0, relief="flat",
                                   highlightthickness=1,
                                   highlightbackground=C["border"],
                                   highlightcolor=C["gold"], width=28)
                    ent.pack(side="right", padx=6)
                    vars_[field] = ("text", var)

            # ── ذخیره ──
            def do_save():
                db = load_db()
                actual = None
                for j, r in enumerate(db.get(db_key,[])):
                    if r.get("slab_id") == sid:
                        if db_key == "scarf_cut" and rec.get("operation") and r.get("operation") != rec.get("operation"):
                            continue
                        actual = j; break
                if actual is None:
                    messagebox.showerror("خطا","رکورد یافت نشد.",parent=win); return

                if db_key == "scarf_cut" and not assert_scarf_cut_allowed(db, sid, parent=win):
                    return

                orig_qc_status = db[db_key][actual].get("qc_status") if db_key == "melts" else None

                for field, info in vars_.items():
                    wtype = info[0]

                    # ── دلیل آخرین انتقال — در movement_log ذخیره می‌شود، نه در رکورد ذوب ──
                    if db_key == "melts" and field == "reason":
                        new_reason = info[1].get().strip()
                        # اگر انتقال به داخلی وجود دارد، دلیل نباید خالی شود
                        mv = _melt_last_transfer_move(db, sid)
                        if mv is not None and (mv.get("to") == "انبار داخلی") and not new_reason:
                            messagebox.showwarning(
                                "خطا", "دلیل انتقال به انبار داخلی الزامی است — نمی‌تواند خالی باشد.",
                                parent=win)
                            return
                        if mv is not None:
                            mv["reason"] = new_reason
                        continue

                    if wtype == "datetime":
                        date_v = info[1].get().strip()
                        time_v = info[2].get().strip()
                        if date_v and time_v:
                            val_save = f"{date_v}  {time_v}"
                        else:
                            val_save = date_v or time_v or ""
                    else:
                        var = info[1]
                        raw = var.get()
                        if wtype == "user":
                            val_save = extract_username(raw)
                        else:
                            val_save = raw

                    # فیلدهای دلیل/توضیح اجباری — خالی مجاز نیست
                    _reason_fields = {
                        ("scrap", "reason"): "دلیل قراضه",
                        ("returns", "reason"): "دلیل برگشت",
                        ("return_log", "reason"): "دلیل برگشت",
                        ("scarf_cut", "note"): "توضیحات",
                        ("transfers_out", "reason"): "دلیل انتقال",
                    }
                    label_req = _reason_fields.get((db_key, field))
                    if label_req and not str(val_save).strip():
                        if db_key == "scarf_cut" and field == "note":
                            reason_txt = ""
                            if "reason" in vars_:
                                reason_txt = str(vars_["reason"][1].get() or "")
                            else:
                                reason_txt = str(db[db_key][actual].get("reason", "") or "")
                            if ("دلایل دیگر" in reason_txt) or ("آسیب‌های دیگر" in reason_txt):
                                messagebox.showwarning(
                                    "خطا", f"لطفاً «{label_req}» را وارد کنید — خالی مجاز نیست.",
                                    parent=win)
                                return
                        elif db_key == "transfers_out" and field == "reason":
                            dest = ""
                            if "destination" in vars_:
                                dest = str(vars_["destination"][1].get() or "")
                            elif "current_location" in vars_:
                                dest = str(vars_["current_location"][1].get() or "")
                            else:
                                dest = str(db[db_key][actual].get("destination", "") or "")
                            if dest == "انبار داخلی":
                                messagebox.showwarning(
                                    "خطا", f"لطفاً «{label_req}» را وارد کنید — خالی مجاز نیست.",
                                    parent=win)
                                return
                        else:
                            messagebox.showwarning(
                                "خطا", f"لطفاً «{label_req}» را وارد کنید — خالی مجاز نیست.",
                                parent=win)
                            return

                    orig = db[db_key][actual].get(field)
                    if isinstance(orig, bool):
                        db[db_key][actual][field] = val_save.lower() in ("true","1","بله","yes")
                    elif isinstance(orig, int):
                        try: db[db_key][actual][field] = int(val_save)
                        except: db[db_key][actual][field] = val_save
                    else:
                        db[db_key][actual][field] = val_save

                # ── اعمال واقعی تغییر وضعیت QC — صرفاً تغییر نوشتاری نیست ──
                if db_key == "melts" and "qc_status" in vars_:
                    new_qc_status = db[db_key][actual].get("qc_status")
                    if new_qc_status != orig_qc_status:
                        _ts_qc = now_str()
                        db[db_key][actual]["updated_at"] = _ts_qc
                        if new_qc_status == "ثبت شده":
                            # بازگشت به «در انتظار تأیید کنترل کیفی» — همهٔ اطلاعات
                            # مراحل بعدی (QC/تأیید مجدد/خروج) خالی می‌شود تا اسلب
                            # دقیقاً به جای اول خودش در تب «ثبت ذوب» برگردد و
                            # دوباره منتظر تأیید یا رد کنترل کیفی بماند.
                            for f in ("qc_by","qc_at","re_approved_by","re_approved_at",
                                      "exit_by","exit_at"):
                                db[db_key][actual][f] = ""
                            db[db_key][actual]["re_approved"]  = False
                            db[db_key][actual]["exit_status"]  = "—"
                            db[db_key][actual]["last_edit_at"] = _ts_qc
                            db[db_key][actual]["last_edit_by"] = self.username
                        else:
                            # تغییر به یک وضعیت تصمیم‌گیری‌شده — تأییدکننده و زمان
                            # واقعی این اصلاح ثبت می‌شود تا با وضعیت جدید هم‌خوان باشد.
                            db[db_key][actual]["qc_by"] = self.username
                            db[db_key][actual]["qc_at"] = _ts_qc
                else:
                    # هر ویرایش ادمین — updated_at برای merge شبکه‌ای
                    try:
                        db[db_key][actual]["updated_at"] = now_str()
                    except Exception:
                        pass

                save_db(db)
                # اسکرول را قبل از بستن پاپ‌آپ قفل کن
                try:
                    tree._stf_locked_yview = float(tree.yview()[0])
                except Exception:
                    pass
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
                win.destroy()
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass

                # melts روی تب ثبت ذوب: درجا — بدون رفرش کامل (پرش اسکرول)
                updated_inplace = False
                if db_key == "melts":
                    inplace = getattr(self, "_melts_inplace_update", None)
                    cols = []
                    try:
                        cols = list(tree["columns"])
                    except Exception:
                        cols = []
                    is_melts_grid = ("chk" in cols and "action" in cols) or (
                        "slab_id" in cols and "qc_status" in cols and "registered_by" in cols
                    )
                    if callable(inplace) and is_melts_grid:
                        try:
                            db2 = load_db()
                            rec2 = next(
                                (r for r in db2.get("melts", []) if r.get("slab_id") == sid),
                                None,
                            )
                            if rec2 is not None:
                                inplace(sid, rec2, db2)
                                updated_inplace = True
                        except Exception:
                            updated_inplace = False
                if not updated_inplace:
                    on_refresh()
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
                invalidate_display_cache()

            # ── حذف ──
            def do_delete():
                if not messagebox.askyesno("⚠️  حذف",
                        f"رکورد اسلب «{sid}» حذف شود?\nبرگشت‌پذیر نیست!",parent=win): return
                try:
                    tree._stf_locked_yview = float(tree.yview()[0])
                except Exception:
                    pass
                db = load_db()
                actual = None
                for j, r in enumerate(db.get(db_key,[])):
                    if r.get("slab_id") == sid:
                        if db_key == "scarf_cut" and rec.get("operation") and r.get("operation") != rec.get("operation"):
                            continue
                        actual = j; break
                if actual is not None:
                    removed = db[db_key].pop(actual)
                    try:
                        from shared.db_merge import note_tombstone
                        note_tombstone(db, db_key, removed if isinstance(removed, dict) else rec, now_str())
                    except Exception:
                        pass
                    save_db(db)
                win.destroy()
                on_refresh()
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass

            btn = tk.Frame(win, bg=C["card"]); btn.pack(fill="x", padx=16, pady=10)
            tk.Button(btn, text="انصراف", command=win.destroy,
                      bg=C["card2"], fg=C["text"], font=(_MAIN_FONT,10,"bold"),
                      bd=0, relief="flat", cursor="hand2", padx=14, pady=8).pack(side="left")
            styled_btn(btn, "🗑  حذف", do_delete, color=C["btn_danger"]).pack(side="right", padx=4)
            styled_btn(btn, "💾  ذخیره", do_save, color=C["gold"]).pack(side="right", padx=4)
            win.bind("<Return>", lambda e: do_save())
            win.bind("<Escape>", lambda e: win.destroy())
            finalize_popup_window(win, self)

        # ── دابل‌کلیک: باز کردن پنجره ویرایش کامل ──
        def on_double(event):
            _close_tooltip()
            region = tree.identify_region(event.x, event.y)
            if region not in ("cell","heading"): return
            row_id = tree.identify_row(event.y)
            if not row_id: return
            open_edit(row_id)

        # ── حذف گروهی چند مورد انتخاب‌شده باهم ──
        def do_bulk_delete(sel_ids):
            targets = []
            for row_id in sel_ids:
                vals = list(tree.item(row_id, "values"))
                if not vals or id_col_idx >= len(vals): continue
                sid_ = str(vals[id_col_idx]).replace("↳","").strip()
                if not sid_ or sid_.startswith("—"): continue
                extra_val = (str(vals[extra_col_idx]).strip()
                             if extra_col_idx is not None and extra_col_idx < len(vals)
                             else None)
                targets.append((sid_, extra_val))
            if not targets: return
            n = len(targets)
            msg = (f"رکورد اسلب «{targets[0][0]}» حذف شود؟"
                   if n == 1 else f"{n} مورد انتخاب‌شده حذف شوند؟")
            try:
                tree._stf_locked_yview = float(tree.yview()[0])
            except Exception:
                pass
            if not messagebox.askyesno("⚠️  حذف", msg + "\nاین عمل برگشت‌پذیر نیست!", parent=self):
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
                return
            db = load_db()
            removed = 0
            _ts_del = now_str()
            for sid_, extra_val in targets:
                for j, r in enumerate(db.get(db_key, [])):
                    if r.get("slab_id") != sid_: continue
                    if extra_key and extra_val and r.get(extra_key) != extra_val: continue
                    popped = db[db_key].pop(j)
                    try:
                        from shared.db_merge import note_tombstone
                        note_tombstone(db, db_key, popped if isinstance(popped, dict) else r, _ts_del)
                    except Exception:
                        pass
                    removed += 1
                    break
            save_db(db)
            on_refresh()
            invalidate_display_cache()
            try:
                if getattr(tree, "_stf_locked_yview", None) is not None:
                    tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            if n > 1:
                messagebox.showinfo("✅  حذف شد", f"{removed} رکورد حذف شد.", parent=self)
            try:
                if getattr(tree, "_stf_locked_yview", None) is not None:
                    tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass

        def _selected_targets(sel_ids):
            targets = []
            seen = set()
            for row_id in sel_ids:
                vals = list(tree.item(row_id, "values"))
                if not vals or id_col_idx >= len(vals):
                    continue
                sid_ = str(vals[id_col_idx]).replace("↳", "").strip()
                if not sid_ or sid_.startswith("—"):
                    continue
                extra_val = (str(vals[extra_col_idx]).strip()
                             if extra_col_idx is not None and extra_col_idx < len(vals)
                             else None)
                key = (sid_, extra_val)
                if key in seen:
                    continue
                seen.add(key)
                targets.append(key)
            return targets

        def _apply_qc_status_fields(rec, new_status):
            """همان منطق ویرایش تکی — روی یک رکورد melts."""
            old = rec.get("qc_status")
            ts = now_str()
            rec["qc_status"] = new_status
            # همیشه updated_at تا merge شبکه‌ای وضعیت را به qc_at قدیمی برنگرداند
            rec["updated_at"] = ts
            if new_status == "ثبت شده":
                for f in ("qc_by", "qc_at", "re_approved_by", "re_approved_at",
                          "exit_by", "exit_at"):
                    rec[f] = ""
                rec["re_approved"] = False
                rec["exit_status"] = "—"
                # qc_at خالی می‌شود ولی last_edit_at مانع برگشت وضعیت می‌شود
                rec["last_edit_at"] = ts
                rec["last_edit_by"] = self.username
            else:
                rec["qc_by"] = self.username
                rec["qc_at"] = ts
            return old != new_status

        def do_bulk_set_qc(sel_ids, new_status):
            """اعمال وضعیت QC روی همهٔ ردیف‌های انتخاب‌شده (مثل حذف گروهی)."""
            if db_key != "melts":
                return
            targets = _selected_targets(sel_ids)
            if not targets:
                return
            n = len(targets)
            msg = (
                f"وضعیت «{new_status}» برای اسلب «{targets[0][0]}» ثبت شود؟"
                if n == 1 else
                f"وضعیت «{new_status}» برای {n} مورد انتخاب‌شده اعمال شود؟"
            )
            try:
                tree._stf_locked_yview = float(tree.yview()[0])
            except Exception:
                pass
            tree._stf_scroll_freeze = True
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            if not messagebox.askyesno("تأیید تغییر وضعیت", msg, parent=self):
                tree._stf_scroll_freeze = False
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
                return
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            db = load_db()
            changed = 0
            for sid_, extra_val in targets:
                for r in db.get("melts", []):
                    if r.get("slab_id") != sid_:
                        continue
                    if _apply_qc_status_fields(r, new_status):
                        changed += 1
                    break
            save_db(db)
            # melts: به‌روزرسانی درجا — اسکرول تکون نمی‌خورد
            inplace = getattr(self, "_melts_inplace_update", None)
            cols = []
            try:
                cols = list(tree["columns"])
            except Exception:
                cols = []
            is_melts_grid = ("chk" in cols and "action" in cols) or (
                "slab_id" in cols and "qc_status" in cols and "registered_by" in cols
            )
            if db_key == "melts" and callable(inplace) and is_melts_grid:
                by_sid = {r.get("slab_id"): r for r in db.get("melts", [])}
                for sid_, _extra in targets:
                    rec = by_sid.get(sid_)
                    if rec:
                        try:
                            inplace(sid_, rec, db)
                        except Exception:
                            pass
            else:
                on_refresh()
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            tree._stf_scroll_freeze = False
            invalidate_display_cache()
            if n > 1:
                messagebox.showinfo(
                    "✅  انجام شد",
                    f"وضعیت «{new_status}» برای {changed} اسلب ثبت شد.",
                    parent=self,
                )
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass

        def _select_all_rows(event=None):
            tree.selection_set(tree.get_children())
            return "break"

        # ── کلیک راست: منوی زمینه — ویرایش تکی / وضعیت گروهی / حذف گروهی ──
        def on_right_click(event):
            _close_tooltip()
            row_id = tree.identify_row(event.y)
            cur_sel = tree.selection()
            if row_id and row_id not in cur_sel:
                tree.selection_set(row_id)
            sel = tree.selection()

            menu = tk.Menu(tree, tearoff=0, bg=C["card"], fg=C["text"],
                            activebackground=C["accent"], activeforeground="#ffffff",
                            font=(_MAIN_FONT, 10, "bold"))
            nsel = len(sel)
            if nsel >= 1 and db_key == "melts":
                qc_menu = tk.Menu(menu, tearoff=0, bg=C["card"], fg=C["text"],
                                   activebackground=C["accent"], activeforeground="#ffffff",
                                   font=(_MAIN_FONT, 10, "bold"))
                suffix = f" ({nsel} مورد)" if nsel > 1 else ""
                for st, label in (
                    ("کنترل کیفی شده", "✅  کنترل کیفی شده"),
                    ("عدم تایید کنترل کیفی", "⛔  عدم تایید کنترل کیفی"),
                    ("ثبت شده", "🔄  ثبت شده (بازگشت)"),
                    ("قراضه", "♻  قراضه"),
                ):
                    qc_menu.add_command(
                        label=f"{label}{suffix}",
                        command=lambda s=list(sel), status=st: do_bulk_set_qc(s, status),
                    )
                menu.add_cascade(
                    label=f"وضعیت QC — اعمال روی {nsel} مورد" if nsel > 1 else "وضعیت QC",
                    menu=qc_menu,
                )
                menu.add_separator()
            if nsel > 1:
                menu.add_command(label=f"🗑  حذف {nsel} مورد انتخاب‌شده",
                                  command=lambda s=list(sel): do_bulk_delete(s))
            elif nsel == 1:
                menu.add_command(label="✏️  ویرایش کامل", command=lambda r=sel[0]: open_edit(r))
                menu.add_command(label="🗑  حذف", command=lambda s=list(sel): do_bulk_delete(s))
            if sel:
                menu.add_separator()
            menu.add_command(label="☑️  انتخاب همه", command=_select_all_rows)
            if sel:
                menu.add_command(label="❌  لغو انتخاب",
                                  command=lambda: tree.selection_remove(*tree.get_children()))
            try:
                _popup_context_menu(tree, menu, event.x_root, event.y_root)
            finally:
                pass

        tree.bind("<Double-Button-1>", on_double)
        tree.bind("<Button-3>",        on_right_click)
        tree.bind("<Control-a>", _select_all_rows)
        tree.bind("<Control-A>", _select_all_rows)
        try:
            tree.configure(selectmode="extended")
        except Exception:
            pass

    def _admin_popup_direct(self, tree, db_key, result, on_refresh, sid):
        """ویرایش مستقیم اینلاین — بدون پنجره جداگانه"""
        if self.role != "admin": return
        # این متد دیگه پنجره باز نمی‌کنه
        # فقط اولین سلول رکورد رو انتخاب می‌کنه تا کاربر دابل‌کلیک کنه
        # (حفظ می‌شه برای سازگاری با کدهای قدیمی‌تر که مستقیم صداش می‌زنن)
        sel = tree.selection()
        if sel:
            try:
                cols = tree["columns"]
                if cols:
                    tree.focus(sel[0])
            except Exception:
                pass



    def _admin_bar(self, parent, db_key, get_selected_idx, get_selected_rec, on_refresh):
        """سازگاری با کدهای قدیمی — فقط _bind_admin_popup رو صدا میزنه"""
        pass  # خالی — حالا همه چیز از طریق دابل‌کلیک کار میکنه

    def _admin_inline_edit(self, db_key, record, on_save):
        pass  # خالی — از _admin_popup استفاده میشه

    # ═══════════════════════════════════════════════════════
    #  تب نمای کلی (فقط ادمین) — نمای کلی داشبورد
    # ═══════════════════════════════════════════════════════
    def _build_home_tab(self, tab):
        tab.configure(bg=C["panel"])

        # اسکرول
        canvas = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                            bg="#707070", troughcolor="#1a1a1a", activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="left", fill="y")
        canvas.pack(side="right", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=C["panel"])
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_configure(e):
            canvas.itemconfig(win_id, width=e.width)
        inner.bind("<Configure>", _on_inner_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # اسکرول ماوس از طریق هندلر سراسری (_global_mousewheel)
        register_scroll_canvas(canvas, inner)

        # کلیدهای صفحه‌کلید
        canvas.bind("<Up>",    lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Down>",  lambda e: canvas.yview_scroll( 1, "units"))
        canvas.bind("<Prior>", lambda e: canvas.yview_scroll(-5, "units"))
        canvas.bind("<Next>",  lambda e: canvas.yview_scroll( 5, "units"))
        canvas.bind("<Home>",  lambda e: canvas.yview_moveto(0))
        canvas.bind("<End>",   lambda e: canvas.yview_moveto(1))
        canvas.bind("<Button-1>", lambda e: canvas.focus_set())

        def refresh_home():
            for w in inner.winfo_children():
                w.destroy()
            self._build_home_content(inner)
            canvas.yview_moveto(0)

        self._build_home_content(inner)

    def _build_home_content(self, parent):
        db = load_db()

        def stat_card(parent_row, val, lbl, icon, accent, sub=""):
            cf = tk.Frame(parent_row, bg=C["card"],
                          highlightthickness=1, highlightbackground=accent)
            cf.pack(side="left", fill="both", expand=True, padx=5, pady=2)
            tk.Frame(cf, bg=accent, height=3).pack(fill="x")
            inner2 = tk.Frame(cf, bg=C["card"])
            inner2.pack(fill="both", padx=14, pady=12)
            tk.Label(inner2, text=icon, bg=C["card"], fg=accent,
                     font=("Segoe UI Emoji", 18, "bold")).pack(anchor="w")
            tk.Label(inner2, text=str(val), bg=C["card"], fg=accent,
                     font=("B Nazanin", 26, "bold")).pack(anchor="w")
            tk.Label(inner2, text=lbl, bg=C["card"], fg=C["text_dim"],
                     font=(_MAIN_FONT, 9, "bold")).pack(anchor="w", pady=(2,0))
            if sub:
                tk.Label(inner2, text=sub, bg=C["card"], fg=C["text_dim"],
                         font=(_MAIN_FONT, 8, "bold")).pack(anchor="w")

        def get_cur_loc(sid):
            tr = next((t for t in db.get("transfers_out",[]) if t.get("slab_id")==sid), None)
            if not tr: return "داخلی"
            return tr.get("current_location") or tr.get("destination") or "روباز"

        total  = len(db.get("melts", []))
        qc     = sum(1 for r in db["melts"] if r.get("qc_status")=="کنترل کیفی شده")
        rej    = sum(1 for r in db["melts"] if r.get("qc_status")=="عدم تایید کنترل کیفی")
        scarfs = sum(1 for r in db.get("scarf_cut",[]) if r.get("operation")=="اسکارفی")
        cuts   = sum(1 for r in db.get("scarf_cut",[]) if r.get("operation")=="برشی")
        scraps = len(db.get("scrap",[]))
        exited = sum(1 for r in db["melts"] if r.get("exit_status")=="خروج زده شده")
        out1   = sum(1 for r in db["melts"] if "۱" in get_cur_loc(r["slab_id"]) and "داخلی" not in get_cur_loc(r["slab_id"]))
        out2   = sum(1 for r in db["melts"] if "۲" in get_cur_loc(r["slab_id"]) and "داخلی" not in get_cur_loc(r["slab_id"]))
        inside = sum(1 for r in db["melts"] if get_cur_loc(r["slab_id"]) == "انبار داخلی")
        total_melts = len(db.get("melts", []))   # کل ذوب‌های ثبت شده

        pad = dict(padx=16, pady=4)

        # ─── عنوان ───
        hdr_f = tk.Frame(parent, bg=C["header_bg"])
        hdr_f.pack(fill="x", **pad)
        tk.Frame(hdr_f, bg=C["accent"], height=2).pack(fill="x")
        hdr_in = tk.Frame(hdr_f, bg=C["header_bg"])
        hdr_in.pack(fill="x", padx=16, pady=10)
        tk.Label(hdr_in, text="🏠  نمای کلی سیستم", bg=C["header_bg"],
                 fg=C["text_bright"], font=(_MAIN_FONT, 15, "bold")).pack(side="right")

        # ─── ردیف ۱: کل، QC، رد، خروج ───
        r1 = tk.Frame(parent, bg=C["panel"])
        r1.pack(fill="x", **pad)
        stat_card(r1, total,  "کل اسلب‌ها",         "🔥", C["accent"])
        stat_card(r1, qc,     "کنترل کیفی شده",      "✅", "#3a9070")
        stat_card(r1, rej,    "تایید نشده",           "⛔", "#b03040")
        stat_card(r1, exited, "خروج زده شده",         "🚪", "#7a7a4a")

        # ─── ردیف ۲: اسکارف، برش، قراضه، کل ذوب ─── 
        r2 = tk.Frame(parent, bg=C["panel"])
        r2.pack(fill="x", **pad)
        stat_card(r2, scarfs,      "اسکارف",                "⚙",  "#8a7840")
        stat_card(r2, cuts,        "برش",                    "✂",  C["accent2"])
        stat_card(r2, scraps,      "قراضه",                 "♻️", "#7a6a40")
        stat_card(r2, total_melts, "کل ذوب‌های ثبت شده",   "🔢", "#5a70a0")

        # ─── ردیف ۳: مکان‌ها ───
        r3 = tk.Frame(parent, bg=C["panel"])
        r3.pack(fill="x", **pad)
        stat_card(r3, out1,   "انبار روباز ۱",  "🏭", "#4a7060")
        stat_card(r3, out2,   "انبار روباز ۲",  "🏭", "#4a6a70")
        stat_card(r3, inside, "انبار داخلی",    "🏠", "#5a6070")

        # ─── پنل ضریب اصلاح ───
        ratio_panel = tk.Frame(parent, bg=C["card"],
                                highlightthickness=1, highlightbackground=C["border"])
        ratio_panel.pack(fill="x", padx=16, pady=8)
        tk.Frame(ratio_panel, bg=C["accent"], height=2).pack(fill="x")
        ratio_hdr = tk.Frame(ratio_panel, bg=C["card"])
        ratio_hdr.pack(fill="x", padx=16, pady=(10,4))
        tk.Label(ratio_hdr, text="📊  ضریب اصلاح — فیلتر بازه شمسی",
                 bg=C["card"], fg=C["text_bright"], font=(_MAIN_FONT, 13, "bold")).pack(side="right")

        # فیلد تاریخ
        _today_ratio = get_first_report_date_sh()
        from_var, to_var = _make_dt_filter(ratio_panel, C["card"], _today_ratio)

        # نمایش ضریب‌ها
        ratio_result = tk.Frame(ratio_panel, bg=C["card"])
        ratio_result.pack(fill="x", padx=16, pady=8)

        def render_ratio(fsh=None, tsh=None):
            for w in ratio_result.winfo_children():
                w.destroy()

            def in_range(at):
                if not fsh or not tsh: return True
                a = (at or "").replace("\u200c", "").strip()
                return fsh <= a <= tsh

            qc_n  = sum(1 for r in db["melts"] if r.get("qc_status")=="کنترل کیفی شده" and in_range(r.get("qc_at","")))
            sc_n  = sum(1 for r in db.get("scarf_cut",[]) if r.get("operation")=="اسکارفی" and in_range(r.get("registered_at","")))
            ct_n  = sum(1 for r in db.get("scarf_cut",[]) if r.get("operation")=="برشی" and in_range(r.get("registered_at","")))
            to_n  = sc_n + ct_n

            def pct(n, d):
                return f"{(n/d*100):.1f}%" if d > 0 else "—"

            lbl_txt = f"{fsh}  تا  {tsh}" if fsh else "کل دوره"
            tk.Label(ratio_result, text=f"بازه: {lbl_txt}   ·   QC شده: {qc_n} اسلب",
                     bg=C["card"], fg=C["text_dim"], font=(_MAIN_FONT, 9, "bold")).pack(anchor="e", pady=(0,8))

            cols_f = tk.Frame(ratio_result, bg=C["card"])
            cols_f.pack(fill="x")

            for val_n, val_d, lbl, icon, color in [
                (sc_n, qc_n, "ضریب اسکارف",   "⚙", "#8a7840"),
                (ct_n, qc_n, "ضریب برش",       "✂", C["accent2"]),
                (to_n, qc_n, "ضریب اصلاح کل", "📊", C["accent"]),
            ]:
                rc = tk.Frame(cols_f, bg=C["card2"],
                               highlightthickness=1, highlightbackground=color)
                rc.pack(side="left", fill="both", expand=True, padx=6)
                tk.Frame(rc, bg=color, height=2).pack(fill="x")
                ri = tk.Frame(rc, bg=C["card2"])
                ri.pack(fill="x", padx=14, pady=12)
                tk.Label(ri, text=f"{icon}  {lbl}", bg=C["card2"], fg=color,
                         font=(_MAIN_FONT, 9, "bold")).pack(anchor="e")
                tk.Label(ri, text=pct(val_n, val_d), bg=C["card2"], fg=color,
                         font=("B Nazanin", 28, "bold")).pack(anchor="e")

                # نوار پیشرفت
                bar_pct = (val_n/val_d*100) if val_d > 0 else 0
                bar_f = tk.Frame(ri, bg=C["bg"], height=5)
                bar_f.pack(fill="x", pady=(6,0))
                if bar_pct > 0:
                    fill_pct = min(bar_pct * 1.5, 100) / 100
                    fill_f = tk.Frame(bar_f, bg=color, height=5)
                    fill_f.place(relwidth=fill_pct, relheight=1.0, anchor="ne", relx=1.0, rely=0)

        def calc_ratio():
            f = from_var.get().strip()
            t = to_var.get().strip()
            if not f or not t:
                render_ratio()
                return
            fsh = f if "  " in f else f + "  00:00:00"
            tsh = t if "  " in t else t + "  23:59:59"
            render_ratio(fsh, tsh)

        btn_row2 = tk.Frame(ratio_panel, bg=C["card"])
        btn_row2.pack(fill="x", padx=16, pady=(0,12))
        styled_btn(btn_row2, "محاسبه", calc_ratio, color=C["btn_primary"], width=120).pack(side="right", padx=4)
        def clear_ratio():
            from_var.set("")
            to_var.set("")
            render_ratio()
        styled_btn(btn_row2, "کل دوره", clear_ratio, color=C["btn_ghost"], width=100).pack(side="right", padx=4)

        render_ratio()

        # ─── نمودار میله‌ای روزانه ───
        chart_panel = tk.Frame(parent, bg=C["card"],
                               highlightthickness=1, highlightbackground=C["border"])
        chart_panel.pack(fill="both", expand=True, padx=16, pady=8)
        tk.Frame(chart_panel, bg=C["accent"], height=2).pack(fill="x")
        chart_hdr = tk.Frame(chart_panel, bg=C["card"])
        chart_hdr.pack(fill="x", padx=16, pady=(10,4))
        tk.Label(chart_hdr, text="📊  نمودار روزانه — اسکارف / برش / کنترل کیفی",
                 bg=C["card"], fg=C["text_bright"], font=(_MAIN_FONT, 13, "bold")).pack(side="right")

        # ── انتخاب بازه تاریخ ──
        date_ctrl = tk.Frame(chart_panel, bg=C["card"])
        date_ctrl.pack(fill="x", padx=16, pady=6)

        _today_sh = to_shamsi(datetime.datetime.now()).split("  ")[0]

        def _mk_dt_entry(parent, var, w):
            return tk.Entry(parent, textvariable=var, width=w,
                            bg=C["entry_bg"], fg=C["accent"], font=("B Nazanin", 11, "bold"),
                            bd=0, relief="flat", highlightthickness=1,
                            highlightbackground=C["border"], highlightcolor=C["accent"],
                            justify="center")

        def _mk_dt_pair(parent, lbl, var, w):
            """لیبل سمت راست، باکس سمت چپ لیبل"""
            f = tk.Frame(parent, bg=C["card"])
            f.pack(side="right", padx=(0, 10))
            tk.Label(f, text=lbl, bg=C["card"], fg=C["text_dim"],
                     font=FONT_NORM).pack(side="right")
            _mk_dt_entry(f, var, w).pack(side="right", padx=(4, 0))

        tk.Label(date_ctrl, text="(فرمت: ۱۴۰۴/۰۳/۱۵  —  ساعت: ۱۴:۳۰:۰۰)", bg=C["card"],
                 fg=C["text_dim"], font=(_MAIN_FONT, 9, "bold")).pack(anchor="e", pady=(0,2))

        row_dates = tk.Frame(date_ctrl, bg=C["card"])
        row_dates.pack(fill="x", pady=(0,3))
        chart_from_date = tk.StringVar(value=get_first_report_date_sh())
        chart_to_date   = tk.StringVar(value=_today_sh)
        _mk_dt_pair(row_dates, "از تاریخ", chart_from_date, 13)
        _mk_dt_pair(row_dates, "تا تاریخ", chart_to_date,   13)

        row_times = tk.Frame(date_ctrl, bg=C["card"])
        row_times.pack(fill="x", pady=(0,4))
        chart_from_time = tk.StringVar(value="00:00:00")
        chart_to_time   = tk.StringVar(value="23:59:59")
        _mk_dt_pair(row_times, "از ساعت", chart_from_time, 13)
        _mk_dt_pair(row_times, "تا ساعت", chart_to_time,   13)

        # ── متغیرهای ترکیبی ──
        chart_from_var = tk.StringVar()
        chart_to_var   = tk.StringVar()
        def _sync_chart_dt(*_):
            chart_from_var.set(f"{chart_from_date.get().strip()}  {chart_from_time.get().strip()}")
            chart_to_var.set(f"{chart_to_date.get().strip()}  {chart_to_time.get().strip()}")
        for _v in (chart_from_date, chart_from_time, chart_to_date, chart_to_time):
            _v.trace_add("write", _sync_chart_dt)
        _sync_chart_dt()

        chart_status = tk.Label(date_ctrl, text="", bg=C["card"],
                                 fg=C["danger"], font=(_MAIN_FONT, 9, "bold"))
        chart_status.pack(anchor="e", pady=2)

        # ── بوم نمودار ──
        chart_canvas_frame = tk.Frame(chart_panel, bg=C["card"])
        chart_canvas_frame.pack(fill="both", expand=True, padx=16, pady=6)

        chart_cv = tk.Canvas(chart_canvas_frame, bg=C["card2"],
                              highlightthickness=1, highlightbackground=C["border"],
                              height=280)
        chart_cv.pack(fill="both", expand=True)

        # راهنمای رنگ
        legend_f = tk.Frame(chart_panel, bg=C["card"])
        legend_f.pack(anchor="center", pady=(2, 4))
        for _lc, _ll in [("#C0392B","اسکارف"), ("#E8EAED","برش"), ("#1A8A4A","کنترل کیفی تایید")]:
            _lf = tk.Frame(legend_f, bg=C["card"])
            _lf.pack(side="left", padx=14)
            tk.Frame(_lf, bg=_lc, width=20, height=12).pack(side="left")
            tk.Label(_lf, text=f"  {_ll}", bg=C["card"], fg=C["text"],
                     font=(_MAIN_FONT, 9, "bold")).pack(side="left")

        # ── توابع کمکی تاریخ شمسی ──
        def _is_leap_sh(y):
            return (y % 33) in (1, 5, 9, 13, 17, 22, 26, 30)

        def _days_in_month_sh(y, m):
            md = [0,31,31,31,31,31,31,30,30,30,30,30,29]
            if m == 12 and _is_leap_sh(y):
                return 30
            return md[m]

        def _next_day_sh(y, m, d):
            d += 1
            if d > _days_in_month_sh(y, m):
                d = 1
                m += 1
                if m > 12:
                    m = 1
                    y += 1
            return y, m, d

        def parse_date_range():
            def _psd(s):
                s = normalize_digits(s.strip())
                mt = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
                if not mt:
                    raise ValueError(f"فرمت اشتباه: {s}")
                return int(mt.group(1)), int(mt.group(2)), int(mt.group(3))

            fy, fm, fd = _psd(chart_from_var.get())
            ty, tm, td = _psd(chart_to_var.get())
            from_str = f"{fy:04d}/{fm:02d}/{fd:02d}"
            to_str   = f"{ty:04d}/{tm:02d}/{td:02d}"
            if from_str > to_str:
                raise ValueError("تاریخ شروع باید قبل از پایان باشد")
            days = []
            cy, cm, cd = fy, fm, fd
            for _ in range(400):
                ds = f"{cy:04d}/{cm:02d}/{cd:02d}"
                days.append(ds)
                if ds == to_str:
                    break
                cy, cm, cd = _next_day_sh(cy, cm, cd)
            return days, from_str, to_str

        def _get_chart_data():
            """داده‌های نمودار را می‌سازد — با فیلتر تاریخ + ساعت"""
            days, from_str, to_str = parse_date_range()
            # بازه کامل با ساعت
            _from_full = chart_from_var.get().strip()
            _to_full   = chart_to_var.get().strip()
            if "  " not in _from_full:
                _from_full += "  00:00:00"
            if "  " not in _to_full:
                _to_full += "  23:59:59"

            def _in_range(at):
                a = (at or "").strip()
                return _from_full <= a <= _to_full

            _db = load_db()
            scarfs_d = {d: 0 for d in days}
            cuts_d   = {d: 0 for d in days}
            qc_d     = {d: 0 for d in days}
            for r in _db.get("scarf_cut", []):
                at_full = (r.get("registered_at") or "").strip()
                at_date = at_full.split("  ")[0].strip()
                if not _in_range(at_full): continue
                if r.get("operation") == "اسکارفی" and at_date in scarfs_d:
                    scarfs_d[at_date] += 1
                elif r.get("operation") == "برشی" and at_date in cuts_d:
                    cuts_d[at_date] += 1
            for r in _db.get("melts", []):
                if r.get("qc_status") == "کنترل کیفی شده":
                    at_full = (r.get("qc_at") or "").strip()
                    at_date = at_full.split("  ")[0].strip()
                    if _in_range(at_full) and at_date in qc_d:
                        qc_d[at_date] += 1
            return days, from_str, to_str, scarfs_d, cuts_d, qc_d

        def draw_chart():
            chart_cv.delete("all")
            chart_status.config(text="")
            try:
                days, from_str, to_str, scarfs_d, cuts_d, qc_d = _get_chart_data()
            except ValueError as e:
                chart_status.config(text=f"⚠️  {e}")
                return

            chart_cv.update_idletasks()
            W = chart_cv.winfo_width() or 800
            H = chart_cv.winfo_height() or 280

            PAD_L = 48
            PAD_R = 16
            PAD_T = 22
            PAD_B = 54

            plot_w = W - PAD_L - PAD_R
            plot_h = H - PAD_T - PAD_B
            n_days = len(days)
            if n_days == 0:
                return

            all_vals = list(scarfs_d.values()) + list(cuts_d.values()) + list(qc_d.values())
            max_val = max(all_vals) if any(v > 0 for v in all_vals) else 1

            # پس‌زمینه
            chart_cv.create_rectangle(PAD_L, PAD_T, W - PAD_R, H - PAD_B,
                                       fill=C["card2"], outline=C["border"])

            # خطوط راهنما
            steps = 5
            for i in range(steps + 1):
                yv = max_val * i / steps
                y  = H - PAD_B - (plot_h * i / steps)
                chart_cv.create_line(PAD_L, y, W - PAD_R, y,
                                     fill=C["border"], dash=(3, 4))
                chart_cv.create_text(PAD_L - 4, y, text=str(int(round(yv))),
                                     anchor="e", font=("B Nazanin", 8, "bold"),
                                     fill=C["text_dim"])

            group_w = plot_w / n_days
            bar_gap = max(1.0, group_w * 0.1)
            bar_w   = max(2.0, (group_w - bar_gap * 2) / 3.0 - 1.0)
            # رنگ‌ها: سبز / قرمز / سفید — بدون هاشور، فقط رنگ توپر
            COLORS   = ["#1A8A4A", "#C0392B", "#E8EAED"]
            OUTLINES = ["#0D4A28", "#6E1F17", "#555555"]

            for i, day in enumerate(days):
                vals = [qc_d[day], scarfs_d[day], cuts_d[day]]
                gx   = PAD_L + i * group_w
                for j, (v, col, oln) in enumerate(zip(vals, COLORS, OUTLINES)):
                    bx1 = gx + bar_gap + j * (bar_w + 1)
                    bx2 = bx1 + bar_w
                    if v > 0:
                        by1 = H - PAD_B - (plot_h * v / max_val)
                        by2 = H - PAD_B
                        chart_cv.create_rectangle(bx1, by1, bx2, by2,
                                                   fill=col, outline=oln, width=1)
                        if bar_w > 12:
                            chart_cv.create_text((bx1 + bx2) / 2, by1 - 5,
                                                  text=str(v),
                                                  font=("B Nazanin", 7, "bold"),
                                                  fill=oln)
                cx = gx + group_w / 2
                chart_cv.create_text(cx, H - PAD_B + 8,
                                     text=day.split("/")[-1],
                                     font=(_MAIN_FONT, 7, "bold"), fill=C["text_dim"], anchor="n")
                if day.endswith("/01") or i == 0:
                    ym = "/".join(day.split("/")[:2])
                    chart_cv.create_text(cx, H - PAD_B + 22, text=ym,
                                         font=(_MAIN_FONT, 7, "bold"), fill=C["accent"], anchor="n")

            _from_disp = chart_from_var.get().strip()
            _to_disp   = chart_to_var.get().strip()
            chart_cv.create_text(W // 2, PAD_T // 2,
                                  text=f"بازه: {_from_disp}  تا  {_to_disp}   |   {n_days} روز",
                                  font=(_MAIN_FONT, 9, "bold"), fill=C["text_dim"])

        def export_excel():
            """خروجی اکسل شش‌شیتی:
            ۱) کلی - کل دوره   ۲) کلی - بازه انتخابی   ۳) نمودار روزانه (حرفه‌ای)
            ۴) کلی - اسکارف/برش نسبت به کنترل کیفی (کل دوره)
            ۵) فعالیت اپراتورها - بازه انتخابی   ۶) فعالیت اپراتورها - کل دوره"""
            chart_status.config(text="")
            try:
                days, from_str, to_str, scarfs_d, cuts_d, qc_d = _get_chart_data()
            except ValueError as e:
                chart_status.config(text=f"⚠️  {e}")
                return

            if not XLSX:
                messagebox.showerror("خطا", "کتابخانه openpyxl نصب نیست.\npip install openpyxl", parent=self)
                return

            fname = f"گزارش_کلی_واحد_اسلب_یارد_{shamsi_date_for_filename()}.xlsx"
            path = self._resolve_report_save_path(
                "overview", fname, [("Excel", "*.xlsx")], ".xlsx")
            if not path:
                return

            try:
                from openpyxl.chart import BarChart, Reference
                from openpyxl.chart.label import DataLabelList
                from openpyxl.utils import get_column_letter

                # ── بازه‌ی انتخابی همان فیلتر «ضریب اصلاح» بالای صفحه — اگر خالی باشد یعنی کل دوره ──
                _f = from_var.get().strip()
                _t = to_var.get().strip()
                _has_range = bool(_f and _t)
                _ov_from_sh, _ov_to_sh = (_f, _t) if _has_range else (None, None)
                _ov_range_label = f"{_ov_from_sh}   تا   {_ov_to_sh}" if _has_range else "کل دوره"
                _full_range_label = f"{get_first_report_date_sh()}   تا   {to_shamsi(datetime.datetime.now()).split('  ')[0]}"

                def _always_true(at):
                    return True

                def _in_ov_range(at):
                    if not _has_range:
                        return True
                    a = (at or "").replace("\u200c", "").strip()
                    return bool(a) and _ov_from_sh <= a <= _ov_to_sh

                _db = load_db()
                def _cur_loc(sid):
                    tr = next((t for t in _db.get("transfers_out", []) if t.get("slab_id") == sid), None)
                    if not tr: return "داخلی"
                    return tr.get("current_location") or tr.get("destination") or "روباز"

                def _pct(n, d):
                    return round(n / d * 100, 1) if d > 0 else None

                def _compute_overview(in_range_fn):
                    melts_in_range = [r for r in _db.get("melts", []) if in_range_fn(r.get("registered_at", ""))]
                    total  = len(melts_in_range)
                    qc     = sum(1 for r in _db["melts"] if r.get("qc_status") == "کنترل کیفی شده" and in_range_fn(r.get("qc_at", "")))
                    rej    = sum(1 for r in _db["melts"] if r.get("qc_status") == "عدم تایید کنترل کیفی" and in_range_fn(r.get("qc_at", "")))
                    scarfs = sum(1 for r in _db.get("scarf_cut", []) if r.get("operation") == "اسکارفی" and in_range_fn(r.get("registered_at", "")))
                    cuts   = sum(1 for r in _db.get("scarf_cut", []) if r.get("operation") == "برشی" and in_range_fn(r.get("registered_at", "")))
                    scraps = sum(1 for r in _db.get("scrap", []) if in_range_fn(r.get("registered_at", "")))
                    exited = sum(1 for r in _db["melts"] if r.get("exit_status") == "خروج زده شده" and in_range_fn(r.get("exit_at", "")))
                    out1   = sum(1 for r in melts_in_range if "۱" in _cur_loc(r["slab_id"]) and "داخلی" not in _cur_loc(r["slab_id"]))
                    out2   = sum(1 for r in melts_in_range if "۲" in _cur_loc(r["slab_id"]) and "داخلی" not in _cur_loc(r["slab_id"]))
                    inside = sum(1 for r in melts_in_range if _cur_loc(r["slab_id"]) == "انبار داخلی")
                    qc_n, sc_n, ct_n = qc, scarfs, cuts
                    return dict(total=total, qc=qc, rej=rej, scarfs=scarfs, cuts=cuts, scraps=scraps,
                                exited=exited, out1=out1, out2=out2, inside=inside,
                                qc_n=qc_n, sc_n=sc_n, ct_n=ct_n, to_n=sc_n + ct_n)

                _tot_range = _compute_overview(_in_ov_range)
                _tot_full  = _compute_overview(_always_true)

                # ── دلایل شناخته‌شده‌ی اسکارف/برش (برای جدول تفکیک اپراتورها) ──
                SCARF_REASONS = ["ترک طولی", "ترک عرضی", "کمربند", "پلاک", "حفره", "آسیب‌های دیگر"]
                CUT_REASONS   = ["طول مازاد / خارج از سفارش", "اسلب با سر", "اسلب با ته", "دلایل دیگر"]

                def _operator_breakdown(in_range_fn):
                    ops = [(u, d) for u, d in _db.get("users", {}).items() if d.get("role") == "scarf"]
                    result = []
                    for uname, udata in ops:
                        recs = [r for r in _db.get("scarf_cut", [])
                                if r.get("registered_by") == uname and in_range_fn(r.get("registered_at", ""))]
                        cut_recs = [r for r in recs if r.get("operation") == "برشی"]
                        scarf_recs = [r for r in recs if r.get("operation") == "اسکارفی"]

                        def reason_tally(records, known):
                            tally = {k: 0 for k in known}
                            for rr in records:
                                for rs in [x.strip() for x in (rr.get("reason") or "").replace("،", "،\n").split("\n")]:
                                    rs = rs.strip("، ").strip()
                                    if not rs:
                                        continue
                                    tally[rs] = tally.get(rs, 0) + 1
                            return tally

                        result.append({
                            "display": udata.get("display", uname),
                            "cut_total": len(cut_recs),
                            "cut_tally": reason_tally(cut_recs, CUT_REASONS),
                            "bauman_n": sum(1 for rr in cut_recs if rr.get("bauman_done")),
                            "scarf_total": len(scarf_recs),
                            "scarf_tally": reason_tally(scarf_recs, SCARF_REASONS),
                        })
                    return result

                _ops_range = _operator_breakdown(_in_ov_range)
                _ops_full  = _operator_breakdown(_always_true)

                # ════════════════════════════════════════════════════════
                #  ساخت Workbook شش‌شیتی
                # ════════════════════════════════════════════════════════
                wb = openpyxl.Workbook()

                NAVY   = "0D1E2E"
                NAVY2  = "16314A"
                GOLD   = "D4A043"
                LIGHT  = "F4F6F9"
                LIGHT2 = "E8EDF3"
                WHITE  = "FFFFFF"
                FN     = "B Nazanin"

                title_font   = Font(name=FN, bold=True, size=11, color=WHITE)
                sub_font     = Font(name=FN, bold=True, size=11, color=GOLD)
                section_font = Font(name=FN, bold=True, size=11, color=WHITE)
                label_font   = Font(name=FN, bold=True, size=11, color="222222")
                value_font   = Font(name=FN, bold=True, size=11, color=NAVY2)
                center       = Alignment(horizontal="center", vertical="center", readingOrder=2)
                right        = Alignment(horizontal="right", vertical="center", readingOrder=2)
                thin = Side(style="thin", color="C9D2DC")
                bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

                def banner(ws, row, span, text, fill=NAVY, font=title_font, height=26):
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
                    c = ws.cell(row=row, column=1, value=text)
                    c.font = font
                    c.alignment = center
                    for col in range(1, span + 1):
                        cc = ws.cell(row=row, column=col)
                        cc.fill = PatternFill("solid", fgColor=fill)
                        cc.border = Border(left=Side(style="medium", color=GOLD),
                                            right=Side(style="medium", color=GOLD),
                                            top=Side(style="medium", color=GOLD),
                                            bottom=Side(style="medium", color=GOLD))
                    ws.row_dimensions[row].height = height

                def kv_row(ws, row, label, value, label_col=1, value_col=2, fill=None):
                    lc = ws.cell(row=row, column=label_col, value=label)
                    lc.font = label_font; lc.alignment = right; lc.border = bdr
                    vc = ws.cell(row=row, column=value_col, value=value)
                    vc.font = value_font; vc.alignment = center; vc.border = bdr
                    if fill:
                        lc.fill = PatternFill("solid", fgColor=fill)
                        vc.fill = PatternFill("solid", fgColor=fill)
                    ws.row_dimensions[row].height = 20

                def setup_print(ws, last_row, span=2):
                    ws.page_setup.orientation = "portrait"
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 1
                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                    ws.print_area = f"A1:{get_column_letter(span)}{last_row}"
                    ws.page_margins.left = 0.4
                    ws.page_margins.right = 0.4
                    ws.page_margins.top = 0.5
                    ws.page_margins.bottom = 0.5

                # ──────────────────────────────────────────────────────
                #  شیت‌ساز خلاصه‌ی کلی — برای «کل دوره» و «بازه انتخابی» هر دو استفاده می‌شود
                # ──────────────────────────────────────────────────────
                def write_overview_sheet(ws, range_label, t):
                    ws.sheet_view.rightToLeft = True
                    ws.sheet_view.showGridLines = False
                    SPAN = 2

                    banner(ws, 1, SPAN, "📋  گزارش واحد اسلب یارد به ناظر پروژه", height=28)
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SPAN)
                    c2 = ws.cell(row=2, column=1, value="شرکت سازه پیشگام مدیسه  ·  فولاد سفید دشت")
                    c2.font = sub_font; c2.alignment = center
                    c2.fill = PatternFill("solid", fgColor=NAVY2)
                    ws.row_dimensions[2].height = 20

                    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=SPAN)
                    c3 = ws.cell(row=3, column=1,
                                 value=f"تاریخ تهیه گزارش: {to_shamsi(datetime.datetime.now())}   |   بازه‌ی گزارش: {range_label}")
                    c3.font = Font(name=FN, bold=True, size=11, color="666666")
                    c3.alignment = center
                    ws.row_dimensions[3].height = 18

                    r = 5
                    banner(ws, r, SPAN, "🔢  خلاصه‌ی کلی تولید", fill=NAVY2, font=section_font); r += 1
                    overview_rows = [
                        ("کل اسلب‌های ثبت‌شده", t["total"]),
                        ("کنترل کیفی شده", t["qc"]),
                        ("تایید نشده (رد QC)", t["rej"]),
                        ("خروج زده شده", t["exited"]),
                        ("اسکارف", t["scarfs"]),
                        ("برش", t["cuts"]),
                        ("قراضه", t["scraps"]),
                        ("کل ذوب‌های ثبت‌شده", t["total"]),
                    ]
                    for i, (lbl, val) in enumerate(overview_rows):
                        kv_row(ws, r, lbl, val, fill=(LIGHT if i % 2 == 0 else WHITE)); r += 1

                    r += 1
                    banner(ws, r, SPAN, "🏭  موجودی انبار", fill=NAVY2, font=section_font); r += 1
                    loc_rows = [
                        ("انبار روباز ۱", t["out1"]),
                        ("انبار روباز ۲", t["out2"]),
                        ("انبار داخلی", t["inside"]),
                    ]
                    for i, (lbl, val) in enumerate(loc_rows):
                        kv_row(ws, r, lbl, val, fill=(LIGHT if i % 2 == 0 else WHITE)); r += 1

                    r += 1
                    banner(ws, r, SPAN, "📊  ضریب اصلاح", fill=NAVY2, font=section_font); r += 1
                    ratio_rows = [
                        ("ضریب اسکارف", _pct(t["sc_n"], t["qc_n"])),
                        ("ضریب برش", _pct(t["ct_n"], t["qc_n"])),
                        ("ضریب اصلاح کل", _pct(t["to_n"], t["qc_n"])),
                    ]
                    for i, (lbl, val) in enumerate(ratio_rows):
                        val_disp = f"{val}%" if val is not None else "—"
                        kv_row(ws, r, lbl, val_disp, fill=(LIGHT if i % 2 == 0 else WHITE)); r += 1
                    r += 1

                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
                    foot = ws.cell(row=r, column=1,
                                    value="این گزارش به‌صورت سیستمی از سامانه مدیریت تختال تولید شده است")
                    foot.font = Font(name=FN, bold=True, size=11, color="999999")
                    foot.alignment = center

                    ws.column_dimensions["A"].width = 60
                    ws.column_dimensions["B"].width = 32
                    setup_print(ws, r, SPAN)

                # ──────────────────────────────────────────────────────
                #  شیت ۱: گزارش کلی — کل دوره (از اولین اسلب تا امروز)
                # ──────────────────────────────────────────────────────
                ws_full = wb.active
                ws_full.title = "کلی - کل دوره"
                write_overview_sheet(ws_full, _full_range_label, _tot_full)

                # ──────────────────────────────────────────────────────
                #  شیت ۲: گزارش کلی — بازه‌ی انتخابی کاربر
                # ──────────────────────────────────────────────────────
                ws_range = wb.create_sheet("کلی - بازه انتخابی")
                write_overview_sheet(ws_range, _ov_range_label, _tot_range)

                # ──────────────────────────────────────────────────────
                #  شیت‌ساز فعالیت اپراتورها — برای «بازه» و «کل دوره» هر دو استفاده می‌شود
                # ──────────────────────────────────────────────────────
                def write_operator_sheet(ws, range_label, ops_data):
                    ws.sheet_view.rightToLeft = True
                    ws.sheet_view.showGridLines = False
                    SPAN = 2
                    banner(ws, 1, SPAN, "👷  گزارش فعالیت اپراتورهای برش و اسکارف", height=28)
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SPAN)
                    c2 = ws.cell(row=2, column=1, value=f"بازه‌ی گزارش: {range_label}")
                    c2.font = sub_font; c2.alignment = center
                    c2.fill = PatternFill("solid", fgColor=NAVY2)
                    ws.row_dimensions[2].height = 20
                    r = 4
                    if not ops_data:
                        kv_row(ws, r, "هیچ اپراتور برش/اسکارفی در سیستم تعریف نشده", "—", fill=LIGHT); r += 1
                    for op in ops_data:
                        banner(ws, r, SPAN, f"👤  {op['display']}", fill=NAVY2, font=section_font); r += 1
                        kv_row(ws, r, "▶ کل اسلب‌های برش‌خورده", op["cut_total"], fill=LIGHT2); r += 1
                        for reason, cnt in op["cut_tally"].items():
                            kv_row(ws, r, f"—  {reason}", cnt, fill=WHITE); r += 1
                        kv_row(ws, r, "—  تست باومن", op["bauman_n"], fill=LIGHT); r += 1
                        r += 1
                        kv_row(ws, r, "▶ کل اسلب‌های اسکارف‌شده", op["scarf_total"], fill=LIGHT2); r += 1
                        for reason, cnt in op["scarf_tally"].items():
                            kv_row(ws, r, f"—  {reason}", cnt, fill=WHITE); r += 1
                        r += 2
                    ws.column_dimensions["A"].width = 62
                    ws.column_dimensions["B"].width = 28
                    setup_print(ws, r, SPAN)

                # (شیت‌های فعالیت اپراتورها بعد از نمودار و شیت کلی برش/اسکارف ساخته می‌شوند — پایین‌تر)

                # ──────────────────────────────────────────────────────
                #  شیت ۳: نمودار روزانه (جدول + چارت واقعی اکسل، حرفه‌ای‌تر)
                # ──────────────────────────────────────────────────────
                ws2 = wb.create_sheet("گزارش روزانه اسلب یارد")
                ws2.sheet_view.rightToLeft = True

                hdr_fill   = PatternFill("solid", fgColor=NAVY)
                hdr_font   = Font(name="B Nazanin", bold=True, color="FFFFFF", size=11)
                hdr_align  = Alignment(horizontal="center", vertical="center")
                thin_border = bdr
                alt_fill   = PatternFill("solid", fgColor="EEF4FB")
                num_font   = Font(name="B Nazanin", size=11)
                date_font  = Font(name="B Nazanin", size=11)

                ws2.merge_cells("A1:E1")
                title_cell = ws2["A1"]
                title_cell.value = f"نمودار روزانه فولاد سفید دشت - گزارش روزانه اسلب یارد  |  بازه: {from_str}  تا  {to_str}"
                title_cell.font = Font(name="B Nazanin", bold=True, size=13, color="003366")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill("solid", fgColor="E8EDF3")
                ws2.row_dimensions[1].height = 28

                headers = ["ردیف", "تاریخ", "کنترل کیفی تایید", "اسکارف", "برش"]
                for ci, h in enumerate(headers, 1):
                    cell = ws2.cell(row=2, column=ci, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = hdr_align
                    cell.border = thin_border
                ws2.row_dimensions[2].height = 24

                for i, day in enumerate(days, 1):
                    row_idx = i + 2
                    row_fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

                    c0 = ws2.cell(row=row_idx, column=1, value=i)
                    c0.font = num_font; c0.alignment = hdr_align; c0.border = thin_border; c0.fill = row_fill

                    c1 = ws2.cell(row=row_idx, column=2, value=day)
                    c1.font = date_font; c1.alignment = hdr_align; c1.border = thin_border; c1.fill = row_fill

                    c2 = ws2.cell(row=row_idx, column=3, value=qc_d[day])
                    c2.font = Font(name="B Nazanin", size=11, color="1A7A50", bold=(qc_d[day] > 0))
                    c2.alignment = hdr_align; c2.border = thin_border
                    c2.fill = PatternFill("solid", fgColor="E8F5EE") if qc_d[day] > 0 else row_fill

                    c3 = ws2.cell(row=row_idx, column=4, value=scarfs_d[day])
                    c3.font = Font(name="B Nazanin", size=11, color="8A7840", bold=(scarfs_d[day] > 0))
                    c3.alignment = hdr_align; c3.border = thin_border
                    c3.fill = PatternFill("solid", fgColor="F5F0E8") if scarfs_d[day] > 0 else row_fill

                    c4 = ws2.cell(row=row_idx, column=5, value=cuts_d[day])
                    c4.font = Font(name="B Nazanin", size=11, color="155882", bold=(cuts_d[day] > 0))
                    c4.alignment = hdr_align; c4.border = thin_border
                    c4.fill = PatternFill("solid", fgColor="E8F0F8") if cuts_d[day] > 0 else row_fill

                    ws2.row_dimensions[row_idx].height = 20

                sum_row = len(days) + 3
                ws2.cell(row=sum_row, column=1, value="جمع").font = Font(name="B Nazanin", bold=True, size=11)
                ws2.cell(row=sum_row, column=1).alignment = hdr_align
                ws2.cell(row=sum_row, column=1).fill = PatternFill("solid", fgColor="D0DCE8")
                ws2.cell(row=sum_row, column=1).border = thin_border
                ws2.cell(row=sum_row, column=2, value="—").alignment = hdr_align
                ws2.cell(row=sum_row, column=2).fill = PatternFill("solid", fgColor="D0DCE8")
                ws2.cell(row=sum_row, column=2).border = thin_border
                for ci, data_dict in [(3, qc_d), (4, scarfs_d), (5, cuts_d)]:
                    sc = ws2.cell(row=sum_row, column=ci, value=sum(data_dict.values()))
                    sc.font = Font(name="B Nazanin", bold=True, size=12)
                    sc.alignment = hdr_align
                    sc.fill = PatternFill("solid", fgColor="C0D0E0")
                    sc.border = thin_border
                ws2.row_dimensions[sum_row].height = 24

                for col, w in [(1, 14), (2, 26), (3, 24), (4, 24), (5, 32)]:
                    ws2.column_dimensions[get_column_letter(col)].width = w

                # ── چارت میله‌ای واقعی اکسل — حرفه‌ای‌تر: تیتر محورها، برچسب مقدار روی هر میله، سبک شیک ──
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "clustered"
                from openpyxl.chart.text import RichText
                from openpyxl.drawing.text import (
                    CharacterProperties, Paragraph, ParagraphProperties, Font as DrawFont,
                    RichTextProperties
                )
                from openpyxl.chart.axis import ChartLines
                from openpyxl.chart.shapes import GraphicalProperties
                from openpyxl.drawing.line import LineProperties
                from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
                from openpyxl.chart.marker import DataPoint

                # ── همه فونت‌های نمودار: ب نازنین، بولد، سایز ۱۱ — یکدست در کل گزارش ──
                CHART_FONT_SZ = 1100

                def _label_txpr(size=CHART_FONT_SZ, bold=True, color="1A2B3C", rotate=False):
                    cp = CharacterProperties(sz=size, b=bold,
                                              latin=DrawFont(typeface="B Nazanin"),
                                              solidFill=color)
                    body = RichTextProperties(rot=-2700000) if rotate else RichTextProperties()
                    return RichText(bodyPr=body,
                                     p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

                def _title_cp(color="0D1E2E"):
                    return CharacterProperties(sz=CHART_FONT_SZ, b=True,
                                                latin=DrawFont(typeface="B Nazanin"), solidFill=color)

                # ── هاشور به‌جای رنگ یکدست — برای میله‌های هر نمودار (الگوهای متفاوت برای تشخیص بهتر) ──
                # الگوهای هاشور مناسب پرینت سیاه‌وسفید — کاملاً متفاوت و خوانا (افقی / مورب بالا / مورب پایین)
                _HATCH_PATTERNS = ["narHorz", "wdUpDiag", "wdDnDiag", "narVert", "dashDnDiag", "dotGrid"]
                _HATCH_COLORS   = ["1A7A50", "D4A043", "155882", "8E3B46", "5B3E96", "2E7D8C"]
                _N_HATCH = len(_HATCH_PATTERNS)

                def _hatch_fill(idx):
                    i = idx % _N_HATCH
                    return GraphicalProperties(
                        pattFill=PatternFillProperties(
                            prst=_HATCH_PATTERNS[i],
                            fgClr=ColorChoice(srgbClr=_HATCH_COLORS[i]),
                            bgClr=ColorChoice(srgbClr="FFFFFF")),
                        ln=LineProperties(solidFill=_HATCH_COLORS[i], w=9000))

                chart.style = 10
                chart.title = "نمودار روزانه فولاد سفید دشت - گزارش روزانه اسلب یارد"
                chart.title.tx.rich.p[0].r[0].rPr = _title_cp()

                # محور Y (عمودی - تعداد)
                chart.y_axis.title = "تعداد (عدد)"
                chart.y_axis.title.tx.rich.p[0].r[0].rPr = _title_cp("155882")
                chart.y_axis.delete = False
                chart.y_axis.majorGridlines = ChartLines(
                    spPr=GraphicalProperties(ln=LineProperties(solidFill="D9E2EC", w=4000)))
                chart.y_axis.txPr = _label_txpr(color="333333")
                chart.y_axis.numFmt = "0"

                # محور X (افقی - تاریخ) — بدون تیتر اضافه تا با برچسب تاریخ‌های چرخیده قاطی نشود
                chart.x_axis.title = None
                chart.x_axis.delete = False
                chart.x_axis.txPr = _label_txpr(color="333333", rotate=True)
                chart.x_axis.lblAlgn = "ctr"
                chart.x_axis.lblOffset = 100

                chart.gapWidth = 60
                chart.overlap = -8
                chart.height = 14
                chart.width = 32
                chart.roundedCorners = True

                data_ref = Reference(ws2, min_col=3, max_col=5, min_row=2, max_row=2 + len(days))
                cats_ref = Reference(ws2, min_col=2, min_row=3, max_row=2 + len(days))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                _series_names = ["کنترل کیفی تایید", "اسکارف", "برش"]
                for i in range(3):
                    chart.series[i].graphicalProperties = _hatch_fill(i)
                    # برچسب روی هر میله: هم نوعش (اسکارف/برش/QC) هم مقدارش — تا قاطی نشود
                    chart.series[i].dLbls = DataLabelList()
                    chart.series[i].dLbls.showVal = True
                    chart.series[i].dLbls.showLegendKey = False
                    chart.series[i].dLbls.showCatName = False
                    chart.series[i].dLbls.showSerName = True
                    chart.series[i].dLbls.separator = ": "
                    chart.series[i].dLbls.numFmt = "0"
                    chart.series[i].dLbls.txPr = _label_txpr(color=_HATCH_COLORS[i])
                    chart.series[i].dLbls.dLblPos = "outEnd"

                chart.dataLabels = None  # دیتالیبل سراسری حذف؛ هر سری لیبل اختصاصی خودش را دارد (بالا تنظیم شد)
                chart.legend.position = "b"
                chart.legend.overlay = False
                chart.legend.txPr = _label_txpr(color="0D1E2E")

                chart_anchor_row = sum_row + 2
                ws2.add_chart(chart, f"A{chart_anchor_row}")
                setup_print(ws2, chart_anchor_row + 28, 14)

                # ──────────────────────────────────────────────────────
                #  شیت ۴: کلی - اسکارف/برش نسبت به کنترل کیفی (کل دوره)
                #  از تاریخ ثبت اولین اسلب در نرم‌افزار تا امروز — تجمیعی، نه روزانه
                # ──────────────────────────────────────────────────────
                ws4 = wb.create_sheet("کلی - اسکارف و برش (کل دوره)")
                ws4.sheet_view.rightToLeft = True
                ws4.sheet_view.showGridLines = False

                ws4.merge_cells("A1:B1")
                t4 = ws4.cell(row=1, column=1,
                    value="📊  گزارش کلی اسکارف و برش نسبت به کنترل کیفی  —  کل دوره")
                t4.font = title_font; t4.alignment = center
                ws4.cell(row=1, column=1).fill = PatternFill("solid", fgColor=NAVY)
                for col in (1, 2):
                    ws4.cell(row=1, column=col).fill = PatternFill("solid", fgColor=NAVY)
                ws4.row_dimensions[1].height = 28

                ws4.merge_cells("A2:B2")
                t4b = ws4.cell(row=2, column=1, value=f"بازه: {_full_range_label}")
                t4b.font = sub_font; t4b.alignment = center
                ws4.cell(row=2, column=1).fill = PatternFill("solid", fgColor=NAVY2)
                ws4.row_dimensions[2].height = 20

                r4 = 4
                banner(ws4, r4, 2, "🔢  وضعیت کلی تصحیح اسلب‌های QC شده", fill=NAVY2, font=section_font); r4 += 1
                rows4 = [
                    ("کل اسلب‌های کنترل کیفی شده (مبنا)", _tot_full["qc_n"]),
                    ("کل اسکارف خورده", _tot_full["sc_n"]),
                    ("کل برش خورده", _tot_full["ct_n"]),
                    ("جمع کل اصلاح (اسکارف + برش)", _tot_full["to_n"]),
                ]
                for i, (lbl, val) in enumerate(rows4):
                    kv_row(ws4, r4, lbl, val, fill=(LIGHT if i % 2 == 0 else WHITE)); r4 += 1

                r4 += 1
                banner(ws4, r4, 2, "📈  درصد نسبت به کنترل کیفی شده‌ها", fill=NAVY2, font=section_font); r4 += 1
                _p_sc = _pct(_tot_full["sc_n"], _tot_full["qc_n"])
                _p_ct = _pct(_tot_full["ct_n"], _tot_full["qc_n"])
                _p_to = _pct(_tot_full["to_n"], _tot_full["qc_n"])
                ratio_rows4 = [
                    ("درصد اسکارف", f"{_p_sc}%" if _p_sc is not None else "—"),
                    ("درصد برش", f"{_p_ct}%" if _p_ct is not None else "—"),
                    ("درصد کل اصلاح", f"{_p_to}%" if _p_to is not None else "—"),
                ]
                for i, (lbl, val) in enumerate(ratio_rows4):
                    kv_row(ws4, r4, lbl, val, fill=(LIGHT if i % 2 == 0 else WHITE)); r4 += 1
                r4 += 1

                # جدول کوچک برای چارت — شامل کل + تفکیک هر اپراتور، همه در یک نمودار واحد
                chart4_hdr_row = r4
                ws4.cell(row=chart4_hdr_row, column=1, value="شاخص").font = hdr_font
                ws4.cell(row=chart4_hdr_row, column=1).fill = hdr_fill
                ws4.cell(row=chart4_hdr_row, column=1).alignment = hdr_align
                ws4.cell(row=chart4_hdr_row, column=1).border = thin_border
                ws4.cell(row=chart4_hdr_row, column=2, value="تعداد").font = hdr_font
                ws4.cell(row=chart4_hdr_row, column=2).fill = hdr_fill
                ws4.cell(row=chart4_hdr_row, column=2).alignment = hdr_align
                ws4.cell(row=chart4_hdr_row, column=2).border = thin_border
                chart4_data = [("کنترل کیفی تایید شده", _tot_full["qc_n"]), ("اسکارف", _tot_full["sc_n"]),
                                ("برش", _tot_full["ct_n"])]
                n_cats4 = len(chart4_data)
                for i, (lbl, val) in enumerate(chart4_data):
                    rr = chart4_hdr_row + 1 + i
                    cA = ws4.cell(row=rr, column=1, value=lbl); cA.border = thin_border; cA.alignment = hdr_align
                    cB = ws4.cell(row=rr, column=2, value=val); cB.border = thin_border; cB.alignment = hdr_align

                chart4 = BarChart()
                chart4.type = "col"
                chart4.style = 10
                chart4.title = "مقایسه کلی اسکارف / برش / کنترل کیفی تایید (کل دوره)"
                chart4.title.tx.rich.p[0].r[0].rPr = _title_cp()
                chart4.y_axis.title = "تعداد (عدد)"
                chart4.y_axis.title.tx.rich.p[0].r[0].rPr = _title_cp("155882")
                # محور افقی بدون تیتر اضافه — خود نام شاخص‌ها (QC/اسکارف/برش/هر اپراتور) زیر هر میله کافی و خواناست
                chart4.x_axis.title = None
                chart4.x_axis.delete = False
                chart4.y_axis.delete = False
                chart4.y_axis.majorGridlines = ChartLines(
                    spPr=GraphicalProperties(ln=LineProperties(solidFill="D9E2EC", w=4000)))
                chart4.x_axis.txPr = _label_txpr(color="333333", rotate=(n_cats4 > 3))
                chart4.y_axis.txPr = _label_txpr(color="333333")
                chart4.gapWidth = 55
                chart4.height = 12
                chart4.width = max(22, min(42, 6 + n_cats4 * 3))
                chart4.legend = None

                data4_ref = Reference(ws4, min_col=2, min_row=chart4_hdr_row, max_row=chart4_hdr_row + n_cats4)
                cats4_ref = Reference(ws4, min_col=1, min_row=chart4_hdr_row + 1, max_row=chart4_hdr_row + n_cats4)
                chart4.add_data(data4_ref, titles_from_data=True)
                chart4.set_categories(cats4_ref)
                pt0 = chart4.series[0]
                pt0.graphicalProperties = _hatch_fill(0)
                pt0.data_points = [
                    DataPoint(idx=i, spPr=_hatch_fill(i)) for i in range(n_cats4)
                ]
                pt0.dLbls = DataLabelList()
                pt0.dLbls.showVal = True
                pt0.dLbls.showLegendKey = False
                # فقط عدد روی میله — اسم هر شاخص همین‌جوری هم زیر محور افقی نوشته شده، تکرارش باعث شلوغی می‌شد
                pt0.dLbls.showCatName = False
                pt0.dLbls.showSerName = False
                pt0.dLbls.numFmt = "0"
                pt0.dLbls.txPr = _label_txpr(size=1200, color="0D1E2E")
                pt0.dLbls.dLblPos = "outEnd"

                # ── نمودار دقیقاً زیر همین جدول کوچک قرار می‌گیرد، نه کنارش ──
                chart4_anchor_row = chart4_hdr_row + n_cats4 + 3
                ws4.add_chart(chart4, f"A{chart4_anchor_row}")
                final_row = chart4_anchor_row + max(24, n_cats4 + 10)


                ws4.column_dimensions["A"].width = 62
                ws4.column_dimensions["B"].width = 30
                ws4.column_dimensions["C"].width = 30
                setup_print(ws4, final_row, 14)

                # ──────────────────────────────────────────────────────
                #  شیت ۵: فعالیت اپراتورها — بازه‌ی انتخابی
                # ──────────────────────────────────────────────────────
                ws_op_range = wb.create_sheet("فعالیت اپراتورها - بازه")
                write_operator_sheet(ws_op_range, _ov_range_label, _ops_range)

                # ──────────────────────────────────────────────────────
                #  شیت ۶: فعالیت اپراتورها — کل دوره (برای پایش کلی عملکرد)
                # ──────────────────────────────────────────────────────
                ws_op_full = wb.create_sheet("فعالیت اپراتورها - کل دوره")
                write_operator_sheet(ws_op_full, _full_range_label, _ops_full)

                wb.save(path)
                chart_status.config(text=f"✔  فایل اکسل ذخیره شد", fg=C["success"])
                messagebox.showinfo("موفق", f"فایل اکسل با موفقیت ذخیره شد:\n{path}", parent=self)
            except Exception as ex:
                messagebox.showerror("خطا", f"خطا در ذخیره فایل:\n{ex}", parent=self)

        # ── دکمه‌های عملیات ──
        btn_row3 = tk.Frame(chart_panel, bg=C["card"])
        btn_row3.pack(fill="x", padx=16, pady=(4, 12))
        styled_btn(btn_row3, "📊  نمایش نمودار", draw_chart,
                   color=C["btn_primary"], width=150).pack(side="right", padx=6)
        styled_btn(btn_row3, "⬇  دانلود اکسل", export_excel,
                   color=C["btn_success"], width=150).pack(side="right", padx=6)

        # رسم اولیه
        chart_cv.after(200, draw_chart)

    def _build_melts_tab(self, tab):
        tab.configure(bg=C["panel"])
        style = ttk.Style()
        style.configure("Melts.TNotebook", background=C["panel"], bordercolor=C["border"])
        style.configure("Melts.TNotebook.Tab",
            background=C["tab_inactive"], foreground=C["text"],
            font=(_MAIN_FONT, 11, "bold"), padding=[14, 7])
        style.map("Melts.TNotebook.Tab",
            background=[("selected", C["accent"])],
            foreground=[("selected", "#ffffff")])
        sub_nb = ttk.Notebook(tab, style="Melts.TNotebook")
        sub_nb.pack(fill="both", expand=True, padx=6, pady=6)
        tab_manual = tk.Frame(sub_nb, bg=C["panel"])
        tab_file   = tk.Frame(sub_nb, bg=C["panel"])
        sub_nb.add(tab_manual, text="➕  ثبت دستی")
        sub_nb.add(tab_file,   text="📂  ثبت از طریق فایل")
        self._build_melts_manual_tab(tab_manual)
        self._build_file_melt_tab(tab_file)

    def _build_melts_manual_tab(self, tab):
        tab.configure(bg=C["panel"])
        form = card_frame(tab)
        form.pack(fill="x", padx=16, pady=12)
        inner = tk.Frame(form, bg=C["card"])
        inner.pack(padx=16, pady=12, fill="x")
        tk.Label(inner, text="➕  ثبت ذوب جدید", bg=C["card"],
                 fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0,8))

        row1 = tk.Frame(inner, bg=C["card"])
        row1.pack(fill="x", pady=4)
        tk.Label(row1, text="شماره اسلب (۱۱ رقم):", bg=C["card"],
                 fg=C["text"], font=FONT_NORM).pack(side="right", padx=(0,6))
        slab_var = tk.StringVar()
        slab_ent = tk.Entry(row1, textvariable=slab_var,
                             bg=C["entry_bg"], fg=C["text"],
                             insertbackground=C["accent"], font=FONT_MONO,
                             justify="right", bd=0, relief="flat", highlightthickness=1,
                             highlightbackground=C["border"],
                             highlightcolor=C["accent"], width=18)
        slab_ent.pack(side="right")

        note_var = tk.StringVar()  # دیگر در فرم نمایش داده نمی‌شود

        status_lbl = tk.Label(inner, text="", bg=C["card"],
                               fg=C["success"], font=FONT_SMALL)
        status_lbl.pack(anchor="e", pady=2)

        def do_register():
            ok, msg, sid = validate_slab_id(slab_var.get())
            if not ok:
                messagebox.showerror("خطای اعتبارسنجی", msg, parent=self)
                return
            db = load_db()
            if check_duplicate(db, "melts", sid):
                messagebox.showerror("⛔  ثبت تکراری",
                    f"اسلب {sid} قبلاً توسط شخص دیگری ثبت شده است.\n\nبرای پیدا کردن از جستجو استفاده کنید.\nدر غیر این صورت با سرپرست کارگاه تماس بگیرید.", parent=self)
                return
            rec = {
                "slab_id": sid,
                "note": note_var.get().strip(),
                "qc_status": "ثبت شده",
                "registered_by": self.username,
                "registered_at": now_str(),
                "updated_at": now_str(),
            }
            db["melts"].append(rec)
            save_db(db)
            slab_var.set("")
            note_var.set("")
            status_lbl.config(text=f"✔  اسلب {sid} با موفقیت ثبت شد.")
            refresh_tree()

        btn_row = tk.Frame(inner, bg=C["card"])
        btn_row.pack(fill="x", pady=6)
        styled_btn(btn_row, "✔  ثبت اسلب", do_register).pack(side="right")
        slab_ent.bind("<Return>", lambda e: do_register())

        separator(tab)

        # ══════════════════════════════════════════
        #  پنل دسته‌جمعی کنترل کیفی — همه نقش‌ها (ثبت انتخاب‌شده‌ها)
        # ══════════════════════════════════════════
        QC_OPTIONS = ["کنترل کیفی شده", "عدم تایید کنترل کیفی"]

        bulk_bar = tk.Frame(tab, bg=C["header_bg"],
                            highlightthickness=2,
                            highlightbackground=C["gold"])
        bulk_bar.pack(fill="x", padx=16, pady=(4,0))

        top_row = tk.Frame(bulk_bar, bg=C["header_bg"])
        top_row.pack(fill="x", padx=14, pady=(8,4))
        tk.Label(top_row,
                 text="✔  ثبت دسته‌جمعی کنترل کیفی (انتخاب با Ctrl یا ☑)",
                 bg=C["header_bg"], fg=C["gold"],
                 font=(_MAIN_FONT, 12, "bold")).pack(side="right")
        sel_count_lbl = tk.Label(top_row,
                                  text="هیچ ردیفی انتخاب نشده",
                                  bg=C["header_bg"], fg=C["text_dim"],
                                  font=(_MAIN_FONT, 10, "bold"))
        sel_count_lbl.pack(side="left", padx=6)

        btn_row2 = tk.Frame(bulk_bar, bg=C["header_bg"])
        btn_row2.pack(fill="x", padx=14, pady=(0,10))

        tk.Label(btn_row2, text="وضعیت:",
                 bg=C["header_bg"], fg=C["text"],
                 font=FONT_NORM).pack(side="right", padx=(0,4))
        bulk_status_cb = make_combo(btn_row2, QC_OPTIONS, width=18)
        bulk_status_cb.set("")
        bulk_status_cb.pack(side="right", padx=(0,12))

        tk.Frame(btn_row2, bg=C["border"], width=1).pack(side="right", fill="y", pady=4, padx=4)

        def _make_sel_btn(parent, text, cmd, color):
            f = tk.Frame(parent, bg=color, cursor="hand2")
            f.pack(side="right", padx=3)
            lbl = tk.Label(f, text=text, bg=color, fg="#ffffff",
                           font=(_MAIN_FONT, 10, "bold"),
                           padx=12, pady=8, cursor="hand2")
            lbl.pack()
            lighter = _lighten(color, 20)
            f.bind("<Button-1>",   lambda e: cmd())
            lbl.bind("<Button-1>", lambda e: cmd())
            f.bind("<Enter>", lambda e: [f.config(bg=lighter), lbl.config(bg=lighter)])
            f.bind("<Leave>", lambda e: [f.config(bg=color),   lbl.config(bg=color)])
            return f

        bulk_status_lbl = tk.Label(bulk_bar, text="", bg=C["header_bg"],
                                    fg=C["success"],
                                    font=(_MAIN_FONT, 9, "bold"))
        bulk_status_lbl.pack(anchor="e", padx=14, pady=(0,6))

        def do_select_all_pending():
            tree.selection_set([])
            for iid in tree.get_children():
                vals = tree.item(iid, "values")
                if vals and vals[2] == "ثبت شده":
                    tree.selection_add(iid)
                    _checked[iid] = True
                    tree.set(iid, "chk", "☑")
            n = len(tree.selection())
            sel_count_lbl.config(text=f"▶  {n} ردیف انتخاب شد",
                                  fg=C["warning"])
            bulk_status_lbl.config(
                text=f"✔  {n} اسلب با وضعیت «ثبت شده» انتخاب شد",
                fg=C["warning"])

        def do_select_all():
            for iid in tree.get_children():
                _checked[iid] = True
                tree.set(iid, "chk", "☑")
            tree.selection_set(tree.get_children())
            n = len(tree.get_children())
            sel_count_lbl.config(text=f"▶  {n} ردیف انتخاب شد",
                                  fg=C["warning"])
            bulk_status_lbl.config(
                text=f"✔  {n} اسلب انتخاب شد", fg=C["warning"])

        def do_deselect_all():
            for iid in tree.get_children():
                _checked[iid] = False
                tree.set(iid, "chk", "○")
            tree.selection_set([])
            sel_count_lbl.config(text="هیچ ردیفی انتخاب نشده",
                                  fg=C["text_dim"])
            bulk_status_lbl.config(text="")

        def _selected_iids():
            sel = list(tree.selection())
            if sel:
                return sel
            return [iid for iid, on in _checked.items() if on]

        def do_bulk_submit():
            sel = _selected_iids()
            if not sel:
                messagebox.showwarning("خطا",
                    "ابتدا اسلب‌ها را انتخاب کنید.\n"
                    "با Ctrl چند ردیف را انتخاب کنید، یا روی ☑ کلیک کنید.",
                    parent=self)
                return
            new_status = (bulk_status_cb.get() or "").strip()
            if not new_status or new_status not in QC_OPTIONS:
                messagebox.showerror("خطا",
                    "لطفا یکی از موارد را انتخاب کنید",
                    parent=self)
                return
            # فقط اسلب‌های «ثبت شده» برای کاربر عادی؛ ادمین همه
            usable = []
            for iid in sel:
                vals = tree.item(iid, "values")
                if not vals or len(vals) < 3:
                    continue
                if self.role != "admin" and vals[2] != "ثبت شده":
                    continue
                usable.append(iid)
            if not usable:
                messagebox.showwarning("خطا",
                    "هیچ ردیف قابل ثبتی در انتخاب‌ها نیست.\n"
                    "فقط اسلب‌های با وضعیت «ثبت شده» قابل تأیید هستند.",
                    parent=self)
                return
            try:
                tree._stf_locked_yview = float(tree.yview()[0])
            except Exception:
                pass
            tree._stf_scroll_freeze = True
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            if not messagebox.askyesno("تأیید دسته‌جمعی",
                f"وضعیت {len(usable)} اسلب به «{new_status}» تغییر می‌کند.\n"
                "آیا مطمئنید؟", parent=self):
                tree._stf_scroll_freeze = False
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
                return
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            db = load_db()
            ts = now_str(); changed = 0
            for iid in usable:
                vals = tree.item(iid, "values")
                if not vals or len(vals) < 2: continue
                sid = vals[1]
                for rec in db["melts"]:
                    if rec["slab_id"] == sid:
                        rec["qc_status"] = new_status
                        rec["qc_by"]     = self.username
                        rec["qc_at"]     = ts
                        changed += 1; break
            save_db(db)
            bulk_status_lbl.config(
                text=f"✔  {changed} اسلب به «{new_status}» تغییر یافت",
                fg=C["success"])
            sel_count_lbl.config(text="هیچ ردیفی انتخاب نشده",
                                  fg=C["text_dim"])
            # بدون رفرش کامل — فقط ردیف‌های تغییرکرده (اسکرول ثابت)
            db2 = load_db()
            by_sid = {r.get("slab_id"): r for r in db2.get("melts", [])}
            for iid in usable:
                try:
                    vals = tree.item(iid, "values")
                    if not vals or len(vals) < 2:
                        continue
                    sid = vals[1]
                    rec = by_sid.get(sid)
                    if rec:
                        _apply_row_inplace(sid, rec, db2)
                except Exception:
                    pass
            for iid in tree.get_children(""):
                _checked[iid] = False
                try:
                    tree.set(iid, "chk", "○")
                except Exception:
                    pass
            tree.selection_set([])
            try:
                tree.yview_moveto(float(tree._stf_locked_yview))
            except Exception:
                pass
            tree._stf_scroll_freeze = False
            def _deferred_other():
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
                if hasattr(self, "_qc_refresh_fn"):
                    try: self._qc_refresh_fn()
                    except: pass
                if hasattr(self, "_rejected_refresh_fn"):
                    try: self._rejected_refresh_fn()
                    except: pass
                try:
                    tree.yview_moveto(float(tree._stf_locked_yview))
                except Exception:
                    pass
            try:
                self.after(300, _deferred_other)
            except Exception:
                pass

        submit_f = tk.Frame(btn_row2, bg=C["btn_success"], cursor="hand2")
        submit_f.pack(side="left", padx=(0,6))
        submit_l = tk.Label(submit_f,
                            text="  ✔  ثبت انتخاب‌شده‌ها  ",
                            bg=C["btn_success"], fg="#ffffff",
                            font=(_MAIN_FONT, 11, "bold"),
                            padx=10, pady=10, cursor="hand2")
        submit_l.pack()
        _gs = _lighten(C["btn_success"], 20)
        submit_f.bind("<Button-1>",   lambda e: do_bulk_submit())
        submit_l.bind("<Button-1>",   lambda e: do_bulk_submit())
        submit_f.bind("<Enter>", lambda e: [submit_f.config(bg=_gs), submit_l.config(bg=_gs)])
        submit_f.bind("<Leave>", lambda e: [submit_f.config(bg=C["btn_success"]),
                                             submit_l.config(bg=C["btn_success"])])

        _make_sel_btn(btn_row2, "☑ انتخاب ثبت‌شده‌ها",
                      do_select_all_pending, C["btn_primary"])
        _make_sel_btn(btn_row2, "☑ انتخاب همه",
                      do_select_all, "#4a6080")
        _make_sel_btn(btn_row2, "○ لغو انتخاب",
                      do_deselect_all, C["btn_ghost"])

        def _update_sel_count(*_):
            n = len(tree.selection()) or sum(1 for v in _checked.values() if v)
            if n == 0:
                sel_count_lbl.config(text="هیچ ردیفی انتخاب نشده",
                                      fg=C["text_dim"])
            else:
                sel_count_lbl.config(text=f"▶  {n} ردیف انتخاب شده",
                                      fg=C["warning"])

        # ── جدول اسلب‌ها ──
        hdr_row = tk.Frame(tab, bg=C["panel"])
        hdr_row.pack(fill="x", padx=16, pady=(4,0))
        tk.Label(hdr_row, text="📋  کلیه اسلب‌های ثبت‌شده",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(side="right")

        # چک‌باکس column در ستون اول
        cols = ("chk","slab_id","qc_status","registered_by","registered_date","registered_time","note","action")
        heads = ("✔","شماره اسلب","وضعیت QC","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","توضیحات","تغییر وضعیت")
        tree_frame, tree = scrolled_tree(tab, cols, heads, height=12)
        tree.configure(selectmode="extended")
        tree_frame.pack(fill="both", expand=True, padx=16, pady=2)
        tree.column("chk",    width=34,  anchor="center", minwidth=34)
        tree.column("action", width=190, anchor="center")

        # دیکشنری وضعیت چک‌باکس هر ردیف
        _checked = {}   # iid -> bool

        def _toggle_check(iid):
            _checked[iid] = not _checked.get(iid, False)
            tree.set(iid, "chk", "☑" if _checked[iid] else "○")
            if _checked[iid]:
                tree.selection_add(iid)
            else:
                tree.selection_remove(iid)
            _update_sel_count()

        def _on_chk_click(event):
            row = tree.identify_row(event.y)
            col = tree.identify_column(event.x)
            if not row: return
            # شناسه ستون chk در displaycolumns
            try:
                col_name = tree.column(col, "id")
            except:
                col_name = ""
            if col_name == "chk":
                _toggle_check(row)

        tree.bind("<Button-1>", _on_chk_click)

        # ── نوار جستجو + مرتب‌سازی (یک خط پایین جدول) ──
        tool_row = tk.Frame(tab, bg=C["panel"])
        tool_row.pack(fill="x", padx=16, pady=(0,2))

        sb = search_bar(tool_row, tree, col_indices=[1])
        sb.pack(side="right", padx=4)
        sort_toolbar(tool_row, tree, slab_col="slab_id", bg=C["panel"]).pack(side="right", padx=4)

        QC_OPTIONS = ["کنترل کیفی شده", "عدم تایید کنترل کیفی"]
        action_widgets = {}

        def clear_action_widgets():
            for fr in action_widgets.values():
                try: fr.destroy()
                except: pass
            action_widgets.clear()

        def build_action_widget(iid, sid, cur_status):
            locked = cur_status != "ثبت شده"
            frame = tk.Frame(tree_frame, bg=C["card2"])
            if self.role == "admin" or not locked:
                cb = make_combo(frame, QC_OPTIONS, width=16)
                cb.set(cur_status if cur_status in QC_OPTIONS else "")
                cb.pack(side="right", padx=2, pady=2)
                def _do_submit(_sid=sid, _cs=cur_status, _cb=cb, _lk=locked):
                    submit_status(_sid, _cs, _cb, _lk)
                btn = styled_btn(frame, "✔ ثبت",
                                  _do_submit,
                                  width=46, height=22)
                btn.pack(side="right", padx=2)
            else:
                tk.Label(frame, text="🔒 قفل شده", bg=C["card2"],
                         fg=C["text_dim"], font=FONT_SMALL).pack(expand=True)
            return frame

        def sync_action_widgets():
            """کشوی «تغییر وضعیت» را روی ستون action قرار بده.
            مهم: فقط مینیمایز را رد کن — ادمین پنجره‌ای/_ui_heavy_ok=False هم باید کشو ببیند.
            (باگ قبلی: return روی not _ui_heavy_ok → کشو برای ادمین حذف می‌شد)
            """
            if not tab.winfo_exists():
                return
            try:
                minimized = (
                    _hwnd_is_iconic(self)
                    or getattr(self, "_minimizing_authorized", False)
                    or self.state() == "iconic"
                )
            except Exception:
                minimized = False
            delay = 12000 if minimized else 700
            try:
                if minimized:
                    tab.after(delay, sync_action_widgets)
                    return
                try:
                    if not tab.winfo_viewable():
                        tab.after(delay, sync_action_widgets)
                        return
                except Exception:
                    pass
                # فقط ردیف‌های واقعاً روی صفحه
                tree_frame.update_idletasks()
                tx = tree.winfo_x()
                ty = tree.winfo_y()
                for iid, frame in list(action_widgets.items()):
                    try:
                        bbox = tree.bbox(iid, "action")
                    except tk.TclError:
                        bbox = None
                    if bbox:
                        x, y, w, h = bbox
                        frame.place(x=tx + x, y=ty + y, width=max(w, 160), height=h)
                        try:
                            frame.lift()
                        except Exception:
                            pass
                    else:
                        try:
                            frame.place_forget()
                        except Exception:
                            pass
            except Exception:
                pass
            tab.after(delay, sync_action_widgets)

        def refresh_tree(focus_sid=None):
            """بازسازی کامل جدول — فقط وقتی لازم است (ثبت اسلب جدید)."""
            _tree_begin_rebuild(tree)
            saved_y = getattr(tree, "_stf_locked_yview", None)
            clear_action_widgets()
            _checked.clear()
            tree.delete(*tree.get_children())
            db = load_db()
            # فقط جدیدترین‌ها در UI — دادهٔ کامل در DB/بک‌آپ می‌ماند
            melts_view = list(reversed(db.get("melts") or []))[:TREE_ROW_LIMIT]
            for rec in melts_view:
                status = rec.get("qc_status", "ثبت شده")
                tag = "qc" if status == "کنترل کیفی شده" else (
                       "rej" if status == "عدم تایید کنترل کیفی" else "")
                iid = tree.insert("", "end", values=(
                    "○",
                    rec["slab_id"],
                    status,
                    (get_display_name(rec.get("registered_by", "—"), db) if self.role == "admin" else "شخص دیگر"),
                    *split_dt(rec.get("registered_at", "—")),
                    rec.get("note", "—"), ""
                ), tags=(tag,))
                _checked[iid] = False
                action_widgets[iid] = build_action_widget(iid, rec["slab_id"], status)
            tree.tag_configure("qc",  background="#3a5040", foreground=C["success"])
            tree.tag_configure("rej", background="#5a3838", foreground=C["danger"])
            # اسکرول را همان‌جا نگه دار — بدون انتخاب/see
            if saved_y is not None:
                tree._stf_locked_yview = saved_y
            try:
                tree.after_idle(lambda: _tree_end_rebuild(tree))
            except Exception:
                _tree_end_rebuild(tree)

        def _find_iid_by_sid(sid):
            sid = str(sid)
            for iid in tree.get_children(""):
                try:
                    vals = tree.item(iid, "values")
                    if vals and len(vals) > 1 and str(vals[1]) == sid:
                        return iid
                except Exception:
                    pass
            return None

        def _pin_scroll():
            """اسکرول را روی مقدار قفل‌شده نگه دار."""
            yv = getattr(tree, "_stf_locked_yview", None)
            if yv is None:
                return
            try:
                tree.yview_moveto(float(yv))
            except Exception:
                pass

        def _refresh_action_frame(iid, sid, status):
            """محتوای ویجت اکشن را عوض کن — خودِ فریم destroy نشود (عامل پرش فوکوس)."""
            frame = action_widgets.get(iid)
            if frame is None:
                action_widgets[iid] = build_action_widget(iid, sid, status)
                return
            for w in list(frame.winfo_children()):
                try:
                    w.destroy()
                except Exception:
                    pass
            locked = status != "ثبت شده"
            if self.role == "admin" or not locked:
                cb = make_combo(frame, QC_OPTIONS, width=16)
                cb.set(status if status in QC_OPTIONS else "")
                cb.pack(side="right", padx=2, pady=2)
                def _do_submit(_sid=sid, _cs=status, _cb=cb, _lk=locked):
                    submit_status(_sid, _cs, _cb, _lk)
                btn = styled_btn(frame, "✔ ثبت", _do_submit, width=46, height=22)
                btn.pack(side="right", padx=2)
            else:
                tk.Label(frame, text="🔒 قفل شده", bg=C["card2"],
                         fg=C["text_dim"], font=FONT_SMALL).pack(expand=True)

        def _apply_row_inplace(sid, rec, db):
            """فقط همان ردیف — بدون رفرش جدول و بدون destroy فریم اکشن."""
            iid = _find_iid_by_sid(sid)
            if not iid:
                return False
            yv = getattr(tree, "_stf_locked_yview", None)
            if yv is None:
                try:
                    yv = float(tree.yview()[0])
                    tree._stf_locked_yview = yv
                except Exception:
                    yv = None
            status = rec.get("qc_status", "ثبت شده")
            tag = "qc" if status == "کنترل کیفی شده" else (
                   "rej" if status == "عدم تایید کنترل کیفی" else "")
            chk = "☑" if _checked.get(iid) else "○"
            try:
                tree.item(iid, values=(
                    chk,
                    rec["slab_id"],
                    status,
                    (get_display_name(rec.get("registered_by", "—"), db) if self.role == "admin" else "شخص دیگر"),
                    *split_dt(rec.get("registered_at", "—")),
                    rec.get("note", "—"), ""
                ), tags=(tag,))
            except Exception:
                return False
            _refresh_action_frame(iid, rec["slab_id"], status)
            _pin_scroll()
            return True

        def submit_status(sid, cur_status, cb, locked_for_others):
            new_status = (cb.get() or "").strip()
            if not new_status or new_status not in QC_OPTIONS:
                messagebox.showerror("خطا",
                    "لطفا یکی از موارد را انتخاب کنید",
                    parent=self)
                return
            if new_status == cur_status:
                messagebox.showinfo("اطلاع", "وضعیتی تغییر نکرده است.", parent=self)
                return
            # اسکرول را قبل از دیالوگ قفل کن — askyesno فوکوس را می‌دزدد و جدول می‌پرد
            try:
                tree._stf_locked_yview = float(tree.yview()[0])
            except Exception:
                tree._stf_locked_yview = getattr(tree, "_stf_user_yview", None)
            tree._stf_scroll_freeze = True
            _pin_scroll()
            ok = messagebox.askyesno(
                "تایید ثبت",
                "آیا از ثبت مطمئنید؟\nدر صورت تایید دیگر قادر به ویرایش نخواهید بود.",
                parent=self,
            )
            _pin_scroll()
            if not ok:
                tree._stf_scroll_freeze = False
                _pin_scroll()
                return
            db = load_db()
            rec_hit = None
            _ts = now_str()
            for rec in db["melts"]:
                if rec["slab_id"] == sid:
                    rec["qc_status"] = new_status
                    rec["qc_by"] = self.username
                    rec["qc_at"] = _ts
                    rec["updated_at"] = _ts
                    if new_status == "ثبت شده":
                        rec["last_edit_at"] = _ts
                        rec["last_edit_by"] = self.username
                    rec_hit = rec
                    break
            save_db(db)
            if rec_hit is not None:
                _apply_row_inplace(sid, rec_hit, db)
            _pin_scroll()
            tree._stf_scroll_freeze = False
            _pin_scroll()

        def _on_sel_sync(*_):
            # همگام‌سازی Ctrl+کلیک با چک‌باکس‌ها
            sel = set(tree.selection())
            for iid in tree.get_children():
                on = iid in sel
                _checked[iid] = on
                try:
                    tree.set(iid, "chk", "☑" if on else "○")
                except Exception:
                    pass
            _update_sel_count()

        refresh_tree()
        sync_action_widgets()
        tab.refresh_melts = refresh_tree
        self._melts_inplace_update = _apply_row_inplace
        tree.bind("<<TreeviewSelect>>", _on_sel_sync, add="+")
        if self.role == "admin":
            self._bind_admin_popup(tree, "melts", refresh_tree, id_col_idx=1)
            tk.Label(tab, text="👑 دابل‌کلیک روی هر ردیف برای ویرایش/حذف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)

    # ═══════════════════════
    #  تب ۲: کنترل کیفی‌شده‌ها
    # ═══════════════════════
    def _build_qc_tab_OLD(self, tab):
        pass

    def _build_qc_tab(self, tab):
        tab.configure(bg=C["panel"])
        self._build_qc_manual_tab(tab)

    def _build_qc_manual_tab(self, tab):
        tab.configure(bg=C["panel"])

        # ── هدر ──
        hdr = card_frame(tab)
        hdr.pack(fill="x", padx=16, pady=(10,4))
        tk.Frame(hdr, bg=C["accent"], height=3).pack(fill="x")
        hdr_in = tk.Frame(hdr, bg=C["card"])
        hdr_in.pack(padx=16, pady=10, fill="x")
        tk.Label(hdr_in, text="📋  اسلب‌های کنترل کیفی شده",
                 bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e")
        tk.Label(hdr_in,
                 text="روی آیکون «🚚 انتقال» در هر ردیف کلیک کنید",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(2,0))

        status_lbl = tk.Label(tab, text="", bg=C["panel"],
                               fg=C["success"], font=FONT_SMALL)
        status_lbl.pack(anchor="e", padx=16, pady=2)

        # ── جدول اصلی — ستون‌ها با همان نام‌هایی که build_rows استفاده می‌کند ──
        cols = ("slab_id","qc_by","qc_date","qc_time","cur_loc","reason","act")
        heads = ("شماره اسلب","تأییدکننده QC","تاریخ QC","ساعت QC","مکان فعلی","دلیل انتقال به داخلی","🚚 انتقال")
        tf, tree = scrolled_tree(tab, cols, heads, height=18)
        tf.pack(fill="both", expand=True, padx=16, pady=4)
        tree.column("slab_id",  width=140, anchor="center")
        tree.column("qc_by",    width=120, anchor="center")
        tree.column("qc_date",  width=120, anchor="center")
        tree.column("qc_time",  width=90,  anchor="center")
        tree.column("cur_loc",  width=140, anchor="center")
        tree.column("reason",   width=170, anchor="center")
        tree.column("act",      width=110, anchor="center")

        search_bar(tab, tree, col_indices=[0]).pack(anchor="e", padx=16, pady=2)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        def _fmt(t):
            if not t: return "—"
            parts = t.replace("  "," ").split(" ")
            if len(parts) < 2: return t
            dp, tp_ = parts[0], parts[1]
            tp2 = tp_.split(":")
            if len(tp2) >= 2:
                tp_ = f"{tp2[0].zfill(2)}:{tp2[1].zfill(2)}:{(tp2[2].zfill(2) if len(tp2)>2 else '00')}"
            return f"{dp}  {tp_}"

        def _ret_label(n):
            """یک‌بار برگشت داده شده / دوبار / سه‌بار / ..."""
            words = ["یک‌بار","دوبار","سه‌بار","چهاربار","پنج‌بار",
                     "شش‌بار","هفت‌بار","هشت‌بار","نه‌بار","ده‌بار"]
            if n == 0: return "—"
            w = words[n-1] if n <= len(words) else f"{n} بار"
            return f"↩ {w} برگشت داده شده"

        # ════════ ثبت انتقال ════════
        def do_transfer(sid, cur, new_dest, popup_win, reason="", reason_detail=""):
            if not new_dest or new_dest == "—":
                messagebox.showwarning("خطا", "مقصد انتخاب نشده", parent=popup_win); return
            if cur == new_dest:
                messagebox.showwarning("خطا", f"اسلب هم‌اکنون در {new_dest} است", parent=popup_win); return
            # انتقال به داخلی → دلیل اجباری
            if new_dest == "انبار داخلی" and not reason:
                messagebox.showwarning("خطا", "لطفاً دلیل انتقال به انبار داخلی را انتخاب کنید.", parent=popup_win)
                return
            if new_dest == "انبار داخلی" and reason == "موارد دیگر" and not reason_detail.strip():
                messagebox.showwarning("خطا", "لطفاً توضیحات «موارد دیگر» را بنویسید.", parent=popup_win)
                return
            # تعیین دلیل نهایی
            final_reason = ""
            if new_dest == "انبار داخلی":
                final_reason = reason_detail.strip() if reason == "موارد دیگر" else reason

            if not messagebox.askyesno("تأیید انتقال",
                    f"اسلب: {sid}\nاز: {cur}\nبه: {new_dest}"
                    + (f"\nدلیل: {final_reason}" if final_reason else "")
                    + "\n\nآیا مطمئن هستید؟",
                    parent=popup_win):
                return
            db = load_db(); ts = now_str()
            prev = [m for m in db.get("movement_log",[]) if m.get("slab_id")==sid]
            db.setdefault("movement_log",[]).append({
                "slab_id":     sid,
                "move_number": len(prev)+1,
                "operation":   "انتقال",
                "from":        cur,
                "to":          new_dest,
                "reason":      final_reason,
                "by":          self.username,
                "at":          ts,
            })
            tr = next((r for r in db.get("transfers_out",[]) if r["slab_id"]==sid), None)
            if tr:
                tr["current_location"] = new_dest
                tr["destination"]      = new_dest
                tr["location"]         = new_dest
                tr["to"]               = new_dest
                tr["transferred_by"]   = self.username
                tr["transferred_at"]   = ts
                tr["updated_at"]       = ts
                tr["at"]               = ts
            else:
                db.setdefault("transfers_out",[]).append({
                    "slab_id":          sid,
                    "destination":      new_dest,
                    "current_location": new_dest,
                    "location":         new_dest,
                    "to":               new_dest,
                    "transferred_by":   self.username,
                    "transferred_at":   ts,
                    "updated_at":       ts,
                    "at":               ts,
                    "source":           "تب QC",
                })
            # همگام‌سازی محل در رکورد ذوب — تا فیلتر موجودی/رد شده درست بماند
            melt = next((r for r in db.get("melts",[]) if r.get("slab_id")==sid), None)
            if melt is not None:
                melt["location"] = new_dest
                melt["rej_location"] = new_dest
                melt["updated_at"] = ts
            save_db(db)
            reason_txt = f"  |  دلیل: {final_reason}" if final_reason else ""
            status_lbl.config(
                text=f"✔  انتقال اسلب {sid}  از {cur}  به {new_dest}{reason_txt}  |  {_fmt(ts)}",
                fg=C["success"])
            popup_win.destroy()
            build_rows()

        def open_transfer_popup(sid, cur):
            """
            پنجره انتقال:
            - داخلی → روباز: بدون دلیل
            - روباز → داخلی: دلیل اجباری (کشویی + توضیح اختیاری/اجباری)
            """
            edit_key = f"transfer:{sid}"
            if _acquire_edit_popup(self, edit_key):
                return

            opts = [w for w in WAREHOUSE_LOCATIONS if w != cur]
            if not opts:
                _release_edit_popup_claim(self, edit_key)
                messagebox.showinfo("انتقال", "مقصد دیگری موجود نیست.", parent=self)
                return

            pop = tk.Toplevel(self)
            prepare_popup_window(pop, self)
            _register_edit_popup(self, edit_key, pop)
            pop.title(f"انتقال اسلب  {sid}")
            pop.configure(bg=C["card"])
            pop.resizable(False, False)
            pop.focus_force()

            # اگه روباز است → ممکن است مقصد داخلی باشد → ارتفاع بیشتر
            is_from_outside = cur != "انبار داخلی"
            pw = 460
            ph = 360 if is_from_outside else 260
            self.update_idletasks()
            sx = self.winfo_screenwidth(); sy = self.winfo_screenheight()
            pop.geometry(f"{pw}x{ph}+{(sx-pw)//2}+{(sy-ph)//2}")

            # ── هدر ──
            tk.Frame(pop, bg=C["accent"], height=3).pack(fill="x")
            hf = tk.Frame(pop, bg=C["header_bg"])
            hf.pack(fill="x")
            tk.Label(hf, text="🚚  انتقال اسلب", bg=C["header_bg"],
                     fg=C["accent"], font=FONT_HEAD).pack(side="right", padx=16, pady=10)
            tk.Label(hf, text=f"شماره: {sid}", bg=C["header_bg"],
                     fg=C["text_dim"], font=FONT_SMALL).pack(side="left", padx=16)

            body = tk.Frame(pop, bg=C["card"])
            body.pack(fill="both", expand=True, padx=20, pady=12)

            # مکان فعلی
            r0 = tk.Frame(body, bg=C["card"]); r0.pack(fill="x", pady=4)
            tk.Label(r0, text="مکان فعلی:", bg=C["card"], fg=C["text_dim"],
                     font=FONT_NORM, width=14, anchor="e").pack(side="right")
            tk.Label(r0, text=cur, bg=C["card"], fg=C["warning"],
                     font=(_MAIN_FONT,12,"bold")).pack(side="right", padx=10)

            tk.Frame(body, bg=C["border"], height=1).pack(fill="x", pady=6)

            # مقصد
            r1 = tk.Frame(body, bg=C["card"]); r1.pack(fill="x", pady=4)
            tk.Label(r1, text="انتقال به:", bg=C["card"], fg=C["text"],
                     font=FONT_NORM, width=14, anchor="e").pack(side="right")
            dest_cb = make_combo(r1, opts, width=20)
            dest_cb.set(opts[0])
            dest_cb.pack(side="right", padx=10)

            # ── بخش دلیل (فقط اگه از روباز است) ──
            reason_frame = tk.Frame(body, bg=C["card"])
            reason_var    = tk.StringVar(value="")
            detail_var    = tk.StringVar()
            detail_frame  = tk.Frame(body, bg=C["card"])

            REASON_OPTS = ["مسائل مربوط به کنترل کیفی", "موارد دیگر"]

            def on_dest_change(e=None):
                dest = dest_cb.get()
                if dest == "انبار داخلی":
                    reason_frame.pack(fill="x", pady=4)
                    if reason_var.get() == "موارد دیگر":
                        detail_frame.pack(fill="x", pady=4)
                    # ارتفاع بیشتر
                    pop.geometry(f"{pw}x{ph}+{(sx-pw)//2}+{(sy-ph)//2}")
                else:
                    reason_frame.pack_forget()
                    detail_frame.pack_forget()
                    pop.geometry(f"{pw}x260+{(sx-pw)//2}+{(sy-ph)//2}")

            def on_reason_change(e=None):
                if reason_var.get() == "موارد دیگر":
                    detail_frame.pack(fill="x", pady=4)
                else:
                    detail_frame.pack_forget()
                    detail_var.set("")

            dest_cb.bind("<<ComboboxSelected>>", on_dest_change)

            # فریم دلیل
            tk.Label(reason_frame, text="دلیل انتقال به داخلی:", bg=C["card"],
                     fg=C["danger"], font=FONT_NORM, width=14, anchor="e").pack(side="right")
            reason_cb = make_combo(reason_frame, REASON_OPTS, width=24)
            reason_cb.set(REASON_OPTS[0])
            reason_var.set(REASON_OPTS[0])
            reason_cb.pack(side="right", padx=10)
            reason_cb.bind("<<ComboboxSelected>>",
                            lambda e: (reason_var.set(reason_cb.get()), on_reason_change()))

            # فریم توضیح اضافی
            tk.Label(detail_frame, text="توضیحات:", bg=C["card"],
                     fg=C["danger"], font=FONT_NORM, width=14, anchor="e").pack(side="right")
            detail_ent = tk.Entry(detail_frame, textvariable=detail_var,
                                   bg=C["entry_bg"], fg=C["text"],
                                   insertbackground=C["accent"],
                                   font=FONT_NORM, bd=0, relief="flat",
                                   highlightthickness=1,
                                   highlightbackground=C["danger"],
                                   highlightcolor=C["danger"], width=24)
            detail_ent.pack(side="right", padx=10, ipady=4)

            # نمایش اولیه بر اساس مقصد پیش‌فرض
            on_dest_change()

            # ── دکمه‌ها ──
            tk.Frame(body, bg=C["border"], height=1).pack(fill="x", pady=8)
            btn_row = tk.Frame(body, bg=C["card"])
            btn_row.pack(fill="x")
            tk.Button(btn_row, text="انصراف", command=pop.destroy,
                      bg=C["card2"], fg=C["text"], font=(_MAIN_FONT,11, "bold"),
                      bd=0, relief="flat", cursor="hand2", padx=16, pady=8
                      ).pack(side="left")

            def on_submit():
                do_transfer(
                    sid, cur,
                    dest_cb.get(),
                    pop,
                    reason=reason_cb.get() if dest_cb.get()=="انبار داخلی" else "",
                    reason_detail=detail_var.get()
                )

            styled_btn(btn_row, "✔  ثبت انتقال", on_submit,
                       color=C["btn_success"]).pack(side="right")
            pop.bind("<Escape>", lambda e: pop.destroy())

        def on_tree_click(event):
            row_id = tree.identify_row(event.y)
            if not row_id: return
            col_id = tree.identify_column(event.x)
            vals = tree.item(row_id, "values")
            if not vals: return
            sid = str(vals[0]).split(" ")[0].strip()
            db = load_db()
            cur = get_current_location(db, sid)
            # ستون آخر displaycolumns (که اولین ستون چپ در RTL است) = 🚚
            # شناسه ستون col_id = "#N" — اگه text آن 🚚 بود popup باز کن
            try:
                col_name = tree["displaycolumns"][int(col_id.replace("#",""))-1]
                if col_name == "act":
                    open_transfer_popup(sid, cur)
            except (IndexError, ValueError):
                pass

        tree.bind("<Button-1>", on_tree_click)

        def build_rows():
            db = load_db()
            qc_list = [r for r in db.get("melts",[])
                       if r.get("qc_status")=="کنترل کیفی شده"]

            # ── جلوگیری از بازسازی بی‌دلیل جدول (که باعث پرش/فلیکر می‌شد) ──
            # اگر داده‌های مربوط به این جدول از آخرین بار تغییر نکرده باشد،
            # کل جدول دوباره ساخته نمی‌شود؛ فقط وقتی واقعاً چیزی تغییر کرده
            # (ثبت QC جدید، انتقال جدید و ...) بازسازی انجام می‌شود.
            _slab_ids_now = {r.get("slab_id") for r in qc_list}
            _relevant_moves = [m for m in db.get("movement_log", [])
                                if m.get("slab_id") in _slab_ids_now]
            try:
                _sig = hashlib.sha1(
                    json.dumps([qc_list, _relevant_moves], sort_keys=True, default=str)
                    .encode("utf-8")
                ).hexdigest()
            except Exception:
                _sig = None
            if _sig is not None and getattr(tree, "_qc_build_sig", None) == _sig:
                return
            tree._qc_build_sig = _sig

            _tree_begin_rebuild(tree)
            tree.delete(*tree.get_children())

            def get_moves(sid):
                """فقط انتقال‌های واقعی بین انبارها — تایید مجدد QC (که یک تغییر وضعیت
                است نه یک جابجایی فیزیکی) از این لیست حذف می‌شود تا در ستون‌های
                «انتقال» نمایش داده نشود."""
                return sorted(
                    [m for m in db.get("movement_log",[])
                     if m.get("slab_id")==sid and m.get("operation","انتقال") == "انتقال"],
                    key=lambda m: m.get("at","")
                )

            def get_re_approve(sid):
                """رکورد تایید مجدد QC (رد شده → کنترل کیفی شده) برای این اسلب، اگر وجود داشته باشد"""
                matches = [m for m in db.get("movement_log",[])
                           if m.get("slab_id")==sid and m.get("operation")=="تأیید مجدد QC"]
                return matches[-1] if matches else None

            max_tr = max((len(get_moves(r["slab_id"])) for r in qc_list), default=0)

            # ── ستون‌های ثابت + داینامیک ──
            fixed_cols  = ("slab_id","qc_by","qc_date","qc_time","cur_loc","reason","re_by","re_date","re_time")
            fixed_heads = ("شماره اسلب","تأییدکننده QC","تاریخ QC","ساعت QC","مکان فعلی",
                            "دلیل انتقال به داخلی","تأییدکننده مجدد پس از رد QC","تاریخ تأیید مجدد","ساعت تأیید مجدد")
            # هر انتقال: مسیر / تاریخ / ساعت / انتقال‌دهنده
            tr_cols  = tuple(c for i in range(1, max_tr+1)
                              for c in (f"tr_{i}", f"tr_{i}_date", f"tr_{i}_time", f"tr_{i}_by"))
            tr_heads = tuple(h for i in range(1, max_tr+1)
                              for h in (f"انتقال {i}", f"تاریخ انتقال {i}", f"ساعت انتقال {i}", f"انتقال‌دهنده {i}"))
            act_col     = ("act",)
            act_head    = ("🚚 انتقال",)

            all_cols  = fixed_cols + tr_cols + act_col
            all_heads = fixed_heads + tr_heads + act_head

            tree.configure(columns=all_cols)
            tree["displaycolumns"] = list(reversed(all_cols))

            for col, head in zip(all_cols, all_heads):
                tree.heading(col, text=head, anchor="center")
                tree.column(col, width=120, anchor="center",
                            minwidth=80, stretch=True)

            # عرض ثابت ستون‌های اصلی
            tree.column("slab_id",  width=140, minwidth=120)
            tree.column("qc_by",    width=130, minwidth=100)
            tree.column("qc_date",  width=110, minwidth=90)
            tree.column("qc_time",  width=85,  minwidth=75)
            tree.column("cur_loc",  width=140, minwidth=120)
            tree.column("reason",   width=170, minwidth=130)
            tree.column("re_by",    width=130, minwidth=100)
            tree.column("re_date",  width=110, minwidth=90)
            tree.column("re_time",  width=85,  minwidth=75)
            tree.column("act",      width=90,  minwidth=70)
            for i in range(1, max_tr+1):
                tree.column(f"tr_{i}",      width=180, anchor="center", minwidth=150)
                tree.column(f"tr_{i}_date", width=110, anchor="center", minwidth=90)
                tree.column(f"tr_{i}_time", width=85,  anchor="center", minwidth=75)
                tree.column(f"tr_{i}_by",   width=120, anchor="center", minwidth=100)

            if not qc_list:
                tree.insert("","end", values=("— هیچ اسلبی کنترل کیفی نشده —",) + ("",)*(len(all_cols)-1))
                _tree_end_rebuild(tree)
                return

            # فقط جدیدترین‌ها در UI — دادهٔ کامل در DB می‌ماند
            qc_list = list(reversed(qc_list))[:TREE_ROW_LIMIT]

            for rec in qc_list:
                sid   = rec["slab_id"]
                cur   = get_current_location(db, sid)
                who   = get_display_name(rec.get("qc_by","—"), db) if self.role=="admin" else "شخص دیگر"
                is_in = (cur == "انبار داخلی")
                tag   = "inside" if is_in else "outside"
                moves = get_moves(sid)

                tr_vals = []
                last_reason = "—"
                for i, m in enumerate(moves, 1):
                    frm    = m.get("from","—")
                    to_    = m.get("to","—")
                    _td, _tt = split_dt(_fmt(m.get("at","")))
                    by_    = get_display_name(m.get("by","—"), db) if self.role=="admin" else "شخص دیگر"
                    reason = m.get("reason","")
                    if reason:
                        last_reason = reason
                    tr_vals.extend([f"از {frm} به {to_}", _td, _tt, by_])
                while len(tr_vals) < max_tr*4:
                    tr_vals.append("—")

                # تایید مجدد (رد شده → کنترل کیفی شده)
                re_by_txt = "—"
                re_date_txt = "—"
                re_time_txt = "—"
                ra = get_re_approve(sid)
                if ra:
                    re_by_txt   = get_display_name(ra.get("by","—"), db) if self.role=="admin" else "شخص دیگر"
                    re_date_txt, re_time_txt = split_dt(_fmt(ra.get("at","")))

                _qd, _qt = split_dt(_fmt(rec.get("qc_at","")))
                row_vals = (sid, who, _qd, _qt, cur, last_reason, re_by_txt, re_date_txt, re_time_txt) \
                           + tuple(tr_vals) + ("🚚",)
                tree.insert("","end", values=row_vals, tags=(tag,))

            # رنگ‌بندی با طوسی
            tree.tag_configure("inside",
                background="#4e5560",
                foreground="#f0c060")
            tree.tag_configure("outside",
                background="#404a46",
                foreground="#4caf80")
            _tree_end_rebuild(tree)


        build_rows()

        # ── دابل‌کلیک / راست‌کلیک ادمین برای ویرایش/حذف ──
        if self.role == "admin":
            self._bind_admin_popup(tree, "melts", build_rows)
            tk.Label(tab,
                     text="👑  دابل‌کلیک روی هر سلول برای ویرایش مستقیم  |  Ctrl+Delete برای حذف ردیف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=16, pady=4)
        tab.refresh_qc    = build_rows
        self._qc_refresh_fn = build_rows

        # ── بروزرسانی خودکار — فقط وقتی تب دیده می‌شود و پنجره فعال است ──
        def _qc_auto_refresh_loop():
            try:
                if not tab.winfo_exists():
                    return
                delay = _ui_idle_ms(self, 4000, 10000, 20000)
                if _ui_heavy_ok(self) and getattr(self, "_current_tab", None) == "qc":
                    try:
                        if tab.winfo_viewable():
                            build_rows()
                    except Exception:
                        build_rows()
                tab.after(delay, _qc_auto_refresh_loop)
            except Exception:
                pass
        tab.after(4000, _qc_auto_refresh_loop)

    def _build_rejected_tab(self, tab):
        tab.configure(bg=C["panel"])
        tk.Label(tab, text="⛔  اسلب‌های رد شده از کنترل کیفی",
                 bg=C["panel"], fg=C["danger"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=10)
        tk.Label(tab,
                 text="اسلب‌هایی که QC رد کرده — تأیید مجدد / قراضه / انتقال (مثل کنترل کیفی)",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        cols  = ("slab_id","qc_by","qc_date","qc_time","re_approved","re_by","re_at","cur_loc","note","act")
        heads = ("شماره اسلب","رد‌کننده","تاریخ رد","ساعت رد","وضعیت تأیید مجدد",
                 "تأییدکننده مجدد","تاریخ تأیید مجدد","مکان فعلی","توضیحات","🚚 انتقال")
        tf, tree = scrolled_tree(tab, cols, heads, height=16)
        tf.pack(fill="both", expand=True, padx=16, pady=8)
        tree.column("slab_id",     width=130, anchor="center")
        tree.column("qc_by",       width=110, anchor="center")
        tree.column("qc_date",     width=100, anchor="center")
        tree.column("qc_time",     width=80,  anchor="center")
        tree.column("re_approved", width=140, anchor="center")
        tree.column("re_by",       width=110, anchor="center")
        tree.column("re_at",       width=140, anchor="center")
        tree.column("cur_loc",     width=120, anchor="center")
        tree.column("note",        width=140, anchor="center")
        tree.column("act",         width=100, anchor="center")
        sb = search_bar(tab, tree, col_indices=[0])
        sb.pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        status_lbl = tk.Label(tab, text="", bg=C["panel"], fg=C["success"], font=FONT_SMALL)
        status_lbl.pack(anchor="e", padx=16)

        def _mask_who(uname, db):
            return get_display_name(uname or "—", db) if self.role == "admin" else "شخص دیگر"

        def refresh():
            db = load_db()
            rows = []
            for rec in db["melts"]:
                if rec.get("qc_status") not in ("عدم تایید کنترل کیفی", "کنترل کیفی شده"):
                    continue
                if rec.get("qc_status") == "کنترل کیفی شده" and not rec.get("re_approved"):
                    continue
                rows.append(rec)
            try:
                _sig = hashlib.sha1(
                    json.dumps(
                        [(r.get("slab_id"), r.get("qc_status"), r.get("qc_at"),
                          r.get("re_approved"), r.get("re_approved_at"), r.get("note"))
                         for r in rows],
                        sort_keys=True, default=str,
                    ).encode("utf-8")
                ).hexdigest()
            except Exception:
                _sig = None
            if _sig is not None and getattr(tree, "_rej_build_sig", None) == _sig:
                return
            tree._rej_build_sig = _sig
            _tree_begin_rebuild(tree)
            tree.delete(*tree.get_children())
            # فقط جدیدترین‌ها در UI — دادهٔ کامل در DB می‌ماند
            rows = list(reversed(rows))[:TREE_ROW_LIMIT]
            for rec in rows:
                sid = rec["slab_id"]
                display_who = _mask_who(rec.get("qc_by","—"), db)
                re_by  = _mask_who(rec.get("re_approved_by","—"), db)
                re_at  = rec.get("re_approved_at","—")
                re_txt = "✔ تأیید مجدد شده" if rec.get("re_approved") else "⛔ رد شده"
                tag    = "reapproved" if rec.get("re_approved") else "rej"
                loc    = get_current_location(db, sid)
                tree.insert("","end", values=(
                    sid,
                    display_who,
                    *split_dt(rec.get("qc_at","—")),
                    re_txt,
                    re_by if rec.get("re_approved") else "—",
                    re_at if rec.get("re_approved") else "—",
                    loc,
                    rec.get("note","—") or "—",
                    "🚚 انتقال",
                ), tags=(tag,))
            tree.tag_configure("rej",        background="#5a3838", foreground=C["danger"])
            tree.tag_configure("reapproved", background="#3a5040", foreground=C["success"])
            _tree_end_rebuild(tree)

        def do_re_approve():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک اسلب انتخاب کنید.", parent=self)
                return
            vals = tree.item(sel[0], "values")
            sid  = vals[0]
            db   = load_db()
            melt = next((r for r in db["melts"] if r["slab_id"]==sid), None)
            if not melt:
                messagebox.showerror("خطا", f"اسلب {sid} یافت نشد.", parent=self)
                return
            if melt.get("re_approved"):
                who = _mask_who(melt.get("re_approved_by","—"), db)
                messagebox.showwarning("⚠️  قبلاً تأیید شده",
                    f"اسلب {sid} قبلاً توسط {who} تأیید مجدد شده است.",
                    parent=self)
                return
            first_rej_by  = _mask_who(melt.get("qc_by","—"), db)
            first_rej_at  = melt.get("qc_at","—")
            if not messagebox.askyesno("تأیید مجدد اسلب رد شده",
                f"اسلب: {sid}\n"
                f"رد شده توسط: {first_rej_by}\n"
                f"تاریخ رد: {first_rej_at}\n\n"
                f"آیا مطمئن هستید که این اسلب تأیید شود؟\n"
                f"تاریخچه رد شدن ثبت و نگهداری می‌شود.",
                parent=self):
                return
            ts = now_str()
            db.setdefault("qc_history", []).append({
                "slab_id": sid, "event": "رد شده",
                "by": melt.get("qc_by","—"), "at": melt.get("qc_at","—"),
                "note": "رد اولیه کنترل کیفی",
            })
            db["qc_history"].append({
                "slab_id": sid, "event": "تأیید مجدد",
                "by": self.username, "at": ts,
                "note": "تأیید مجدد پس از رد اولیه",
            })
            melt["qc_status"]       = "کنترل کیفی شده"
            melt["re_approved"]     = True
            melt["re_approved_by"]  = self.username
            melt["re_approved_at"]  = ts
            melt["location"]        = melt.get("location","انبار داخلی")
            db.setdefault("movement_log",[]).append({
                "slab_id": sid, "operation": "تأیید مجدد QC",
                "from": "رد شده", "to": "کنترل کیفی شده",
                "by": self.username, "at": ts,
                "note": f"ابتدا رد شده توسط {melt.get('qc_by','—')} در {melt.get('qc_at','—')}",
            })
            save_db(db)
            status_lbl.config(text=f"✔  اسلب {sid} تأیید مجدد شد  |  {ts}", fg=C["success"])
            refresh()

        def do_register_as_scrap():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک اسلب انتخاب کنید.", parent=self)
                return
            vals = tree.item(sel[0], "values")
            sid  = vals[0]
            db   = load_db()
            melt = next((r for r in db["melts"] if r["slab_id"]==sid), None)
            if not melt:
                messagebox.showerror("خطا", f"اسلب {sid} یافت نشد.", parent=self)
                return
            if melt.get("re_approved"):
                messagebox.showwarning("⚠️  قبلاً تأیید شده",
                    f"اسلب {sid} قبلاً تأیید مجدد شده و قابل ثبت به‌عنوان قراضه نیست.", parent=self)
                return
            if check_duplicate(db, "scrap", sid):
                messagebox.showerror("⛔  ثبت تکراری",
                    f"اسلب {sid} قبلاً به عنوان قراضه ثبت شده است.", parent=self)
                return
            scrap_reason = simpledialog.askstring(
                "ثبت قراضه", f"دلیل ثبت اسلب {sid} به عنوان قراضه را وارد کنید:", parent=self)
            if scrap_reason is None:
                return
            if not scrap_reason.strip():
                messagebox.showwarning("خطا", "دلیل قراضه الزامی است — نمی‌تواند خالی باشد.", parent=self)
                return
            if not messagebox.askyesno("⚠️  تأیید ثبت قراضه",
                f"اسلب: {sid}\nاین اسلب به عنوان «قراضه» ثبت می‌شود.\nآیا مطمئن هستید؟", parent=self):
                return
            ts = now_str()
            melt["qc_status"] = "قراضه"
            melt["updated_at"] = ts
            db.setdefault("scrap", []).append({
                "slab_id": sid, "reason": scrap_reason.strip(),
                "registered_by": self.username, "registered_at": ts,
                "updated_at": ts,
            })
            db.setdefault("qc_history", []).append({
                "slab_id": sid, "event": "ثبت قراضه از بخش رد شده‌ها",
                "by": self.username, "at": ts,
                "note": f"رد شده توسط {melt.get('qc_by','—')} در {melt.get('qc_at','—')}",
            })
            save_db(db)
            status_lbl.config(text=f"♻️  اسلب {sid} به عنوان قراضه ثبت شد  |  {ts}", fg=C["warning"])
            refresh()

        def on_tree_click(event):
            row_id = tree.identify_row(event.y)
            if not row_id:
                return
            col_id = tree.identify_column(event.x)
            vals = tree.item(row_id, "values")
            if not vals:
                return
            sid = str(vals[0]).split(" ")[0].strip()
            try:
                col_name = tree["displaycolumns"][int(col_id.replace("#", "")) - 1]
                if col_name == "act":
                    db = load_db()
                    cur = get_current_location(db, sid)
                    open_slab_transfer_popup(
                        self, sid, cur,
                        status_lbl=status_lbl, on_done=refresh,
                        source="تب رد شده‌ها")
            except (IndexError, ValueError):
                pass

        tree.bind("<Button-1>", on_tree_click)

        act_card = card_frame(tab)
        act_card.pack(fill="x", padx=16, pady=6)
        tk.Frame(act_card, bg=C["success"], height=2).pack(fill="x")
        act_in = tk.Frame(act_card, bg=C["card"])
        act_in.pack(padx=16, pady=10, fill="x")
        tk.Label(act_in,
                 text="⚠️  تأیید مجدد: تاریخچه رد شدن ثبت می‌شود — برای جابجایی روی «🚚 انتقال» کلیک کنید",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(0,8))
        styled_btn(act_in, "✅  تأیید مجدد اسلب انتخابی",
                   do_re_approve, color=C["btn_success"]).pack(side="right", padx=4)
        styled_btn(act_in, "♻️  ثبت به عنوان قراضه",
                   do_register_as_scrap, color=C["warning"]).pack(side="right", padx=4)

        refresh()
        tab.refresh_rejected = refresh
        self._rejected_refresh_fn = refresh
        self._bind_admin_popup(tree, "melts", refresh)
        if self.role == "admin":
            tk.Label(tab, text="👑 دابل‌کلیک یا راست‌کلیک برای ویرایش/حذف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)

        def _auto_refresh_loop():
            try:
                if not tab.winfo_exists():
                    return
                delay = _ui_idle_ms(self, 5000, 12000, 25000)
                if _ui_heavy_ok(self) and getattr(self, "_current_tab", None) == "rejected":
                    try:
                        if tab.winfo_viewable():
                            refresh()
                    except Exception:
                        refresh()
                tab.after(delay, _auto_refresh_loop)
            except Exception:
                pass
        tab.after(5000, _auto_refresh_loop)

    # ═══════════════════════
    #  تب ۳: انتقال به انبار روباز
    # ═══════════════════════
    def _build_transfer_tab(self, tab):
        """موجودی انبار — سه زیرتب"""
        tab.configure(bg=C["panel"])
        style = ttk.Style()
        style.configure("Dark.TNotebook", background=C["panel"], bordercolor=C["border"])
        style.configure("Dark.TNotebook.Tab",
            background=C["tab_inactive"], foreground=C["text"],
            font=(_MAIN_FONT, 10, "bold"), padding=[12, 6])
        style.map("Dark.TNotebook.Tab",
            background=[("selected", C["accent"])],
            foreground=[("selected", "#ffffff")])
        sub = ttk.Notebook(tab, style="Dark.TNotebook")
        sub.pack(fill="both", expand=True, padx=4, pady=4)
        wh_frames = []
        for wh_name, wh_icon, wh_color in [
            ("انبار داخلی",   "🏠", C["accent2"]),
            ("انبار روباز ۱", "🏭", "#3a8060"),
            ("انبار روباز ۲", "🏭", "#3a6080"),
        ]:
            wt = tk.Frame(sub, bg=C["panel"])
            sub.add(wt, text=f"{wh_icon}  {wh_name}")
            self._build_warehouse_subtab(wt, wh_name, wh_color)
            wh_frames.append(wt)

        def refresh_all_warehouses():
            for wt in wh_frames:
                fn = getattr(wt, "refresh_warehouse", None)
                if callable(fn):
                    try:
                        fn()
                    except Exception:
                        pass

        tab.refresh_warehouse = refresh_all_warehouses
        tab.refresh_all = refresh_all_warehouses

    def _build_warehouse_subtab(self, tab, warehouse_name, accent_color):
        tab.configure(bg=C["panel"])
        hdr = card_frame(tab)
        hdr.pack(fill="x", padx=12, pady=(10,4))
        tk.Frame(hdr, bg=accent_color, height=3).pack(fill="x")
        hdr_in = tk.Frame(hdr, bg=C["card"])
        hdr_in.pack(padx=14, pady=10, fill="x")
        tk.Label(hdr_in, text=f"موجودی {warehouse_name}",
                 bg=C["card"], fg=accent_color, font=FONT_HEAD).pack(anchor="e")
        tk.Label(hdr_in,
                 text="روی «تاریخچه» کلیک کنید برای مشاهده سابقه کامل جابجایی‌های هر اسلب",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(2,0))

        cols = ("slab_id","qc_by","qc_date","qc_time","ret_count","last_reason","history_summary","cur_status")
        heads = ("شماره اسلب","تأییدکننده QC","تاریخ QC","ساعت QC","برگشت","دلیل انتقال","خلاصه تاریخچه","وضعیت")
        tf, tree = scrolled_tree(tab, cols, heads, height=16)
        tf.pack(fill="both", expand=True, padx=12, pady=6)
        tree.column("slab_id",        width=140, anchor="center")
        tree.column("qc_by",          width=120, anchor="center")
        tree.column("qc_date",        width=110, anchor="center")
        tree.column("qc_time",        width=85,  anchor="center")
        tree.column("ret_count",      width=70,  anchor="center")
        tree.column("last_reason",    width=160, anchor="center")
        tree.column("history_summary",width=240, anchor="center")
        tree.column("cur_status",     width=190, anchor="center")
        cnt_lbl = tk.Label(tab,"",bg=C["panel"],fg=C["text_dim"],font=FONT_SMALL)
        cnt_lbl.pack(anchor="e",padx=12)
        sb = search_bar(tab,tree,col_indices=[0]); sb.pack(anchor="e",padx=12,pady=2)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=12, pady=2)

        def build_hist_summary(db, sid):
            """تاریخچه انتقال‌ها به صورت متنی واضح — انتقال ۱، انتقال ۲، ...
            فقط مسیر (از کجا به کجا) نمایش داده می‌شود — تاریخ/ساعت و دلیل
            هرکدام ستون مجزای خودشان را در همین جدول دارند، پس اینجا تکرار نمی‌شوند.
            تایید مجدد QC (رد شده → کنترل کیفی شده) یک انتقال فیزیکی نیست و اینجا نمی‌آید."""
            moves = sorted(
                [m for m in db.get("movement_log",[])
                 if m.get("slab_id")==sid and m.get("operation","انتقال") == "انتقال"],
                key=lambda m: m.get("at","")
            )
            if not moves:
                return "بدون جابجایی"
            parts = []
            for i, m in enumerate(moves, 1):
                frm = m.get("from","—")
                to  = m.get("to","—")
                parts.append(f"انتقال {i}: از {frm} به {to}")
            return "  ●  ".join(parts)

        def last_reason_for(db, sid):
            """آخرین دلیل انتقال ثبت‌شده برای این اسلب (ستون مجزا)"""
            moves = sorted(
                [m for m in db.get("movement_log",[])
                 if m.get("slab_id")==sid and m.get("reason")
                 and m.get("operation","انتقال") == "انتقال"],
                key=lambda m: m.get("at","")
            )
            return moves[-1]["reason"] if moves else "—"

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            n = 0
            for rec in db.get("melts",[]):
                if rec.get("qc_status") != "کنترل کیفی شده": continue
                sid = rec["slab_id"]
                cur = get_current_location(db, sid)
                if cur != warehouse_name: continue
                ret_n = len([r for r in db.get("return_log",[]) if r.get("slab_id")==sid])
                who = get_display_name(rec.get("qc_by","—"), db) if self.role=="admin" else "شخص دیگر"
                hist = build_hist_summary(db, sid)
                reason_disp = last_reason_for(db, sid)
                exit_s = rec.get("exit_status","")
                # وضعیت: نام انبار فعلی (همان منبع تب) — نه فقط «موجود»
                status = "خارج شد" if exit_s=="خروج زده شده" else f"موجود در {cur}"
                tag = "exited" if exit_s=="خروج زده شده" else ("returned" if ret_n>0 else "normal")
                tree.insert("","end", values=(
                    sid, who, *split_dt(rec.get("qc_at","—")),
                    str(ret_n)+" بار" if ret_n else "—",
                    reason_disp, hist, status
                ), tags=(tag,))
                n += 1
            tree.tag_configure("normal",background=C["card2"],foreground=C["text"])
            tree.tag_configure("returned",background="#4a4830",foreground="#e0aa40")
            tree.tag_configure("exited",background="#3a5040",foreground="#4caf80")
            cnt_lbl.config(text=f"موجودی: {n} اسلب در {warehouse_name}")

        def show_detail():
            sel = tree.selection()
            if not sel: return
            sid = tree.item(sel[0],"values")[0]
            db = load_db()
            win = tk.Toplevel(self)
            prepare_popup_window(win, self)
            win.title(f"تاریخچه کامل — {sid}")
            win.configure(bg=C["card"])
            win.geometry("820x460")
            self._center(win, 820, 460)

            # هدر
            tk.Frame(win, bg=accent_color, height=3).pack(fill="x")
            tk.Label(win, text=f"📋  تاریخچه کامل اسلب  {sid}",
                     bg=C["card"], fg=accent_color, font=FONT_HEAD).pack(pady=(12,4), padx=14, anchor="e")

            # جدول
            cols_d  = ("i","event","frm","to","by","at")
            heads_d = ("ردیف","رویداد","از","به","توسط","تاریخ و ساعت")
            tf_d, tr_d = scrolled_tree(win, cols_d, heads_d, height=12)
            tf_d.pack(fill="both", expand=True, padx=12, pady=6)
            tr_d.column("i",     width=45,  anchor="center")
            tr_d.column("event", width=180, anchor="center")
            tr_d.column("frm",   width=140, anchor="center")
            tr_d.column("to",    width=140, anchor="center")
            tr_d.column("by",    width=120, anchor="center")
            tr_d.column("at",    width=165, anchor="center")

            melt = next((r for r in db["melts"] if r["slab_id"]==sid), {})
            events = []

            # ثبت ذوب
            who_r = get_display_name(melt.get("registered_by","—"), db) if self.role=="admin" else "شخص دیگر"
            events.append((
                "ثبت ذوب",
                "—", "انبار داخلی",
                who_r, melt.get("registered_at","—")
            ))
            # تأیید QC
            if melt.get("qc_at"):
                who_q = get_display_name(melt.get("qc_by","—"), db) if self.role=="admin" else "شخص دیگر"
                events.append((
                    "تأیید کنترل کیفی",
                    "—", "انبار داخلی",
                    who_q, melt.get("qc_at","—")
                ))
            # انتقال‌ها — به ترتیب زمان (فقط جابجایی‌های فیزیکی واقعی شماره‌گذاری می‌شوند)
            moves = sorted(
                [m for m in db.get("movement_log",[]) if m.get("slab_id")==sid],
                key=lambda m: m.get("at","")
            )
            tr_counter = 1
            for m in moves:
                by_mv = m.get("by","—") if self.role=="admin" else "شخص دیگر"
                if m.get("operation") == "تأیید مجدد QC":
                    events.append((
                        "تأیید مجدد پس از رد",
                        m.get("from","—"), m.get("to","—"),
                        by_mv, m.get("at","—")
                    ))
                else:
                    events.append((
                        f"انتقال {tr_counter}",
                        m.get("from","—"), m.get("to","—"),
                        by_mv, m.get("at","—")
                    ))
                    tr_counter += 1
            # اسکارف/برش
            for sc in db.get("scarf_cut",[]):
                if sc.get("slab_id")==sid:
                    by_sc = get_display_name(sc.get("registered_by","—"), db) if self.role=="admin" else "شخص دیگر"
                    events.append((
                        sc.get("operation","اسکارف/برش"),
                        "—", "—",
                        by_sc, sc.get("registered_at","—")
                    ))
            # خروج
            if melt.get("exit_at"):
                by_ex = get_display_name(melt.get("exit_by","—"), db) if self.role=="admin" else "شخص دیگر"
                events.append((
                    "خروج از فولاد",
                    "—", "خارج شده",
                    by_ex, melt.get("exit_at","—")
                ))

            # مرتب بر اساس زمان
            events.sort(key=lambda x: x[4])

            for i, ev in enumerate(events, 1):
                tr_d.insert("","end", values=(i,)+ev)

            tk.Button(win, text="بستن", command=win.destroy,
                      bg=C["btn_primary"], fg="white", font=FONT_NORM,
                      bd=0, relief="flat", cursor="hand2",
                      padx=24, pady=8).pack(pady=10)

        refresh()
        tab.refresh_warehouse = refresh
        ctrl = tk.Frame(tab,bg=C["panel"]); ctrl.pack(fill="x",padx=12,pady=4)
        styled_btn(ctrl,"📋  تاریخچه اسلب انتخابی",show_detail,color=C["accent2"]).pack(side="right",padx=4)
        self._bind_admin_popup(tree, "melts", refresh)
        if self.role == "admin":
            tk.Label(tab, text="👑  دابل‌کلیک یا راست‌کلیک روی هر ردیف برای ویرایش/حذف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=12, pady=2)


    def _build_lab_tab(self, tab):
        tab.configure(bg=C["panel"])
        tk.Label(tab, text="🧪  تحویل اسلب به آزمایشگاه",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=10)
        tk.Label(tab,
                 text="اسلب‌هایی که برش‌کار «تست باومن» را ثبت کرده — شیفت تأیید تحویل می‌کند",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        cols = ("slab_id","cut_by","cut_date","cut_time","lab_status","delivered_by","del_date","del_time")
        heads = ("شماره اسلب","برش‌کار","تاریخ برش","ساعت برش","وضعیت","تأیید‌کننده","تاریخ تحویل","ساعت تحویل")
        tf, tree = scrolled_tree(tab, cols, heads, height=14)
        tf.pack(fill="both", expand=True, padx=16, pady=8)
        search_bar(tab, tree, col_indices=[0]).pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        status_lbl = tk.Label(tab, text="", bg=C["panel"], fg=C["success"], font=FONT_NORM)
        status_lbl.pack(anchor="e", padx=16)

        def do_deliver():
            # فقط شیفت و ادمین مجاز به تأیید هستند
            if self.role not in ("shift", "admin"):
                messagebox.showerror("خطا", "فقط پرسنل شیفت می‌توانند تحویل را تأیید کنند.", parent=self)
                return
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک اسلب انتخاب کنید.", parent=self)
                return
            vals = tree.item(sel[0], "values")
            sid = vals[0]
            cur_status = vals[3]
            if cur_status == "تحویل داده شده":
                messagebox.showerror("⛔  قبلاً ثبت شده",
                    f"اسلب {sid} قبلاً به آزمایشگاه تحویل داده شده است.", parent=self)
                return
            if not messagebox.askyesno("تأیید تحویل",
                f"اسلب: {sid}\n\nتأیید می‌کنید که این اسلب به آزمایشگاه تحویل داده شد؟",
                parent=self):
                return
            db = load_db()
            # ثبت در lab_deliveries
            if not check_duplicate(db, "lab_deliveries", sid):
                db["lab_deliveries"].append({
                    "slab_id": sid,
                    "delivered_by": self.username,
                    "delivered_at": now_str()
                })
            # آپدیت وضعیت در bauman اگه وجود داشت
            for r in db.get("bauman", []):
                if r.get("slab_id") == sid:
                    r["lab_status"] = "تحویل داده شده"
                    r["delivered_by"] = self.username
                    r["delivered_at"] = now_str()
                    break
            save_db(db)
            status_lbl.config(text=f"✔  اسلب {sid} به آزمایشگاه تحویل داده شد.")
            refresh()

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=16, pady=4)
        if self.role in ("shift", "admin"):
            styled_btn(ctrl, "🧪  تأیید تحویل به آزمایشگاه", do_deliver,
                       color=C["btn_success"]).pack(side="right")
        else:
            tk.Label(ctrl, text="👁  فقط مشاهده — تأیید تحویل توسط شیفت انجام می‌شود",
                     bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(side="right")

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            delivered_map = {r["slab_id"]: r for r in db.get("lab_deliveries", [])}

            # ── فقط اسلب‌هایی که برش‌کار تست باومن را ثبت کرده ──
            # منبع اصلی: scarf_cut با operation=برشی و bauman_done=True
            # منبع ثانوی: جدول bauman (هر دو مسیر)
            seen = set()
            candidates = []

            # مسیر ۱: برش با تیک باومن
            for r in db.get("scarf_cut", []):
                if r.get("operation") == "برشی" and r.get("bauman_done"):
                    sid = r["slab_id"]
                    if sid not in seen:
                        seen.add(sid)
                        candidates.append({
                            "slab_id":       sid,
                            "cut_by":        get_display_name(r.get("registered_by","—"), db),
                            "cut_at":        r.get("registered_at","—"),
                        })

            # مسیر ۲: جدول bauman (ثبت مستقیم توسط برش‌کار)
            for bm in db.get("bauman", []):
                sid = bm["slab_id"]
                if sid not in seen:
                    seen.add(sid)
                    candidates.append({
                        "slab_id": sid,
                        "cut_by":  get_display_name(bm.get("cut_by", bm.get("registered_by","—")), db),
                        "cut_at":  bm.get("cut_at", bm.get("registered_at","—")),
                    })

            for rec in candidates:
                sid = rec["slab_id"]
                dl  = delivered_map.get(sid)
                status = "تحویل داده شده" if dl else "در انتظار تحویل"
                clr    = "done" if dl else "wait"
                cut_who = get_display_name(rec.get("cut_by","—"), db) if self.role == "admin" else "شخص دیگر"
                del_who = (get_display_name(dl.get("delivered_by","—"), db) if self.role == "admin" else "شخص دیگر") if dl else "—"
                tree.insert("", "end", values=(
                    sid,
                    cut_who,
                    *split_dt(rec.get("cut_at","—")),
                    status,
                    del_who,
                    *split_dt(dl.get("delivered_at","—") if dl else "—"),
                ), tags=(clr,))
            tree.tag_configure("done", background="#3a5040", foreground=C["success"])
            tree.tag_configure("wait", background="#4a4830", foreground=C["warning"])
        refresh()

        if self.role == "admin":
            tk.Label(tab, text="👑 دابل‌کلیک روی ردیف برای ویرایش",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)
            self._bind_admin_popup(tree, "lab_deliveries", refresh)

    # ═══════════════════════
    #  تب ۵: اسلب قراضه
    # ═══════════════════════
    def _build_scrap_tab(self, tab):
        tab.configure(bg=C["panel"])
        hdr = card_frame(tab)
        hdr.pack(fill="x", padx=16, pady=(10, 4))
        tk.Frame(hdr, bg=C["warning"], height=3).pack(fill="x")
        hdr_in = tk.Frame(hdr, bg=C["card"])
        hdr_in.pack(padx=16, pady=10, fill="x")
        tk.Label(hdr_in, text="♻️  اسلب‌های قراضه",
                 bg=C["card"], fg=C["warning"], font=FONT_HEAD).pack(anchor="e")
        tk.Label(hdr_in,
                 text="ثبت قراضه فقط از بخش «اسلب‌های رد شده» انجام می‌شود — برای جابجایی روی «🚚 انتقال» کلیک کنید",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(2, 0))

        status_lbl = tk.Label(tab, text="", bg=C["panel"], fg=C["warning"], font=FONT_SMALL)
        status_lbl.pack(anchor="e", padx=16, pady=2)

        cols = ("slab_id", "reason", "registered_by", "registered_date", "registered_time", "cur_loc", "act")
        heads = ("شماره اسلب", "دلیل", "ثبت‌کننده", "تاریخ ثبت", "ساعت ثبت", "مکان فعلی", "🚚 انتقال")
        tf, tree = scrolled_tree(tab, cols, heads, height=16)
        tf.pack(fill="both", expand=True, padx=16, pady=4)
        tree.column("slab_id", width=130, anchor="center")
        tree.column("reason", width=200, anchor="center")
        tree.column("registered_by", width=120, anchor="center")
        tree.column("registered_date", width=110, anchor="center")
        tree.column("registered_time", width=90, anchor="center")
        tree.column("cur_loc", width=130, anchor="center")
        tree.column("act", width=100, anchor="center")
        sb = search_bar(tab, tree, col_indices=[0])
        sb.pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            for rec in reversed(db.get("scrap", [])):
                sid = rec["slab_id"]
                who = (get_display_name(rec.get("registered_by", "—"), db)
                       if self.role == "admin" else "شخص دیگر")
                tree.insert("", "end", values=(
                    sid,
                    rec.get("reason", "—"),
                    who,
                    *split_dt(rec.get("registered_at", "—")),
                    get_current_location(db, sid),
                    "🚚 انتقال",
                ))

        def on_tree_click(event):
            row_id = tree.identify_row(event.y)
            if not row_id:
                return
            col_id = tree.identify_column(event.x)
            vals = tree.item(row_id, "values")
            if not vals:
                return
            sid = str(vals[0]).split(" ")[0].strip()
            try:
                col_name = tree["displaycolumns"][int(col_id.replace("#", "")) - 1]
                if col_name == "act":
                    db = load_db()
                    cur = get_current_location(db, sid)
                    open_slab_transfer_popup(
                        self, sid, cur,
                        status_lbl=status_lbl, on_done=refresh,
                        source="تب قراضه")
            except (IndexError, ValueError):
                pass

        tree.bind("<Button-1>", on_tree_click)
        refresh()
        tab.refresh = refresh
        self._bind_admin_popup(tree, "scrap", refresh)
        if self.role == "admin":
            tk.Label(tab, text="👑 دابل‌کلیک یا راست‌کلیک روی هر ردیف برای ویرایش/حذف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)

    # ═══════════════════════
    #  تب: اسکارف / برش
    # ══════════════════════════════════════
    #  تب: اسکارف
    # ══════════════════════════════════════
    def _build_scarf_tab(self, tab):
        tab.configure(bg=C["panel"])
        self._build_operation_tab(tab, "اسکارفی",
            items=["ترک طولی","ترک عرضی","کمربند","پلاک","حفره","آسیب‌های دیگر"],
            other_key="آسیب‌های دیگر",
            active_color=C["warning"],
            row_bg=C["card2"],
            header_icon="⚙",
            header_text="دلایل اسکارف",
            on_bg="#fff8e8",
            center_text=True,
        )

    # ══════════════════════════════════════
    #  تب: برش
    # ══════════════════════════════════════
    def _build_cut_tab(self, tab):
        tab.configure(bg=C["panel"])
        self._build_operation_tab(tab, "برشی",
            items=["طول مازاد / خارج از سفارش","اسلب با سر","اسلب با ته","دلایل دیگر"],
            other_key="دلایل دیگر",
            active_color=C["accent"],
            row_bg=C["card2"],
            header_icon="✂",
            header_text="دلایل برش",
            on_bg="#eaf4ff",
            center_text=True,
        )




    # ══════════════════════════════════════════════
    #  موتور مشترک تب‌های اسکارف و برش
    # ══════════════════════════════════════════════
    def _build_operation_tab(self, tab, op_name, items, other_key,
                              active_color, row_bg, header_icon, header_text, on_bg,
                              center_text=False):
        BG = C["panel"]

        # ── اسکرول ──
        canvas = tk.Canvas(tab, bg=BG, highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a",
                           activebackground=active_color, width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        sf = tk.Frame(canvas, bg=BG)
        _win = canvas.create_window((0,0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_win, width=e.width))

        register_scroll_canvas(canvas, sf)

        # ── هدر فشرده ──
        hero = tk.Frame(sf, bg=C["card"], highlightthickness=1, highlightbackground=active_color)
        hero.pack(fill="x", padx=16, pady=8)
        tk.Frame(hero, bg=active_color, height=3).pack(fill="x")
        hero_in = tk.Frame(hero, bg=C["card"])
        hero_in.pack(fill="x", padx=16, pady=10)
        big_icon = "🔧" if op_name=="اسکارفی" else "✂️"
        # عنوان دلایل وسط‌چین (دلایل اسکارف / دلایل برش)
        tk.Label(hero_in, text=f"{big_icon}  {header_text}",
                 bg=C["card"], fg=active_color,
                 font=(_MAIN_FONT,13,"bold"),
                 anchor="center", justify="center").pack(fill="x")
        tk.Label(hero_in, text="چند دلیل می‌توانید انتخاب کنید",
                 bg=C["card"], fg=C["text_dim"], font=(_MAIN_FONT,9, "bold"),
                 anchor="center", justify="center").pack(fill="x", pady=(2, 0))

        # ── شماره اسلب ──
        sid_card = tk.Frame(sf, bg=C["card"],
                            highlightthickness=1, highlightbackground=C["border"])
        sid_card.pack(fill="x", padx=16, pady=(0,8))
        tk.Frame(sid_card, bg=active_color, height=2).pack(fill="x")
        sid_in = tk.Frame(sid_card, bg=C["card"])
        sid_in.pack(fill="x", padx=20, pady=14)
        sid_row = tk.Frame(sid_in, bg=C["card"])
        sid_row.pack(fill="x")
        tk.Label(sid_row, text="شماره اسلب  (۱۱ رقم):",
                 bg=C["card"], fg=C["text"], font=FONT_NORM).pack(side="right", padx=(0,6))
        sid_var = tk.StringVar()
        sid_ent = tk.Entry(sid_row, textvariable=sid_var,
                           bg=C["entry_bg"], fg=C["text"],
                           insertbackground=C["accent"],
                           font=FONT_MONO,
                           justify="right", bd=0, relief="flat",
                           highlightthickness=1,
                           highlightbackground=C["border"],
                           highlightcolor=C["accent"], width=18)
        sid_ent.pack(side="right")
        sid_ent.focus_set()

        # ── لیست دلایل ──
        reasons_card = tk.Frame(sf, bg=C["card"],
                                highlightthickness=1, highlightbackground=C["border"])
        reasons_card.pack(fill="x", padx=16, pady=(0,8))
        tk.Frame(reasons_card, bg=active_color, height=2).pack(fill="x")
        rc_in = tk.Frame(reasons_card, bg=C["card"])
        rc_in.pack(fill="x", padx=16, pady=14)

        rc_head = tk.Frame(rc_in, bg=C["card"])
        rc_head.pack(fill="x", pady=(0,10))
        if center_text:
            tk.Label(rc_head, text="دلیل / دلایل:",
                     bg=C["card"], fg=C["text_dim"], font=(_MAIN_FONT,9, "bold"),
                     anchor="center", justify="center").pack(fill="x")
            sel_count_lbl = tk.Label(rc_head, text="", bg=C["card"],
                                      fg=active_color, font=(_MAIN_FONT,9,"bold"),
                                      anchor="center")
            sel_count_lbl.pack(fill="x")
        else:
            tk.Label(rc_head, text="دلیل / دلایل:",
                     bg=C["card"], fg=C["text_dim"], font=(_MAIN_FONT,9, "bold")).pack(side="right")
            sel_count_lbl = tk.Label(rc_head, text="", bg=C["card"],
                                      fg=active_color, font=(_MAIN_FONT,9,"bold"))
            sel_count_lbl.pack(side="left")

        # فیلد توضیحات «موارد دیگر» — فقط وقتی گزینهٔ «دلایل دیگر / آسیب‌های دیگر» انتخاب شود نمایش داده می‌شود
        note_var = tk.StringVar()
        note_frame = tk.Frame(rc_in, bg=on_bg,
                              highlightthickness=1, highlightbackground=active_color)
        note_in = tk.Frame(note_frame, bg=on_bg)
        note_in.pack(fill="x", padx=14, pady=12)
        nt_h = tk.Frame(note_in, bg=on_bg)
        nt_h.pack(fill="x", pady=(0,6))
        tk.Label(nt_h, text="توضیحات:  (الزامی)", bg=on_bg,
                 fg=active_color, font=(_MAIN_FONT,10,"bold")).pack(side="right")
        tk.Label(nt_h, text="✏", bg=on_bg, fg=active_color,
                 font=("Segoe UI Symbol",12, "bold")).pack(side="right", padx=(0,6))
        nw = tk.Frame(note_in, bg=active_color, padx=1, pady=1)
        nw.pack(fill="x")
        tk.Entry(nw, textvariable=note_var, bg=C["entry_bg"], fg=C["text"],
                 insertbackground=active_color, font=(_MAIN_FONT,11, "bold"),
                 bd=0, relief="flat").pack(fill="x", ipady=9, padx=2)
        # توجه: note_frame.pack() اینجا فراخوانی نمی‌شود — به‌صورت داینامیک توسط toggle آیتم «other_key» کنترل می‌شود

        ICONS = {
            "ترک طولی":"━━","ترک عرضی":"┃┃","کمربند":"◻","پلاک":"▪",
            "حفره":"●","آسیب‌های دیگر":"…",
            "طول مازاد / خارج از سفارش":"↔","اسلب با سر":"▲",
            "اسلب با ته":"▼","دلایل دیگر":"…",
        }

        checks = {}
        selected_set = set()

        def update_note_visibility():
            if other_key in selected_set:
                if not note_frame.winfo_ismapped():
                    note_frame.pack(fill="x", pady=(4,0))
            else:
                if note_frame.winfo_ismapped():
                    note_frame.pack_forget()
                note_var.set("")



        def update_count():
            n = len(selected_set)
            sel_count_lbl.config(text=f"{n} مورد انتخاب شده" if n else "")

        def _render_reason_card(idx, item):
            """یک کارت دلیل (چک‌باکس) می‌سازد — برای استفاده در ترتیب دلخواه"""
            var = tk.BooleanVar()
            checks[item] = var
            is_other = (item == other_key)
            ic_txt = ICONS.get(item,"◆")

            card = tk.Frame(rc_in, bg=row_bg, cursor="hand2",
                            highlightthickness=1, highlightbackground=C["border2"])
            card.pack(fill="x", pady=2)
            card_in = tk.Frame(card, bg=row_bg)
            card_in.pack(fill="x", padx=10, pady=6)

            # چک‌باکس سمت راست
            chk = tk.Label(card_in, text="○", bg=row_bg, fg=C["text_dim"],
                           font=("Segoe UI Symbol",15, "bold"), cursor="hand2", width=2)

            # متن دلیل
            nm = tk.Label(card_in, text=item, bg=row_bg, fg=C["text"],
                          font=(_MAIN_FONT,10, "bold"), cursor="hand2",
                          anchor=("center" if center_text else "e"),
                          justify=("center" if center_text else "right"))

            # شماره و آیکون — در RTL سمت چپ نمایش
            num_b = tk.Label(card_in, text=f"{idx+1}", bg=C["border2"],
                             fg=C["text_dim"], font=("B Nazanin",8,"bold"),
                             width=2, padx=2, pady=1)

            ic_lbl = tk.Label(card_in, text=ic_txt, bg=row_bg,
                              fg=C["text_dim"], font=("B Nazanin",11, "bold"), width=2)

            if center_text:
                # شماره/آیکون چپ، چک‌باکس راست، متن دقیقاً وسط فضای باقی‌مانده
                # (side=right روی متن باعث راست‌چین شدن روی کلاینت می‌شد)
                num_b.pack(side="left", padx=(0, 6))
                ic_lbl.pack(side="left", padx=(0, 4))
                chk.pack(side="right")
                nm.pack(fill="both", expand=True, padx=8)
            else:
                chk.pack(side="right")
                nm.pack(side="right", padx=(0,8))
                num_b.pack(side="left", padx=(0,6))
                ic_lbl.pack(side="left", padx=(0,4))

            def mk(v=var,c=chk,cd=card,ci=card_in,n=nm,il=ic_lbl,nb=num_b,
                   nf=note_frame,other=is_other,iname=item):
                def tog(e=None):
                    v.set(not v.get()); on=v.get()
                    if on:
                        selected_set.add(iname)
                        c.config(text="◉",fg=active_color)
                        cd.config(bg=on_bg,highlightbackground=active_color,highlightthickness=1)
                        ci.config(bg=on_bg); n.config(bg=on_bg,fg=active_color)
                        il.config(bg=on_bg,fg=active_color); c.config(bg=on_bg)
                        nb.config(bg=active_color,
                                  fg="#000" if active_color==C["accent"] else C["card"])
                    else:
                        selected_set.discard(iname)
                        c.config(text="○",fg=C["text_dim"])
                        cd.config(bg=row_bg,highlightbackground=C["border2"],highlightthickness=1)
                        ci.config(bg=row_bg); n.config(bg=row_bg,fg=C["text"])
                        il.config(bg=row_bg,fg=C["text_dim"]); c.config(bg=row_bg)
                        nb.config(bg=C["border2"],fg=C["text_dim"])
                    update_count()
                    update_note_visibility()
                return tog
            t=mk()
            for w in [card,card_in,chk,nm,ic_lbl,num_b]: w.bind("<Button-1>",t)
            def bh(cd=card,v=var):
                cd.bind("<Enter>", lambda e: cd.config(highlightbackground=active_color) if not v.get() else None)
                cd.bind("<Leave>", lambda e: cd.config(highlightbackground=C["border2"]) if not v.get() else None)
            bh()

        # ── ترتیب نمایش: همه دلایل عادی، سپس (برای برش) تست باومن، و در آخر «دلایل دیگر/آسیب‌های دیگر» ──
        _normal_items = [it for it in items if it != other_key]
        for idx, item in enumerate(_normal_items):
            _render_reason_card(idx, item)

        # ── باومن (فقط برش) — بین دلایل عادی و «دلایل دیگر» قرار می‌گیرد ──
        bauman_done_var = tk.BooleanVar()
        if op_name == "برشی":
            bm_card = tk.Frame(rc_in, bg=row_bg, cursor="hand2",
                               highlightthickness=1, highlightbackground=C["border2"])
            bm_card.pack(fill="x", pady=2)
            bm_ci = tk.Frame(bm_card, bg=row_bg)
            bm_ci.pack(fill="x", padx=10, pady=6)

            # چک‌باکس سمت راست
            bm_chk = tk.Label(bm_ci, text="○", bg=row_bg, fg=C["text_dim"],
                              font=("Segoe UI Symbol",15, "bold"), cursor="hand2", width=2)
            bm_chk.pack(side="right")

            # شماره و آیکون سمت چپ + چک‌باکس راست، متن وسط
            bm_nb = tk.Label(bm_ci, text="◈", bg=C["border2"],
                             fg=C["text_dim"], font=("B Nazanin",8,"bold"),
                             width=2, padx=2, pady=1)
            bm_il = tk.Label(bm_ci, text="🔬", bg=row_bg,
                             fg=C["text_dim"], font=("Segoe UI Emoji",10, "bold"), width=2)
            bm_nm = tk.Label(bm_ci, text="تست باومن", bg=row_bg, fg=C["text"],
                             font=(_MAIN_FONT,10, "bold"), cursor="hand2",
                             anchor=("center" if center_text else "e"),
                             justify=("center" if center_text else "right"))
            if center_text:
                bm_nb.pack(side="left", padx=(0, 6))
                bm_il.pack(side="left")
                # bm_chk قبلاً pack شده — متن فضای وسط را پر می‌کند
                bm_nm.pack(fill="both", expand=True, padx=8)
            else:
                bm_nm.pack(side="right", padx=(0, 8))
                bm_nb.pack(side="left", padx=(0, 6))
                bm_il.pack(side="left")

            _bm_on_bg = "#fff8d0"
            def tog_bm(e=None):
                bauman_done_var.set(not bauman_done_var.get()); on=bauman_done_var.get()
                bm_chk.config(text="◉" if on else "○", fg=C["gold"] if on else C["text_dim"])
                bg_ = _bm_on_bg if on else row_bg; hl_ = C["gold"] if on else C["border2"]
                bm_card.config(bg=bg_, highlightbackground=hl_, highlightthickness=1)
                bm_ci.config(bg=bg_); bm_chk.config(bg=bg_); bm_nm.config(bg=bg_,fg=C["gold"] if on else C["text"])
                bm_nb.config(bg=C["gold"] if on else C["border2"], fg=C["card"] if on else C["text_dim"])
                bm_il.config(bg=bg_)
            for w in [bm_card, bm_ci, bm_chk, bm_nm, bm_nb, bm_il]:
                w.bind("<Button-1>", tog_bm)
            bm_card.bind("<Enter>", lambda e: bm_card.config(highlightbackground=C["gold"]) if not bauman_done_var.get() else None)
            bm_card.bind("<Leave>", lambda e: bm_card.config(highlightbackground=C["border2"]) if not bauman_done_var.get() else None)

        # ── «دلایل دیگر / آسیب‌های دیگر» — همیشه آخرین گزینه ──
        _render_reason_card(len(_normal_items), other_key)

        # ── تعداد برش (فقط برشی) + دکمه ویرایش از جدول ──
        cut_count_var = tk.IntVar(value=1)
        if op_name == "برشی":
            cc_row = tk.Frame(sf, bg=BG)
            cc_row.pack(fill="x", padx=16, pady=(0,4))
            tk.Label(cc_row, text="تعداد برش:", bg=BG, fg=C["text"],
                     font=FONT_NORM).pack(side="right", padx=(0,6))
            cc_frame = tk.Frame(cc_row, bg=C["border2"], highlightthickness=1,
                                 highlightbackground=C["border"])
            cc_frame.pack(side="right")
            tk.Button(cc_frame, text=" ＋ ", command=lambda: cut_count_var.set(cut_count_var.get()+1),
                      bg=C["btn_success"], fg="white", font=(_MAIN_FONT,11,"bold"),
                      bd=0, relief="flat", cursor="hand2", padx=6).pack(side="left")
            tk.Label(cc_frame, textvariable=cut_count_var, bg=C["card2"], fg=active_color,
                     font=("B Nazanin",14,"bold"), width=4, anchor="center").pack(side="left", padx=4)
            tk.Button(cc_frame, text=" － ", command=lambda: cut_count_var.set(max(1,cut_count_var.get()-1)),
                      bg=C["btn_danger"], fg="white", font=(_MAIN_FONT,11,"bold"),
                      bd=0, relief="flat", cursor="hand2", padx=6).pack(side="left")
            tk.Label(cc_row, text="(پیش‌فرض: ۱)", bg=BG, fg=C["text_dim"],
                     font=FONT_SMALL).pack(side="right", padx=8)

        # ── دکمه ویرایش ردیف انتخابی از جدول ──
        def load_for_edit():
            """بارگذاری اطلاعات ردیف انتخابی برای ویرایش"""
            sel = tree2.selection()
            if not sel:
                messagebox.showwarning("خطا","یک ردیف از جدول انتخاب کنید.",parent=self); return
            vals = tree2.item(sel[0],"values")
            sid_val = vals[0]
            db = load_db()
            existing = next((r for r in db.get("scarf_cut",[])
                             if r["slab_id"]==sid_val and r.get("operation")==op_name), None)
            if not existing:
                messagebox.showerror("خطا","رکورد یافت نشد.",parent=self); return
            # پر کردن فیلد شماره
            sid_var.set(sid_val)
            # تیک کردن دلایل قبلی
            prev_reasons = set(existing.get("reason","").split("، ")) if existing.get("reason") else set()
            for item, var in checks.items():
                var.set(item in prev_reasons)
            selected_set.clear()
            selected_set.update(prev_reasons)
            update_count()
            update_note_visibility()
            # توضیحات
            note_var.set(existing.get("note",""))
            # باومن
            if op_name=="برشی" and existing.get("bauman_done"):
                bauman_done_var.set(True)
                bm_chk.config(text="◉", fg=C["gold"])
            # تعداد برش
            if op_name=="برشی":
                cut_count_var.set(existing.get("cut_count",1))
            status_lbl.config(text=f"✏  اسلب {sid_val} بارگذاری شد — تغییرات را اعمال کنید",
                              fg=C["warning"])

        # ── وضعیت + دکمه ──
        status_lbl=tk.Label(sf,text="",bg=BG,fg=C["success"],font=(_MAIN_FONT,10, "bold"))
        status_lbl.pack(anchor="e",padx=20,pady=(4,0))
        bot=tk.Frame(sf,bg=BG)
        bot.pack(fill="x",padx=16,pady=12)

        def do_register():
            sid=sid_var.get().strip()
            ok,msg,sid=validate_slab_id(sid)
            if not ok: messagebox.showerror("خطا",msg,parent=self); return
            sel=[k for k,v in checks.items() if v.get()]
            if not sel and selected_set:
                sel = list(selected_set)
            has_bm = (op_name=="برشی" and bauman_done_var.get())
            if not sel and not has_bm:
                messagebox.showerror("خطا",
                    "لطفاً یکی از موارد یا چند مورد را انتخاب کنید",
                    parent=self)
                return
            db=load_db()
            if not any(r["slab_id"]==sid for r in db["melts"]):
                messagebox.showerror("خطا",f"اسلب {sid} در سیستم ثبت نشده.",parent=self); return
            if not assert_scarf_cut_allowed(db, sid, parent=self):
                return
            reason="، ".join(sel)
            note=note_var.get().strip()

            if other_key in sel and not note:
                messagebox.showerror("خطا",
                    f"برای گزینهٔ «{other_key}» باید توضیحات را وارد کنید.",parent=self); return

            def _ensure_bauman():
                if not any(r.get("slab_id")==sid for r in db.get("bauman",[])):
                    db.setdefault("bauman",[]).append({
                        "slab_id": sid, "cut_by": self.username,
                        "cut_at": now_str(), "lab_status": "در انتظار تحویل",
                        "auto_from_cut": True
                    })

            if check_duplicate(db,"scarf_cut",sid):
                # اجازه اضافه کردن مشکلات جدید — ولی نه ثبت کامل دوباره
                if not messagebox.askyesno("⚠️  اسلب قبلاً ثبت شده",
                    f"اسلب {sid} قبلاً در این بخش ثبت شده است.\n\n"
                    "می‌توانید دلایل جدید به آن اضافه یا از آن حذف کنید.\n\n"
                    "برای ویرایش، روی دکمه «بله» کلیک کنید.",
                    parent=self):
                    return
                # ویرایش: merge دلایل جدید با قدیمی
                existing = next((r for r in db["scarf_cut"] if r["slab_id"]==sid and r.get("operation")==op_name), None)
                if existing:
                    new_reasons = set(reason.split("، ")) if reason else set()
                    old_reasons = set(existing.get("reason","").split("، ")) if existing.get("reason") else set()
                    merged = "، ".join(sorted(old_reasons | new_reasons))
                    existing["reason"] = merged
                    existing["last_edit_by"] = self.username
                    existing["last_edit_at"] = now_str()
                    existing["updated_at"] = existing["last_edit_at"]
                    if note: existing["note"] = note
                    if has_bm:
                        existing["bauman_done"] = True
                        _ensure_bauman()
                    if op_name=="برشی":
                        existing["cut_count"] = cut_count_var.get()
                    save_db(db)
                    sid_var.set(""); note_var.set("")
                    for v in checks.values(): v.set(False)
                    selected_set.clear(); update_count(); bauman_done_var.set(False)
                    update_note_visibility()
                    status_lbl.config(text=f"✔  اسلب {sid} ویرایش شد — دلایل: {merged}")
                    refresh()
                    return
            reason_display = reason if reason else "—  (فقط تست باومن)"
            if not messagebox.askyesno("تأیید ثبت",
                f"اسلب: {sid}\nنوع: {op_name}\nدلایل: {reason_display}\n\nآیا اطمینان دارید؟",parent=self): return
            db["scarf_cut"].append({"slab_id":sid,"operation":op_name,"reason":reason,"note":note,
                "bauman_done":has_bm,"cut_count":cut_count_var.get() if op_name=="برشی" else 1,
                "registered_by":self.username,"registered_at":now_str(),"updated_at":now_str()})
            if has_bm:
                _ensure_bauman()
            save_db(db)
            sid_var.set(""); note_var.set("")
            for v in checks.values(): v.set(False)
            selected_set.clear(); update_count(); bauman_done_var.set(False)
            update_note_visibility()
            status_lbl.config(text=f"✔  اسلب {sid}  ({op_name})  ثبت شد.")
            refresh()

        btn_wrap=tk.Frame(bot,bg=active_color,cursor="hand2")
        btn_wrap.pack(side="right")
        btn_lbl=tk.Label(btn_wrap,text=f"  ✔  ثبت {op_name}  ",bg=active_color,
                         fg="#000000" if active_color==C["accent"] else C["card"],
                         font=(_MAIN_FONT,12,"bold"),padx=20,pady=12,cursor="hand2")
        btn_lbl.pack()
        btn_wrap.bind("<Button-1>",lambda e:do_register())
        btn_lbl.bind("<Button-1>",lambda e:do_register())
        btn_wrap.bind("<Enter>",lambda e:[btn_wrap.config(bg=_lighten(active_color,20)),btn_lbl.config(bg=_lighten(active_color,20))])
        btn_wrap.bind("<Leave>",lambda e:[btn_wrap.config(bg=active_color),btn_lbl.config(bg=active_color)])

        # ── جدول تاریخچه ──
        clr_tag = "scarf_row" if op_name=="اسکارفی" else "cut_row"
        clr_bg2  = "#100e00" if op_name=="اسکارفی" else "#00080e"

        # ── جدول تاریخچه با Treeview ──
        tk.Frame(sf, bg=C["border"], height=1).pack(fill="x", padx=16, pady=(10,4))
        tk.Label(sf, text="📋  تاریخچه ثبت‌شده", bg=BG,
                 fg=C["text_dim"], font=(_MAIN_FONT,10,"bold")).pack(anchor="e", padx=18, pady=(0,4))

        if op_name == "برشی":
            tc2 = ("slab_id","reason","cut_count","bauman","registered_by","reg_date","reg_time","edit_col")
            th2 = ("شماره اسلب","دلایل","تعداد برش","باومن","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","ویرایش")
            cw2 = [130, 240, 75, 60, 110, 115, 70, 120]
        else:
            tc2 = ("slab_id","reason","registered_by","reg_date","reg_time","edit_col")
            th2 = ("شماره اسلب","دلایل","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","ویرایش")
            cw2 = [130, 300, 110, 115, 70, 120]

        tf2, tree2 = scrolled_tree(sf, tc2, th2, height=9)
        tf2.pack(fill="both", expand=True, padx=16, pady=4)
        for col, w in zip(tc2, cw2):
            tree2.column(col, width=w, anchor="center", minwidth=max(w-20,50))
        tree2.column("edit_col", width=120, anchor="center", minwidth=100)
        # دلایل اسکارف/برش — وسط‌چین
        if op_name == "برشی":
            tree2.column("reason", width=240, anchor="center")
        else:
            tree2.column("reason", width=300, anchor="center")
        search_bar(sf, tree2, col_indices=[0]).pack(anchor="e", padx=16, pady=4)
        sort_toolbar(sf, tree2, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        # dummy برای سازگاری با open_edit_popup
        rows_f = None
        hdr_f  = None

        def open_edit_popup(sid, rec_ref):
            """پنجره ویرایش — حداکثر ۳ بار مجاز — اسکرول‌دار و یکسان برای اسکارف و برش"""
            edit_key = f"scarf_cut:{op_name}:{sid}"
            if _acquire_edit_popup(self, edit_key):
                return

            db = load_db()
            rec = next((r for r in db.get("scarf_cut",[])
                        if r["slab_id"]==sid and r.get("operation")==op_name), None)
            if not rec:
                _release_edit_popup_claim(self, edit_key)
                messagebox.showerror("خطا","رکورد یافت نشد.",parent=self); return
            if not assert_scarf_cut_allowed(db, sid, parent=self):
                _release_edit_popup_claim(self, edit_key)
                return
            edit_count = rec.get("edit_count", 1 if rec.get("edited") else 0)
            MAX_EDITS = 3
            is_admin_editor = (self.role == "admin")
            # محدودیت ۳ بار فقط برای کاربران عادی — مدیریت سیستم بدون سقف
            if (not is_admin_editor) and edit_count >= MAX_EDITS:
                _release_edit_popup_claim(self, edit_key)
                messagebox.showwarning("⚠️  ویرایش مجاز نیست",
                    f"اسلب {sid} سه بار ویرایش شده است.\n"
                    "برای تغییرات بیشتر با سرپرست کارگاه تماس بگیرید.",
                    parent=self); return

            remaining = None if is_admin_editor else (MAX_EDITS - edit_count)

            pop = tk.Toplevel(self)
            prepare_popup_window(pop, self)
            _register_edit_popup(self, edit_key, pop)
            pop.title(f"ویرایش  {op_name}  —  اسلب {sid}")
            pop.configure(bg=C["bg"])
            pop.resizable(True, True)
            pop.focus_force()
            pw, ph = 560, 620
            sx = self.winfo_screenwidth(); sy = self.winfo_screenheight()
            pop.geometry(f"{pw}x{ph}+{(sx-pw)//2}+{(sy-ph)//2}")
            pop.minsize(480, 500)

            # ── نوار رنگی بالا ──
            tk.Frame(pop, bg=active_color, height=4).pack(fill="x")

            # ── هدر ──
            hf2 = tk.Frame(pop, bg=C["header_bg"])
            hf2.pack(fill="x")
            tk.Label(hf2, text=f"✏  ویرایش {op_name}",
                     bg=C["header_bg"], fg=active_color,
                     font=(_MAIN_FONT, 14, "bold")).pack(side="right", padx=16, pady=12)
            # نشانگر ویرایش‌های باقی‌مانده (ادمین: بدون محدودیت)
            if is_admin_editor:
                remaining_color = C["gold"]
                badge_txt = "مدیریت — بدون محدودیت ویرایش"
            else:
                remaining_color = C["success"] if remaining > 1 else C["warning"] if remaining == 1 else C["danger"]
                badge_txt = f"{remaining} ویرایش باقی‌مانده از ۳"
            badge_f = tk.Frame(hf2, bg=remaining_color, padx=8, pady=4)
            badge_f.pack(side="left", padx=16, pady=10)
            tk.Label(badge_f,
                     text=badge_txt,
                     bg=remaining_color, fg="#000000" if (is_admin_editor or remaining > 0) else "#ffffff",
                     font=(_MAIN_FONT, 10, "bold")).pack()
            tk.Label(hf2, text=f"اسلب: {sid}",
                     bg=C["header_bg"], fg=C["text_dim"],
                     font=FONT_SMALL).pack(side="left", padx=4, pady=12)

            # ── بدنه اسکرول‌دار ──
            body_outer = tk.Frame(pop, bg=C["card"])
            body_outer.pack(fill="both", expand=True, padx=0, pady=0)

            body_canvas = tk.Canvas(body_outer, bg=C["card"], highlightthickness=0)
            body_vsb = tk.Scrollbar(body_outer, orient="vertical",
                                     command=body_canvas.yview,
                                     bg="#707070", troughcolor="#1a1a1a",
                                     activebackground=active_color, width=16)
            body_canvas.configure(yscrollcommand=body_vsb.set)
            body_vsb.pack(side="left", fill="y")
            body_canvas.pack(side="right", fill="both", expand=True)
            body = tk.Frame(body_canvas, bg=C["card"])
            _bwin = body_canvas.create_window((0,0), window=body, anchor="nw")
            body.bind("<Configure>", lambda e: body_canvas.configure(
                scrollregion=body_canvas.bbox("all")))
            body_canvas.bind("<Configure>", lambda e: body_canvas.itemconfig(
                _bwin, width=e.width))
            register_scroll_canvas(body_canvas, body)

            inner = tk.Frame(body, bg=C["card"])
            inner.pack(fill="both", padx=20, pady=14)

            # ── دلایل فعلی ──
            cur_reason = rec.get("reason","—")
            tk.Label(inner, text="دلایل فعلی:", bg=C["card"],
                     fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(0,4))
            cur_lbl_f = tk.Frame(inner, bg=C["card2"],
                                  highlightthickness=1,
                                  highlightbackground=active_color)
            cur_lbl_f.pack(fill="x", pady=(0,12))
            tk.Label(cur_lbl_f, text=cur_reason if cur_reason else "—",
                     bg=C["card2"], fg=active_color,
                     font=(_MAIN_FONT, 11, "bold"),
                     anchor="e", justify="right",
                     wraplength=480, padx=12, pady=8).pack(fill="x")

            # ── انتخاب دلایل ──
            sep_lbl = tk.Frame(inner, bg=active_color, height=2)
            sep_lbl.pack(fill="x", pady=(0,6))
            tk.Label(inner, text="دلایل را انتخاب کنید  (وضعیت فعلی نشان داده شده):",
                     bg=C["card"], fg=C["text"], font=FONT_NORM,
                     anchor=("center" if center_text else "e"),
                     justify=("center" if center_text else "right")).pack(
                         fill="x" if center_text else None,
                         anchor=("center" if center_text else "e"),
                         pady=(0, 6))

            chk_frame = tk.Frame(inner, bg=C["card"])
            chk_frame.pack(fill="x")
            new_checks = {}
            prev_set = set(cur_reason.split("، ")) if cur_reason and cur_reason not in ("—","") else set()

            def _add_edit_check_row(i, item):
                is_on = item in prev_set
                v2 = tk.BooleanVar(value=is_on)
                rbg = C["card2"] if i % 2 == 0 else C["card"]
                row_f = tk.Frame(chk_frame, bg=rbg if not is_on else _lighten(rbg, 8),
                                  cursor="hand2",
                                  highlightthickness=1,
                                  highlightbackground=active_color if is_on else C["border"])
                row_f.pack(fill="x", pady=2)
                inner_r = tk.Frame(row_f, bg=row_f["bg"])
                inner_r.pack(fill="x", padx=10, pady=5)

                # چک‌باکس
                chk2 = tk.Label(inner_r,
                                 text="◉" if is_on else "○",
                                 bg=inner_r["bg"],
                                 fg=active_color if is_on else C["text_dim"],
                                 font=("Segoe UI Symbol", 16, "bold"),
                                 cursor="hand2", width=2)
                chk2.pack(side="right", padx=6)

                # متن — روی کلاینت/center_text وسط‌چین
                nm2 = tk.Label(inner_r, text=item,
                                bg=inner_r["bg"],
                                fg=active_color if is_on else C["text"],
                                font=(_MAIN_FONT, 11, "bold"),
                                cursor="hand2",
                                anchor=("center" if center_text else "e"),
                                justify=("center" if center_text else "right"))
                if center_text:
                    nm2.pack(fill="both", expand=True, padx=4)
                else:
                    nm2.pack(side="right", padx=4)

                def mk(v=v2, c=chk2, rf=row_f, ir=inner_r, n=nm2, on_bg=rbg):
                    def tog(e=None):
                        v.set(not v.get())
                        on = v.get()
                        _bg = _lighten(on_bg, 8) if on else on_bg
                        c.config(text="◉" if on else "○",
                                 fg=active_color if on else C["text_dim"],
                                 bg=_bg)
                        rf.config(bg=_bg,
                                  highlightbackground=active_color if on else C["border"])
                        ir.config(bg=_bg)
                        n.config(bg=_bg, fg=active_color if on else C["text"])
                    return tog

                t2 = mk()
                chk2.bind("<Button-1>", t2)
                row_f.bind("<Button-1>", t2)
                inner_r.bind("<Button-1>", t2)
                nm2.bind("<Button-1>", t2)
                new_checks[item] = v2

            _normal_items_edit = [it for it in items if it != other_key]
            for i, item in enumerate(_normal_items_edit):
                _add_edit_check_row(i, item)

            # ── تست باومن (فقط برش) ──
            edit_bauman_var = tk.BooleanVar(value=bool(rec.get("bauman_done"))) if op_name == "برشی" else None
            if op_name == "برشی":
                _bm_is_on = edit_bauman_var.get()
                _bm_i = len(_normal_items_edit)
                _bm_rbg = C["card2"] if _bm_i % 2 == 0 else C["card"]
                bm_row = tk.Frame(chk_frame,
                                   bg=_lighten(_bm_rbg, 8) if _bm_is_on else _bm_rbg,
                                   cursor="hand2",
                                   highlightthickness=1,
                                   highlightbackground=C["gold"] if _bm_is_on else C["border"])
                bm_row.pack(fill="x", pady=2)
                bm_inner = tk.Frame(bm_row, bg=bm_row["bg"])
                bm_inner.pack(fill="x", padx=10, pady=5)
                bm_chk2 = tk.Label(bm_inner,
                                    text="◉" if _bm_is_on else "○",
                                    bg=bm_inner["bg"],
                                    fg=C["gold"] if _bm_is_on else C["text_dim"],
                                    font=("Segoe UI Symbol", 16, "bold"),
                                    cursor="hand2", width=2)
                bm_chk2.pack(side="right", padx=6)
                bm_nm2 = tk.Label(bm_inner, text="🔬  تست باومن",
                                   bg=bm_inner["bg"],
                                   fg=C["gold"] if _bm_is_on else C["text"],
                                   font=(_MAIN_FONT, 11, "bold"), cursor="hand2", anchor="e")
                bm_nm2.pack(side="right", padx=4)

                def _tog_bm2(e=None, on_bg=_bm_rbg):
                    edit_bauman_var.set(not edit_bauman_var.get())
                    on = edit_bauman_var.get()
                    _bg = _lighten(on_bg, 8) if on else on_bg
                    bm_row.config(bg=_bg, highlightbackground=C["gold"] if on else C["border"])
                    bm_inner.config(bg=_bg)
                    bm_chk2.config(text="◉" if on else "○",
                                   fg=C["gold"] if on else C["text_dim"], bg=_bg)
                    bm_nm2.config(bg=_bg, fg=C["gold"] if on else C["text"])

                for w in [bm_row, bm_inner, bm_chk2, bm_nm2]:
                    w.bind("<Button-1>", _tog_bm2)

            # ── «دلایل دیگر / آسیب‌های دیگر» ──
            _add_edit_check_row(len(_normal_items_edit) + (1 if op_name == "برشی" else 0),
                                other_key)

            # فیلد توضیحات برای «دلایل دیگر / آسیب‌های دیگر» — اجباری وقتی انتخاب شود
            edit_note_var = tk.StringVar(value=rec.get("note", "") or "")
            edit_note_frame = tk.Frame(inner, bg=on_bg,
                                       highlightthickness=1, highlightbackground=active_color)
            _en_in = tk.Frame(edit_note_frame, bg=on_bg)
            _en_in.pack(fill="x", padx=12, pady=10)
            tk.Label(_en_in, text="توضیحات:  (الزامی)", bg=on_bg,
                     fg=active_color, font=(_MAIN_FONT, 10, "bold")).pack(anchor="e", pady=(0, 4))
            tk.Entry(_en_in, textvariable=edit_note_var, bg=C["entry_bg"], fg=C["text"],
                     insertbackground=active_color, font=(_MAIN_FONT, 11, "bold"),
                     bd=0, relief="flat", justify="right").pack(fill="x", ipady=8)

            def _sync_edit_note_vis():
                if new_checks.get(other_key) and new_checks[other_key].get():
                    if not edit_note_frame.winfo_ismapped():
                        edit_note_frame.pack(fill="x", pady=(8, 0))
                else:
                    if edit_note_frame.winfo_ismapped():
                        edit_note_frame.pack_forget()

            # اتصال به toggle دلایل دیگر
            if other_key in new_checks:
                _old_tog = None
                for w in chk_frame.winfo_children():
                    pass
                _prev_other = new_checks[other_key]

                def _watch_other(*_a):
                    _sync_edit_note_vis()

                try:
                    _prev_other.trace_add("write", _watch_other)
                except Exception:
                    _prev_other.trace("w", _watch_other)
                _sync_edit_note_vis()

            # ── تعداد برش (فقط برش) ──
            edit_cut_count = tk.IntVar(value=rec.get("cut_count", 1))
            if op_name == "برشی":
                cc2_row = tk.Frame(inner, bg=C["card"])
                cc2_row.pack(fill="x", pady=(12, 4))
                tk.Label(cc2_row, text="تعداد برش:", bg=C["card"],
                         fg=C["text"], font=FONT_NORM).pack(side="right", padx=(0, 8))
                cc2_f = tk.Frame(cc2_row, bg=C["border2"])
                cc2_f.pack(side="right")
                tk.Button(cc2_f, text=" ＋ ", bg=C["btn_success"], fg="white",
                          font=(_MAIN_FONT, 11, "bold"), bd=0, relief="flat",
                          cursor="hand2", padx=8,
                          command=lambda: edit_cut_count.set(edit_cut_count.get()+1)).pack(side="left")
                tk.Label(cc2_f, textvariable=edit_cut_count, bg=C["card2"],
                         fg=active_color, font=("B Nazanin", 14, "bold"),
                         width=4, anchor="center").pack(side="left", padx=6)
                tk.Button(cc2_f, text=" － ", bg=C["btn_danger"], fg="white",
                          font=(_MAIN_FONT, 11, "bold"), bd=0, relief="flat",
                          cursor="hand2", padx=8,
                          command=lambda: edit_cut_count.set(max(1, edit_cut_count.get()-1))).pack(side="left")

            # ── دکمه ثبت ──
            def do_edit_save():
                sel_new = set(k for k, v in new_checks.items() if v.get())
                has_bm_edit = (op_name == "برشی" and edit_bauman_var is not None
                               and edit_bauman_var.get())
                if not sel_new and not has_bm_edit:
                    messagebox.showerror("خطا",
                        "حداقل یک دلیل انتخاب کنید یا تیک باومن را بزنید.",
                        parent=pop); return
                if other_key in sel_new and not edit_note_var.get().strip():
                    messagebox.showerror("خطا",
                        f"برای گزینهٔ «{other_key}» باید توضیحات را وارد کنید — خالی مجاز نیست.",
                        parent=pop); return
                merged = "، ".join(sorted(sel_new))
                reason_disp = merged if merged else "—  (فقط تست باومن)"
                if is_admin_editor:
                    _confirm_msg = (
                        f"اسلب: {sid}\nدلایل: {reason_disp}\n\n"
                        "آیا مطمئن هستید؟"
                    )
                else:
                    _confirm_msg = (
                        f"اسلب: {sid}\nدلایل: {reason_disp}\n\n"
                        f"بعد از این ویرایش، {remaining - 1} ویرایش باقی می‌ماند.\n"
                        "آیا مطمئن هستید؟"
                    )
                if not messagebox.askyesno("تأیید ویرایش", _confirm_msg, parent=pop):
                    return
                _db3 = load_db()
                if not assert_scarf_cut_allowed(_db3, sid, parent=pop):
                    return
                _rec = next((r for r in _db3.get("scarf_cut", [])
                              if r["slab_id"] == sid and r.get("operation") == op_name), None)
                if _rec:
                    _rec["reason"]       = merged
                    _rec["note"]         = edit_note_var.get().strip() if other_key in sel_new else _rec.get("note", "")
                    if other_key not in sel_new:
                        _rec["note"] = ""
                    _rec["last_edit_by"] = self.username
                    _rec["last_edit_at"] = now_str()
                    _new_ec = _rec.get("edit_count", 1 if _rec.get("edited") else 0) + 1
                    _rec["edit_count"]   = _new_ec
                    # قفل ۳ بار فقط برای کاربران عادی؛ ادمین هرگز قفل نمی‌شود
                    _rec["edited"]       = (False if is_admin_editor else (_new_ec >= MAX_EDITS))
                    if op_name == "برشی":
                        _rec["cut_count"]   = edit_cut_count.get()
                        _rec["bauman_done"] = bool(edit_bauman_var.get())
                        if edit_bauman_var.get():
                            if not any(r.get("slab_id") == sid
                                       for r in _db3.get("bauman", [])):
                                _db3.setdefault("bauman", []).append({
                                    "slab_id": sid, "cut_by": self.username,
                                    "cut_at": now_str(),
                                    "lab_status": "در انتظار تحویل",
                                    "auto_from_cut": True
                                })
                save_db(_db3)
                pop.destroy()
                refresh()

            # ── نوار دکمه‌ها (ثابت، پایین پنجره) ──
            tk.Frame(pop, bg=C["border"], height=1).pack(fill="x")
            btn_bar = tk.Frame(pop, bg=C["header_bg"])
            btn_bar.pack(fill="x")
            tk.Button(btn_bar, text="  ✕  انصراف  ",
                      command=pop.destroy,
                      bg=C["btn_danger"], fg="#ffffff",
                      font=(_MAIN_FONT, 11, "bold"),
                      bd=0, relief="flat", cursor="hand2",
                      padx=14, pady=12).pack(side="left", padx=12, pady=10)

            save_btn_f = tk.Frame(btn_bar, bg=C["btn_success"])
            save_btn_f.pack(side="right", padx=12, pady=10)
            save_btn_l = tk.Label(save_btn_f,
                                   text="  ✔  ثبت ویرایش  ",
                                   bg=C["btn_success"], fg="#ffffff",
                                   font=(_MAIN_FONT, 13, "bold"),
                                   padx=20, pady=12, cursor="hand2")
            save_btn_l.pack()
            save_btn_f.bind("<Button-1>", lambda e: do_edit_save())
            save_btn_l.bind("<Button-1>", lambda e: do_edit_save())
            lighter_sv = _lighten(C["btn_success"], 20)
            save_btn_f.bind("<Enter>", lambda e: [save_btn_f.config(bg=lighter_sv),
                                                   save_btn_l.config(bg=lighter_sv)])
            save_btn_f.bind("<Leave>", lambda e: [save_btn_f.config(bg=C["btn_success"]),
                                                   save_btn_l.config(bg=C["btn_success"])])
            pop.bind("<Escape>", lambda e: pop.destroy())
            pop.bind("<Return>",  lambda e: do_edit_save())
            finalize_popup_window(pop, self)

        def refresh():
            tree2.delete(*tree2.get_children())
            _db2 = load_db()
            for rec in reversed(_db2.get("scarf_cut",[])):
                if rec.get("operation") != op_name: continue
                who    = get_display_name(rec.get("registered_by","—"), _db2) if self.role=="admin" else "شخص دیگر"
                _ec    = rec.get("edit_count", 1 if rec.get("edited") else 0)
                _MAX   = 3
                if self.role == "admin":
                    edited = False
                    edit_txt = f"✏  ویرایش  ({_ec})"
                else:
                    edited = (_ec >= _MAX)
                    if edited:
                        edit_txt = "🔒 قفل (۳/۳)"
                    else:
                        edit_txt = f"✏  ویرایش  ({_ec}/{_MAX})"
                if op_name == "برشی":
                    vals = (
                        rec["slab_id"],
                        rec.get("reason","—"),
                        str(rec.get("cut_count",1)),
                        "✔" if rec.get("bauman_done") else "—",
                        who,
                        *split_dt(rec.get("registered_at","—")),
                        edit_txt,
                    )
                else:
                    vals = (
                        rec["slab_id"],
                        rec.get("reason","—"),
                        who,
                        *split_dt(rec.get("registered_at","—")),
                        edit_txt,
                    )
                tag = "edited_tag" if edited else clr_tag
                tree2.insert("","end", values=vals, tags=(tag,))
            tree2.tag_configure(clr_tag,    background=clr_bg2, foreground=active_color)
            tree2.tag_configure("edited_tag", background=clr_bg2, foreground=C["text_dim"])

        def on_tree2_click(event):
            row_id = tree2.identify_row(event.y)
            if not row_id: return
            try:
                clicked = tree2.column(tree2.identify_column(event.x), "id")
            except: return
            if clicked == "edit_col":
                vals = tree2.item(row_id, "values")
                open_edit_popup(vals[0], None)

        tree2.bind("<Button-1>", on_tree2_click)

        refresh()
        tab.refresh = refresh
        self._bind_admin_popup(tree2, "scarf_cut", refresh,
                                extra_key="operation", extra_col_idx=None)
        if self.role == "admin":
            tk.Label(sf, text="👑 ادمین: دابل‌کلیک روی هر ردیف برای ویرایش/حذف کامل",
                     bg=BG, fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=18, pady=(4,8))



    # ═══════════════════════════════════════════════════════════
    #  تب صندوق نامه (تیکت)
    # ═══════════════════════════════════════════════════════════
    def _build_ticket_tab(self, tab):
        """
        صندوق نامه داخلی:
        - کاربران عادی: ارسال نامه به مدیریت + نامه‌های دریافتی
        - ادمین: مشاهده همه نامه‌ها + ارسال نامه به هر کاربر
        """
        tab.configure(bg=C["panel"])
        is_admin = (self.role == "admin")

        # ── اسکرول ──
        canvas = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a",
                           activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        sf = tk.Frame(canvas, bg=C["panel"])
        _wid = canvas.create_window((0, 0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_wid, width=e.width))
        register_scroll_canvas(canvas, sf)

        # ── عنوان ──
        tk.Label(sf, text="✉  صندوق نامه",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=(12,2))
        if is_admin:
            tk.Label(sf, text="مشاهده همه نامه‌ها  |  ارسال نامه به هر کاربر",
                     bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)
        else:
            tk.Label(sf, text="ارسال نامه به مدیریت سیستم  |  مشاهده نامه‌های دریافتی",
                     bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        # ══════════════════════════════════════════
        #  فرم ارسال نامه
        # ══════════════════════════════════════════
        compose_card = card_frame(sf)
        compose_card.pack(fill="x", padx=16, pady=10)
        tk.Frame(compose_card, bg=C["accent"], height=3).pack(fill="x")
        compose_in = tk.Frame(compose_card, bg=C["card"])
        compose_in.pack(padx=16, pady=12, fill="x")
        tk.Label(compose_in, text="📝  نوشتن نامه جدید",
                 bg=C["card"], fg=C["accent"], font=(_MAIN_FONT, 11, "bold")).pack(anchor="e", pady=(0,8))

        # گیرنده (ادمین: کشویی، بقیه: ثابت = مدیریت سیستم)
        to_row = tk.Frame(compose_in, bg=C["card"])
        to_row.pack(fill="x", pady=4)
        tk.Label(to_row, text="گیرنده:", bg=C["card"], fg=C["text"],
                 font=FONT_NORM, width=10, anchor="e").pack(side="right", padx=(0,6))

        if is_admin:
            db0 = load_db()
            _recipients = ["همه کاربران"] + [
                f"{ud.get('display', un)}  [{un}]"
                for un, ud in db0["users"].items()
                if un != "admin"
            ]
            to_var = tk.StringVar(value=_recipients[0] if _recipients else "")
            to_cb = make_combo(to_row, _recipients, width=28)
            to_cb.set(_recipients[0] if _recipients else "")
            to_cb.pack(side="right")
        else:
            to_var = tk.StringVar(value="مدیریت سیستم")
            tk.Label(to_row, text="مدیریت سیستم",
                     bg=C["card2"], fg=C["gold"],
                     font=(_MAIN_FONT, 11, "bold"),
                     padx=10, pady=4).pack(side="right")

        # موضوع
        subj_row = tk.Frame(compose_in, bg=C["card"])
        subj_row.pack(fill="x", pady=4)
        tk.Label(subj_row, text="موضوع:", bg=C["card"], fg=C["text"],
                 font=FONT_NORM, width=10, anchor="e").pack(side="right", padx=(0,6))
        subj_var = tk.StringVar()
        tk.Entry(subj_row, textvariable=subj_var,
                 bg=C["entry_bg"], fg=C["text"],
                 insertbackground=C["accent"], font=FONT_NORM,
                 justify="right", bd=0, relief="flat",
                 highlightthickness=1, highlightbackground=C["border"],
                 highlightcolor=C["accent"], width=40).pack(side="right", fill="x", expand=True)

        # متن نامه
        tk.Label(compose_in, text="متن نامه:", bg=C["card"],
                 fg=C["text"], font=FONT_NORM).pack(anchor="e", pady=(8,2))
        txt_frame = tk.Frame(compose_in, bg=C["accent"], padx=1, pady=1)
        txt_frame.pack(fill="x")
        body_txt = tk.Text(txt_frame, height=5,
                            bg=C["entry_bg"], fg=C["text"],
                            insertbackground=C["accent"],
                            font=(_MAIN_FONT, 11, "bold"),
                            bd=0, relief="flat",
                            wrap="word")
        body_txt.pack(fill="x", padx=2, pady=2)

        send_status = tk.Label(compose_in, text="", bg=C["card"],
                                fg=C["success"], font=FONT_SMALL)
        send_status.pack(anchor="e", pady=4)

        def do_send():
            subj = subj_var.get().strip()
            body = body_txt.get("1.0", "end").strip()
            if not subj:
                messagebox.showwarning("خطا", "موضوع نامه را وارد کنید.", parent=self)
                return
            if not body:
                messagebox.showwarning("خطا", "متن نامه را وارد کنید.", parent=self)
                return

            db = load_db()
            ts  = now_str()

            if is_admin:
                sel = to_cb.get().strip()
                if sel == "همه کاربران":
                    # ارسال به همه
                    targets = [un for un in db["users"] if un != "admin"]
                else:
                    # استخراج نام کاربری از "نام [username]"
                    import re as _re
                    m = _re.search(r'\[(\w+)\]', sel)
                    targets = [m.group(1)] if m else []
            else:
                targets = ["admin"]

            if not targets:
                messagebox.showwarning("خطا", "گیرنده‌ای انتخاب نشده.", parent=self)
                return

            ticket_base = {
                "from":    self.username,
                "from_display": self.udata.get("display", self.username),
                "subject": subj,
                "body":    body,
                "sent_at": ts,
                "read":    False,
            }

            for to_un in targets:
                t = dict(ticket_base)
                t["to"] = to_un
                t["to_display"] = db["users"].get(to_un, {}).get("display", to_un)
                db.setdefault("tickets", []).append(t)

            save_db(db)
            subj_var.set("")
            body_txt.delete("1.0", "end")
            n = len(targets)
            send_status.config(
                text=f"✔  نامه برای {n} نفر ارسال شد  |  {ts}")
            refresh_inbox()
            refresh_sent()
            if is_admin:
                refresh_all()

        # دکمه ارسال
        send_btn_f = tk.Frame(compose_in, bg=C["btn_success"], cursor="hand2")
        send_btn_f.pack(anchor="w", pady=(4,0))
        send_btn_l = tk.Label(send_btn_f, text="  ✉  ارسال نامه  ",
                               bg=C["btn_success"], fg="#ffffff",
                               font=(_MAIN_FONT, 12, "bold"),
                               padx=16, pady=10, cursor="hand2")
        send_btn_l.pack()
        _sbg = _lighten(C["btn_success"], 20)
        send_btn_f.bind("<Button-1>", lambda e: do_send())
        send_btn_l.bind("<Button-1>", lambda e: do_send())
        send_btn_f.bind("<Enter>", lambda e: [send_btn_f.config(bg=_sbg), send_btn_l.config(bg=_sbg)])
        send_btn_f.bind("<Leave>", lambda e: [send_btn_f.config(bg=C["btn_success"]), send_btn_l.config(bg=C["btn_success"])])

        # ══════════════════════════════════════════
        #  صندوق دریافت (نامه‌های دریافتی)
        # ══════════════════════════════════════════
        inbox_card = card_frame(sf)
        inbox_card.pack(fill="x", padx=16, pady=(8,4))
        tk.Frame(inbox_card, bg=C["success"], height=3).pack(fill="x")
        inbox_in = tk.Frame(inbox_card, bg=C["card"])
        inbox_in.pack(padx=16, pady=10, fill="x")
        inbox_hdr = tk.Frame(inbox_in, bg=C["card"])
        inbox_hdr.pack(fill="x", pady=(0,6))
        tk.Label(inbox_hdr, text="📥  نامه‌های دریافتی",
                 bg=C["card"], fg=C["success"],
                 font=(_MAIN_FONT, 11, "bold")).pack(side="right")
        inbox_cnt = tk.Label(inbox_hdr, text="",
                              bg=C["card"], fg=C["warning"],
                              font=(_MAIN_FONT, 9, "bold"))
        inbox_cnt.pack(side="left")

        inbox_cols  = ("from_d","subject","sent_at","body_preview","status")
        inbox_heads = ("فرستنده","موضوع","تاریخ ارسال","پیش‌نمایش متن","وضعیت")
        tf_in2, tree_inbox = scrolled_tree(inbox_in, inbox_cols, inbox_heads, height=6)
        tf_in2.pack(fill="x", pady=4)
        tree_inbox.column("from_d",       width=130, anchor="center")
        tree_inbox.column("subject",      width=200, anchor="e")
        tree_inbox.column("sent_at",      width=170, anchor="center")
        tree_inbox.column("body_preview", width=280, anchor="e")
        tree_inbox.column("status",       width=80,  anchor="center")

        # نمایش متن کامل با دابل‌کلیک
        def on_inbox_dbl(event):
            sel = tree_inbox.selection()
            if not sel: return
            vals = tree_inbox.item(sel[0], "values")
            # شماره ایندکس از tag
            iid = sel[0]
            idx = tree_inbox.item(iid, "tags")
            if idx:
                try:
                    _show_ticket_detail(int(idx[0]))
                except (ValueError, IndexError):
                    pass

        tree_inbox.bind("<Double-Button-1>", on_inbox_dbl)

        def _show_ticket_detail(ticket_idx):
            db = load_db()
            tickets = db.get("tickets", [])
            if ticket_idx >= len(tickets): return
            t = tickets[ticket_idx]
            # علامت‌گذاری به عنوان خوانده‌شده
            if not t.get("read") and t.get("to") == self.username:
                tickets[ticket_idx]["read"] = True
                save_db(db)
                refresh_inbox()

            pop = tk.Toplevel(self)
            prepare_popup_window(pop, self)
            pop.title("نامه داخلی")
            pop.configure(bg=C["card"])
            pop.focus_force()
            pop.geometry("560x420")
            self._center(pop, 560, 420)
            tk.Frame(pop, bg=C["accent"], height=3).pack(fill="x")
            hf = tk.Frame(pop, bg=C["header_bg"])
            hf.pack(fill="x")
            tk.Label(hf, text="✉  نامه داخلی",
                     bg=C["header_bg"], fg=C["accent"],
                     font=FONT_HEAD).pack(side="right", padx=16, pady=10)
            tk.Button(hf, text="  ✕  ",
                      command=pop.destroy,
                      bg=C["header_bg"], fg=C["text_dim"],
                      font=FONT_NORM, bd=0, relief="flat",
                      cursor="hand2").pack(side="left", padx=8)

            body_f = tk.Frame(pop, bg=C["card"])
            body_f.pack(fill="both", expand=True, padx=20, pady=14)

            for lbl, val in [
                ("از:", t.get("from_display", t.get("from","—"))),
                ("به:", t.get("to_display", t.get("to","—"))),
                ("موضوع:", t.get("subject","—")),
                ("تاریخ:", t.get("sent_at","—")),
            ]:
                r = tk.Frame(body_f, bg=C["card"])
                r.pack(fill="x", pady=2)
                tk.Label(r, text=lbl, bg=C["card"], fg=C["text_dim"],
                         font=FONT_SMALL, width=8, anchor="e").pack(side="right")
                tk.Label(r, text=val, bg=C["card2"], fg=C["text"],
                         font=FONT_NORM, anchor="e", justify="right",
                         padx=8, pady=3, wraplength=420).pack(side="right", fill="x", expand=True)

            tk.Frame(body_f, bg=C["border"], height=1).pack(fill="x", pady=8)
            tk.Label(body_f, text="متن نامه:", bg=C["card"],
                     fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e")
            txt_disp = tk.Text(body_f, height=8, bg=C["entry_bg"], fg=C["text"],
                                font=(_MAIN_FONT, 11, "bold"),
                                bd=0, relief="flat", wrap="word",
                                state="normal")
            txt_disp.insert("end", t.get("body",""))
            txt_disp.configure(state="disabled")
            txt_disp.pack(fill="both", expand=True)
            pop.bind("<Escape>", lambda e: pop.destroy())

        def refresh_inbox():
            tree_inbox.delete(*tree_inbox.get_children())
            db = load_db()
            tickets = db.get("tickets", [])
            unread = 0
            for i, t in enumerate(reversed(tickets)):
                real_idx = len(tickets) - 1 - i
                if t.get("to") != self.username: continue
                is_unread = not t.get("read", False)
                if is_unread: unread += 1
                status = "🔵 جدید" if is_unread else "✔ خوانده"
                preview = t.get("body","")[:40].replace("\n"," ")
                tag = "unread" if is_unread else "read"
                iid = tree_inbox.insert("", "end", values=(
                    t.get("from_display", t.get("from","—")),
                    t.get("subject","—"),
                    t.get("sent_at","—"),
                    preview,
                    status,
                ), tags=(str(real_idx), tag))
            tree_inbox.tag_configure("unread", foreground=C["warning"],
                                      background="#1a1500")
            tree_inbox.tag_configure("read",   foreground=C["text"],
                                      background=C["card2"])
            inbox_cnt.config(
                text=f"📬 {unread} نامه خوانده‌نشده" if unread else "")

        # ══════════════════════════════════════════
        #  صندوق ارسال‌شده‌ها
        # ══════════════════════════════════════════
        sent_card = card_frame(sf)
        sent_card.pack(fill="x", padx=16, pady=(4,8))
        tk.Frame(sent_card, bg=C["btn_primary"], height=3).pack(fill="x")
        sent_in = tk.Frame(sent_card, bg=C["card"])
        sent_in.pack(padx=16, pady=10, fill="x")
        tk.Label(sent_in, text="📤  نامه‌های ارسال‌شده",
                 bg=C["card"], fg=C["accent"],
                 font=(_MAIN_FONT, 11, "bold")).pack(anchor="e", pady=(0,6))

        sent_cols  = ("to_d","subject","sent_at","body_preview")
        sent_heads = ("گیرنده","موضوع","تاریخ ارسال","پیش‌نمایش")
        tf_sent, tree_sent = scrolled_tree(sent_in, sent_cols, sent_heads, height=5)
        tf_sent.pack(fill="x", pady=4)
        tree_sent.column("to_d",         width=130, anchor="center")
        tree_sent.column("subject",      width=200, anchor="e")
        tree_sent.column("sent_at",      width=170, anchor="center")
        tree_sent.column("body_preview", width=280, anchor="e")

        def on_sent_dbl(event):
            sel = tree_sent.selection()
            if not sel: return
            idx = tree_sent.item(sel[0], "tags")
            if idx:
                try: _show_ticket_detail(int(idx[0]))
                except: pass
        tree_sent.bind("<Double-Button-1>", on_sent_dbl)

        def refresh_sent():
            tree_sent.delete(*tree_sent.get_children())
            db = load_db()
            tickets = db.get("tickets", [])
            for i, t in enumerate(reversed(tickets)):
                real_idx = len(tickets) - 1 - i
                if t.get("from") != self.username: continue
                preview = t.get("body","")[:40].replace("\n"," ")
                tree_sent.insert("", "end", values=(
                    t.get("to_display", t.get("to","—")),
                    t.get("subject","—"),
                    t.get("sent_at","—"),
                    preview,
                ), tags=(str(real_idx),))

        # ══════════════════════════════════════════
        #  همه نامه‌ها (فقط ادمین)
        # ══════════════════════════════════════════
        if is_admin:
            all_card = card_frame(sf)
            all_card.pack(fill="x", padx=16, pady=(4,8))
            tk.Frame(all_card, bg=C["gold"], height=3).pack(fill="x")
            all_in = tk.Frame(all_card, bg=C["card"])
            all_in.pack(padx=16, pady=10, fill="x")
            all_hdr = tk.Frame(all_in, bg=C["card"])
            all_hdr.pack(fill="x", pady=(0,6))
            tk.Label(all_hdr, text="👑  همه نامه‌های سیستم",
                     bg=C["card"], fg=C["gold"],
                     font=(_MAIN_FONT, 11, "bold")).pack(side="right")
            all_cnt = tk.Label(all_hdr, text="", bg=C["card"],
                                fg=C["warning"], font=FONT_SMALL)
            all_cnt.pack(side="left")

            all_cols  = ("from_d","to_d","subject","sent_at","status","body_preview")
            all_heads = ("فرستنده","گیرنده","موضوع","تاریخ","وضعیت","پیش‌نمایش")
            tf_all, tree_all = scrolled_tree(all_in, all_cols, all_heads, height=7)
            tf_all.pack(fill="x", pady=4)
            tree_all.column("from_d",       width=120, anchor="center")
            tree_all.column("to_d",         width=120, anchor="center")
            tree_all.column("subject",      width=180, anchor="e")
            tree_all.column("sent_at",      width=160, anchor="center")
            tree_all.column("status",       width=80,  anchor="center")
            tree_all.column("body_preview", width=250, anchor="e")

            def on_all_dbl(event):
                sel = tree_all.selection()
                if not sel: return
                idx = tree_all.item(sel[0], "tags")
                if idx:
                    try: _show_ticket_detail(int(idx[0]))
                    except: pass
            tree_all.bind("<Double-Button-1>", on_all_dbl)

            def refresh_all():
                tree_all.delete(*tree_all.get_children())
                db = load_db()
                tickets = db.get("tickets", [])
                unread_total = 0
                for i, t in enumerate(reversed(tickets)):
                    real_idx = len(tickets) - 1 - i
                    is_unread = not t.get("read", False)
                    if is_unread: unread_total += 1
                    status = "🔵 جدید" if is_unread else "✔ خوانده"
                    preview = t.get("body","")[:35].replace("\n"," ")
                    tag = "unread_a" if is_unread else "read_a"
                    tree_all.insert("", "end", values=(
                        t.get("from_display", t.get("from","—")),
                        t.get("to_display",   t.get("to","—")),
                        t.get("subject","—"),
                        t.get("sent_at","—"),
                        status,
                        preview,
                    ), tags=(str(real_idx), tag))
                tree_all.tag_configure("unread_a", foreground=C["warning"],
                                        background="#1a1500")
                tree_all.tag_configure("read_a",   foreground=C["text"],
                                        background=C["card2"])
                all_cnt.config(
                    text=f"📬 {unread_total} نامه خوانده‌نشده  |  مجموع: {len(tickets)}"
                    if tickets else "")

            def delete_ticket():
                sel = tree_all.selection()
                if not sel: return
                idx = tree_all.item(sel[0], "tags")
                if not idx: return
                try:
                    real_idx = int(idx[0])
                except: return
                if not messagebox.askyesno("حذف نامه",
                        "آیا مطمئنید این نامه حذف شود؟", parent=self): return
                db = load_db()
                tickets = db.get("tickets", [])
                if 0 <= real_idx < len(tickets):
                    tickets.pop(real_idx)
                    save_db(db)
                    refresh_all()
                    refresh_inbox()
                    refresh_sent()

            del_btn = tk.Frame(all_in, bg=C["btn_danger"], cursor="hand2")
            del_btn.pack(anchor="w", pady=(6,0))
            del_lbl = tk.Label(del_btn, text="  🗑  حذف نامه انتخابی  ",
                                bg=C["btn_danger"], fg="#ffffff",
                                font=(_MAIN_FONT, 10, "bold"),
                                padx=10, pady=6, cursor="hand2")
            del_lbl.pack()
            del_btn.bind("<Button-1>", lambda e: delete_ticket())
            del_lbl.bind("<Button-1>", lambda e: delete_ticket())

            refresh_all()

        # ── بارگذاری اولیه ──
        refresh_inbox()
        refresh_sent()

        # ── بروزرسانی خودکار هر ۱۰ ثانیه ──
        def _auto_refresh():
            try:
                if tab.winfo_exists():
                    refresh_inbox()
                    refresh_sent()
                    if is_admin:
                        refresh_all()
                    tab.after(10000, _auto_refresh)
            except Exception:
                pass
        tab.after(10000, _auto_refresh)

    def _build_bauman_tab(self, tab):
        tab.configure(bg=C["panel"])
        form = card_frame(tab)
        form.pack(fill="x", padx=16, pady=12)
        inner = tk.Frame(form, bg=C["card"])
        inner.pack(padx=16, pady=12, fill="x")
        tk.Label(inner, text="🔬  ثبت سر اسلب باومن (برش تست)",
                 bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0,8))
        tk.Label(inner, text="فقط اسلب‌هایی ثبت شوند که دلیل برش = باومن بوده است.",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e")

        r1 = tk.Frame(inner, bg=C["card"])
        r1.pack(fill="x", pady=6)
        tk.Label(r1, text="شماره اسلب:", bg=C["card"], fg=C["text"],
                 font=FONT_NORM).pack(side="right", padx=(0,6))
        sid_var = tk.StringVar()
        sid_ent = tk.Entry(r1, textvariable=sid_var, bg=C["entry_bg"], fg=C["text"],
                            insertbackground=C["accent"], font=FONT_MONO, justify="right", bd=0, relief="flat",
                            highlightthickness=1, highlightbackground=C["border"],
                            highlightcolor=C["accent"], width=16)
        sid_ent.pack(side="right")

        tk.Label(inner, text="یا انتخاب از برش‌های باومن ثبت‌شده:",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(4,0))
        bauman_cb = make_combo(inner, [], width=24)
        bauman_cb.pack(anchor="e", pady=2)

        def load_bauman_list():
            db = load_db()
            opts = [r["slab_id"] for r in db["scarf_cut"] if r.get("reason") == "باومن"]
            bauman_cb.config(values=opts)
        load_bauman_list()

        def pick_bauman(e):
            sid_var.set(bauman_cb.get())
        bauman_cb.bind("<<ComboboxSelected>>", pick_bauman)

        status_lbl = tk.Label(inner, text="", bg=C["card"], fg=C["success"], font=FONT_SMALL)
        status_lbl.pack(anchor="e")

        def do_register():
            sid = sid_var.get().strip()
            ok, msg, sid = validate_slab_id(sid)
            if not ok:
                messagebox.showerror("خطا", msg, parent=self)
                return
            db = load_db()
            if check_duplicate(db, "bauman", sid):
                messagebox.showerror("⛔  ثبت تکراری",
                    f"اسلب {sid} قبلاً توسط شخص دیگری ثبت شده است.\n\nبرای پیدا کردن از جستجو استفاده کنید.\nدر غیر این صورت با سرپرست کارگاه تماس بگیرید.", parent=self)
                return
            sc = next((r for r in db["scarf_cut"] if r["slab_id"]==sid and r.get("reason")=="باومن"), None)
            if not sc and not messagebox.askyesno("هشدار",
                f"اسلب {sid} در لیست برش‌های باومن یافت نشد.\nآیا مطمئن هستید که ادامه دهید?", parent=self):
                return
            db["bauman"].append({
                "slab_id": sid, "cut_by": self.username,
                "cut_at": now_str(), "lab_status": "در انتظار"
            })
            save_db(db)
            sid_var.set("")
            status_lbl.config(text=f"✔  اسلب {sid} در لیست باومن ثبت شد.")
            load_bauman_list()
            refresh()

        styled_btn(inner, "🔬  ثبت باومن", do_register).pack(anchor="e", pady=6)

        separator(tab)
        cols = ("slab_id","cut_by","cut_date","cut_time","lab_status","delivered_by","del_date","del_time")
        heads = ("شماره اسلب","برش‌کار","تاریخ برش","ساعت برش","وضعیت","تحویل‌دهنده","تاریخ تحویل","ساعت تحویل")
        tf, tree = scrolled_tree(tab, cols, heads, height=12)
        tf.pack(fill="both", expand=True, padx=16, pady=4)
        sb = search_bar(tab, tree)
        sb.pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            for rec in db["bauman"]:
                status = rec.get("lab_status","—")
                clr = "done" if status=="تحویل داده شده" else "wait"
                who_cut = get_display_name(rec.get("cut_by","—"), db) if self.role=="admin" else "شخص دیگر"
                who_del = get_display_name(rec.get("delivered_by","—"), db) if self.role=="admin" else "شخص دیگر"
                tree.insert("", "end", values=(
                    rec["slab_id"], who_cut, *split_dt(rec.get("cut_at","—")),
                    status, who_del, *split_dt(rec.get("delivered_at","—"))
                ), tags=(clr,))
            tree.tag_configure("done", background="#3a5040", foreground=C["success"])
            tree.tag_configure("wait", background="#1a1a00", foreground=C["warning"])
        refresh()
        self._bind_admin_popup(tree, "bauman", refresh)
        if self.role == "admin":
            tk.Label(tab, text="👑 دابل‌کلیک یا راست‌کلیک روی هر ردیف برای ویرایش/حذف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)

    # ═══════════════════════════════════════════════════════════
    #  تب: نوبت‌کار  (خروج اسلب — دو سربرگ)
    # ═══════════════════════════════════════════════════════════
    def _build_nobat_tab(self, tab):
        tab.configure(bg=C["panel"])

        # ── ساخت Sub-Notebook با دو سربرگ ──
        style = ttk.Style()
        style.configure("Nobat.TNotebook", background=C["panel"], bordercolor=C["border"])
        style.configure("Nobat.TNotebook.Tab",
            background=C["tab_inactive"], foreground=C["text"],
            font=(_MAIN_FONT, 11, "bold"), padding=[14, 7])
        style.map("Nobat.TNotebook.Tab",
            background=[("selected", C["accent"])],
            foreground=[("selected", "#ffffff")])

        sub_nb = ttk.Notebook(tab, style="Nobat.TNotebook")
        sub_nb.pack(fill="both", expand=True, padx=6, pady=6)

        tab_manual = tk.Frame(sub_nb, bg=C["panel"])
        tab_file   = tk.Frame(sub_nb, bg=C["panel"])
        sub_nb.add(tab_manual, text="🚀  خروج دستی")
        sub_nb.add(tab_file,   text="📂  خروج از طریق فایل")

        # ══════════════════════════════════════
        #  سربرگ ۱: خروج دستی (کد اصلی قبلی)
        # ══════════════════════════════════════
        tk.Label(tab_manual, text="🔄  خروج اسلب‌های کنترل کیفی شده از انبار",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=10)
        tk.Label(tab_manual, text="فقط اسلب‌هایی که وضعیت کنترل کیفی = کنترل کیفی شده دارند نمایش داده می‌شوند.",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        cols = ("slab_id","qc_by_anon","qc_date","qc_time","warehouse","exit_status","exit_by","exit_date","exit_time")
        heads = ("شماره اسلب","تأییدکننده QC","تاریخ QC","ساعت QC","انبار","وضعیت خروج","خروج‌دهنده","تاریخ خروج","ساعت خروج")
        tf, tree = scrolled_tree(tab_manual, cols, heads, height=14)
        tf.pack(fill="both", expand=True, padx=16, pady=8)
        sb = search_bar(tab_manual, tree)
        sb.pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab_manual, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        status_lbl = tk.Label(tab_manual, text="", bg=C["panel"], fg=C["success"], font=FONT_NORM)
        status_lbl.pack(anchor="e", padx=16)

        def do_exit():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک اسلب انتخاب کنید.", parent=self)
                return
            iid = sel[0]
            vals = tree.item(iid, "values")
            sid = vals[0]
            if vals[5] == "خروج زده شده":
                messagebox.showerror("⛔  ثبت تکراری", f"اسلب {sid} قبلاً توسط شخص دیگری ثبت شده است.\n\nبرای پیدا کردن از جستجو استفاده کنید.\nدر غیر این صورت با سرپرست کارگاه تماس بگیرید.", parent=self)
                return
            db = load_db()
            for rec in db["melts"]:
                if rec["slab_id"] == sid:
                    rec["exit_status"] = "خروج زده شده"
                    rec["exit_by"] = self.username
                    rec["exit_at"] = now_str()
                    break
            save_db(db)
            status_lbl.config(text=f"✔  خروج اسلب {sid} ثبت شد.")
            refresh()

        ctrl = tk.Frame(tab_manual, bg=C["panel"])
        ctrl.pack(fill="x", padx=16, pady=4)
        styled_btn(ctrl, "🚀  خروج اسلب", do_exit, color=C["btn_success"]).pack(side="right")

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            for rec in db["melts"]:
                if rec.get("qc_status") == "کنترل کیفی شده":
                    qc_who = get_display_name(rec.get("qc_by","—"), db) if self.role == "admin" else "شخص دیگر"
                    exit_who = get_display_name(rec.get("exit_by","—"), db) if self.role == "admin" else "شخص دیگر"
                    wh = get_current_location(db, rec["slab_id"])
                    exit_s = rec.get("exit_status","در انبار")
                    if exit_s != "خروج زده شده":
                        exit_s = f"در {wh}"
                    clr = "exited" if exit_s=="خروج زده شده" else "instock"
                    tree.insert("", "end", values=(
                        rec["slab_id"], qc_who, *split_dt(rec.get("qc_at","—")), wh,
                        exit_s, exit_who, *split_dt(rec.get("exit_at","—"))
                    ), tags=(clr,))
            tree.tag_configure("exited",  background="#3a5040", foreground=C["success"])
            tree.tag_configure("instock", background="#1a2000", foreground=C["warning"])
        refresh()
        self._bind_admin_popup(tree, "melts", refresh)
        if self.role == "admin":
            tk.Label(tab_manual, text="👑 دابل‌کلیک یا راست‌کلیک روی هر ردیف برای ویرایش/حذف",
                     bg=C["panel"], fg=C["gold"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=2)

        # ══════════════════════════════════════════════════════
        #  سربرگ ۲: خروج از طریق فایل اکسل
        # ══════════════════════════════════════════════════════
        self._build_file_exit_tab(tab_file)

    # ─────────────────────────────────────────────────────────
    def _build_file_exit_tab(self, tab):
        """سربرگ خروج از طریق فایل — با تاریخچه کامل و گزارش PDF"""
        tab.configure(bg=C["panel"])

        tk.Label(tab, text="📂  خروج از طریق فایل اکسل",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=(10,2))
        tk.Label(tab,
                 text="فایل اکسل با ستون‌های ردیف و شماره اسلب را آپلود کنید — خروج اسلب‌های QC‌شده به صورت دسته‌ای ثبت می‌شود.",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        # ── دکمه‌های بالا ──
        top_bar = tk.Frame(tab, bg=C["panel"])
        top_bar.pack(fill="x", padx=16, pady=8)

        status_upload = tk.Label(top_bar, text="", bg=C["panel"],
                                 fg=C["success"], font=FONT_SMALL, anchor="e")
        status_upload.pack(side="right", padx=(8,0))

        def _download_template():
            """ساخت فایل اکسل الگو برای دانلود"""
            if not XLSX:
                messagebox.showerror("خطا","کتابخانه openpyxl نصب نیست.\npip install openpyxl", parent=self)
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile="الگوی_خروج_اسلب.xlsx",
                title="ذخیره فایل الگو",
                parent=self)
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "خروج اسلب"
            ws.sheet_view.rightToLeft = True

            hdr_fill = PatternFill("solid", fgColor="003366")
            hdr_font = Font(name="B Nazanin", bold=True, color="FFFFFF", size=12)
            cell_font = Font(name="B Nazanin", size=11)
            center    = Alignment(horizontal="center", vertical="center", readingOrder=2)
            right_al  = Alignment(horizontal="right",  vertical="center", readingOrder=2)
            thin      = Side(style="thin", color="AAAAAA")
            border    = Border(left=thin,right=thin,top=thin,bottom=thin)

            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 22
            ws.row_dimensions[1].height = 28

            for col, (val, al) in enumerate(
                    [("ردیف", center), ("شماره اسلب", right_al)], start=1):
                c = ws.cell(row=1, column=col, value=val)
                c.fill   = hdr_fill
                c.font   = hdr_font
                c.alignment = al
                c.border = border

            for i in range(1, 6):
                for col, (val, al) in enumerate(
                        [(i, center), ("", right_al)], start=1):
                    c = ws.cell(row=i+1, column=col, value=val)
                    c.font      = cell_font
                    c.alignment = al
                    c.border    = border
                    if col == 1:
                        c.fill = PatternFill("solid", fgColor="EEF2FF")

            wb.save(path)
            messagebox.showinfo("موفق", f"فایل الگو ذخیره شد:\n{path}", parent=self)

        styled_btn(top_bar, "⬇️  دانلود فایل الگو", _download_template,
                   color="#4a6080", width=160).pack(side="right", padx=(0,4))

        def _upload_and_process():
            """آپلود فایل اکسل و پردازش خروج اسلب‌ها"""
            if not XLSX:
                messagebox.showerror("خطا","کتابخانه openpyxl نصب نیست.\npip install openpyxl", parent=self)
                return
            path = filedialog.askopenfilename(
                filetypes=[("Excel","*.xlsx *.xls")],
                title="انتخاب فایل اکسل خروج اسلب",
                parent=self)
            if not path:
                return

            # ── خواندن شماره اسلب‌ها از فایل ──
            try:
                wb_in = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws_in = wb_in.active
                rows_data = list(ws_in.iter_rows(min_row=2, values_only=True))
                wb_in.close()
            except Exception as ex:
                messagebox.showerror("خطای خواندن فایل", str(ex), parent=self)
                return

            # ستون دوم = شماره اسلب (ایندکس 1)
            slabs_in_file = []
            for row in rows_data:
                if len(row) >= 2 and row[1] is not None:
                    val = str(row[1]).strip()
                    if val:
                        slabs_in_file.append(val)

            if not slabs_in_file:
                messagebox.showwarning("فایل خالی",
                    "هیچ شماره اسلبی در ستون دوم فایل یافت نشد.", parent=self)
                return

            # ── پردازش ──
            db          = load_db()
            now_ts      = now_str()
            file_name   = os.path.basename(path)

            # ذخیره محتوای فایل آپلودی (base64) برای دانلود آینده
            try:
                import base64
                with open(path, "rb") as _f:
                    file_b64 = base64.b64encode(_f.read()).decode("utf-8")
            except Exception:
                file_b64 = ""

            registered_slabs = []   # ثبت موفق
            duplicate_slabs  = []   # قبلاً خروج خورده
            not_found_slabs  = []   # در QC نبود

            qc_slab_ids = {r["slab_id"] for r in db["melts"]
                           if r.get("qc_status") == "کنترل کیفی شده"}

            for sid in slabs_in_file:
                if sid not in qc_slab_ids:
                    not_found_slabs.append(sid)
                    continue
                for rec in db["melts"]:
                    if rec["slab_id"] == sid:
                        if rec.get("exit_status") == "خروج زده شده":
                            duplicate_slabs.append(sid)
                        else:
                            rec["exit_status"] = "خروج زده شده"
                            rec["exit_by"]     = self.username
                            rec["exit_at"]     = now_ts
                            rec["exit_source"] = "فایل"   # نشانگر منبع خروج
                            registered_slabs.append(sid)
                        break

            # ── ثبت در تاریخچه (file_exit_log) ──
            log_entry = {
                "batch_id":        f"FE_{now_ts.replace('/','').replace(' ','_').replace(':','')}",
                "uploaded_at":     now_ts,
                "uploaded_by":     self.username,
                "file_name":       file_name,
                "file_b64":        file_b64,
                "slabs_total":     len(slabs_in_file),
                "registered":      registered_slabs,
                "duplicates":      duplicate_slabs,
                "not_found":       not_found_slabs,
            }
            if "file_exit_log" not in db:
                db["file_exit_log"] = []
            db["file_exit_log"].append(log_entry)
            save_db(db)

            status_upload.config(
                text=f"✔  پردازش شد — ثبت: {len(registered_slabs)}  |  تکراری: {len(duplicate_slabs)}  |  پیدا نشد: {len(not_found_slabs)}")
            refresh_history()

            # پیشنهاد گزارش PDF
            msg = (f"نتیجه پردازش فایل:\n\n"
                   f"✅  خروج ثبت شد:  {len(registered_slabs)} اسلب\n"
                   f"⚠️  تکراری (قبلاً خروج خورده):  {len(duplicate_slabs)} اسلب\n"
                   f"❌  در بخش QC یافت نشد:  {len(not_found_slabs)} اسلب\n\n"
                   f"آیا گزارش PDF دریافت کنید؟")
            if messagebox.askyesno("نتیجه پردازش", msg, parent=self):
                _generate_pdf_for_batch(log_entry)

        styled_btn(top_bar, "📤  آپلود فایل اکسل", _upload_and_process,
                   color=C["btn_success"], width=160).pack(side="right", padx=(0,4))

        # ── جدول تاریخچه ──
        tk.Label(tab, text="📋  تاریخچه خروج از طریق فایل",
                 bg=C["panel"], fg=C["accent2"], font=FONT_NORM).pack(anchor="e", padx=16, pady=(8,2))

        h_cols  = ("uploaded_at","file_name","slabs_total","registered","duplicates","not_found","result","uploaded_by")
        h_heads = ("تاریخ آپلود","نام فایل","کل اسلب","ثبت شد","تکراری","پیدا نشد","نتیجه","آپلود توسط")
        hf, htree = scrolled_tree(tab, h_cols, h_heads, height=10)
        hf.pack(fill="both", expand=True, padx=16, pady=4)
        search_bar(tab, htree).pack(anchor="e", padx=16, pady=2)

        def refresh_history():
            htree.delete(*htree.get_children())
            db  = load_db()
            for entry in reversed(db.get("file_exit_log", [])):
                reg  = len(entry.get("registered", []))
                dup  = len(entry.get("duplicates",  []))
                nf   = len(entry.get("not_found",   []))
                tot  = entry.get("slabs_total", reg+dup+nf)
                result_txt = f"ثبت:{reg}  تکراری:{dup}  نیافت:{nf}"
                who  = get_display_name(entry.get("uploaded_by","—"), db) if self.role=="admin" else "شخص دیگر"
                tag  = "ok" if nf==0 and dup==0 else ("warn" if nf==0 else "err")
                htree.insert("", "end", iid=entry["batch_id"], values=(
                    entry.get("uploaded_at","—"),
                    entry.get("file_name","—"),
                    tot, reg, dup, nf,
                    result_txt,
                    who
                ), tags=(tag,))
            htree.tag_configure("ok",   background="#1a3a1a", foreground=C["success"])
            htree.tag_configure("warn", background="#3a3a00", foreground=C["warning"])
            htree.tag_configure("err",  background="#3a1a1a", foreground="#ff6b6b")

        # ── دکمه‌های پایین (روی ردیف انتخاب شده از تاریخچه) ──
        btn_bar = tk.Frame(tab, bg=C["panel"])
        btn_bar.pack(fill="x", padx=16, pady=(4,8))

        def _get_selected_entry():
            sel = htree.selection()
            if not sel:
                messagebox.showwarning("انتخاب", "یک ردیف از تاریخچه انتخاب کنید.", parent=self)
                return None
            db = load_db()
            batch_id = sel[0]
            for e in db.get("file_exit_log", []):
                if e["batch_id"] == batch_id:
                    return e
            return None

        def _download_original_file():
            """دانلود مجدد فایل اکسل آپلود‌شده"""
            entry = _get_selected_entry()
            if not entry:
                return
            if not entry.get("file_b64"):
                messagebox.showinfo("اطلاع", "محتوای فایل برای این ردیف ذخیره نشده است.", parent=self)
                return
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=entry.get("file_name","فایل_خروج.xlsx"),
                title="ذخیره فایل اصلی",
                parent=self)
            if not save_path:
                return
            try:
                import base64
                with open(save_path, "wb") as _f:
                    _f.write(base64.b64decode(entry["file_b64"]))
                messagebox.showinfo("موفق", f"فایل ذخیره شد:\n{save_path}", parent=self)
            except Exception as ex:
                messagebox.showerror("خطا", str(ex), parent=self)

        def _generate_pdf_for_batch(entry=None):
            """گزارش PDF برای یک دسته از تاریخچه — با همان فونت و bidi بقیه گزارشات"""
            if entry is None:
                entry = _get_selected_entry()
            if not entry:
                return

            # ── بارگذاری فونت و ابزارهای فارسی — دقیقاً مثل بقیه گزارشات ──
            rl = self._pdf_font_tools()
            if not rl:
                return

            A4             = rl["A4"]
            rl_colors      = rl["rl_colors"]
            SimpleDocTemplate = rl["SimpleDocTemplate"]
            Table          = rl["Table"]
            TableStyle     = rl["TableStyle"]
            Paragraph      = rl["Paragraph"]
            Spacer         = rl["Spacer"]
            HRFlowable     = rl["HRFlowable"]
            ParagraphStyle = rl["ParagraphStyle"]
            cm             = rl["cm"]
            FONT           = rl["FONT"]
            rt             = rl["rt"]

            from reportlab.lib.enums import TA_CENTER, TA_RIGHT

            def ps(name, size, lead_mult=1.4, color="#000000"):
                return ParagraphStyle(name + "_feb",
                    fontName=FONT, fontSize=size,
                    leading=size*lead_mult,
                    textColor=rl_colors.HexColor(color),
                    alignment=TA_RIGHT)

            def P(txt, sty):
                return Paragraph(rt(str(txt)), sty)

            fname = f"گزارش_خروج_فایل_{shamsi_date_for_filename()}.pdf"
            path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF","*.pdf")],
                initialfile=fname,
                title="ذخیره گزارش PDF",
                parent=self)
            if not path:
                return

            reg  = entry.get("registered",  [])
            dup  = entry.get("duplicates",  [])
            nf   = entry.get("not_found",   [])
            who_disp = get_display_name(entry.get("uploaded_by","—"), load_db()) if self.role=="admin" else "شخص دیگر"

            doc = SimpleDocTemplate(path, pagesize=A4,
                rightMargin=1.5*cm, leftMargin=1.5*cm,
                topMargin=1.8*cm, bottomMargin=1.5*cm)

            story = []

            # ── عنوان ──
            story.append(P("گزارش خروج از طریق فایل اکسل", ps("t",16,1.3,"#003366")))
            story.append(Spacer(1, .3*cm))
            story.append(P(
                f"تاریخ آپلود: {entry.get('uploaded_at','—')}   |   "
                f"نام فایل: {entry.get('file_name','—')}   |   "
                f"آپلود توسط: {who_disp}",
                ps("i",10,1.3,"#555555")))
            story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#003366")))
            story.append(Spacer(1, .4*cm))

            # ── جدول خلاصه ──
            summary_data = [
                [P("تعداد", ps("sh0",11,1.2,"#ffffff")),
                 P("وضعیت", ps("sh1",11,1.2,"#ffffff"))],
                [P(str(len(reg)), ps("s0",11,1.2,"#1a7a1a")),
                 P("اسلب هایی که خروج ثبت شد", ps("s1",11,1.2,"#1a7a1a"))],
                [P(str(len(dup)), ps("s2",11,1.2,"#7a6a00")),
                 P("اسلب هایی که قبلا خروج خورده اند / تکراری", ps("s3",11,1.2,"#7a6a00"))],
                [P(str(len(nf)),  ps("s4",11,1.2,"#9a1a1a")),
                 P("اسلب هایی که در بخش کنترل کیفی یافت نشدند", ps("s5",11,1.2,"#9a1a1a"))],
            ]
            sum_tbl = Table(summary_data, colWidths=[2*cm, 14*cm])
            sum_tbl.setStyle(TableStyle([
                ("BACKGROUND",      (0,0), (-1,0),  rl_colors.HexColor("#003366")),
                ("FONTNAME",        (0,0), (-1,-1),  FONT),
                ("ALIGN",           (0,0), (0,-1),   "CENTER"),
                ("ALIGN",           (1,0), (1,-1),   "RIGHT"),
                ("VALIGN",          (0,0), (-1,-1),  "MIDDLE"),
                ("ROWBACKGROUNDS",  (0,1), (-1,-1),
                    [rl_colors.HexColor("#e8f5e9"),
                     rl_colors.HexColor("#fff9c4"),
                     rl_colors.HexColor("#ffebee")]),
                ("GRID",            (0,0), (-1,-1),  .5, rl_colors.HexColor("#cccccc")),
                ("TOPPADDING",      (0,0), (-1,-1),  6),
                ("BOTTOMPADDING",   (0,0), (-1,-1),  6),
            ]))
            story.append(sum_tbl)
            story.append(Spacer(1, .5*cm))

            # ── جداول اسلب‌ها ──
            _ps_counter = [0]
            def _slab_table(title_txt, slab_list, row_color, title_color):
                if not slab_list:
                    return
                _ps_counter[0] += 1
                story.append(P(title_txt, ps(f"sec{_ps_counter[0]}", 13, 1.3, title_color)))
                story.append(Spacer(1, .2*cm))

                # ── هدر جدول — ترتیب از راست به چپ (RTL): ردیف | شماره اسلب | ردیف | شماره اسلب ──
                # چون PDF از چپ به راست رندر می‌شه ولی فارسی RTL است،
                # col0=ردیف چپ | col1=اسلب چپ | col2=ردیف راست | col3=اسلب راست
                _ps_counter[0] += 1
                hdr_ps = ps(f"th{_ps_counter[0]}", 10, 1.2, "#ffffff")
                hdr_row = [
                    P("ردیف",       hdr_ps),
                    P("شماره اسلب", hdr_ps),
                    P("ردیف",       hdr_ps),
                    P("شماره اسلب", hdr_ps),
                ]
                # col0=ردیف چپ | col1=اسلب چپ | col2=ردیف راست | col3=اسلب راست
                col_w = [1.5*cm, 6.5*cm, 1.5*cm, 6.5*cm]

                rows_table = [hdr_row]
                # شماره‌گذاری ستون‌محور:
                # ستون راست (col3,col2): ۱، ۳، ۵، ... (فرد)   ستون چپ (col1,col0): ۲، ۴، ۶، ... (زوج)
                n = len(slab_list)
                half = (n + 1) // 2          # تعداد ردیف‌ها = تعداد اسلب‌های ستون راست
                right_col = slab_list[:half]  # ستون راست — شماره‌های فرد
                left_col  = slab_list[half:]  # ستون چپ  — شماره‌های زوج

                for row_i in range(half):
                    _ps_counter[0] += 1
                    cell_ps = ps(f"c{_ps_counter[0]}", 10, 1.3, "#222222")
                    idx_r = row_i * 2 + 1          # شماره فرد: ۱، ۳، ۵، ...
                    sid_r = right_col[row_i]
                    if row_i < len(left_col):
                        idx_l = row_i * 2 + 2      # شماره زوج: ۲، ۴، ۶، ...
                        sid_l = left_col[row_i]
                        _ps_counter[0] += 1
                        cell_ps2 = ps(f"c{_ps_counter[0]}", 10, 1.3, "#222222")
                        row_cells = [
                            P(str(idx_l), cell_ps2),
                            P(sid_l,      cell_ps2),
                            P(str(idx_r), cell_ps),
                            P(sid_r,      cell_ps),
                        ]
                    else:
                        _ps_counter[0] += 1
                        ep = ps(f"ce{_ps_counter[0]}", 10)
                        row_cells = [
                            P("", ep),
                            P("", ep),
                            P(str(idx_r), cell_ps),
                            P(sid_r,      cell_ps),
                        ]
                    rows_table.append(row_cells)

                t = Table(rows_table, colWidths=col_w)
                t.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,0),   rl_colors.HexColor(title_color)),
                    ("FONTNAME",      (0,0), (-1,-1),  FONT),
                    ("ALIGN",         (0,0), (-1,-1),  "CENTER"),   # همه وسط‌چین
                    ("VALIGN",        (0,0), (-1,-1),  "MIDDLE"),
                    ("BACKGROUND",    (0,1), (-1,-1),  rl_colors.HexColor(row_color)),
                    # خط جداکننده وسط بین دو جفت
                    ("LINEAFTER",     (1,0), (1,-1),   1.5, rl_colors.HexColor("#888888")),
                    ("GRID",          (0,0), (-1,-1),  .4,  rl_colors.HexColor("#cccccc")),
                    ("TOPPADDING",    (0,0), (-1,-1),  6),
                    ("BOTTOMPADDING", (0,0), (-1,-1),  6),
                    ("LEFTPADDING",   (0,0), (-1,-1),  6),
                    ("RIGHTPADDING",  (0,0), (-1,-1),  6),
                ]))
                story.append(t)
                story.append(Spacer(1, .4*cm))

            _slab_table("اسلب هایی که خروج ثبت شد",                      reg, "#e8f5e9", "#1a5a1a")
            _slab_table("اسلب هایی که قبلا خروج خورده اند / تکراری",     dup, "#fff9c4", "#7a6000")
            _slab_table("اسلب هایی که در بخش کنترل کیفی یافت نشدند",     nf,  "#ffebee", "#9a1a1a")

            story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#003366")))
            story.append(P(
                "این گزارش به صورت سیستمی تولید شده است — سامانه مدیریت تختال، شرکت سازه پیشگام مدیسه",
                ps("ft", 6, 1, "#999999")))

            try:
                doc.build(story)
                messagebox.showinfo("موفق", f"گزارش PDF ذخیره شد:\n{path}", parent=self)
            except Exception as ex:
                messagebox.showerror("خطا در ساخت PDF", str(ex), parent=self)

        styled_btn(btn_bar, "📄  گزارش این ردیف", _generate_pdf_for_batch,
                   color=C["accent2"], width=170).pack(side="right", padx=(0,4))
        styled_btn(btn_bar, "⬇️  دانلود فایل اصلی", _download_original_file,
                   color="#4a6080", width=160).pack(side="right", padx=(0,4))

        refresh_history()

    # ─────────────────────────────────────────────────────────
    def _build_file_melt_tab(self, tab):
        """ثبت ذوب از طریق فایل — سه حالت + انتخاب انبار + گزارش PDF دقیقاً مثل بخش خروج"""
        tab.configure(bg=C["panel"])

        # ── اسکرول ──
        canvas = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a",
                           activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        sf = tk.Frame(canvas, bg=C["panel"])
        _wid = canvas.create_window((0, 0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_wid, width=e.width))
        register_scroll_canvas(canvas, sf)

        tk.Label(sf, text="📂  ثبت ذوب از طریق فایل اکسل",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=(10,2))
        tk.Label(sf,
                 text="فایل اکسل با ستون ردیف و شماره اسلب — ثبت دسته‌ای با انتخاب وضعیت و مکان",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        # ── کارت تنظیمات: حالت ثبت + انبار ──
        cfg_card = card_frame(sf)
        cfg_card.pack(fill="x", padx=16, pady=8)
        cfg_in = tk.Frame(cfg_card, bg=C["card"])
        cfg_in.pack(padx=16, pady=12, fill="x")

        # -- حالت ثبت --
        tk.Label(cfg_in, text="حالت ثبت:", bg=C["card"],
                 fg=C["accent"], font=(_MAIN_FONT,10,"bold")).pack(anchor="e", pady=(0,6))
        mode_var = tk.StringVar(value="only_melt")
        mode_row = tk.Frame(cfg_in, bg=C["card"])
        mode_row.pack(anchor="e", pady=(0,4))

        MODES = [
            ("only_melt",    "🔥  فقط ثبت ذوب",                         C["accent"]),
            ("melt_qc_ok",   "✅  ثبت ذوب + کنترل کیفی شده",             C["success"]),
            ("melt_qc_fail", "❌  ثبت ذوب + عدم تایید کنترل کیفی",       C["danger"]),
        ]
        for val, lbl, col in MODES:
            tk.Radiobutton(mode_row, text=lbl, variable=mode_var, value=val,
                bg=C["card"], fg=col, selectcolor=C["card"],
                activebackground=C["card"], activeforeground=col,
                font=(_MAIN_FONT, 11, "bold"), cursor="hand2",
                command=lambda: _on_mode_change()).pack(side="right", padx=10)

        # -- انتخاب انبار (فقط وقتی حالت QC باشد) --
        wh_frame = tk.Frame(cfg_in, bg=C["card"])
        wh_frame.pack(anchor="e", pady=(6,0))
        tk.Label(wh_frame, text="انبار مقصد:", bg=C["card"],
                 fg=C["text"], font=FONT_NORM).pack(side="right", padx=(0,8))
        wh_var = tk.StringVar(value="انبار داخلی")
        wh_btns = []
        for wh, icon, col in [
            ("انبار داخلی",   "🏠", C["accent2"]),
            ("انبار روباز ۱", "🏭", "#3a8060"),
            ("انبار روباز ۲", "🏭", "#3a6080"),
        ]:
            rb = tk.Radiobutton(wh_frame, text=f"{icon} {wh}", variable=wh_var, value=wh,
                bg=C["card"], fg=col, selectcolor=C["card"],
                activebackground=C["card"], activeforeground=col,
                font=(_MAIN_FONT, 10, "bold"), cursor="hand2")
            rb.pack(side="right", padx=8)
            wh_btns.append(rb)

        wh_note = tk.Label(cfg_in, text="", bg=C["card"],
                           fg=C["text_dim"], font=FONT_SMALL)
        wh_note.pack(anchor="e", pady=(4,0))

        def _on_mode_change(*_):
            m = mode_var.get()
            if m == "only_melt":
                for rb in wh_btns: rb.configure(state="disabled")
                wh_var.set("انبار داخلی")
                wh_note.config(text="📍  ثبت ذوب ساده — مکان: انبار داخلی (ثابت)")
            else:
                for rb in wh_btns: rb.configure(state="normal")
                wh_note.config(text="📍  انبار مقصد را انتخاب کنید")
        _on_mode_change()

        # -- دکمه‌های بالا --
        top_bar = tk.Frame(sf, bg=C["panel"])
        top_bar.pack(fill="x", padx=16, pady=6)
        status_upload = tk.Label(top_bar, text="", bg=C["panel"],
                                 fg=C["success"], font=FONT_SMALL, anchor="e")
        status_upload.pack(side="right", padx=(8,0))

        def _download_template():
            if not XLSX:
                messagebox.showerror("خطا","کتابخانه openpyxl نصب نیست.", parent=self); return
            path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")], initialfile="الگوی_ثبت_ذوب.xlsx", parent=self)
            if not path: return
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "ثبت ذوب"
            ws.sheet_view.rightToLeft = True
            hf = PatternFill("solid", fgColor="003366")
            hfont = Font(name="B Nazanin", bold=True, color="FFFFFF", size=12)
            cfont = Font(name="B Nazanin", size=11)
            ca = Alignment(horizontal="center", vertical="center", readingOrder=2)
            ra = Alignment(horizontal="right",  vertical="center", readingOrder=2)
            th = Side(style="thin", color="AAAAAA")
            br = Border(left=th, right=th, top=th, bottom=th)
            ws.column_dimensions["A"].width = 10; ws.column_dimensions["B"].width = 22
            ws.row_dimensions[1].height = 28
            for col, (val, al) in enumerate([("ردیف", ca), ("شماره اسلب", ra)], start=1):
                c = ws.cell(row=1, column=col, value=val)
                c.fill = hf; c.font = hfont; c.alignment = al; c.border = br
            for i in range(1, 6):
                for col, (val, al) in enumerate([(i, ca), ("", ra)], start=1):
                    c = ws.cell(row=i+1, column=col, value=val)
                    c.font = cfont; c.alignment = al; c.border = br
                    if col == 1: c.fill = PatternFill("solid", fgColor="EEF2FF")
            wb.save(path)
            messagebox.showinfo("موفق", f"فایل الگو ذخیره شد:\n{path}", parent=self)

        styled_btn(top_bar, "⬇️  دانلود فایل الگو", _download_template,
                   color="#4a6080", width=160).pack(side="right", padx=(0,4))

        def _upload_and_process():
            if not XLSX:
                messagebox.showerror("خطا","کتابخانه openpyxl نصب نیست.", parent=self); return
            m = mode_var.get()
            dest = wh_var.get()
            # اگر QC هم هست، انبار داخلی پیش‌فرض نیست — باید انتخاب شده باشد
            if m != "only_melt" and dest == "انبار داخلی":
                if not messagebox.askyesno("تأیید",
                    "انبار مقصد «انبار داخلی» انتخاب شده.\nادامه می‌دهید؟", parent=self):
                    return
            path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls")],
                title="انتخاب فایل اکسل ثبت ذوب", parent=self)
            if not path: return
            try:
                wb_in = openpyxl.load_workbook(path, read_only=True, data_only=True)
                rows_data = list(wb_in.active.iter_rows(min_row=2, values_only=True))
                wb_in.close()
            except Exception as ex:
                messagebox.showerror("خطای خواندن فایل", str(ex), parent=self); return

            slabs_in_file = []
            for row in rows_data:
                if len(row) >= 2 and row[1] is not None:
                    val = str(row[1]).strip()
                    if val: slabs_in_file.append(val)
            if not slabs_in_file:
                messagebox.showwarning("فایل خالی","هیچ شماره اسلبی در ستون دوم یافت نشد.", parent=self); return

            db = load_db(); now_ts = now_str(); file_name = os.path.basename(path)
            try:
                import base64
                with open(path, "rb") as _f:
                    file_b64 = base64.b64encode(_f.read()).decode("utf-8")
            except Exception:
                file_b64 = ""

            registered_slabs  = []  # ثبت موفق
            duplicate_slabs   = []  # قبلاً ثبت شده
            invalid_slabs     = []  # شماره نامعتبر
            existing_ids = {r["slab_id"] for r in db["melts"]}

            for sid in slabs_in_file:
                ok, _, sid = validate_slab_id(sid)
                if not ok:
                    invalid_slabs.append(sid); continue
                if sid in existing_ids:
                    duplicate_slabs.append(sid); continue

                # ── ۱. ثبت ذوب ──
                rec = {
                    "slab_id":       sid,
                    "note":          "",
                    "qc_status":     "ثبت شده",
                    "registered_by": self.username,
                    "registered_at": now_ts,
                    "source":        "فایل",
                }

                # ── ۲. اعمال وضعیت QC (در صورت انتخاب) ──
                if m in ("melt_qc_ok", "melt_qc_fail"):
                    qc_status = "کنترل کیفی شده" if m == "melt_qc_ok" else "عدم تایید کنترل کیفی"
                    rec["qc_status"] = qc_status
                    rec["qc_by"]     = self.username
                    rec["qc_at"]     = now_ts

                db["melts"].append(rec)
                existing_ids.add(sid)

                # ── ۳. انتقال به انبار مقصد (اگر QC شده و انبار غیر داخلی) ──
                if m == "melt_qc_ok" and dest != "انبار داخلی":
                    db.setdefault("transfers_out", []).append({
                        "slab_id":          sid,
                        "destination":      dest,
                        "current_location": dest,
                        "transferred_by":   self.username,
                        "transferred_at":   now_ts,
                        "reason":           "ثبت از طریق فایل",
                    })
                    db.setdefault("movement_log", []).append({
                        "slab_id":   sid,
                        "operation": "انتقال",
                        "from":      "انبار داخلی",
                        "to":        dest,
                        "by":        self.username,
                        "at":        now_ts,
                        "reason":    "ثبت از طریق فایل",
                    })
                # ── اگر رد شده و انبار مشخص ──
                elif m == "melt_qc_fail":
                    rec["rej_location"] = dest
                    rec["rej_location_locked"] = True
                    rec["rej_location_by"] = self.username
                    rec["rej_location_at"] = now_ts
                    if dest != "انبار داخلی":
                        dest_key = dest.replace("انبار روباز ", "انتقال به انبار روباز ")
                        db.setdefault("movement_log", []).append({
                            "slab_id":   sid,
                            "operation": "انتقال",
                            "from":      "انبار داخلی",
                            "to":        dest,
                            "by":        self.username,
                            "at":        now_ts,
                            "reason":    "رد شده — ثبت از طریق فایل",
                        })

                registered_slabs.append(sid)

            # ── ثبت لاگ ──
            mode_label = {"only_melt":"فقط ثبت ذوب",
                          "melt_qc_ok":"ثبت ذوب + کنترل کیفی شده",
                          "melt_qc_fail":"ثبت ذوب + عدم تایید کنترل کیفی"}[m]
            log_entry = {
                "batch_id":    f"FM_{now_ts.replace('/','').replace(' ','_').replace(':','')}",
                "uploaded_at": now_ts, "uploaded_by": self.username,
                "file_name":   file_name, "file_b64": file_b64,
                "mode":        mode_label,
                "warehouse":   dest if m != "only_melt" else "انبار داخلی",
                "slabs_total": len(slabs_in_file),
                "registered":  registered_slabs,
                "duplicates":  duplicate_slabs,
                "not_valid":   invalid_slabs,
            }
            if "file_melt_log" not in db: db["file_melt_log"] = []
            db["file_melt_log"].append(log_entry)
            save_db(db)

            status_upload.config(
                text=f"✔  ثبت: {len(registered_slabs)}  |  تکراری: {len(duplicate_slabs)}  |  نامعتبر: {len(invalid_slabs)}")
            refresh_history()

            msg = (f"نتیجه پردازش فایل\n"
                   f"حالت: «{mode_label}»   |   مکان: «{log_entry['warehouse']}»\n\n"
                   f"✅  ثبت شد:  {len(registered_slabs)} اسلب\n"
                   f"⚠️  تکراری (قبلاً ثبت شده):  {len(duplicate_slabs)} اسلب\n"
                   f"❌  شماره نامعتبر:  {len(invalid_slabs)} اسلب\n\n"
                   f"آیا گزارش PDF دریافت کنید؟")
            if messagebox.askyesno("نتیجه پردازش", msg, parent=self):
                _generate_pdf_for_batch(log_entry)

        styled_btn(top_bar, "📤  آپلود و ثبت", _upload_and_process,
                   color=C["btn_success"], width=160).pack(side="right", padx=(0,4))

        # ── تاریخچه ──
        tk.Label(sf, text="📋  تاریخچه ثبت از طریق فایل",
                 bg=C["panel"], fg=C["accent2"], font=FONT_NORM).pack(anchor="e", padx=16, pady=(10,2))

        h_cols  = ("uploaded_at","file_name","mode","warehouse","slabs_total","registered","duplicates","not_valid","uploaded_by")
        h_heads = ("تاریخ آپلود","نام فایل","حالت ثبت","انبار","کل","ثبت شد","تکراری","نامعتبر","آپلود توسط")
        hf_frame, htree = scrolled_tree(sf, h_cols, h_heads, height=8)
        hf_frame.pack(fill="both", expand=True, padx=16, pady=4)

        def refresh_history():
            htree.delete(*htree.get_children())
            db = load_db()
            for entry in reversed(db.get("file_melt_log", [])):
                reg = len(entry.get("registered", []))
                dup = len(entry.get("duplicates",  []))
                nv  = len(entry.get("not_valid",   []))
                tot = entry.get("slabs_total", reg+dup+nv)
                who = get_display_name(entry.get("uploaded_by","—"), db) if self.role=="admin" else "شخص دیگر"
                tag = "ok" if nv==0 and dup==0 else ("warn" if nv==0 else "err")
                htree.insert("", "end", iid=entry["batch_id"], values=(
                    entry.get("uploaded_at","—"), entry.get("file_name","—"),
                    entry.get("mode","—"), entry.get("warehouse","—"),
                    tot, reg, dup, nv, who), tags=(tag,))
            htree.tag_configure("ok",   background="#1a3a1a", foreground=C["success"])
            htree.tag_configure("warn", background="#3a3a00", foreground=C["warning"])
            htree.tag_configure("err",  background="#3a1a1a", foreground="#ff6b6b")

        btn_bar = tk.Frame(sf, bg=C["panel"])
        btn_bar.pack(fill="x", padx=16, pady=(4,10))

        def _get_selected_entry():
            sel = htree.selection()
            if not sel:
                messagebox.showwarning("انتخاب","یک ردیف از تاریخچه انتخاب کنید.", parent=self); return None
            db = load_db()
            for e in db.get("file_melt_log", []):
                if e["batch_id"] == sel[0]: return e
            return None

        def _download_original_file():
            entry = _get_selected_entry()
            if not entry or not entry.get("file_b64"):
                messagebox.showinfo("اطلاع","محتوای فایل ذخیره نشده است.", parent=self); return
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")], initialfile=entry.get("file_name","فایل.xlsx"), parent=self)
            if not save_path: return
            try:
                import base64
                with open(save_path, "wb") as _f:
                    _f.write(base64.b64decode(entry["file_b64"]))
                messagebox.showinfo("موفق", f"فایل ذخیره شد:\n{save_path}", parent=self)
            except Exception as ex:
                messagebox.showerror("خطا", str(ex), parent=self)

        def _generate_pdf_for_batch(entry=None):
            if entry is None: entry = _get_selected_entry()
            if not entry: return
            rl = self._pdf_font_tools()
            if not rl: return
            A4=rl["A4"]; rl_colors=rl["rl_colors"]; SimpleDocTemplate=rl["SimpleDocTemplate"]
            Table=rl["Table"]; TableStyle=rl["TableStyle"]; Paragraph=rl["Paragraph"]
            Spacer=rl["Spacer"]; HRFlowable=rl["HRFlowable"]; ParagraphStyle=rl["ParagraphStyle"]
            cm=rl["cm"]; FONT=rl["FONT"]; rt=rl["rt"]
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT

            def ps(name, size, lead_mult=1.4, color="#000000"):
                return ParagraphStyle(name+"_fm2", fontName=FONT, fontSize=size,
                    leading=size*lead_mult, textColor=rl_colors.HexColor(color), alignment=TA_RIGHT)
            def P(txt, sty):
                return Paragraph(rt(str(txt)), sty)

            mode_lbl = entry.get("mode","ثبت ذوب")
            wh_lbl   = entry.get("warehouse","انبار داخلی")
            who_disp = get_display_name(entry.get("uploaded_by","—"), load_db()) if self.role=="admin" else "شخص دیگر"
            fname = f"گزارش_ثبت_ذوب_{shamsi_date_for_filename()}.pdf"
            path = filedialog.asksaveasfilename(defaultextension=".pdf",
                filetypes=[("PDF","*.pdf")], initialfile=fname, parent=self)
            if not path: return

            reg = entry.get("registered", []); dup = entry.get("duplicates", [])
            nv  = entry.get("not_valid",  [])

            doc = SimpleDocTemplate(path, pagesize=A4,
                rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.8*cm, bottomMargin=1.5*cm)
            story = []
            story.append(P(f"گزارش ثبت ذوب از طریق فایل اکسل", ps("t",16,1.3,"#003366")))
            story.append(Spacer(1, .3*cm))
            story.append(P(
                f"تاریخ آپلود: {entry.get('uploaded_at','—')}   |   "
                f"نام فایل: {entry.get('file_name','—')}   |   آپلود توسط: {who_disp}",
                ps("i",10,1.3,"#555555")))
            story.append(P(f"حالت ثبت: «{mode_lbl}»   |   انبار مقصد: «{wh_lbl}»",
                ps("i2",10,1.3,"#336633")))
            story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#003366")))
            story.append(Spacer(1, .4*cm))

            summary_data = [
                [P("تعداد", ps("sh0",11,1.2,"#ffffff")), P("وضعیت", ps("sh1",11,1.2,"#ffffff"))],
                [P(str(len(reg)), ps("s0",11,1.2,"#1a7a1a")), P("اسلب هایی که با موفقیت ثبت شدند", ps("s1",11,1.2,"#1a7a1a"))],
                [P(str(len(dup)), ps("s2",11,1.2,"#7a6a00")), P("اسلب هایی که قبلاً در سیستم ثبت شده بودند (تکراری)", ps("s3",11,1.2,"#7a6a00"))],
                [P(str(len(nv)),  ps("s4",11,1.2,"#9a1a1a")), P("اسلب هایی با شماره نامعتبر", ps("s5",11,1.2,"#9a1a1a"))],
            ]
            sum_tbl = Table(summary_data, colWidths=[2*cm, 14*cm])
            sum_tbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#003366")),
                ("FONTNAME",(0,0),(-1,-1),FONT),
                ("ALIGN",(0,0),(0,-1),"CENTER"), ("ALIGN",(1,0),(1,-1),"RIGHT"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#e8f5e9"),
                    rl_colors.HexColor("#fff9c4"), rl_colors.HexColor("#ffebee")]),
                ("GRID",(0,0),(-1,-1),.5,rl_colors.HexColor("#cccccc")),
                ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
            ]))
            story.append(sum_tbl); story.append(Spacer(1, .5*cm))

            _ps_c = [0]
            def _slab_table(title_txt, slab_list, row_color, t_color):
                if not slab_list: return
                _ps_c[0] += 1
                story.append(P(title_txt, ps(f"sec{_ps_c[0]}", 13, 1.3, t_color)))
                story.append(Spacer(1, .2*cm))
                _ps_c[0] += 1
                hdr_ps = ps(f"th{_ps_c[0]}", 10, 1.2, "#ffffff")
                hdr_row = [P("ردیف",hdr_ps), P("شماره اسلب",hdr_ps), P("ردیف",hdr_ps), P("شماره اسلب",hdr_ps)]
                col_w = [1.5*cm, 6.5*cm, 1.5*cm, 6.5*cm]
                rows_table = [hdr_row]
                n = len(slab_list); half = (n+1)//2
                right_col = slab_list[:half]; left_col = slab_list[half:]
                for row_i in range(half):
                    _ps_c[0] += 1
                    cp = ps(f"c{_ps_c[0]}", 10, 1.3, "#222222")
                    idx_r = row_i*2+1; sid_r = right_col[row_i]
                    if row_i < len(left_col):
                        idx_l = row_i*2+2; sid_l = left_col[row_i]
                        _ps_c[0] += 1; cp2 = ps(f"c{_ps_c[0]}", 10, 1.3, "#222222")
                        row_cells = [P(str(idx_l),cp2), P(sid_l,cp2), P(str(idx_r),cp), P(sid_r,cp)]
                    else:
                        _ps_c[0] += 1; ep = ps(f"ce{_ps_c[0]}", 10)
                        row_cells = [P("",ep), P("",ep), P(str(idx_r),cp), P(sid_r,cp)]
                    rows_table.append(row_cells)
                t_tbl = Table(rows_table, colWidths=col_w)
                t_tbl.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor(t_color)),
                    ("FONTNAME",(0,0),(-1,-1),FONT), ("ALIGN",(0,0),(-1,-1),"CENTER"),
                    ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                    ("BACKGROUND",(0,1),(-1,-1),rl_colors.HexColor(row_color)),
                    ("LINEAFTER",(1,0),(1,-1),1.5,rl_colors.HexColor("#888888")),
                    ("GRID",(0,0),(-1,-1),.4,rl_colors.HexColor("#cccccc")),
                    ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
                    ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
                ]))
                story.append(t_tbl); story.append(Spacer(1, .4*cm))

            _slab_table("اسلب هایی که با موفقیت ثبت شدند", reg, "#e8f5e9", "#1a5a1a")
            _slab_table("اسلب هایی که قبلاً ثبت شده بودند (تکراری)", dup, "#fff9c4", "#7a6000")
            _slab_table("اسلب هایی با شماره نامعتبر",                  nv,  "#ffebee", "#9a1a1a")
            story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#003366")))
            story.append(P("این گزارش به صورت سیستمی تولید شده است — سامانه مدیریت تختال، شرکت سازه پیشگام مدیسه",
                ps("ft", 6, 1, "#999999")))
            try:
                doc.build(story)
                messagebox.showinfo("موفق", f"گزارش PDF ذخیره شد:\n{path}", parent=self)
            except Exception as ex:
                messagebox.showerror("خطا در ساخت PDF", str(ex), parent=self)

        styled_btn(btn_bar, "📄  گزارش این ردیف", _generate_pdf_for_batch,
                   color=C["accent2"], width=170).pack(side="right", padx=(0,4))
        styled_btn(btn_bar, "⬇️  دانلود فایل اصلی", _download_original_file,
                   color="#4a6080", width=160).pack(side="right", padx=(0,4))
        refresh_history()


    def _center(self, win, w, h):
        """پنجره را در وسط صفحه قرار می‌دهد"""
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")
        _bring_popup_to_front(win)

    # ═══════════════════════
    #  مسیر ذخیره گزارش‌ها — برای ادمین آزاد (با پیش‌فرض پوشه گزارش)،
    #  برای بقیه‌ی کاربران کاملاً قفل روی همان پوشه‌ی مخصوص گزارش
    # ═══════════════════════
    def _resolve_report_save_path(self, key, filename, filetypes, defaultext):
        folder = get_report_dir(key)
        # اگر گزارشی با همین نام (همان نوع گزارش + همان تاریخ) از قبل در این
        # پوشه وجود داشته باشد، به‌جای جایگزینی، نام جدید با پسوند -2, -3 و...
        # ساخته می‌شود تا هیچ گزارش قبلی پاک/بازنویسی نشود.
        filename = make_unique_filename(folder, filename)
        if self.role == "admin":
            return filedialog.asksaveasfilename(
                defaultextension=defaultext, filetypes=filetypes,
                initialfile=filename, initialdir=folder, parent=self)
        # غیر ادمین: بدون پنجره انتخاب مسیر — مستقیم و قفل‌شده روی پوشه مخصوص گزارش
        return os.path.join(folder, filename)

    # ═══════════════════════
    #  نمایش/پرینت آخرین گزارش PDF تولیدشده — بدون دادن دسترسی به
    #  فایل‌سیستم یا پوشه‌ها؛ فقط همان یک فایل PDF مشخص باز می‌شود
    #  تا کاربر بتواند آن را ببیند/پرینت بگیرد.
    # ═══════════════════════
    def _remember_last_report(self, key, path):
        db = load_db()
        db.setdefault("settings", {}).setdefault("last_report_paths", {})[key] = path
        save_db(db)

    def _open_last_report(self, key, label):
        db = load_db()
        path = db.get("settings", {}).get("last_report_paths", {}).get(key)
        if not path or not os.path.exists(path):
            messagebox.showinfo("اطلاع", f"هنوز گزارش {label} تولید نشده است.", parent=self)
            return
        try:
            if os.name == "nt":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as ex:
            messagebox.showerror("خطا", f"امکان بازکردن فایل نبود:\n{ex}", parent=self)

    # ═══════════════════════
    #  ردیف تغییر رمز Minimize — رمز فعلی نمایش داده می‌شود (با دکمه‌ی
    #  چشم برای مخفی/آشکار کردن) و رمز جدید در کنار آن وارد می‌شود، تا
    #  همیشه مشخص باشد رمز فعلی چیست.
    # ═══════════════════════
    def _build_minimize_pw_row(self, parent):
        card = card_frame(parent)
        card.pack(fill="x", padx=16, pady=(10, 12))
        inner = tk.Frame(card, bg=C["card"])
        inner.pack(fill="x", padx=14, pady=10)

        tk.Label(inner, text="🔒  رمز عبور (مینیمایز / بستن / کوچک‌کردن)",
                 bg=C["card"], fg=C["accent"], font=(_MAIN_FONT, 11, "bold")).pack(anchor="e")

        row = tk.Frame(inner, bg=C["card"])
        row.pack(fill="x", pady=(10, 0))

        # ── رمز فعلی + چشم ──
        cur_box = tk.Frame(row, bg=C["card"])
        cur_box.pack(side="right", padx=(10, 0))
        tk.Label(cur_box, text="رمز فعلی", bg=C["card"], fg=C["text_dim"],
                 font=FONT_SMALL).pack(anchor="e")
        cur_sub = tk.Frame(cur_box, bg=C["card"])
        cur_sub.pack()
        pw_disabled = not is_lock_password_required()
        stored_pw = "" if pw_disabled else get_lock_password()
        pw_show = tk.BooleanVar(value=False)

        cur_entry = make_entry(cur_sub, width=16, show="●")
        cur_entry.insert(0, "(غیرفعال — بدون رمز)" if pw_disabled else stored_pw)
        cur_entry.configure(state="disabled", disabledbackground=C["entry_bg"])
        cur_entry.pack(side="right")

        def _refresh_cur_display():
            cur_entry.configure(state="normal")
            cur_entry.delete(0, "end")
            if pw_disabled:
                cur_entry.insert(0, "(غیرفعال — بدون رمز)")
                cur_entry.configure(show="")
            else:
                cur_entry.insert(0, stored_pw)
                cur_entry.configure(show="" if pw_show.get() else "●")
            cur_entry.configure(state="disabled", disabledbackground=C["entry_bg"])

        def toggle_show():
            if pw_disabled:
                return
            pw_show.set(not pw_show.get())
            _refresh_cur_display()
            eye_btn.config(text="🙈" if pw_show.get() else "👁")

        eye_btn = tk.Label(cur_sub, text="👁", bg=C["card"], fg=C["accent"],
                            font=(_MAIN_FONT, 12, "bold"), cursor="hand2")
        eye_btn.pack(side="right", padx=(4, 6))
        eye_btn.bind("<Button-1>", lambda e: toggle_show())

        # ── جداکننده ──
        tk.Frame(row, bg=C["border"], width=1).pack(side="right", fill="y", padx=6)

        # ── رمز جدید (بدون تکرار) ──
        new_box = tk.Frame(row, bg=C["card"])
        new_box.pack(side="right", padx=(10, 0))
        tk.Label(new_box, text="رمز جدید", bg=C["card"], fg=C["text_dim"],
                 font=FONT_SMALL).pack(anchor="e")
        new_entry = make_entry(new_box, width=16, show="●")
        new_entry.pack()

        def change_minimize_pw():
            nonlocal stored_pw, pw_disabled
            new = new_entry.get().strip()
            if not new:
                messagebox.showerror("خطا", "رمز جدید را وارد کنید.", parent=parent)
                return

            db2 = load_db()
            db2.setdefault("settings", {})["minimize_password"] = normalize_digits(new)
            db2["settings"]["minimize_password_disabled"] = False
            save_db(db2)
            stored_pw = normalize_digits(new)
            pw_disabled = False
            pw_show.set(False)
            new_entry.delete(0, "end")
            _refresh_cur_display()
            messagebox.showinfo("موفق", "رمز عبور با موفقیت تغییر کرد.", parent=parent)

        def remove_minimize_pw():
            nonlocal stored_pw, pw_disabled
            if not messagebox.askyesno(
                    "حذف رمز",
                    "رمز بستن/مینیمایز/کوچک‌کردن حذف شود؟\n"
                    "پس از آن این عملیات بدون رمز انجام می‌شوند.",
                    parent=parent):
                return
            db2 = load_db()
            db2.setdefault("settings", {})["minimize_password"] = ""
            db2["settings"]["minimize_password_disabled"] = True
            save_db(db2)
            stored_pw = ""
            pw_disabled = True
            pw_show.set(False)
            new_entry.delete(0, "end")
            _refresh_cur_display()
            messagebox.showinfo("موفق", "رمز حذف شد. بستن و مینیمایز بدون رمز انجام می‌شود.", parent=parent)

        btn_row = tk.Frame(inner, bg=C["card"])
        btn_row.pack(fill="x", pady=(10, 0))
        styled_btn(btn_row, "✏️  ذخیره رمز جدید", change_minimize_pw,
                   color=C["accent2"], width=130, height=32).pack(side="right", padx=(6, 0))
        styled_btn(btn_row, "🗑  حذف رمز", remove_minimize_pw,
                   color=C["btn_danger"], width=110, height=32).pack(side="right", padx=(6, 0))
        return card

    def _build_backup_pw_row(self, parent):
        """مدیریت رمز بک‌آپ/آرشیو — دقیقاً مثل ردیف مینیمایز/بستن."""
        card = card_frame(parent)
        card.pack(fill="x", padx=16, pady=(0, 12))
        inner = tk.Frame(card, bg=C["card"])
        inner.pack(fill="x", padx=14, pady=10)

        tk.Label(inner, text="💾  رمز پکیج / توزیع (اختیاری)",
                 bg=C["card"], fg=C["accent"], font=(_MAIN_FONT, 11, "bold")).pack(anchor="e")
        tk.Label(inner,
                 text="بک‌آپ داخل برنامه بدون رمز فولدر است. رمز فقط برای قفل پکیج Final استفاده می‌شود.",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(2, 0))

        row = tk.Frame(inner, bg=C["card"])
        row.pack(fill="x", pady=(10, 0))

        cur_box = tk.Frame(row, bg=C["card"])
        cur_box.pack(side="right", padx=(10, 0))
        tk.Label(cur_box, text="رمز فعلی", bg=C["card"], fg=C["text_dim"],
                 font=FONT_SMALL).pack(anchor="e")
        cur_sub = tk.Frame(cur_box, bg=C["card"])
        cur_sub.pack()
        pw_disabled = not is_backup_vault_password_required()
        stored_pw = "" if pw_disabled else get_backup_vault_password()
        pw_show = tk.BooleanVar(value=False)

        cur_entry = make_entry(cur_sub, width=16, show="●")
        cur_entry.insert(0, "(غیرفعال — بدون رمز)" if pw_disabled else stored_pw)
        cur_entry.configure(state="disabled", disabledbackground=C["entry_bg"])
        cur_entry.pack(side="right")

        def _refresh_cur_display():
            cur_entry.configure(state="normal")
            cur_entry.delete(0, "end")
            if pw_disabled:
                cur_entry.insert(0, "(غیرفعال — بدون رمز)")
                cur_entry.configure(show="")
            else:
                cur_entry.insert(0, stored_pw)
                cur_entry.configure(show="" if pw_show.get() else "●")
            cur_entry.configure(state="disabled", disabledbackground=C["entry_bg"])

        def toggle_show():
            if pw_disabled:
                return
            pw_show.set(not pw_show.get())
            _refresh_cur_display()
            eye_btn.config(text="🙈" if pw_show.get() else "👁")

        eye_btn = tk.Label(cur_sub, text="👁", bg=C["card"], fg=C["accent"],
                            font=(_MAIN_FONT, 12, "bold"), cursor="hand2")
        eye_btn.pack(side="right", padx=(4, 6))
        eye_btn.bind("<Button-1>", lambda e: toggle_show())

        tk.Frame(row, bg=C["border"], width=1).pack(side="right", fill="y", padx=6)

        new_box = tk.Frame(row, bg=C["card"])
        new_box.pack(side="right", padx=(10, 0))
        tk.Label(new_box, text="رمز جدید", bg=C["card"], fg=C["text_dim"],
                 font=FONT_SMALL).pack(anchor="e")
        new_entry = make_entry(new_box, width=16, show="●")
        new_entry.pack()

        def _reseal_backup_dirs(password: str, enabled: bool):
            """دیگر روی backups قفل نمی‌گذارد — فقط پوشه‌ها را مرتب می‌کند."""
            try:
                if r"D:\SteelFactory2-v2" not in sys.path:
                    sys.path.insert(0, r"D:\SteelFactory2-v2")
                from shared.backup_vault import cleanup_extra_backup_dirs, set_vault_password
                set_vault_password(password if enabled else "")
                cleanup_extra_backup_dirs(_app_base_dir())
            except Exception:
                pass

        def change_backup_pw():
            nonlocal stored_pw, pw_disabled
            new = new_entry.get().strip()
            if not new:
                messagebox.showerror("خطا", "رمز جدید را وارد کنید.", parent=parent)
                return
            db2 = load_db()
            db2.setdefault("settings", {})["backup_vault_password"] = new
            db2["settings"]["backup_vault_password_disabled"] = False
            save_db(db2)
            stored_pw = new
            pw_disabled = False
            pw_show.set(False)
            new_entry.delete(0, "end")
            _refresh_cur_display()
            _reseal_backup_dirs(new, True)
            messagebox.showinfo("موفق",
                "رمز ورود فولدر تغییر کرد و روی پوشه‌های بک‌آپ/آرشیو اعمال شد.",
                parent=parent)

        def remove_backup_pw():
            nonlocal stored_pw, pw_disabled
            if not messagebox.askyesno(
                    "حذف رمز",
                    "رمز ورود به فولدر بک‌آپ/آرشیو حذف شود؟\n"
                    "پس از آن فولدرها بدون رمز باز می‌شوند.",
                    parent=parent):
                return
            db2 = load_db()
            db2.setdefault("settings", {})["backup_vault_password"] = ""
            db2["settings"]["backup_vault_password_disabled"] = True
            save_db(db2)
            stored_pw = ""
            pw_disabled = True
            pw_show.set(False)
            new_entry.delete(0, "end")
            _refresh_cur_display()
            _reseal_backup_dirs("", False)
            messagebox.showinfo("موفق",
                "رمز ورود فولدر حذف شد.",
                parent=parent)

        btn_row = tk.Frame(inner, bg=C["card"])
        btn_row.pack(fill="x", pady=(10, 0))
        styled_btn(btn_row, "✏️  ذخیره رمز جدید", change_backup_pw,
                   color=C["accent2"], width=130, height=32).pack(side="right", padx=(6, 0))
        styled_btn(btn_row, "🗑  حذف رمز", remove_backup_pw,
                   color=C["btn_danger"], width=110, height=32).pack(side="right", padx=(6, 0))
        return card

    # ═══════════════════════
    #  تب: مدیریت سیستم
    # ═══════════════════════
    def _build_admin_tab(self, tab):
        tab.configure(bg=C["panel"])
        style = ttk.Style()
        style.configure("Sub.TNotebook", background=C["bg"], bordercolor=C["border"])
        style.configure("Sub.TNotebook.Tab",
            background=C["tab_inactive"], foreground=C["text"],
            font=(_MAIN_FONT, 10, "bold"), padding=[12, 6])
        style.map("Sub.TNotebook.Tab",
            background=[("selected", C["accent"])],
            foreground=[("selected", "#ffffff")])
        sub_nb = ttk.Notebook(tab, style="Sub.TNotebook")
        sub_nb.pack(fill="both", expand=True, padx=4, pady=4)
        self._admin_settings(sub_nb)
        self._admin_users(sub_nb)
        self._admin_edit_records(sub_nb)
        self._admin_reports(sub_nb)
        self._admin_movement_log(sub_nb)
        self._admin_return_log(sub_nb)
        self._admin_import_excel(sub_nb)

    def _admin_reset(self, nb):
        """بازنشانی کامل سیستم — پاک کردن همه‌ی اطلاعات عملیاتی و شروع از نو.
        کاربران و تنظیمات ظاهری (تم/رمز مینیمایز) برای جلوگیری از قفل‌شدن
        کامل برنامه حفظ می‌شوند؛ فقط داده‌های ثبت‌شده پاک می‌شوند."""
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="♻️  بازنشانی سیستم")
        self._build_reset_tab(tab)

    def _build_reset_tab(self, tab):
        """محتوای واقعی صفحه‌ی بازنشانی — جدا شده تا هم به‌صورت سربرگ
        مستقل و هم کنار بخش بک‌آپ در سربرگ‌های مدیریت سیستم قابل استفاده باشد."""
        tab.configure(bg=C["panel"])

        wrap = tk.Frame(tab, bg=C["panel"])
        wrap.pack(fill="both", expand=True, padx=24, pady=24)

        warn_card = tk.Frame(wrap, bg="#3a1414", highlightthickness=2,
                              highlightbackground=C["danger"])
        warn_card.pack(fill="x", pady=(0, 18))
        tk.Label(warn_card, text="⚠️  بازنشانی کامل سیستم",
                 bg="#3a1414", fg="#ff6060", font=FONT_HEAD).pack(anchor="e", padx=18, pady=(14, 6))
        tk.Label(warn_card, justify="right", wraplength=760, bg="#3a1414", fg="#f0d0d0",
                 font=FONT_NORM,
                 text=(
                     "با زدن این دکمه، تمام اطلاعات ثبت‌شده در سیستم برای همیشه پاک می‌شود؛\n"
                     "درست مثل روزی که برنامه را برای اولین بار نصب کرده‌اید:\n"
                     "ثبت ذوب، کنترل کیفی، تأیید/رد، انبار و انتقال‌ها، تحویل باومن، قراضه، "
                     "اسکارف/برش، خروج اسلب‌ها، تیکت‌ها، برگشتی‌ها و کل تاریخچه/گزارش‌ها.\n\n"
                     "این عمل هیچ راه بازگشتی ندارد مگر با بازیابی از فایل بک‌آپ.\n"
                     "(نام‌های کاربری/رمزها و تنظیمات ظاهری برنامه حذف نمی‌شوند تا بعد از "
                     "بازنشانی همچنان بتوانید وارد سیستم شوید.)"
                 )).pack(anchor="e", padx=18, pady=(0, 16))

        def do_reset():
            if not messagebox.askyesno(
                    "⚠️  اخطار جدی — بازنشانی کامل",
                    "آیا مطمئنید می‌خواهید همه‌ی اطلاعات سیستم برای همیشه پاک شود؟\n\n"
                    "قبل از پاک‌سازی، یک بک‌آپ خودکار گرفته می‌شود.\n"
                    "این کار روی داده‌های عملیاتی است؛ فولدر backups حذف نمی‌شود.",
                    icon="warning", parent=self):
                return
            confirm_word = simpledialog.askstring(
                "تأیید نهایی",
                'برای تأیید نهایی فقط بنویسید:\n\nok',
                parent=self)
            if confirm_word is None:
                return
            # فقط ok (انگلیسی) — بدون حساسیت به حروف بزرگ/کوچک و فاصله
            typed = normalize_digits(str(confirm_word)).strip().lower()
            if typed != "ok":
                messagebox.showwarning("لغو شد", "برای تأیید باید دقیقاً ok بنویسید.", parent=self)
                return

            db = load_db()
            # بک‌آپ خودکار قبل از هر بازنشانی — داده سال‌های قبل حفظ شود
            try:
                import json as _json
                from pathlib import Path as _P
                bdir = _P(_app_base_dir()) / "backups"
                bdir.mkdir(parents=True, exist_ok=True)
                stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                bpath = bdir / f"slab_db_before_reset_{stamp}.json"
                bpath.write_text(_json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception as e:
                if not messagebox.askyesno(
                        "بک‌آپ ناموفق",
                        f"بک‌آپ خودکار گرفته نشد:\n{e}\n\nباز هم بازنشانی انجام شود؟",
                        parent=self):
                    return

            reset_list_keys = [
                "melts", "qc_melts", "transfers_out", "lab_deliveries",
                "scarf_cut", "bauman", "scrap", "returns", "return_log",
                "movement_log", "login_log", "qc_history", "tickets",
                "file_exit_log", "file_qc_log", "file_melt_log", "file_warehouse_log",
            ]
            for key in reset_list_keys:
                if key in db:
                    db[key] = []
            save_db(db)
            invalidate_display_cache()
            messagebox.showinfo(
                "✅  بازنشانی انجام شد",
                "همه‌ی اطلاعات عملیاتی سیستم پاک شد.\n"
                "یک بک‌آپ خودکار در پوشه backups ذخیره شده است.\n"
                "برای مشاهده‌ی کامل تغییرات، برنامه را مجدداً اجرا کنید.",
                parent=self)

        styled_btn(warn_card, "♻️  بازنشانی کامل و شروع از نو", do_reset,
                    color=C["btn_danger"]).pack(anchor="e", padx=18, pady=(0, 16))

    def _admin_import_excel(self, nb):
        """ایمپورت داده از اکسل — ثبت دسته‌جمعی اسلب‌ها"""
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="📥  ایمپورت از اکسل")
        tab.configure(bg=C["panel"])

        # ── اسکرول ──
        canvas = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a",
                           activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        sf = tk.Frame(canvas, bg=C["panel"])
        _wid = canvas.create_window((0,0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_wid, width=e.width))
        register_scroll_canvas(canvas, sf)

        tk.Label(sf, text="📥  ایمپورت اطلاعات از فایل اکسل",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=(12,4))
        tk.Label(sf,
                 text="فایل اکسل را با سرستون‌های مناسب انتخاب کنید — سیستم خودکار ستون‌ها را شناسایی و داده را ثبت می‌کند",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL,
                 wraplength=700, justify="right").pack(anchor="e", padx=16)

        # ── راهنمای الگو ──
        guide_card = card_frame(sf)
        guide_card.pack(fill="x", padx=16, pady=8)
        guide_in = tk.Frame(guide_card, bg=C["card"])
        guide_in.pack(padx=16, pady=10, fill="x")
        tk.Label(guide_in, text="📋  الگوی ستون‌های قابل شناسایی:",
                 bg=C["card"], fg=C["gold"], font=(_MAIN_FONT,10,"bold")).pack(anchor="e", pady=(0,6))

        guide_text = (
            "ستون‌های قابل شناسایی (نام سرستون در فایل اکسل):\n\n"
            "• شماره اسلب / slab_id / slab id → شماره اسلب (اجباری، ۱۱ رقم)\n"
            "• تاریخ ثبت / registered_at / date → تاریخ ثبت\n"
            "• وضعیت / qc_status / status → وضعیت کنترل کیفی\n"
            "• ثبت کننده / registered_by → نام کاربری ثبت‌کننده\n"
            "• توضیحات / note → توضیحات\n\n"
            "⚠️  ستون «شماره اسلب» اجباری است. بقیه اختیاری هستند.\n"
            "اسلب‌هایی که قبلاً ثبت شده‌اند رد می‌شوند."
        )
        tk.Label(guide_in, text=guide_text, bg=C["card"], fg=C["text_dim"],
                 font=FONT_SMALL, justify="right", wraplength=680).pack(anchor="e")

        # ── دانلود الگو ──
        def download_template():
            if not XLSX:
                messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
                initialfile="الگوی_ایمپورت_اسلب.xlsx", parent=self)
            if not path: return
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "ثبت اسلب"
            ws.sheet_view.rightToLeft = True
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            headers = ["شماره اسلب","تاریخ ثبت","وضعیت","ثبت کننده","توضیحات"]
            fn = "B Nazanin"
            h_font = Font(name=fn, bold=True, size=11, color="FFFFFF")
            h_fill = PatternFill("solid", fgColor="0D1E3C")
            h_align = Alignment(horizontal="center", vertical="center", readingOrder=2)
            thin = Side(style="thin", color="BBBBBB")
            bdr = Border(left=thin,right=thin,top=thin,bottom=thin)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font=h_font; c.fill=h_fill; c.alignment=h_align; c.border=bdr
                ws.column_dimensions[c.column_letter].width = 20
            ws.row_dimensions[1].height = 28
            # یک ردیف نمونه
            sample = ["12345678901","1404/03/15  08:00:00","ثبت شده","shift1",""]
            for ci, v in enumerate(sample, 1):
                c = ws.cell(row=2, column=ci, value=v)
                c.font = Font(name=fn, size=10, color="888888", italic=True)
                c.alignment = h_align
            wb.save(path)
            messagebox.showinfo("موفق",f"فایل الگو ذخیره شد:\n{path}",parent=self)

        styled_btn(guide_in, "⬇  دانلود فایل الگو", download_template,
                   color=C["btn_primary"]).pack(anchor="w", pady=(8,0))

        # ── انتخاب فایل ──
        file_card = card_frame(sf)
        file_card.pack(fill="x", padx=16, pady=8)
        file_in = tk.Frame(file_card, bg=C["card"])
        file_in.pack(padx=16, pady=10, fill="x")
        tk.Label(file_in, text="📂  انتخاب فایل اکسل:",
                 bg=C["card"], fg=C["accent"], font=FONT_NORM).pack(anchor="e", pady=(0,6))

        file_path_var = tk.StringVar(value="")
        file_row = tk.Frame(file_in, bg=C["card"])
        file_row.pack(fill="x")
        file_ent = tk.Entry(file_row, textvariable=file_path_var,
                             bg=C["entry_bg"], fg=C["text"],
                             insertbackground=C["accent"], font=FONT_SMALL,
                             justify="right", bd=0, relief="flat",
                             highlightthickness=1, highlightbackground=C["border"],
                             highlightcolor=C["accent"], state="readonly", width=52)
        file_ent.pack(side="right", padx=(0,6), fill="x", expand=True)

        def pick_file():
            path = filedialog.askopenfilename(
                title="انتخاب فایل اکسل",
                filetypes=[("Excel","*.xlsx *.xls"),("همه","*.*")],
                parent=self)
            if path:
                file_path_var.set(path)
                do_preview()

        tk.Button(file_row, text="📂 انتخاب", command=pick_file,
                  bg=C["btn_primary"], fg="#ffffff", font=FONT_SMALL,
                  bd=0, relief="flat", cursor="hand2", padx=10, pady=6).pack(side="right")

        # ── پیش‌نمایش ──
        preview_lbl = tk.Label(file_in, text="", bg=C["card"],
                                fg=C["warning"], font=FONT_SMALL)
        preview_lbl.pack(anchor="e", pady=4)

        cols_pv = ("row","slab_id","date","status","by","note","result")
        heads_pv = ("ردیف","شماره اسلب","تاریخ","وضعیت","ثبت‌کننده","توضیحات","نتیجه")
        tf_pv, tree_pv = scrolled_tree(sf, cols_pv, heads_pv, height=10)
        tf_pv.pack(fill="both", expand=True, padx=16, pady=4)
        for c in cols_pv:
            tree_pv.column(c, width=120, anchor="center")
        tree_pv.column("slab_id", width=150)
        tree_pv.column("result",  width=180)

        _preview_data = []

        SLAB_ID_ALIASES = {"شماره اسلب","slab_id","slab id","slab","شماره تختال","کد اسلب"}
        DATE_ALIASES    = {"تاریخ ثبت","registered_at","date","تاریخ"}
        STATUS_ALIASES  = {"وضعیت","qc_status","status","وضعیت qc"}
        BY_ALIASES      = {"ثبت کننده","registered_by","by","ثبت‌کننده"}
        NOTE_ALIASES    = {"توضیحات","note","یادداشت"}

        def _match_col(header_lower, aliases):
            return any(a in header_lower or header_lower in a for a in aliases)

        def do_preview():
            _preview_data.clear()
            tree_pv.delete(*tree_pv.get_children())
            path = file_path_var.get().strip()
            if not path or not os.path.exists(path):
                preview_lbl.config(text="⚠️  فایل انتخاب نشده"); return
            if not XLSX:
                preview_lbl.config(text="⚠️  openpyxl نصب نیست"); return
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                # یافتن سرستون‌ها
                header_row = [str(c.value or "").strip().lower() for c in list(ws.rows)[0]]
                col_map = {}
                for ci, h in enumerate(header_row):
                    if _match_col(h, SLAB_ID_ALIASES): col_map["slab_id"] = ci
                    elif _match_col(h, DATE_ALIASES):   col_map["date"] = ci
                    elif _match_col(h, STATUS_ALIASES): col_map["status"] = ci
                    elif _match_col(h, BY_ALIASES):     col_map["by"] = ci
                    elif _match_col(h, NOTE_ALIASES):   col_map["note"] = ci

                if "slab_id" not in col_map:
                    preview_lbl.config(text="❌  ستون «شماره اسلب» یافت نشد"); return

                db = load_db()
                existing_ids = {r["slab_id"] for r in db["melts"]}
                valid = 0; dup = 0; invalid = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if all(c is None for c in row): continue
                    sid = str(row[col_map["slab_id"]] or "").strip()
                    date_v  = str(row[col_map.get("date", -1)] or "").strip() if "date"   in col_map and col_map["date"] < len(row) else ""
                    stat_v  = str(row[col_map.get("status",-1)] or "").strip() if "status" in col_map and col_map["status"] < len(row) else "ثبت شده"
                    by_v    = str(row[col_map.get("by",-1)] or "").strip() if "by" in col_map and col_map["by"] < len(row) else ""
                    note_v  = str(row[col_map.get("note",-1)] or "").strip() if "note" in col_map and col_map["note"] < len(row) else ""

                    import re as _re
                    ok_sid = bool(_re.fullmatch(r'\d{11}', sid))
                    if not ok_sid:
                        result = "❌ شماره اسلب نامعتبر"
                        tag = "invalid"
                        invalid += 1
                    elif sid in existing_ids:
                        result = "⚠️ قبلاً ثبت شده"
                        tag = "dup"
                        dup += 1
                    else:
                        result = "✔ آماده ثبت"
                        tag = "ok"
                        valid += 1
                        _preview_data.append({
                            "slab_id": sid,
                            "registered_at": date_v or now_str(),
                            "qc_status": stat_v if stat_v in ("ثبت شده","کنترل کیفی شده","عدم تایید کنترل کیفی") else "ثبت شده",
                            "registered_by": by_v or self.username,
                            "note": note_v,
                        })
                    tree_pv.insert("","end", values=(
                        row_idx-1, sid, date_v, stat_v or "ثبت شده", by_v, note_v, result
                    ), tags=(tag,))

                tree_pv.tag_configure("ok",      background="#1a3020", foreground=C["success"])
                tree_pv.tag_configure("dup",     background="#3a3000", foreground=C["warning"])
                tree_pv.tag_configure("invalid", background="#3a1010", foreground=C["danger"])
                preview_lbl.config(
                    text=f"✔ آماده ثبت: {valid}  |  ⚠️ تکراری: {dup}  |  ❌ نامعتبر: {invalid}")
            except Exception as ex:
                preview_lbl.config(text=f"❌ خطا در خواندن فایل: {ex}")

        # ── ثبت دسته‌جمعی ──
        import_status = tk.Label(sf, text="", bg=C["panel"],
                                  fg=C["success"], font=FONT_NORM)
        import_status.pack(anchor="e", padx=16, pady=4)

        def do_import():
            if not _preview_data:
                messagebox.showwarning("خطا",
                    "ابتدا یک فایل اکسل معتبر انتخاب کنید.", parent=self); return
            if not messagebox.askyesno("تأیید ایمپورت",
                f"تعداد {len(_preview_data)} اسلب ثبت می‌شود.\nآیا مطمئنید؟",
                parent=self): return
            db = load_db()
            existing_ids = {r["slab_id"] for r in db["melts"]}
            added = 0
            for rec in _preview_data:
                if rec["slab_id"] in existing_ids: continue
                db["melts"].append({
                    "slab_id":       rec["slab_id"],
                    "note":          rec.get("note",""),
                    "qc_status":     rec.get("qc_status","ثبت شده"),
                    "registered_by": rec.get("registered_by", self.username),
                    "registered_at": rec.get("registered_at", now_str()),
                })
                existing_ids.add(rec["slab_id"])
                added += 1
            save_db(db)
            import_status.config(
                text=f"✔  {added} اسلب با موفقیت ایمپورت شد  |  {now_str()}")
            messagebox.showinfo("موفق",f"{added} اسلب ثبت شد.",parent=self)
            do_preview()

        ctrl_imp = tk.Frame(sf, bg=C["panel"])
        ctrl_imp.pack(fill="x", padx=16, pady=8)
        styled_btn(ctrl_imp, "🔍  پیش‌نمایش", do_preview,
                   color=C["btn_primary"]).pack(side="right", padx=4)
        styled_btn(ctrl_imp, "📥  ثبت دسته‌جمعی", do_import,
                   color=C["btn_success"]).pack(side="right", padx=4)

    def _admin_users(self, nb):
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="👥  مدیریت کاربران")
        tk.Label(tab, text="👥  مدیریت کاربران سیستم",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        tk.Label(tab,
                 text="✏️  روی هر سلول دابل‌کلیک کنید تا مستقیم ویرایش کنید",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        # ── مدیریت رمز Minimize — فقط با کلیک روی آیکون تغییر می‌کند ──
        self._build_minimize_pw_row(tab)
        self._build_backup_pw_row(tab)

        cols  = ("uname","display","role","status","visible","pw_show")
        heads = ("نام کاربری","نام نمایشی","نقش","وضعیت","نمایش در لاگین","رمز عبور")
        tf, tree = scrolled_tree(tab, cols, heads, height=10)
        tf.pack(fill="both", expand=True, padx=10, pady=8)
        tree.column("uname",    width=130, anchor="center")
        tree.column("display",  width=160, anchor="center")
        tree.column("role",     width=100, anchor="center")
        tree.column("status",   width=90,  anchor="center")
        tree.column("visible",  width=120, anchor="center")
        tree.column("pw_show",  width=140, anchor="center")

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            for uname, ud in db["users"].items():
                status = "🔴 تعلیق" if ud.get("suspended") else "🟢 فعال"
                vis_txt = "🚫 عدم نمایش" if ud.get("hidden") else "🟢 نمایش"
                pw_disp = ud.get("initial_pw", "••••")
                tag = "susp" if ud.get("suspended") else ("hidden" if ud.get("hidden") else ("adm" if ud.get("role")=="admin" else ""))
                tree.insert("", "end", values=(
                    uname, ud.get("display","—"), ud.get("role","—"), status, vis_txt, pw_disp,
                ), tags=(tag,))
            tree.tag_configure("susp", background="#4a3030", foreground=C["danger"])
            tree.tag_configure("adm",  background="#3a3820", foreground=C["gold"])
            tree.tag_configure("hidden", background="#3a3a3a", foreground=C["text_dim"])

        refresh()

        # ── ویرایش اینلاین مثل اکسل ──
        _inline_editor = [None]  # [widget] برای بستن ویرایشگر قبلی

        def _close_inline():
            if _inline_editor[0]:
                try: _inline_editor[0].destroy()
                except: pass
                _inline_editor[0] = None

        def on_double_click(event):
            _close_inline()
            region = tree.identify_region(event.x, event.y)
            if region != "cell": return
            col_id = tree.identify_column(event.x)
            row_id = tree.identify_row(event.y)
            if not row_id: return

            # شماره ستون (col_id = "#1", "#2", ...)
            # cols معکوس نمایش داده می‌شه (RTL)، باید از displaycolumns بگیریم
            disp_cols = list(tree["displaycolumns"])
            disp_idx  = int(col_id.replace("#","")) - 1
            if disp_idx < 0 or disp_idx >= len(disp_cols): return
            col_name = disp_cols[disp_idx]
            all_cols = list(tree["columns"])
            values   = tree.item(row_id, "values")
            # slab_id (uname) همیشه اولین ستون داخلی
            uname = values[0]

            # ستون نام کاربری و وضعیت با ورودی معمولی ویرایش نمی‌شن
            if col_name == "uname":
                messagebox.showinfo("راهنما", "نام کاربری قابل تغییر نیست.", parent=tab)
                return
            if col_name == "status":
                # toggle تعلیق با دابل‌کلیک روی ستون وضعیت
                if uname == "admin":
                    messagebox.showwarning("خطا", "نمی‌توان مدیر را تعلیق کرد.", parent=tab)
                    return
                db = load_db()
                db["users"][uname]["suspended"] = not db["users"][uname].get("suspended", False)
                save_db(db)
                refresh()
                return
            if col_name == "visible":
                # toggle نمایش/عدم‌نمایش در صفحه لاگین با دابل‌کلیک روی ستون نمایش
                if uname == "admin":
                    messagebox.showwarning("خطا", "نمی‌توان مدیر را از صفحه ورود مخفی کرد.", parent=tab)
                    return
                db = load_db()
                db["users"][uname]["hidden"] = not db["users"][uname].get("hidden", False)
                save_db(db)
                refresh()
                return

            # موقعیت سلول برای نمایش ویرایشگر روی آن
            x, y, w, h = tree.bbox(row_id, col_id)
            if not x: return

            try:
                internal_idx = all_cols.index(col_name)
                cur_val = str(values[internal_idx]) if internal_idx < len(values) else ""
            except (ValueError, IndexError):
                cur_val = ""

            if col_name == "role":
                # کمبوباکس برای نقش
                role_var = tk.StringVar(value=cur_val)
                role_opts = ["shift", "scarf", "shift_n", "admin"]
                cb = ttk.Combobox(tree, textvariable=role_var, values=role_opts,
                                  font=(_MAIN_FONT, 11, "bold"), state="readonly", width=w//10)
                cb.place(x=x, y=y, width=w, height=h)
                cb.set(cur_val)
                cb.focus_set()
                _inline_editor[0] = cb

                def save_role(e=None):
                    new_role = role_var.get().strip()
                    if new_role and new_role in role_opts:
                        db = load_db()
                        db["users"][uname]["role"] = new_role
                        save_db(db)
                    _close_inline()
                    refresh()

                cb.bind("<<ComboboxSelected>>", save_role)
                cb.bind("<Escape>", lambda e: (_close_inline(), refresh()))
                cb.bind("<FocusOut>", save_role)

            else:
                # ورودی متنی برای نام نمایشی و رمز عبور
                is_pw = (col_name == "pw_show")
                ent_var = tk.StringVar(value=cur_val)
                ent = tk.Entry(tree, textvariable=ent_var,
                               show="●" if is_pw else "",
                               bg=C["accent"], fg="#1a1a1a",
                               insertbackground="#1a1a1a",
                               font=(_MAIN_FONT, 11, "bold"),
                               justify="right", bd=0, relief="flat",
                               selectbackground=C["accent2"])
                ent.place(x=x, y=y, width=w, height=h)
                ent.select_range(0, "end")
                ent.focus_set()
                _inline_editor[0] = ent

                def save_inline(e=None):
                    new_val = ent_var.get().strip()
                    if new_val:
                        db = load_db()
                        if col_name == "display":
                            db["users"][uname]["display"] = new_val
                        elif col_name == "pw_show":
                            if not new_val:
                                messagebox.showerror("خطا", "رمز نمی‌تواند خالی باشد.", parent=tab)
                                _close_inline()
                                return
                            db["users"][uname]["password"] = hash_pw(new_val)
                            db["users"][uname]["initial_pw"] = new_val
                        save_db(db)
                    _close_inline()
                    refresh()

                ent.bind("<Return>", save_inline)
                ent.bind("<Tab>", save_inline)
                ent.bind("<Escape>", lambda e: (_close_inline(), refresh()))
                ent.bind("<FocusOut>", save_inline)

        tree.bind("<Double-Button-1>", on_double_click)
        # کلیک روی جای دیگه ویرایشگر رو می‌بنده
        tree.bind("<Button-1>", lambda e: _close_inline())

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=10, pady=4)

        def add_user():
            win = tk.Toplevel(self)
            prepare_popup_window(win, self)
            win.title("افزودن کاربر جدید")
            win.configure(bg=C["card"])
            win.geometry("400x320")
            self._center(win, 400, 320)
            tk.Label(win, text="➕  افزودن کاربر جدید",
                     bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(pady=(16,10))
            fields = {}
            for lbl, key in [("نام کاربری (انگلیسی):", "uname"),
                               ("نام نمایشی (فارسی):", "display"),
                               ("رمز عبور:", "password")]:
                row = tk.Frame(win, bg=C["card"])
                row.pack(fill="x", padx=30, pady=4)
                tk.Label(row, text=lbl, bg=C["card"], fg=C["text"],
                         font=FONT_NORM, anchor="e", width=20).pack(side="right")
                var = tk.StringVar()
                tk.Entry(row, textvariable=var, show="●" if key=="password" else "",
                         bg=C["entry_bg"], fg=C["text"], insertbackground=C["accent"],
                         font=FONT_NORM, justify="right", bd=0, relief="flat", highlightthickness=1,
                         highlightbackground=C["border"], highlightcolor=C["accent"], width=18
                         ).pack(side="right", padx=4)
                fields[key] = var
            role_row = tk.Frame(win, bg=C["card"])
            role_row.pack(fill="x", padx=30, pady=4)
            tk.Label(role_row, text="نقش:", bg=C["card"], fg=C["text"],
                     font=FONT_NORM, anchor="e", width=20).pack(side="right")
            role_cb = make_combo(role_row, ["shift (شیفت‌کار)", "scarf (اسکارف/برش)", "shift_n (نوبت‌کار)"], width=22)
            role_cb.set("shift (شیفت‌کار)")
            role_cb.pack(side="right", padx=4)
            err = tk.Label(win, text="", bg=C["card"], fg=C["danger"], font=FONT_SMALL)
            err.pack()
            def do_add():
                uname = fields["uname"].get().strip()
                display = fields["display"].get().strip()
                pw = fields["password"].get().strip()
                role = role_cb.get().split(" ")[0]
                if not uname or not display or not pw:
                    err.config(text="همه فیلدها الزامی هستند")
                    return
                db = load_db()
                if uname in db["users"]:
                    err.config(text="این نام کاربری قبلاً وجود دارد")
                    return
                db["users"][uname] = {
                    "password": hash_pw(pw), "role": role, "display": display,
                    "suspended": False, "hidden": False, "initial_pw": pw
                }
                save_db(db)
                win.destroy()
                refresh()
            styled_btn(win, "✔  افزودن کاربر", do_add, color=C["btn_success"]).pack(pady=12)

        def delete_user():
            sel = tree.selection()
            if not sel: return
            uname = tree.item(sel[0], "values")[0]
            if uname == "admin":
                messagebox.showwarning("خطا", "نمی‌توان مدیر را حذف کرد.", parent=self)
                return
            if messagebox.askyesno("حذف کاربر",
                    f"آیا مطمئنید کاربر «{uname}» حذف شود؟\nاین عمل برگشت‌پذیر نیست.", parent=self):
                db = load_db()
                del db["users"][uname]
                save_db(db)
                refresh()

        styled_btn(ctrl, "➕  افزودن کاربر", add_user, color=C["btn_success"]).pack(side="right", padx=4)
        styled_btn(ctrl, "🗑  حذف کاربر", delete_user, color=C["btn_danger"]).pack(side="right", padx=4)
        tk.Label(ctrl,
                 text="💡 دابل‌کلیک = ویرایش  |  وضعیت = تعلیق/فعال  |  نمایش در لاگین = نمایش/عدم‌نمایش",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(side="right", padx=10)

    def _admin_edit_records(self, nb):
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="✏️  ویرایش / حذف رکوردها")
        top = card_frame(tab)
        top.pack(fill="x", padx=12, pady=8)
        tf = tk.Frame(top, bg=C["card"])
        tf.pack(padx=12, pady=8, fill="x")
        tk.Label(tf, text="انتخاب بخش:", bg=C["card"], fg=C["text"],
                 font=FONT_NORM).pack(side="right", padx=(0,6))
        SECTIONS = {
            "ذوب جدید":           "melts",
            "اسکارف / برش":       "scarf_cut",
            "باومن":               "bauman",
            "انتقال انبار روباز":  "transfers_out",
            "تحویل آزمایشگاه":    "lab_deliveries",
            "قراضه":               "scrap",
            "برگشت‌ها":            "returns",
            "تاریخچه برگشت‌ها":   "return_log",
        }
        section_cb = make_combo(tf, list(SECTIONS.keys()), width=22)
        section_cb.set("ذوب جدید")
        section_cb.pack(side="right")

        cols = ("idx", "slab_id", "f1", "f2", "f3", "f4")
        heads = ("#", "شماره اسلب", "فیلد ۱", "فیلد ۲", "فیلد ۳", "فیلد ۴")
        tree_frame = tk.Frame(tab, bg=C["panel"])
        tree_frame.pack(fill="both", expand=True, padx=12, pady=4)
        style = ttk.Style()
        style.configure("Edit.Treeview", background=C["card"], foreground=C["text"],
            fieldbackground=C["card"], rowheight=28, font=(_MAIN_FONT, 10, "bold"))
        style.configure("Edit.Treeview.Heading", background=C["header_bg"],
            foreground=C["accent"], font=(_MAIN_FONT, 10, "bold"))
        style.map("Edit.Treeview", background=[("selected", C["accent2"])],
                  foreground=[("selected", C["bg"])])
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                            style="Edit.Treeview", height=14)
        for c, h in zip(cols, heads):
            tree.heading(c, text=h)
            tree.column(c, width=80 if c=="idx" else 160, anchor="center")
        tree.column("slab_id", width=140)
        vsb = tk.Scrollbar(tree_frame, orient="vertical", command=tree.yview,
                            bg="#707070", troughcolor="#1a1a1a",
                            activebackground=C["accent"], width=16)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def load_section(*_):
            key = SECTIONS[section_cb.get()]
            db = load_db()
            records = db.get(key, [])
            tree.delete(*tree.get_children())
            for i, rec in enumerate(records):
                sid = rec.get("slab_id", rec.get("slab","—"))
                vals = list(rec.values())
                extra = [str(v) for v in vals if str(v) != sid][:4]
                while len(extra) < 4: extra.append("—")
                tree.insert("", "end", iid=str(i),
                            values=(i+1, sid, extra[0], extra[1], extra[2], extra[3]))

        section_cb.bind("<<ComboboxSelected>>", load_section)
        load_section()

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=12, pady=6)

        def edit_record():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک رکورد انتخاب کنید.", parent=self)
                return
            idx = int(tree.item(sel[0], "values")[0]) - 1
            key = SECTIONS[section_cb.get()]
            db = load_db()
            rec = db[key][idx]
            win = tk.Toplevel(self)
            prepare_popup_window(win, self)
            win.title(f"ویرایش رکورد — {section_cb.get()}")
            win.configure(bg=C["card"])
            win.geometry("460x420")
            self._center(win, 460, 420)
            tk.Label(win, text=f"✏️  ویرایش رکورد #{idx+1}",
                     bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(pady=(14,10))
            vars_ = {}
            entries_frame = tk.Frame(win, bg=C["card"])
            entries_frame.pack(fill="both", expand=True, padx=20)
            for field, val in rec.items():
                row = tk.Frame(entries_frame, bg=C["card"])
                row.pack(fill="x", pady=3)
                tk.Label(row, text=f"{field}:", bg=C["card"], fg=C["text_dim"],
                         font=FONT_SMALL, width=18, anchor="e").pack(side="right")
                var = tk.StringVar(value=str(val))
                tk.Entry(row, textvariable=var, bg=C["entry_bg"], fg=C["text"],
                         insertbackground=C["accent"], font=FONT_NORM, justify="right", bd=0, relief="flat",
                         highlightthickness=1, highlightbackground=C["border"],
                         highlightcolor=C["accent"], width=24).pack(side="right", padx=4)
                vars_[field] = var
            def save_edit():
                for field, var in vars_.items():
                    rec[field] = var.get()
                db[key][idx] = rec
                save_db(db)
                win.destroy()
                load_section()
                messagebox.showinfo("موفق", "رکورد ویرایش شد.", parent=self)
            styled_btn(win, "💾  ذخیره تغییرات", save_edit, color=C["btn_success"]).pack(pady=12)

        def delete_record():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک رکورد انتخاب کنید.", parent=self)
                return
            idx = int(tree.item(sel[0], "values")[0]) - 1
            sid = tree.item(sel[0], "values")[1]
            key = SECTIONS[section_cb.get()]
            if key == "return_log":
                messagebox.showerror("ممنوع",
                    "تاریخچه برگشت‌ها قابل حذف نیست.\nاین اطلاعات دائمی هستند.", parent=self)
                return
            if messagebox.askyesno("حذف رکورد",
                    f"آیا مطمئنید رکورد اسلب «{sid}» حذف شود؟\nاین عمل برگشت‌پذیر نیست.", parent=self):
                db = load_db()
                db[key].pop(idx)
                save_db(db)
                load_section()

        def delete_all():
            key = SECTIONS[section_cb.get()]
            if key == "return_log":
                messagebox.showerror("ممنوع",
                    "تاریخچه برگشت‌ها قابل حذف نیست.\nاین اطلاعات دائمی هستند.", parent=self)
                return
            name = section_cb.get()
            if messagebox.askyesno("حذف همه رکوردها",
                    f"⚠️  آیا مطمئنید همه رکوردهای بخش «{name}» حذف شوند؟\n"
                    "این عمل برگشت‌پذیر نیست!", parent=self):
                db = load_db()
                db[key] = []
                save_db(db)
                load_section()

        styled_btn(ctrl, "✏️  ویرایش رکورد انتخابی", edit_record, color=C["warning"]).pack(side="right", padx=4)
        styled_btn(ctrl, "🗑  حذف رکورد انتخابی", delete_record, color=C["btn_danger"]).pack(side="right", padx=4)
        styled_btn(ctrl, "🗑🗑  حذف همه این بخش", delete_all, color="#6a0000").pack(side="right", padx=4)



    def _pdf_font_tools(self):
        """بارگذاری فونت فارسی + bidi/reshape — در سطح کلاس تا همه متدها بتوانند استفاده کنند"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Paragraph, Spacer, HRFlowable)
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.units import cm
        except ImportError:
            messagebox.showerror("خطا", "reportlab نصب نیست.\npip install reportlab", parent=self)
            return None

        try:
            import arabic_reshaper
            from bidi.algorithm import get_display as bidi_disp
            _reshaper = arabic_reshaper.ArabicReshaper(dict(
                delete_harakat=False, support_ligatures=True))
            def rt(s):
                return bidi_disp(_reshaper.reshape(str(s)))
        except ImportError:
            messagebox.showerror("خطا",
                "arabic_reshaper یا python-bidi نصب نیست.\n"
                "pip install arabic_reshaper python-bidi", parent=self)
            return None

        FONT = "BNazanin"
        font_ok = False
        font_files = ["BNazanin.ttf", "B Nazanin.ttf", "b_nazanin.ttf", "b nazanin.ttf",
                      "BNazanin Bold.ttf", "B Nazanin Bold.ttf"]
        font_dirs  = [
            os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "Fonts"),
            r"C:\Windows\Fonts",
            r"C:\Users\{}\AppData\Local\Microsoft\Windows\Fonts".format(
                os.environ.get("USERNAME", "")),
            _app_base_dir(),
            os.getcwd(),
        ]
        for d in font_dirs:
            for ff in font_files:
                fp = os.path.join(d, ff)
                if os.path.exists(fp):
                    try:
                        pdfmetrics.registerFont(TTFont(FONT, fp))
                        font_ok = True
                        break
                    except Exception:
                        pass
            if font_ok:
                break

        if not font_ok:
            for fb_file in ["tahoma.ttf", "Tahoma.ttf", "arial.ttf", "Arial.ttf", "arialuni.ttf"]:
                for d in [os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "Fonts"),
                          r"C:\Windows\Fonts", _app_base_dir(), os.getcwd()]:
                    fp = os.path.join(d, fb_file)
                    if os.path.exists(fp):
                        try:
                            pdfmetrics.registerFont(TTFont("PersianFB", fp))
                            FONT = "PersianFB"
                            font_ok = True
                            break
                        except Exception:
                            pass
                if font_ok:
                    break

        if not font_ok:
            messagebox.showerror("خطا",
                "فونت فارسی پیدا نشد. BNazanin.ttf را کنار برنامه قرار دهید.",
                parent=self)
            return None

        return dict(
            A4=A4, rl_colors=rl_colors,
            SimpleDocTemplate=SimpleDocTemplate, Table=Table, TableStyle=TableStyle,
            Paragraph=Paragraph, Spacer=Spacer, HRFlowable=HRFlowable,
            ParagraphStyle=ParagraphStyle, cm=cm,
            FONT=FONT, rt=rt
        )

    def _admin_reports(self, nb):
        """گزارشات ادمین — مدیریت کاربران + دو گزارش + تنظیمات تم + بک‌آپ"""
        sub = ttk.Notebook(nb, style="Dark.TNotebook")

        # ── سربرگ ۱: مدیریت کاربران ──
        t_users = tk.Frame(sub, bg=C["panel"])
        sub.add(t_users, text="👥  مدیریت کاربران")

        # ── سربرگ ۲: گزارش ثبت ذوب ──
        t1 = tk.Frame(sub, bg=C["panel"])
        sub.add(t1, text="🔥  گزارش ثبت ذوب")

        # ── سربرگ ۳: گزارش کنترل کیفی ──
        t2 = tk.Frame(sub, bg=C["panel"])
        sub.add(t2, text="✅  گزارش کنترل کیفی")

        # ── سربرگ ۴: تنظیمات تم ──
        t_theme = tk.Frame(sub, bg=C["panel"])
        sub.add(t_theme, text="🎨  تنظیمات تم")

        # ── سربرگ: به‌روزرسانی نرم‌افزار کلاینت‌ها ──
        t_upd = tk.Frame(sub, bg=C["panel"])
        sub.add(t_upd, text="⬆  به‌روزرسانی نرم‌افزار کلاینت‌ها")

        # ── سربرگ ۵: بک‌آپ ──
        t3 = tk.Frame(sub, bg=C["panel"])
        sub.add(t3, text="💾  بک‌آپ")

        # ── سربرگ ۶: بازنشانی سیستم — کنار بک‌آپ ──
        t_reset = tk.Frame(sub, bg=C["panel"])
        sub.add(t_reset, text="♻️  بازنشانی سیستم")

        # ── سربرگ ۷: لاگ ورود کاربران ──
        t_login = tk.Frame(sub, bg=C["panel"])
        sub.add(t_login, text="🔐  لاگ ورود کاربران")

        sub.pack(fill="both", expand=True)

        self._build_users_panel(t_users)
        self._build_melt_report(t1)
        self._build_qc_excel_report(t2)
        self._admin_settings_inline(t_theme)
        self._build_client_update_tab(t_upd)
        self._build_backup_tab(t3)
        self._build_reset_tab(t_reset)
        self._admin_login_log(t_login)

    def _build_users_panel(self, tab):
        """پنل مدیریت کاربران — نمایش در گزارشات"""
        tab.configure(bg=C["panel"])
        tk.Label(tab, text="👥  مدیریت کاربران سیستم",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        tk.Label(tab,
                 text="✏️  روی هر سلول دابل‌کلیک کنید تا مستقیم ویرایش کنید",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        # ── مدیریت رمز Minimize — فقط با کلیک روی آیکون تغییر می‌کند ──
        self._build_minimize_pw_row(tab)
        self._build_backup_pw_row(tab)

        cols  = ("uname","display","role","status","visible","pw_show")
        heads = ("نام کاربری","نام نمایشی","نقش","وضعیت","نمایش در لاگین","رمز عبور")
        tf, tree = scrolled_tree(tab, cols, heads, height=12)
        tf.pack(fill="both", expand=True, padx=12, pady=8)
        tree.column("uname",   width=130, anchor="center")
        tree.column("display", width=180, anchor="center")
        tree.column("role",    width=110, anchor="center")
        tree.column("status",  width=100, anchor="center")
        tree.column("visible", width=120, anchor="center")
        tree.column("pw_show", width=140, anchor="center")

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            for uname, ud in db["users"].items():
                status = "🔴 تعلیق" if ud.get("suspended") else "🟢 فعال"
                vis_txt = "🚫 عدم نمایش" if ud.get("hidden") else "🟢 نمایش"
                pw_disp = ud.get("initial_pw","••••")
                tag = "susp" if ud.get("suspended") else ("hidden" if ud.get("hidden") else ("adm" if ud.get("role")=="admin" else ""))
                tree.insert("","end", values=(
                    uname, ud.get("display","—"), ud.get("role","—"), status, vis_txt, pw_disp
                ), tags=(tag,))
            tree.tag_configure("susp", background="#4a3030", foreground=C["danger"])
            tree.tag_configure("adm",  background="#3a3820", foreground=C["gold"])
            tree.tag_configure("hidden", background="#3a3a3a", foreground=C["text_dim"])

        refresh()

        # ── ویرایش اینلاین مثل اکسل ──
        _inline_editor2 = [None]

        def _close_inline2():
            if _inline_editor2[0]:
                try: _inline_editor2[0].destroy()
                except: pass
                _inline_editor2[0] = None

        def on_double_click2(event):
            _close_inline2()
            region = tree.identify_region(event.x, event.y)
            if region != "cell": return
            col_id = tree.identify_column(event.x)
            row_id = tree.identify_row(event.y)
            if not row_id: return

            # cols معکوس نمایش داده می‌شه (RTL)، باید از displaycolumns بگیریم
            disp_cols = list(tree["displaycolumns"])
            disp_idx  = int(col_id.replace("#","")) - 1
            if disp_idx < 0 or disp_idx >= len(disp_cols): return
            col_name = disp_cols[disp_idx]
            all_cols = list(tree["columns"])
            values   = tree.item(row_id, "values")
            # slab_id (uname) همیشه اولین ستون داخلی
            uname = values[0]

            if col_name == "uname":
                messagebox.showinfo("راهنما", "نام کاربری قابل تغییر نیست.", parent=tab)
                return
            if col_name == "status":
                if uname == "admin":
                    messagebox.showwarning("خطا", "نمی‌توان مدیر را تعلیق کرد.", parent=tab)
                    return
                db = load_db()
                db["users"][uname]["suspended"] = not db["users"][uname].get("suspended", False)
                save_db(db)
                refresh()
                return
            if col_name == "visible":
                if uname == "admin":
                    messagebox.showwarning("خطا", "نمی‌توان مدیر را از صفحه ورود مخفی کرد.", parent=tab)
                    return
                db = load_db()
                db["users"][uname]["hidden"] = not db["users"][uname].get("hidden", False)
                save_db(db)
                refresh()
                return

            x, y, w, h = tree.bbox(row_id, col_id)
            if not x: return
            try:
                internal_idx = all_cols.index(col_name)
                cur_val = str(values[internal_idx]) if internal_idx < len(values) else ""
            except (ValueError, IndexError):
                cur_val = ""

            if col_name == "role":
                role_var = tk.StringVar(value=cur_val)
                role_opts = ["shift", "scarf", "shift_n", "admin"]
                cb = ttk.Combobox(tree, textvariable=role_var, values=role_opts,
                                  font=(_MAIN_FONT, 11, "bold"), state="readonly", width=w//10)
                cb.place(x=x, y=y, width=w, height=h)
                cb.set(cur_val)
                cb.focus_set()
                _inline_editor2[0] = cb

                def save_role2(e=None):
                    new_role = role_var.get().strip()
                    if new_role and new_role in role_opts:
                        db = load_db()
                        db["users"][uname]["role"] = new_role
                        save_db(db)
                    _close_inline2()
                    refresh()

                cb.bind("<<ComboboxSelected>>", save_role2)
                cb.bind("<Escape>", lambda e: (_close_inline2(), refresh()))
                cb.bind("<FocusOut>", save_role2)
            else:
                is_pw = (col_name == "pw_show")
                ent_var = tk.StringVar(value=cur_val)
                ent = tk.Entry(tree, textvariable=ent_var,
                               show="●" if is_pw else "",
                               bg=C["accent"], fg="#1a1a1a",
                               insertbackground="#1a1a1a",
                               font=(_MAIN_FONT, 11, "bold"),
                               justify="right", bd=0, relief="flat",
                               selectbackground=C["accent2"])
                ent.place(x=x, y=y, width=w, height=h)
                ent.select_range(0, "end")
                ent.focus_set()
                _inline_editor2[0] = ent

                def save_inline2(e=None):
                    new_val = ent_var.get().strip()
                    if new_val:
                        db = load_db()
                        if col_name == "display":
                            db["users"][uname]["display"] = new_val
                        elif col_name == "pw_show":
                            if not new_val:
                                messagebox.showerror("خطا", "رمز نمی‌تواند خالی باشد.", parent=tab)
                                _close_inline2()
                                return
                            db["users"][uname]["password"] = hash_pw(new_val)
                            db["users"][uname]["initial_pw"] = new_val
                        save_db(db)
                    _close_inline2()
                    refresh()

                ent.bind("<Return>", save_inline2)
                ent.bind("<Tab>", save_inline2)
                ent.bind("<Escape>", lambda e: (_close_inline2(), refresh()))
                ent.bind("<FocusOut>", save_inline2)

        tree.bind("<Double-Button-1>", on_double_click2)
        tree.bind("<Button-1>", lambda e: _close_inline2())

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=12, pady=6)

        def add_user():
            win = tk.Toplevel(self)
            prepare_popup_window(win, self)
            win.title("افزودن کاربر جدید")
            win.configure(bg=C["card"])
            win.geometry("400x320")
            self._center(win, 400, 320)
            tk.Label(win, text="➕  افزودن کاربر جدید",
                     bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(pady=(16,10))
            fields = {}
            for lbl, key in [("نام کاربری (انگلیسی):","uname"),
                               ("نام نمایشی (فارسی):","display"),
                               ("رمز عبور:","password")]:
                row = tk.Frame(win, bg=C["card"]); row.pack(fill="x", padx=30, pady=4)
                tk.Label(row, text=lbl, bg=C["card"], fg=C["text"],
                         font=FONT_NORM, anchor="e", width=20).pack(side="right")
                var = tk.StringVar()
                tk.Entry(row, textvariable=var, show="●" if key=="password" else "",
                         bg=C["entry_bg"], fg=C["text"], insertbackground=C["accent"],
                         font=FONT_NORM, justify="right", bd=0, relief="flat",
                         highlightthickness=1, highlightbackground=C["border"],
                         highlightcolor=C["accent"], width=18).pack(side="right", padx=4)
                fields[key] = var
            role_row = tk.Frame(win, bg=C["card"]); role_row.pack(fill="x", padx=30, pady=4)
            tk.Label(role_row, text="نقش:", bg=C["card"], fg=C["text"],
                     font=FONT_NORM, anchor="e", width=20).pack(side="right")
            role_cb = make_combo(role_row,
                ["shift (شیفت‌کار)","scarf (اسکارف/برش)","shift_n (نوبت‌کار)"], width=22)
            role_cb.set("shift (شیفت‌کار)")
            role_cb.pack(side="right", padx=4)
            err = tk.Label(win, text="", bg=C["card"], fg=C["danger"], font=FONT_SMALL)
            err.pack()
            def do_add():
                uname = fields["uname"].get().strip()
                display = fields["display"].get().strip()
                pw = fields["password"].get().strip()
                role = role_cb.get().split(" ")[0]
                if not uname or not display or not pw:
                    err.config(text="همه فیلدها الزامی هستند"); return
                db = load_db()
                if uname in db["users"]:
                    err.config(text="این نام کاربری قبلاً وجود دارد"); return
                db["users"][uname] = {
                    "password": hash_pw(pw), "role": role,
                    "display": display, "suspended": False, "hidden": False, "initial_pw": pw
                }
                save_db(db); win.destroy(); refresh()
            styled_btn(win,"✔  افزودن کاربر",do_add,color=C["btn_success"]).pack(pady=12)

        def delete_user():
            sel = tree.selection()
            if not sel: return
            uname = tree.item(sel[0],"values")[0]
            if uname == "admin":
                messagebox.showwarning("خطا","نمی‌توان مدیر را حذف کرد.",parent=self); return
            if messagebox.askyesno("حذف کاربر",
                    f"آیا مطمئنید کاربر «{uname}» حذف شود؟\nبرگشت‌پذیر نیست.",parent=self):
                db = load_db()
                del db["users"][uname]
                save_db(db); refresh()

        styled_btn(ctrl,"➕  افزودن",  add_user,    color=C["btn_success"]).pack(side="right",padx=4)
        styled_btn(ctrl,"🗑  حذف",     delete_user, color=C["btn_danger"]).pack(side="right",padx=4)
        tk.Label(ctrl,
                 text="💡 دابل‌کلیک = ویرایش  |  وضعیت = تعلیق/فعال  |  نمایش در لاگین = نمایش/عدم‌نمایش",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(side="right", padx=10)

    def _date_filter_row(self, parent):
        """فیلتر تاریخ و ساعت — لیبل سمت راست، باکس سمت چپ"""
        ff = card_frame(parent)
        ff.pack(fill="x", padx=14, pady=8)
        fi = tk.Frame(ff, bg=C["card"])
        fi.pack(padx=14, pady=10, fill="x")
        _now = to_shamsi(datetime.datetime.now()).split("  ")[0]
        _first_dt = get_first_report_date_sh()

        def mk_e(parent, var, w=13):
            return tk.Entry(parent, textvariable=var,
                            bg=C["entry_bg"], fg=C["text"],
                            insertbackground=C["accent"], font=FONT_MONO,
                            justify="right", bd=0, relief="flat",
                            highlightthickness=1, highlightbackground=C["border"],
                            highlightcolor=C["accent"], width=w)

        def mk_pair(parent, lbl, var, w):
            f = tk.Frame(parent, bg=C["card"])
            f.pack(side="right", padx=(0, 10))
            tk.Label(f, text=lbl, bg=C["card"], fg=C["accent"],
                     font=(_MAIN_FONT, 11, "bold")).pack(side="right")
            mk_e(f, var, w).pack(side="right", padx=(4, 0))
            return f

        # ردیف تاریخ
        row_d = tk.Frame(fi, bg=C["card"])
        row_d.pack(fill="x", pady=(0, 4))
        from_date_v = tk.StringVar(value=_first_dt)
        to_date_v   = tk.StringVar(value=_now)
        mk_pair(row_d, "از تاریخ", from_date_v, 13)
        mk_pair(row_d, "تا تاریخ", to_date_v,   13)

        # ردیف ساعت
        row_t = tk.Frame(fi, bg=C["card"])
        row_t.pack(fill="x")
        from_time_v = tk.StringVar(value="00:00:00")
        to_time_v   = tk.StringVar(value="23:59:59")
        mk_pair(row_t, "از ساعت", from_time_v, 9)
        mk_pair(row_t, "تا ساعت", to_time_v,   9)

        from_v = tk.StringVar()
        to_v   = tk.StringVar()
        def _sync(*_):
            from_v.set(f"{from_date_v.get().strip()}  {from_time_v.get().strip()}")
            to_v.set(f"{to_date_v.get().strip()}  {to_time_v.get().strip()}")
        for v in (from_date_v, from_time_v, to_date_v, to_time_v):
            v.trace_add("write", _sync)
        _sync()
        return from_v, to_v

    def _in_range(self, at, from_v, to_v):
        f = from_v.get().strip().replace("  ", " ")
        t = to_v.get().strip().replace("  ", " ")
        if not f or not t: return True
        a = (at or "").replace("  ", " ")
        return f <= a <= t

    def _build_melt_report(self, tab):
        """گزارش ثبت ذوب — Excel حرفه‌ای"""
        tab.configure(bg=C["panel"])
        tk.Label(tab, text="🔥  گزارش ثبت ذوب جدید",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        tk.Label(tab, text="خروجی Excel با هدر حرفه‌ای، رنگ‌بندی سطری و اطلاعات کامل شرکت",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)
        from_v, to_v = self._date_filter_row(tab)
        cols  = ("slab_id","registered_by","reg_date","reg_time")
        heads = ("شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت")
        tf, tree = scrolled_tree(tab, cols, heads, height=16)
        tf.pack(fill="both", expand=True, padx=14, pady=6)
        tree.column("slab_id",       width=160, anchor="center")
        tree.column("registered_by", width=160, anchor="center")
        tree.column("reg_date",      width=140, anchor="center")
        tree.column("reg_time",      width=100, anchor="center")
        cnt = tk.Label(tab,"",bg=C["panel"],fg=C["text_dim"],font=FONT_SMALL)
        cnt.pack(anchor="e", padx=14)

        def get_rows():
            db = load_db(); rows = []
            for rec in db["melts"]:
                if not self._in_range(rec.get("registered_at",""), from_v, to_v): continue
                who = get_display_name(rec.get("registered_by","—"), db) if self.role=="admin" else "شخص دیگر"
                _d, _t = split_dt(rec.get("registered_at","—"))
                rows.append((rec["slab_id"], who, _d, _t))
            return rows

        def refresh():
            tree.delete(*tree.get_children())
            rows = get_rows()
            for r in rows: tree.insert("","end", values=r)
            cnt.config(text=f"تعداد: {len(rows)} اسلب")

        def export_xl():
            if not XLSX:
                messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            rows = get_rows()
            if not rows:
                messagebox.showinfo("خطا","داده‌ای در این بازه یافت نشد.",parent=self); return
            path = self._resolve_report_save_path(
                "melt_excel",
                f"گزارش_ذوب_{to_shamsi(datetime.datetime.now()).split()[0].replace('/','')}.xlsx",
                [("Excel","*.xlsx")], ".xlsx")
            if not path: return
            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "ثبت ذوب"
            ws.sheet_view.rightToLeft = True

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            fn   = "B Nazanin"
            thin = Side(style="thin", color="BBBBBB")
            bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
            c_al = Alignment(horizontal="center", vertical="center", wrap_text=False, readingOrder=2)
            ws.merge_cells("A1:E1")
            t1 = ws["A1"]
            t1.value = "شرکت سازه پیشگام مدیسه — فولاد سفید دشت"
            t1.font  = Font(name=fn, bold=True, size=14, color="FFFFFF")
            t1.fill  = PatternFill("solid", fgColor="0D1E3C")
            t1.alignment = c_al
            ws.row_dimensions[1].height = 32

            # ── ردیف ۲: عنوان گزارش (merge A:E) ──
            ws.merge_cells("A2:E2")
            t2 = ws["A2"]
            t2.value = "گزارش ثبت ذوب"
            t2.font  = Font(name=fn, bold=True, size=12, color="003366")
            t2.fill  = PatternFill("solid", fgColor="DCE8F4")
            t2.alignment = c_al
            ws.row_dimensions[2].height = 26

            # ── ردیف ۳: اطلاعات بازه (merge A:E) ──
            f_str = from_v.get().strip(); t_str = to_v.get().strip()
            ws.merge_cells("A3:E3")
            t3 = ws["A3"]
            t3.value = f"بازه زمانی:  از  {f_str}  تا  {t_str}   |   تعداد: {len(rows)} اسلب   |   تاریخ چاپ: {to_shamsi(datetime.datetime.now()).split('  ')[0]}"
            t3.font  = Font(name=fn, size=9, color="555555")
            t3.fill  = PatternFill("solid", fgColor="F0F4F8")
            t3.alignment = c_al
            ws.row_dimensions[3].height = 20

            # ── ردیف ۴: سرتیتر ──
            hdr_cols = ["ردیف","شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت"]
            ws.merge_cells("A4:A4")  # ردیف
            for ci, h in enumerate(hdr_cols, 1):
                c = ws.cell(row=4, column=ci, value=h)
                c.font      = Font(name=fn, bold=True, size=11, color="FFFFFF")
                c.fill      = PatternFill("solid", fgColor="1A3A5C")
                c.alignment = c_al
                c.border    = bdr
            ws.row_dimensions[4].height = 24

            # ── داده‌ها ──
            for ri, (sid, who, _date, _time) in enumerate(rows, 1):
                fill_clr = "FFFFFF" if ri%2==1 else "EEF4FB"
                row_vals = [ri, sid, who, _date, _time]
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=ri+4, column=ci, value=val)
                    c.font      = Font(name=fn, bold=True, size=11)
                    c.fill      = PatternFill("solid", fgColor=fill_clr)
                    c.alignment = c_al
                    c.border    = bdr
                ws.row_dimensions[ri+4].height = 20

            # ── عرض ستون‌ها — auto-fit بر اساس محتوا ──
            col_data = {1: ["ردیف"], 2: ["شماره اسلب"], 3: ["ثبت‌کننده"], 4: ["تاریخ ثبت"], 5: ["ساعت ثبت"]}
            for ri, (sid, who, _date, _time) in enumerate(rows, 1):
                for ci, val in zip([1,2,3,4,5], [str(ri), sid, who, _date, _time]):
                    col_data[ci].append(str(val) if val else "—")
            for ci in range(1, 6):
                max_len = max(len(s) for s in col_data[ci])
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len * 1.3 + 4, 60)

            # ── فوتر ──
            foot_row = len(rows) + 7
            ws.merge_cells(f"A{foot_row}:E{foot_row}")
            fc = ws[f"A{foot_row}"]
            fc.value     = "این گزارش به صورت سیستمی تولید شده است — سامانه مدیریت تختال، شرکت سازه پیشگام مدیسه"
            fc.font      = Font(name=fn, size=8, italic=True, color="999999")
            fc.alignment = c_al

            ws.freeze_panes = "A5"
            wb.save(path)
            messagebox.showinfo("موفق", f"Excel ذخیره شد:\n{path}", parent=self)

        refresh()
        ctrl = tk.Frame(tab, bg=C["panel"]); ctrl.pack(fill="x", padx=14, pady=4)
        styled_btn(ctrl,"📥  خروجی Excel",export_xl,color=C["btn_success"]).pack(side="right",padx=4)
        styled_btn(ctrl,"🔄  بروزرسانی",refresh,color=C["card"]).pack(side="right",padx=4)

    def _build_qc_excel_report(self, tab):
        """گزارش QC: هر اسلب یک سطر — ستون انتقال ۱، انتقال ۲، ... داینامیک"""
        tab.configure(bg=C["panel"])
        tk.Label(tab, text="✅  گزارش جامع کنترل کیفی",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        tk.Label(tab,
                 text="هر اسلب در یک سطر — به ازای هر انتقال یک ستون جداگانه",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)
        from_v, to_v = self._date_filter_row(tab)

        # ستون‌های ثابت — تاریخ و ساعت همه جدا
        FIXED_COLS  = ("slab_id","registered_by","reg_date","reg_time",
                       "qc_by","qc_date","qc_time","qc_status",
                       "scarf","scarf_by","scarf_date","scarf_time",
                       "cut","cut_count","cut_by","cut_date","cut_time",
                       "bauman","lab","lab_by","lab_date","lab_time",
                       "cur_loc","exit_s","exit_by","scrap")
        FIXED_HEADS = ("شماره اسلب","ثبت‌کننده ذوب","تاریخ ثبت","ساعت ثبت",
                       "تأییدکننده QC","تاریخ QC","ساعت QC","وضعیت QC",
                       "اسکارف","ثبت‌کننده اسکارف","تاریخ اسکارف","ساعت اسکارف",
                       "برش","تعداد برش","ثبت‌کننده برش","تاریخ برش","ساعت برش",
                       "تست باومن","تحویل آزمایشگاه","تحویل‌دهنده","تاریخ تحویل","ساعت تحویل",
                       "محل فعلی","وضعیت خروج","خروج توسط","قراضه")

        # پیش‌نمایش — بدون ستون‌های انتقال داینامیک (فقط برای نمایش)
        preview_cols  = FIXED_COLS + ("transfers",)
        preview_heads = FIXED_HEADS + ("انتقال‌ها",)
        tf, tree = scrolled_tree(tab, preview_cols, preview_heads, height=13)
        tf.pack(fill="both", expand=True, padx=14, pady=6)
        for c in preview_cols: tree.column(c, width=100, anchor="center")
        tree.column("slab_id",   width=140)
        tree.column("reg_date",  width=120)
        tree.column("reg_time",  width=80)
        tree.column("qc_date",   width=120)
        tree.column("qc_time",   width=80)
        tree.column("scarf",     width=170)
        tree.column("cut",       width=170)
        tree.column("scrap",     width=170)
        tree.column("transfers", width=320, anchor="w")
        search_bar(tab, tree, col_indices=[0]).pack(anchor="e", padx=14, pady=2)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)
        cnt = tk.Label(tab, "", bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL)
        cnt.pack(anchor="e", padx=14)

        def get_moves(db, sid):
            """فقط انتقال‌های واقعی — تایید مجدد QC از این لیست حذف می‌شود"""
            return sorted(
                [m for m in db.get("movement_log",[])
                 if m.get("slab_id")==sid and m.get("operation","انتقال") == "انتقال"],
                key=lambda m: m.get("at","")
            )

        def moves_summary(db, sid):
            moves = get_moves(db, sid)
            if not moves: return "—"
            parts = []
            for i, m in enumerate(moves, 1):
                frm = m.get("from","—")
                to  = m.get("to","—")
                parts.append(f"انتقال {i}: از {frm} به {to}")
            return "  ●  ".join(parts)

        def get_fixed_row(db, rec):
            sid = rec["slab_id"]
            who = get_display_name(rec.get("qc_by","—"), db) if self.role=="admin" else "شخص دیگر"
            reg_who = get_display_name(rec.get("registered_by","—"), db) if self.role=="admin" else "شخص دیگر"
            sc = next((r for r in db.get("scarf_cut",[])
                       if r["slab_id"]==sid and r.get("operation")=="اسکارفی"), None)
            ct = next((r for r in db.get("scarf_cut",[])
                       if r["slab_id"]==sid and r.get("operation")=="برشی"), None)
            bm = next((r for r in db.get("bauman",[]) if r["slab_id"]==sid), None)
            lb = next((r for r in db.get("lab_deliveries",[]) if r["slab_id"]==sid), None)
            sp = next((r for r in db.get("scrap",[]) if r["slab_id"]==sid), None)
            cur = get_current_location(db, sid)
            exit_who = get_display_name(rec.get("exit_by","—"), db) if self.role=="admin" else "شخص دیگر"
            exit_s = rec.get("exit_status","در انبار")
            if exit_s != "خروج زده شده":
                exit_s = f"در {cur}"   # ذکر نام دقیق انبار به‌جای «در انبار» کلی
            cut_count_val = str(ct.get("cut_count",1)) + " بار" if ct else "—"
            # تفکیک تاریخ و ساعت همه فیلدها
            reg_d,   reg_t   = split_dt(rec.get("registered_at","—"))
            qc_d,    qc_t    = split_dt(rec.get("qc_at","—"))
            scarf_d, scarf_t = split_dt(sc.get("registered_at","—") if sc else "—")
            cut_d,   cut_t   = split_dt(ct.get("registered_at","—") if ct else "—")
            lab_d,   lab_t   = split_dt(lb.get("delivered_at","—") if lb else "—")
            return (
                sid, reg_who, reg_d, reg_t,
                who, qc_d, qc_t, rec.get("qc_status","ثبت شده"),
                ("دارد: "+sc.get("reason","")[:30]) if sc else "ندارد",
                (get_display_name(sc.get("registered_by","—"), db) if self.role=="admin" else "شخص دیگر") if sc else "—",
                scarf_d, scarf_t,
                ("دارد: "+ct.get("reason","")[:30]) if ct else "ندارد",
                cut_count_val,
                (get_display_name(ct.get("registered_by","—"), db) if self.role=="admin" else "شخص دیگر") if ct else "—",
                cut_d, cut_t,
                "✔ دارد" if bm else "ندارد",
                "✔ تحویل شد" if lb else ("آماده" if bm else "ندارد"),
                (get_display_name(lb.get("delivered_by","—"), db) if self.role=="admin" else "شخص دیگر") if lb else "—",
                lab_d, lab_t,
                cur,
                exit_s,
                exit_who if rec.get("exit_status")=="خروج زده شده" else "—",
                ("دارد: "+sp.get("reason","")[:30]) if sp else "ندارد",
            )

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db(); n = 0
            for rec in db["melts"]:
                at = rec.get("qc_at","") or rec.get("registered_at","")
                if not self._in_range(at, from_v, to_v): continue
                fixed = get_fixed_row(db, rec)
                tr_sum = moves_summary(db, rec["slab_id"])
                tree.insert("","end", values=fixed+(tr_sum,))
                n += 1
            cnt.config(text=f"تعداد: {n} اسلب")

        def export_xl():
            if not XLSX:
                messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            db = load_db()
            all_recs = [r for r in db["melts"]
                        if self._in_range(r.get("qc_at","") or r.get("registered_at",""), from_v, to_v)]
            if not all_recs:
                messagebox.showinfo("خطا","داده‌ای یافت نشد.",parent=self); return
            max_tr = max((len(get_moves(db, r["slab_id"])) for r in all_recs), default=0)
            tr_heads = [h for i in range(1, max_tr+1)
                        for h in (f"انتقال {i}", f"تاریخ انتقال {i}", f"ساعت انتقال {i}", f"انتقال‌دهنده {i}")]
            all_heads  = list(FIXED_HEADS) + tr_heads
            rows = []
            for rec in all_recs:
                fixed = get_fixed_row(db, rec)
                moves = get_moves(db, rec["slab_id"])
                tr_vals = []
                for m in moves:
                    frm = m.get("from","—")
                    to  = m.get("to","—")
                    at_d, at_t = split_dt(m.get("at","—"))
                    by  = get_display_name(m.get("by","—"), db) if self.role=="admin" else "شخص دیگر"
                    tr_vals.extend([f"از {frm} به {to}", at_d, at_t, by])
                while len(tr_vals) < max_tr*4:
                    tr_vals.append("—")
                rows.append(fixed + tuple(tr_vals))

            path = self._resolve_report_save_path(
                "qc_excel",
                f"گزارش_QC_{to_shamsi(datetime.datetime.now()).split()[0].replace('/','')}.xlsx",
                [("Excel","*.xlsx")], ".xlsx")
            if not path: return

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.page import PageMargins

            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "کنترل کیفی"
            ws.sheet_view.rightToLeft = True

            # ── Landscape + fit to one page wide ──
            ws.page_setup.orientation       = "landscape"
            ws.page_setup.fitToPage         = True
            ws.page_setup.fitToWidth        = 1   # همه ستون‌ها در یک صفحه عرض
            ws.page_setup.fitToHeight       = 0   # ارتفاع: هر چقدر لازم
            ws.page_setup.paperSize         = 9   # A4
            ws.page_margins                 = PageMargins(
                left=0.4, right=0.4, top=0.8, bottom=0.8,
                header=0.3, footer=0.3)

            fn   = "B Nazanin"
            thin = Side(style="thin",   color="BBBBBB")
            med  = Side(style="medium", color="003366")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            c_al = Alignment(horizontal="center", vertical="center",
                              wrap_text=False, readingOrder=2)
            total_cols = len(all_heads)

            # ── ردیف ۱: عنوان شرکت ──
            ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
            c = ws["A1"]
            c.value     = "شرکت سازه پیشگام مدیسه — فولاد سفید دشت"
            c.font      = Font(name=fn, bold=True, size=14, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="0D1E3C")
            c.alignment = c_al
            ws.row_dimensions[1].height = 30

            # ── ردیف ۲: عنوان گزارش ──
            ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
            c = ws["A2"]
            c.value     = "گزارش جامع کنترل کیفی"
            c.font      = Font(name=fn, bold=True, size=12, color="003366")
            c.fill      = PatternFill("solid", fgColor="DCE8F4")
            c.alignment = c_al
            ws.row_dimensions[2].height = 24

            # ── ردیف ۳: اطلاعات بازه ──
            f_str = from_v.get().strip(); t_str = to_v.get().strip()
            ws.merge_cells(f"A3:{get_column_letter(total_cols)}3")
            c = ws["A3"]
            c.value     = f"بازه:  {f_str}  تا  {t_str}   |   تعداد: {len(rows)} اسلب   |   تاریخ چاپ: {to_shamsi(datetime.datetime.now()).split('  ')[0]}"
            c.font      = Font(name=fn, size=9, color="555555")
            c.fill      = PatternFill("solid", fgColor="F0F4F8")
            c.alignment = c_al
            ws.row_dimensions[3].height = 18

            # ── ردیف ۴: سرتیتر ستون‌ها ──
            for ci, h in enumerate(all_heads, 1):
                c = ws.cell(row=4, column=ci, value=h)
                c.font      = Font(name=fn, bold=True, size=11, color="FFFFFF")
                c.fill      = PatternFill("solid", fgColor="1A3A5C")
                c.alignment = c_al
                c.border    = Border(left=thin,right=thin,
                                      top=Side(style="medium",color="003366"),
                                      bottom=Side(style="medium",color="003366"))
            ws.row_dimensions[4].height = 22
            ws.freeze_panes = "A5"

            # ── داده‌ها ──
            for ri, row in enumerate(rows, 1):
                fill_clr = "FFFFFF" if ri%2==1 else "EEF4FB"
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=ri+4, column=ci, value=str(val) if val else "—")
                    c.font      = Font(name=fn, bold=True, size=11)
                    c.fill      = PatternFill("solid", fgColor=fill_clr)
                    c.alignment = c_al
                    c.border    = bdr
                ws.row_dimensions[ri+4].height = 20

            # ── شکست صفحه هر ۱۵ اسلب — برای پرینت لندسکیپ ──
            from openpyxl.worksheet.pagebreak import Break
            ROWS_PER_PRINT_PAGE = 15
            for ri in range(ROWS_PER_PRINT_PAGE, len(rows), ROWS_PER_PRINT_PAGE):
                ws.row_breaks.append(Break(id=ri+4))

            # ── عرض ستون‌ها — auto-fit بر اساس محتوا ──
            from openpyxl.utils import get_column_letter as gcl2
            for ci in range(1, len(all_heads)+1):
                max_len = len(str(all_heads[ci-1]))
                for row in rows:
                    if ci-1 < len(row):
                        cell_len = len(str(row[ci-1]) if row[ci-1] is not None else "—")
                        if cell_len > max_len:
                            max_len = cell_len
                ws.column_dimensions[gcl2(ci)].width = min(max_len * 1.3 + 4, 60)

            # ── فوتر ──
            foot_row = len(rows) + 7
            ws.merge_cells(f"A{foot_row}:{get_column_letter(total_cols)}{foot_row}")
            fc = ws[f"A{foot_row}"]
            fc.value     = "این گزارش به صورت سیستمی تولید شده است — سامانه مدیریت تختال، شرکت سازه پیشگام مدیسه"
            fc.font      = Font(name=fn, size=8, italic=True, color="999999")
            fc.alignment = c_al

            wb.save(path)
            messagebox.showinfo("موفق", f"Excel ذخیره شد:\n{path}", parent=self)

        refresh()
        ctrl = tk.Frame(tab, bg=C["panel"]); ctrl.pack(fill="x", padx=14, pady=4)
        styled_btn(ctrl,"📥  خروجی Excel",export_xl,color=C["btn_success"]).pack(side="right",padx=4)
        styled_btn(ctrl,"🔄 بروزرسانی",refresh,color=C["card"]).pack(side="right",padx=4)


    def _make_excel_styles(self):
        hf = Font(name="B Nazanin", bold=True, color="1a1a1a", size=11)
        hfill = PatternFill("solid", fgColor="B8CCE4")
        cf = Font(name="B Nazanin", bold=True, color="1a1a1a", size=11)
        cfill = PatternFill("solid", fgColor="FFFFFF")
        afill = PatternFill("solid", fgColor="DCE6F1")
        bdr = Border(
            left=Side(style="thin",color="9DB3CC"),
            right=Side(style="thin",color="9DB3CC"),
            top=Side(style="thin",color="9DB3CC"),
            bottom=Side(style="thin",color="9DB3CC"))
        align = Alignment(horizontal="center", vertical="center", wrap_text=False, readingOrder=2)
        return hf, hfill, cf, cfill, afill, bdr, align

    def _write_sheet(self, ws, headers, rows, col_widths=None):
        hf,hfill,cf,cfill,afill,bdr,align = self._make_excel_styles()
        ws.sheet_view.rightToLeft = True
        for ci,h in enumerate(headers,1):
            c = ws.cell(row=1,column=ci,value=h)
            c.font=hf; c.fill=hfill; c.alignment=align; c.border=bdr
        ws.row_dimensions[1].height = 24
        for ri,row in enumerate(rows,2):
            fill = cfill if ri%2==0 else afill
            for ci,val in enumerate(row,1):
                c = ws.cell(row=ri,column=ci,value=str(val) if val is not None else "—")
                c.font=cf; c.fill=fill; c.alignment=align; c.border=bdr
            ws.row_dimensions[ri].height = 18
        # ── عرض ستون‌ها: همیشه auto-fit بر اساس محتوا ──
        from openpyxl.utils import get_column_letter
        col_count = len(headers)
        for ci in range(1, col_count+1):
            # حداقل عرض = طول سرتیتر
            max_len = len(str(headers[ci-1]))
            for row in rows:
                if ci-1 < len(row):
                    cell_len = len(str(row[ci-1]) if row[ci-1] is not None else "—")
                    if cell_len > max_len:
                        max_len = cell_len
            # ضریب ۱.۳ برای فونت فارسی + ۴ پدینگ
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len * 1.3 + 4, 60)
        ws.freeze_panes = "A2"

    def _build_report_full(self, tab):
        tk.Label(tab, text="📋  گزارش جامع — هر اسلب با کلیه جزئیات",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        tk.Label(tab, text="برای هر اسلب: ثبت، QC، اسکارف، برش، باومن، انتقال، برگشت‌ها، خروج، محل فعلی",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        ff_c = card_frame(tab); ff_c.pack(fill="x", padx=16, pady=8)
        _now_date = get_first_report_date_sh()
        from_v, to_v = _make_dt_filter(ff_c, C["card"], _now_date)
        tk.Label(ff_c, text="(خالی=همه)", bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=14)

        COLS = ("slab_id","registered_by","reg_date","reg_time","qc_status","qc_by","qc_date","qc_time",
                "scarf","cut","bauman_done","lab_delivered","transfer_dest",
                "ret_to_internal","current_location","exit_status","exit_by")
        HEADS = ("شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","وضعیت QC","تأییدکننده QC","تاریخ QC","ساعت QC",
                 "اسکارف","برش","باومن","آزمایشگاه","مقصد انتقال",
                 "برگشت به داخلی","محل فعلی","وضعیت خروج","خروج توسط")
        tf,tree = scrolled_tree(tab, COLS, HEADS, height=12)
        tf.pack(fill="both", expand=True, padx=16, pady=4)
        for col in COLS: tree.column(col, width=100, anchor="center")
        tree.column("slab_id", width=130)
        search_bar(tab, tree, col_indices=[0]).pack(anchor="e", padx=16, pady=2)
        sort_toolbar(tab, tree, date_col="reg_date", bg=C["panel"]).pack(anchor="e", padx=16, pady=2)
        cnt_lbl = tk.Label(tab,"",bg=C["panel"],fg=C["text_dim"],font=FONT_SMALL)
        cnt_lbl.pack(anchor="e", padx=16)

        def get_rows():
            db = load_db()
            fsh = from_v.get().strip().replace("  "," ")
            tsh = to_v.get().strip().replace("  "," ")
            rows = []
            for rec in db["melts"]:
                at = rec.get("registered_at","").replace("  "," ")
                if fsh and tsh and at and not (fsh<=at<=tsh): continue
                info = get_slab_full_info(db, rec["slab_id"])
                reg_d, reg_t = split_dt(info.get("registered_at","—"))
                qc_d,  qc_t  = split_dt(info.get("qc_at","—"))
                row = []
                for c in COLS:
                    if c == "reg_date": row.append(reg_d)
                    elif c == "reg_time": row.append(reg_t)
                    elif c == "qc_date": row.append(qc_d)
                    elif c == "qc_time": row.append(qc_t)
                    else: row.append(str(info.get(c,"—")))
                rows.append(tuple(row))
            return rows

        def refresh():
            tree.delete(*tree.get_children())
            rows = get_rows()
            for row in rows: tree.insert("","end",values=row)
            cnt_lbl.config(text=f"📊 {len(rows)} اسلب")

        refresh()

        def export_full():
            if not XLSX: messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            rows = get_rows()
            path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=f"گزارش_جامع_{now_str().split()[0].replace('/','')}.xlsx",parent=self)
            if not path: return
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "گزارش جامع"
            full_heads = [
                "شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","وضعیت QC","تأییدکننده QC","تاریخ QC","ساعت QC",
                "اسکارف","برش","تست باومن","تحویل آزمایشگاه","مقصد انتقال",
                "برگشت به داخلی","محل فعلی","وضعیت خروج","خروج توسط"
            ]
            self._write_sheet(ws, full_heads, rows,
                col_widths=[16,16,14,12,14,16,14,12,28,28,12,18,18,12,22,16,16])
            # شیت برگشت‌ها
            db = load_db()
            ws2 = wb.create_sheet("تاریخچه برگشت‌ها")
            rh=["شماره اسلب","شماره برگشت","برچسب","از انبار","دلیل","توسط","تاریخ","ساعت"]
            rr=[]
            for r in db.get("return_log",[]):
                rd, rt = split_dt(r.get("returned_at",""))
                rr.append((r.get("slab_id",""),r.get("return_number",""),r.get("label",""),
                           r.get("source",""),r.get("reason",""),r.get("returned_by",""),rd, rt))
            self._write_sheet(ws2,rh,rr,col_widths=[16,10,22,18,30,16,14,12])
            # شیت جابجایی
            ws3 = wb.create_sheet("جابجایی‌ها")
            mh=["شماره اسلب","عملیات","از","به","دلیل","توسط","تاریخ","ساعت"]
            mr=[]
            for r in db.get("movement_log",[]):
                md, mt = split_dt(r.get("at",""))
                mr.append((r.get("slab_id",""),r.get("operation",""),r.get("from",""),
                           r.get("to",""),r.get("reason",""),r.get("by",""),md, mt))
            self._write_sheet(ws3,mh,mr,col_widths=[16,20,18,18,30,16,14,12])
            wb.save(path)
            messagebox.showinfo("موفق",f"Excel جامع ذخیره شد:\n{path}",parent=self)

        ctrl = tk.Frame(tab,bg=C["panel"]); ctrl.pack(fill="x",padx=16,pady=4)
        styled_btn(ctrl,"📥  Excel جامع",export_full,color=C["btn_success"]).pack(side="right",padx=4)

    def _build_report_section(self, tab):
        tk.Label(tab, text="📂  گزارش به تفکیک بخش با فیلتر زمانی",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        ff_c=card_frame(tab); ff_c.pack(fill="x",padx=16,pady=6)
        ff=tk.Frame(ff_c,bg=C["card"]); ff.pack(padx=14,pady=10,fill="x")
        _now_date=to_shamsi(datetime.datetime.now()).split("  ")[0]
        SECS=["ثبت ذوب","کنترل کیفی","رد شده","اسکارف","برش","باومن","انتقال روباز","برگشت‌ها","آزمایشگاه","قراضه","خروج"]
        tk.Label(ff,text="بخش:",bg=C["card"],fg=C["text"],font=FONT_NORM).pack(side="right",padx=(0,6))
        sec_cb=make_combo(ff,SECS,width=18); sec_cb.set("ثبت ذوب"); sec_cb.pack(side="right",padx=4)
        fv, tv = _make_dt_filter(ff_c, C["card"], _now_date)

        COLS_S=("f1","f2","fd","ft","f4","f5","f6","f7")
        tf_s,tree_s=scrolled_tree(tab,COLS_S,("شماره اسلب","فیلد ۱","تاریخ","ساعت","فیلد ۴","فیلد ۵","فیلد ۶","فیلد ۷"),height=14)
        tf_s.pack(fill="both",expand=True,padx=16,pady=4)
        search_bar(tab,tree_s,col_indices=[0]).pack(anchor="e",padx=16,pady=2)
        sort_toolbar(tab, tree_s, date_col="fd", bg=C["panel"]).pack(anchor="e", padx=16, pady=2)
        cnt_s=tk.Label(tab,"",bg=C["panel"],fg=C["text_dim"],font=FONT_SMALL); cnt_s.pack(anchor="e",padx=16)

        # نکته: ستون سوم و چهارم همیشه «تاریخ» و «ساعت» تفکیک‌شده‌اند (نه ترکیبی).
        # برای بخش‌هایی که دو فیلد زمانی دارند (مثل باومن: زمان برش + زمان تحویل)
        # دو ستون آخر هم به‌صورت تاریخ/ساعت دوم استفاده می‌شوند.
        HDR_MAP={
            "ثبت ذوب":["شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","وضعیت QC","محل","توضیحات",""],
            "کنترل کیفی":["شماره اسلب","تأییدکننده","تاریخ QC","ساعت QC","وضعیت","محل","",""],
            "رد شده":["شماره اسلب","رد‌کننده","تاریخ رد","ساعت رد","محل","","",""],
            "اسکارف":["شماره اسلب","ثبت‌کننده","تاریخ","ساعت","دلایل","توضیحات","",""],
            "برش":["شماره اسلب","ثبت‌کننده","تاریخ","ساعت","دلایل","باومن","توضیحات",""],
            "باومن":["شماره اسلب","برش‌کار","تاریخ برش","ساعت برش","وضعیت آزمایشگاه","تحویل‌دهنده","تاریخ تحویل","ساعت تحویل"],
            "انتقال روباز":["شماره اسلب","انتقال‌دهنده","تاریخ انتقال","ساعت انتقال","مقصد","وضعیت فعلی","",""],
            "برگشت‌ها":["شماره اسلب","ثبت‌کننده","تاریخ","ساعت","دلیل","از انبار","برچسب",""],
            "آزمایشگاه":["شماره اسلب","تحویل‌دهنده","تاریخ تحویل","ساعت تحویل","","","",""],
            "قراضه":["شماره اسلب","ثبت‌کننده","تاریخ","ساعت","دلیل","","",""],
            "خروج":["شماره اسلب","خروج‌دهنده","تاریخ خروج","ساعت خروج","محل","","",""],
        }

        def get_sec_rows():
            db=load_db()
            sec=sec_cb.get()
            fsh=fv.get().strip().replace("  "," ")
            tsh=tv.get().strip().replace("  "," ")
            def flt(at): return not (fsh and tsh and at and not (fsh<=at.replace("  "," ")<=tsh))
            rows=[]
            if sec=="ثبت ذوب":
                for r in db["melts"]:
                    if flt(r.get("registered_at","")):
                        d,t = split_dt(r.get("registered_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("registered_by","—"), db),d,t,
                                     r.get("qc_status","—"),r.get("location","—"),r.get("note","—"),""))
            elif sec=="کنترل کیفی":
                for r in db["melts"]:
                    if r.get("qc_status")=="کنترل کیفی شده" and flt(r.get("qc_at","")):
                        d,t = split_dt(r.get("qc_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("qc_by","—"), db),d,t,
                                     r.get("qc_status","—"),r.get("location","—"),"",""))
            elif sec=="رد شده":
                for r in db["melts"]:
                    if r.get("qc_status")=="عدم تایید کنترل کیفی" and flt(r.get("qc_at","")):
                        d,t = split_dt(r.get("qc_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("qc_by","—"), db),d,t,
                                     r.get("rej_location","—"),"","",""))
            elif sec=="اسکارف":
                for r in db["scarf_cut"]:
                    if r.get("operation")=="اسکارفی" and flt(r.get("registered_at","")):
                        d,t = split_dt(r.get("registered_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("registered_by","—"), db),d,t,
                                     r.get("reason","—"),r.get("note","—"),"",""))
            elif sec=="برش":
                for r in db["scarf_cut"]:
                    if r.get("operation")=="برشی" and flt(r.get("registered_at","")):
                        d,t = split_dt(r.get("registered_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("registered_by","—"), db),d,t,
                                     r.get("reason","—"),"✔" if r.get("bauman_done") else "—",r.get("note","—"),""))
            elif sec=="باومن":
                lab_map = {r["slab_id"]: r for r in db.get("lab_deliveries", [])}
                for r in db["scarf_cut"]:
                    if r.get("operation")=="برشی" and r.get("bauman_done") and flt(r.get("registered_at","")):
                        dl = lab_map.get(r["slab_id"])
                        cd,ct = split_dt(r.get("registered_at","—"))
                        dd,dt = split_dt(dl.get("delivered_at","—")) if dl else ("—","—")
                        rows.append((r["slab_id"],get_display_name(r.get("registered_by","—"), db),cd,ct,
                                     "تحویل داده شده" if dl else "در انتظار",
                                     get_display_name(dl.get("delivered_by","—"), db) if dl else "—",
                                     dd, dt))
            elif sec=="انتقال روباز":
                for r in db["transfers_out"]:
                    if flt(r.get("transferred_at","")):
                        d,t = split_dt(r.get("transferred_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("transferred_by","—"), db),d,t,
                                     r.get("destination","—"),r.get("current_location","—"),"",""))
            elif sec=="برگشت‌ها":
                for r in db.get("return_log",[]):
                    if flt(r.get("returned_at","")):
                        d,t = split_dt(r.get("returned_at","—"))
                        rows.append((r.get("slab_id","—"),get_display_name(r.get("returned_by","—"), db),d,t,
                                     r.get("reason","—"),r.get("source","—"),r.get("label","—"),""))
            elif sec=="آزمایشگاه":
                for r in db["lab_deliveries"]:
                    if flt(r.get("delivered_at","")):
                        d,t = split_dt(r.get("delivered_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("delivered_by","—"), db),d,t,"","","",""))
            elif sec=="قراضه":
                for r in db["scrap"]:
                    if flt(r.get("registered_at","")):
                        d,t = split_dt(r.get("registered_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("registered_by","—"), db),d,t,
                                     r.get("reason","—"),"","",""))
            elif sec=="خروج":
                for r in db["melts"]:
                    if r.get("exit_status")=="خروج زده شده" and flt(r.get("exit_at","")):
                        d,t = split_dt(r.get("exit_at","—"))
                        rows.append((r["slab_id"],get_display_name(r.get("exit_by","—"), db),d,t,
                                     r.get("location","—"),"","",""))
            return rows, HDR_MAP.get(sec,["شماره اسلب","فیلد ۱","تاریخ","ساعت","فیلد ۴","فیلد ۵","فیلد ۶","فیلد ۷"])

        def refresh_s(*_):
            tree_s.delete(*tree_s.get_children())
            rows,heads_s=get_sec_rows()
            for i,col in enumerate(COLS_S):
                tree_s.heading(col,text=heads_s[i] if i<len(heads_s) else "")
            for row in rows: tree_s.insert("","end",values=row)
            cnt_s.config(text=f"📊 {len(rows)} رکورد")

        sec_cb.bind("<<ComboboxSelected>>",refresh_s); refresh_s()

        def export_sec():
            if not XLSX: messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            rows,heads_s=get_sec_rows()
            sec=sec_cb.get()
            path=filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=f"گزارش_{sec}_{now_str().split()[0].replace('/','')}.xlsx",parent=self)
            if not path: return
            wb=openpyxl.Workbook(); ws=wb.active; ws.title=sec[:30]
            self._write_sheet(ws,heads_s,rows)
            wb.save(path)
            messagebox.showinfo("موفق",f"Excel ذخیره شد:\n{path}",parent=self)

        ctrl_s=tk.Frame(tab,bg=C["panel"]); ctrl_s.pack(fill="x",padx=16,pady=4)
        styled_btn(ctrl_s,"📥  Excel این بخش",export_sec,color=C["btn_success"]).pack(side="right",padx=4)

    def _build_report_person(self, tab):
        tk.Label(tab, text="👤  گزارش انفرادی پرسنل — همه فعالیت‌ها",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        ff_c=card_frame(tab); ff_c.pack(fill="x",padx=16,pady=6)
        ff=tk.Frame(ff_c,bg=C["card"]); ff.pack(padx=14,pady=10,fill="x")
        _now_date=to_shamsi(datetime.datetime.now()).split("  ")[0]
        db=load_db()
        person_list=[(un,ud.get("display","")) for un,ud in db["users"].items() if un!="admin"]
        disp_names=[f"{d}  [{u}]" for u,d in person_list]
        tk.Label(ff,text="پرسنل:",bg=C["card"],fg=C["text"],font=FONT_NORM).pack(side="right",padx=(0,6))
        pers_cb=make_combo(ff,disp_names,width=26)
        if disp_names: pers_cb.set(disp_names[0])
        pers_cb.pack(side="right",padx=4)
        fvp, tvp = _make_dt_filter(ff_c, C["card"], _now_date)

        COLS_P=("section","slab_id","action","at_date","at_time","detail")
        HEADS_P=("بخش","شماره اسلب","عملیات","تاریخ","ساعت","جزئیات")
        tf_p,tree_p=scrolled_tree(tab,COLS_P,HEADS_P,height=14)
        tf_p.pack(fill="both",expand=True,padx=16,pady=4)
        search_bar(tab,tree_p,col_indices=[0,1]).pack(anchor="e",padx=16,pady=2)
        sort_toolbar(tab, tree_p, date_col="at_date", bg=C["panel"]).pack(anchor="e", padx=16, pady=2)
        cnt_p=tk.Label(tab,"",bg=C["panel"],fg=C["text_dim"],font=FONT_SMALL); cnt_p.pack(anchor="e",padx=16)

        def get_uname():
            s=pers_cb.get()
            m=s.split("[")
            return m[-1].rstrip("]").strip() if len(m)>1 else s

        def get_person_rows():
            uname=get_uname()
            fsh=fvp.get().strip().replace("  "," ")
            tsh=tvp.get().strip().replace("  "," ")
            def flt(at): return not (fsh and tsh and at and not (fsh<=at.replace("  "," ")<=tsh))
            db=load_db(); rows=[]
            # هر سطر: (بخش، شماره اسلب، عملیات، تاریخ، ساعت، جزئیات، کلید_مرتب‌سازی_زمانی)
            for r in db["melts"]:
                if r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    d,t = split_dt(r.get("registered_at","—"))
                    rows.append(("ثبت ذوب",r["slab_id"],"ثبت ذوب",d,t,r.get("qc_status","—"),r.get("registered_at","")))
                if r.get("qc_by")==uname and flt(r.get("qc_at","")):
                    d,t = split_dt(r.get("qc_at","—"))
                    rows.append(("کنترل کیفی",r["slab_id"],"تأیید QC",d,t,r.get("location","—"),r.get("qc_at","")))
                if r.get("exit_by")==uname and flt(r.get("exit_at","")):
                    d,t = split_dt(r.get("exit_at","—"))
                    rows.append(("خروج",r["slab_id"],"خروج",d,t,"خارج شده",r.get("exit_at","")))
            for r in db["scarf_cut"]:
                if r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    d,t = split_dt(r.get("registered_at","—"))
                    rows.append((r.get("operation","عملیات"),r["slab_id"],"ثبت",d,t,r.get("reason","—"),r.get("registered_at","")))
                if r.get("operation")=="برشی" and r.get("bauman_done") and r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    d,t = split_dt(r.get("registered_at","—"))
                    rows.append(("باومن",r["slab_id"],"تست باومن",d,t,"آماده تحویل",r.get("registered_at","")))
            for r in db.get("lab_deliveries", []):
                if r.get("delivered_by")==uname and flt(r.get("delivered_at","")):
                    d,t = split_dt(r.get("delivered_at","—"))
                    rows.append(("آزمایشگاه",r["slab_id"],"تحویل",d,t,"تحویل داده شده",r.get("delivered_at","")))
            for r in db["transfers_out"]:
                if r.get("transferred_by")==uname and flt(r.get("transferred_at","")):
                    d,t = split_dt(r.get("transferred_at","—"))
                    rows.append(("انتقال",r["slab_id"],"انتقال",d,t,r.get("destination","—"),r.get("transferred_at","")))
            for r in db.get("return_log",[]):
                if r.get("returned_by")==uname and flt(r.get("returned_at","")):
                    d,t = split_dt(r.get("returned_at","—"))
                    rows.append(("برگشت",r.get("slab_id","—"),r.get("label","برگشت"),d,t,r.get("reason","—"),r.get("returned_at","")))
            for r in db.get("scrap",[]):
                if r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    d,t = split_dt(r.get("registered_at","—"))
                    rows.append(("قراضه",r["slab_id"],"ثبت قراضه",d,t,r.get("reason","—"),r.get("registered_at","")))
            rows.sort(key=lambda x: x[6])
            # کلید مرتب‌سازی داخلی را قبل از نمایش/خروجی حذف می‌کنیم
            return [row[:6] for row in rows]

        def refresh_p(*_):
            tree_p.delete(*tree_p.get_children())
            rows=get_person_rows()
            for row in rows: tree_p.insert("","end",values=row)
            cnt_p.config(text=f"📊 {len(rows)} عملیات")

        pers_cb.bind("<<ComboboxSelected>>",refresh_p); refresh_p()

        def export_person():
            if not XLSX: messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            disp=pers_cb.get().split("[")[0].strip()
            rows=get_person_rows()
            path=filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=f"پرسنل_{disp}_{now_str().split()[0].replace('/','')}.xlsx",parent=self)
            if not path: return
            wb=openpyxl.Workbook(); ws=wb.active; ws.title=disp[:30]
            self._write_sheet(ws,["بخش","شماره اسلب","عملیات","تاریخ","ساعت","جزئیات"],
                              rows,col_widths=[16,16,18,14,12,28])
            wb.save(path)
            messagebox.showinfo("موفق",f"Excel پرسنل ذخیره شد:\n{path}",parent=self)

        ctrl_p=tk.Frame(tab,bg=C["panel"]); ctrl_p.pack(fill="x",padx=16,pady=4)
        styled_btn(ctrl_p,"📥  Excel این پرسنل",export_person,color=C["btn_success"]).pack(side="right",padx=4)

    def _build_client_update_tab(self, tab):
        """انتشار نرم‌افزار برای کلاینت‌ها — انتخاب نقش + پیشرفت + تاریخچه."""
        BG = C["panel"]
        wrap = tk.Frame(tab, bg=BG)
        wrap.pack(fill="both", expand=True, padx=20, pady=16)

        tk.Label(wrap, text="⬆  به‌روزرسانی نرم‌افزار برای کلاینت‌ها",
                 bg=BG, fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0, 6))
        tk.Label(
            wrap,
            text="با زدن «انتشار»، ابتدا کلاینت(های) هدف را انتخاب کنید.\n"
                 "سپس بسته روی سرور می‌رود و همان کلاینت‌ها صفحهٔ پیشرفت واقعی نشان می‌دهند.\n"
                 "ری‌استارت کلاینت فقط بعد از تأیید کاربر روی کلاینت انجام می‌شود. فلش لازم نیست.",
            bg=BG, fg=C["text_dim"], font=FONT_SMALL, justify="right",
        ).pack(anchor="e", pady=(0, 14))

        card = card_frame(wrap)
        card.pack(fill="x", pady=6)
        inner = tk.Frame(card, bg=C["card"])
        inner.pack(fill="x", padx=14, pady=14)

        status = tk.Label(inner, text="", bg=C["card"], fg=C["text"], font=FONT_SMALL, justify="right")
        status.pack(anchor="e", pady=(0, 8))

        hist_frm = tk.Frame(inner, bg=C["card"])
        hist_frm.pack(fill="x", pady=(0, 10))
        hist_c1 = tk.Label(hist_frm, text="Client1: —", bg=C["card"], fg=C["text"],
                           font=FONT_SMALL, justify="right", anchor="e")
        hist_c1.pack(fill="x", pady=1)
        hist_c2 = tk.Label(hist_frm, text="Client2: —", bg=C["card"], fg=C["text"],
                           font=FONT_SMALL, justify="right", anchor="e")
        hist_c2.pack(fill="x", pady=1)

        def _hist_status_fa(st):
            s = str(st or "").lower()
            if s in ("done", "ok", "success"):
                return "با موفقیت انجام شد"
            if s in ("failed", "error"):
                return "ناموفق"
            if s in ("pending", "waiting", "updating"):
                return "در انتظار / در حال انجام"
            return s or "—"

        def _refresh_status():
            try:
                from server.update_service import (
                    load_manifest, find_client_source, updates_dir, load_update_history,
                )
                man = load_manifest()
                c1 = find_client_source("client1")
                c2 = find_client_source("client2")
                lines = [
                    f"پوشهٔ انتشار روی ادمین: {updates_dir()}",
                    f"منبع Client1: {c1 or '— پیدا نشد —'}",
                    f"منبع Client2: {c2 or '— پیدا نشد —'}",
                ]
                if man:
                    lines.append(
                        f"آخرین انتشار: v{man.get('version')}  build {man.get('build')}  "
                        f"({man.get('published_at', '')})"
                    )
                    clients = man.get("clients") or {}
                    for role in ("client1", "client2"):
                        if role in clients:
                            sz = int(clients[role].get("size") or 0)
                            lines.append(f"  {role}: {max(1, sz // (1024 * 1024))} MB")
                    db = load_db()
                    sig = ((db.get("settings") or {}).get("client_software_update") or {})
                    if sig.get("build") or sig.get("publish_id"):
                        tgt = sig.get("target_roles") or ["client1", "client2"]
                        force_on = bool(sig.get("force"))
                        force_txt = "force روشن (منتظر کلاینت)" if force_on else "force خاموش"
                        lines.append(
                            f"آخرین سیگنال: build {sig.get('build') or '—'} → {', '.join(tgt)}  [{force_txt}]"
                        )
                else:
                    lines.append("هنوز بسته‌ای منتشر نشده.")
                status.config(text="\n".join(lines), fg=C["text"])

                hist = load_update_history()
                for role, lbl in (("client1", hist_c1), ("client2", hist_c2)):
                    h = hist.get(role) or {}
                    when = h.get("finished_at") or "—"
                    st_fa = _hist_status_fa(h.get("status"))
                    name = "Client1" if role == "client1" else "Client2"
                    ver = str(h.get("version") or "").strip()
                    ver_txt = f" — نسخه در حال اجرا: v{ver}" if ver else ""
                    lbl.config(
                        text=f"{name}: آخرین آپدیت {when} — وضعیت: {st_fa}{ver_txt}"
                    )
            except Exception as e:
                status.config(text=f"خطا در خواندن وضعیت: {e}", fg=C["danger"])

        def _pick_target_roles():
            result = {"roles": None}
            dlg = tk.Toplevel(tab)
            dlg.title("انتخاب کلاینت برای انتشار")
            dlg.resizable(False, False)
            try:
                dlg.transient(tab.winfo_toplevel())
                dlg.grab_set()
            except Exception:
                pass
            dlg.configure(bg=C["panel"])
            frm = tk.Frame(dlg, bg=C["panel"], padx=18, pady=14)
            frm.pack(fill="both", expand=True)
            tk.Label(
                frm, text="کدام کلاینت(ها) به‌روزرسانی شوند؟",
                bg=C["panel"], fg=C["accent"], font=FONT_HEAD,
            ).pack(anchor="e", pady=(0, 10))
            v1 = tk.BooleanVar(value=True)
            v2 = tk.BooleanVar(value=True)
            tk.Checkbutton(
                frm, text="Client1", variable=v1, bg=C["panel"], fg=C["text"],
                selectcolor=C["card"], activebackground=C["panel"], font=FONT_SMALL,
                anchor="e",
            ).pack(fill="x", pady=4)
            tk.Checkbutton(
                frm, text="Client2", variable=v2, bg=C["panel"], fg=C["text"],
                selectcolor=C["card"], activebackground=C["panel"], font=FONT_SMALL,
                anchor="e",
            ).pack(fill="x", pady=4)

            def _ok():
                roles = []
                if v1.get():
                    roles.append("client1")
                if v2.get():
                    roles.append("client2")
                if not roles:
                    messagebox.showwarning("انتخاب", "حداقل یک کلاینت را انتخاب کنید.", parent=dlg)
                    return
                result["roles"] = roles
                try:
                    dlg.grab_release()
                except Exception:
                    pass
                dlg.destroy()

            def _cancel():
                result["roles"] = None
                try:
                    dlg.grab_release()
                except Exception:
                    pass
                dlg.destroy()

            br = tk.Frame(frm, bg=C["panel"])
            br.pack(fill="x", pady=(12, 0))
            styled_btn(br, "ادامه", _ok, color=C["btn_success"]).pack(side="right", padx=4)
            styled_btn(br, "انصراف", _cancel, color=C["btn_ghost"]).pack(side="right", padx=4)
            dlg.protocol("WM_DELETE_WINDOW", _cancel)
            try:
                dlg.wait_window()
            except Exception:
                pass
            return result["roles"]

        def do_publish():
            import json as _json
            import threading
            import time as _time
            import urllib.request

            target_roles = _pick_target_roles()
            if not target_roles:
                return

            prog = {
                "pct": 0.0,
                "phase": "آماده‌سازی…",
                "t0": _time.time(),
                "title": "در حال به‌روزرسانی",
                "detail": "",
                "allow_close": False,
                "done_ok": False,
                "outcome": "",  # success | client_error | no_clients | timeout | error
                "live_eta_sec": None,
                "pack_done": False,
                "target_roles": list(target_roles),
            }
            overlay = tk.Toplevel(tab)
            overlay.title("انتشار و به‌روزرسانی کلاینت‌ها")
            overlay.resizable(False, False)
            try:
                overlay.transient(tab.winfo_toplevel())
                overlay.attributes("-topmost", True)
                overlay.grab_set()
            except Exception:
                pass
            overlay.protocol("WM_DELETE_WINDOW", lambda: None)
            bg, accent, fg = "#132033", "#c9a227", "#f0e6d2"
            dlg_w, dlg_h = 560, 500
            try:
                root = tab.winfo_toplevel()
                root.update_idletasks()
                sw = int(root.winfo_screenwidth())
                sh = int(root.winfo_screenheight())
                ox = max(0, (sw - dlg_w) // 2)
                oy = max(0, (sh - dlg_h) // 2)
                overlay.geometry(f"{dlg_w}x{dlg_h}+{ox}+{oy}")
            except Exception:
                overlay.geometry(f"{dlg_w}x{dlg_h}")
            ofrm = tk.Frame(overlay, bg=bg, highlightbackground=accent, highlightthickness=2)
            ofrm.pack(fill="both", expand=True)
            box = tk.Frame(ofrm, bg=bg, padx=18, pady=14)
            box.place(relx=0.5, rely=0.5, anchor="center")
            tk.Label(box, text="⬆", bg=bg, fg=accent, font=("Segoe UI Symbol", 28)).pack()
            otitle = tk.Label(box, text=prog["title"],
                              bg=bg, fg=fg, font=("B Nazanin", 14, "bold"),
                              wraplength=500, justify="center")
            otitle.pack(pady=(4, 6))
            ocan = tk.Canvas(box, width=120, height=120, bg=bg, highlightthickness=0)
            ocan.pack(pady=2)
            _pad = 10
            _bb = (_pad, _pad, 120 - _pad, 120 - _pad)
            ocan.create_oval(*_bb, outline="#2a3a4f", width=10)
            oarc = ocan.create_arc(*_bb, start=90, extent=0, style="arc", outline=accent, width=10)
            opct = ocan.create_text(60, 60, text="0%", fill=fg, font=("B Nazanin", 16, "bold"))
            oeta = tk.Label(box, text="در حال شروع…", bg=bg, fg=accent, font=("B Nazanin", 11))
            oeta.pack(pady=(6, 2))
            ophase = tk.Label(box, text="", bg=bg, fg="#9fb0c3", font=("B Nazanin", 10),
                              wraplength=500, justify="center")
            ophase.pack()
            odetail = tk.Label(box, text="", bg=bg, fg="#c5d0dc", font=("B Nazanin", 10),
                               wraplength=500, justify="right")
            odetail.pack(pady=(8, 0))
            btn_row_o = tk.Frame(box, bg=bg)
            btn_row_o.pack(pady=(12, 0))
            ok_btn = tk.Button(
                btn_row_o, text="تأیید", font=("B Nazanin", 12, "bold"),
                bg=accent, fg="#1a1520", activebackground="#e0c04a",
                relief="flat", padx=28, pady=5,
                command=lambda: prog.update({"allow_close": True}),
            )

            def _eta(pct, t0):
                elapsed = max(0.1, _time.time() - t0)
                el_s = int(elapsed)
                el_txt = f"گذشته: {el_s} ثانیه" if el_s < 60 else f"گذشته: {el_s // 60}:{el_s % 60:02d}"
                if prog.get("done_ok"):
                    oc = str(prog.get("outcome") or "")
                    if oc == "success":
                        return "با موفقیت انجام شد"
                    if oc == "no_clients":
                        return "کلاینت آفلاین است"
                    if oc == "client_error":
                        return "خطا روی کلاینت"
                    if oc == "timeout":
                        return "زمان انتظار تمام شد"
                    return str(prog.get("title") or "پایان انتشار")
                live_eta = prog.get("live_eta_sec")
                if live_eta is not None:
                    try:
                        e = float(live_eta)
                        if e < 45:
                            return f"{el_txt}  |  حدود {max(1, int(e))} ثانیه مانده"
                        return f"{el_txt}  |  حدود {max(1, int(round(e / 60.0)))} دقیقه مانده"
                    except Exception:
                        pass
                if not prog.get("pack_done") and pct < 5.0:
                    return f"{el_txt}  |  در حال شروع…"
                if prog.get("pack_done") and pct < 42.0:
                    return f"{el_txt}  |  منتظر گزارش کلاینت…"
                if pct < 5.0:
                    return f"{el_txt}  |  در حال شروع…"
                remain = elapsed * (100.0 - pct) / max(pct, 0.1)
                if remain < 45:
                    return f"{el_txt}  |  حدود {max(1, int(remain))} ثانیه مانده"
                return f"{el_txt}  |  حدود {max(1, int(round(remain / 60.0)))} دقیقه مانده"

            def _tick_ui():
                try:
                    if not overlay.winfo_exists():
                        return
                except Exception:
                    return
                pct = float(prog.get("pct") or 0)
                try:
                    ocan.itemconfigure(oarc, extent=-max(0.0, min(100.0, pct)) * 3.6)
                    ocan.itemconfigure(opct, text=f"{int(pct)}%")
                    oeta.config(text=_eta(pct, float(prog.get("t0") or _time.time())))
                    ophase.config(text=str(prog.get("phase") or ""))
                    odetail.config(text=str(prog.get("detail") or ""))
                    otitle.config(text=str(prog.get("title") or ""))
                    if prog.get("done_ok") and not ok_btn.winfo_ismapped():
                        oc = str(prog.get("outcome") or "")
                        if oc == "success":
                            ok_btn.config(text="با موفقیت انجام شد — تأیید")
                        elif oc == "no_clients":
                            ok_btn.config(text="تأیید — کلاینت آفلاین است")
                        elif oc == "client_error":
                            ok_btn.config(text="تأیید — خطا روی کلاینت")
                        elif oc == "timeout":
                            ok_btn.config(text="تأیید — بدون تأیید کامل")
                        else:
                            ok_btn.config(text="تأیید")
                        ok_btn.pack()
                    if prog.get("allow_close"):
                        try:
                            overlay.grab_release()
                        except Exception:
                            pass
                        try:
                            overlay.destroy()
                        except Exception:
                            pass
                        return
                except Exception:
                    pass
                try:
                    overlay.after(200, _tick_ui)
                except Exception:
                    pass

            _tick_ui()
            status.config(text="در حال انتشار…", fg=C["gold"])

            def _set_prog(pct, phase=""):
                mapped = max(0.0, min(38.0, float(pct) * 0.38))
                prog["pct"] = mapped
                if phase:
                    prog["phase"] = phase

            def _format_role_line(role_key, st):
                label = "Client1" if role_key == "client1" else "Client2"
                status_s = str((st or {}).get("status") or "waiting")
                pct = float((st or {}).get("percent") or 0)
                phase = str((st or {}).get("phase") or "")
                eta = (st or {}).get("eta_sec")
                if status_s == "waiting":
                    return f"{label}: منتظر شروع آپدیت…"
                if status_s == "done":
                    return f"{label}: انجام شد (100%)"
                if status_s == "failed":
                    return f"{label}: خطا — {phase or 'ناموفق'}"
                eta_txt = ""
                if eta is not None:
                    try:
                        e = float(eta)
                        if e < 45:
                            eta_txt = f" | حدود {max(1, int(e))}ث مانده"
                        else:
                            eta_txt = f" | حدود {max(1, int(round(e / 60)))}د مانده"
                    except Exception:
                        pass
                return f"{label}: در حال به‌روزرسانی — {int(pct)}%{eta_txt}\n    {phase}"

            def _fetch_summary(timeout_sec=180.0):
                try:
                    with urllib.request.urlopen(
                        "http://127.0.0.1:8080/api/v1/updates/progress", timeout=1.5
                    ) as resp:
                        data = _json.loads(resp.read().decode("utf-8"))
                        summary = data.get("summary")
                        if isinstance(summary, dict):
                            return summary
                except Exception:
                    pass
                from server.update_service import summarize_client_progress
                return summarize_client_progress(timeout_sec=timeout_sec)

            def _work():
                err = None
                wait_result = {
                    "timed_out": False, "all_done": False, "any_failed": False, "no_clients": False,
                }
                roles_sel = list(prog.get("target_roles") or ["client1", "client2"])
                try:
                    from server.update_service import (
                        publish_client_packages,
                        reset_client_update_progress,
                        record_client_update_history,
                    )
                    prog["title"] = "در حال به‌روزرسانی"
                    prog["t0"] = _time.time()
                    _set_prog(5, "شروع بسته‌بندی…")
                    man = publish_client_packages(
                        note="published from admin UI",
                        progress_cb=_set_prog,
                        roles=roles_sel,
                    )
                    _set_prog(95, "ثبت سیگنال برای کلاینت‌ها…")
                    pub_id = str(man.get("publish_id") or man.get("build") or "")
                    reset_client_update_progress(pub_id, roles=roles_sel)
                    try:
                        body = _json.dumps({"publish_id": pub_id, "roles": roles_sel}).encode("utf-8")
                        req = urllib.request.Request(
                            "http://127.0.0.1:8080/api/v1/updates/progress/reset",
                            data=body,
                            method="POST",
                            headers={"Content-Type": "application/json"},
                        )
                        with urllib.request.urlopen(req, timeout=2.0) as _:
                            pass
                    except Exception:
                        pass
                    db2 = load_db()
                    s = db2.setdefault("settings", {})
                    force_ttl = 360.0  # ۶ دقیقه — تا Admin منتظر heartbeat بماند و Client فرصت بیدارشدن داشته باشد
                    force_set_at = _time.time()
                    expires_at = force_set_at + force_ttl
                    expires_at_str = _time.strftime("%Y-%m-%d %H:%M:%S", _time.localtime(expires_at))
                    s["client_software_update"] = {
                        "version": man.get("version"),
                        "build": man.get("build"),
                        "publish_id": man.get("publish_id"),
                        "published_at": man.get("published_at") or _time.strftime("%Y-%m-%d %H:%M:%S"),
                        "force": True,
                        "force_set_at": _time.strftime("%Y-%m-%d %H:%M:%S", _time.localtime(force_set_at)),
                        "expires_at": expires_at_str,
                        "target_roles": roles_sel,
                        "clients": man.get("clients") or {},
                    }
                    hist_settings = s.setdefault("client_update_history", {})
                    for rk in roles_sel:
                        prev = dict(hist_settings.get(rk) or {})
                        prev.update({
                            "status": "pending",
                            "publish_id": pub_id,
                            "version": man.get("version") or "",
                            "build": man.get("build") or "",
                            "started_at": _time.strftime("%Y-%m-%d %H:%M:%S"),
                        })
                        hist_settings[rk] = prev
                    save_db(db2)
                    # اطمینان: بعد از merge، force واقعاً در SQL مانده (نژاد merge کهنه)
                    try:
                        db_verify = load_db(force=True)
                        sig_v = ((db_verify.get("settings") or {}).get("client_software_update") or {})
                        if not sig_v.get("force") or str(sig_v.get("publish_id") or "") != str(pub_id):
                            s2 = db_verify.setdefault("settings", {})
                            s2["client_software_update"] = dict(
                                (db2.get("settings") or {}).get("client_software_update") or {}
                            )
                            save_db(db_verify)
                    except Exception:
                        pass

                    prog["pack_done"] = True
                    prog["title"] = "در حال به‌روزرسانی"
                    prog["phase"] = "سیگنال ثبت شد — منتظر کلاینت‌ها…"
                    prog["pct"] = 40.0
                    prog["detail"] = "\n".join(
                        f"{'Client1' if r == 'client1' else 'Client2'}: منتظر…" for r in roles_sel
                    )
                    timeout_sec = 300.0  # پایه؛ اگر clients_online باشد تمدید می‌شود
                    while True:
                        summary = _fetch_summary(timeout_sec)
                        # کلاینت Online (heartbeat) → مهلت طولانی‌تر؛ force را وسط راه پاک نکن
                        if summary.get("clients_online") and not summary.get("participated_count"):
                            timeout_sec = max(timeout_sec, 360.0)
                        roles = summary.get("roles") or {}
                        lines = [_format_role_line(rk, roles.get(rk)) for rk in roles_sel]
                        prog["detail"] = "\n".join(lines)
                        avg = float(summary.get("avg_percent") or 0)
                        prog["pct"] = 40.0 + max(0.0, min(100.0, avg)) * 0.60
                        etas = []
                        for rk in roles_sel:
                            st_r = roles.get(rk) or {}
                            if st_r.get("status") == "updating" and st_r.get("eta_sec") is not None:
                                try:
                                    etas.append(float(st_r["eta_sec"]))
                                except Exception:
                                    pass
                        prog["live_eta_sec"] = max(etas) if etas else None
                        active = int(summary.get("active_count") or 0)
                        done_n = int(summary.get("done_count") or 0)
                        if summary.get("no_clients") and not summary.get("clients_online"):
                            prog["phase"] = "کلاینت آفلاین است"
                        elif active:
                            prog["phase"] = f"کلاینت در حال آپدیت ({done_n} تمام‌شده)…"
                        elif done_n:
                            prog["phase"] = "در حال تکمیل…"
                        else:
                            prog["phase"] = "منتظر شروع آپدیت روی کلاینت‌ها…"
                        # اگر هیچ ACK نیامد و no_clients — فوراً force را خاموش کن (بدون sticky)
                        # اما اگر clients_online است (حضور واقعی)، offline نگو و force را پاک نکن
                        if summary.get("no_clients") and not summary.get("participated_count") and not summary.get("clients_online"):
                            try:
                                db_off = load_db()
                                s_off = db_off.setdefault("settings", {})
                                sig_off = dict(s_off.get("client_software_update") or {})
                                if sig_off.get("force"):
                                    sig_off["force"] = False
                                    s_off["client_software_update"] = sig_off
                                    save_db(db_off)
                            except Exception:
                                pass
                            wait_result["no_clients"] = True
                            wait_result["timed_out"] = False
                            break
                        if summary.get("clients_online") and not summary.get("participated_count"):
                            # کلاینت Online است — منتظر بمان؛ پیام آفلاین نشان نده
                            if summary.get("no_clients"):
                                prog["phase"] = "منتظر شروع آپدیت روی کلاینت‌ها…"
                        if summary.get("complete") or summary.get("timed_out"):
                            wait_result["timed_out"] = bool(summary.get("timed_out"))
                            wait_result["all_done"] = bool(summary.get("all_done"))
                            wait_result["any_failed"] = bool(summary.get("any_failed"))
                            wait_result["no_clients"] = bool(
                                summary.get("no_clients") and not summary.get("clients_online")
                            )
                            break
                        _time.sleep(0.6)

                    try:
                        db3 = load_db()
                        s3 = db3.setdefault("settings", {})
                        hist3 = s3.setdefault("client_update_history", {})
                        summary_final = _fetch_summary(timeout_sec)
                        for rk in roles_sel:
                            stf = (summary_final.get("roles") or {}).get(rk) or {}
                            st_name = str(stf.get("status") or ("done" if wait_result["all_done"] else "pending"))
                            if wait_result["no_clients"] and st_name in ("waiting", "idle", ""):
                                st_name = "pending"
                            entry = {
                                "status": st_name,
                                "publish_id": pub_id,
                                "finished_at": stf.get("finished_at") or _time.strftime("%Y-%m-%d %H:%M:%S"),
                                "phase": stf.get("phase") or "",
                                "version": man.get("version") or "",
                                "build": man.get("build") or "",
                            }
                            hist3[rk] = entry
                            try:
                                record_client_update_history(
                                    rk,
                                    status=st_name,
                                    publish_id=pub_id,
                                    phase=str(entry.get("phase") or ""),
                                    finished_at=str(entry.get("finished_at") or ""),
                                    version=str(man.get("version") or ""),
                                    build=str(man.get("build") or ""),
                                )
                            except Exception:
                                pass
                        # CRITICAL: force را خاموش کن — نباید برای همیشه True بماند
                        sig_keep = dict(s3.get("client_software_update") or {})
                        if sig_keep:
                            sig_keep["force"] = False
                            s3["client_software_update"] = sig_keep
                        save_db(db3)
                        try:
                            from server.update_service import progress_path as _pp
                            import json as _j2
                            idle = {
                                "publish_id": "",
                                "started_at": 0,
                                "roles": {"client1": {}, "client2": {}},
                                "target_roles": [],
                            }
                            _pp().write_text(
                                _j2.dumps(idle, ensure_ascii=False),
                                encoding="utf-8",
                            )
                        except Exception:
                            pass
                    except Exception:
                        try:
                            db_f = load_db()
                            s_f = db_f.setdefault("settings", {})
                            sig_f = dict(s_f.get("client_software_update") or {})
                            if sig_f.get("force"):
                                sig_f["force"] = False
                                s_f["client_software_update"] = sig_f
                                save_db(db_f)
                        except Exception:
                            pass

                    if wait_result["all_done"] and not wait_result["any_failed"] and not wait_result["no_clients"]:
                        prog["pct"] = 100.0
                        prog["phase"] = "آپدیت کلاینت‌ها با موفقیت انجام شد"
                        prog["title"] = "با موفقیت انجام شد"
                        prog["outcome"] = "success"
                        prog["done_ok"] = True
                    elif wait_result["any_failed"]:
                        prog["phase"] = "یکی از کلاینت‌ها با خطا مواجه شد"
                        prog["title"] = "انتشار انجام شد — خطا در کلاینت"
                        prog["outcome"] = "client_error"
                        prog["done_ok"] = True
                    elif wait_result["no_clients"]:
                        prog["phase"] = "کلاینت آفلاین است — force پاک شد؛ انتشار بعدی فقط با کلیک دوباره"
                        prog["title"] = "کلاینت آفلاین است"
                        prog["outcome"] = "no_clients"
                        prog["done_ok"] = True
                    else:
                        prog["phase"] = "زمان انتظار تمام شد — ممکن است کلاینت هنوز در حال آپدیت باشد"
                        prog["title"] = "انتشار انجام شد — بدون تأیید کامل کلاینت"
                        prog["outcome"] = "timeout"
                        prog["done_ok"] = True
                except Exception as e:
                    err = e
                    # حتی در خطا force را پاک کن تا cold start کلاینت آپدیت نشان ندهد
                    try:
                        db_e = load_db()
                        s_e = db_e.setdefault("settings", {})
                        sig_e = dict(s_e.get("client_software_update") or {})
                        if sig_e.get("force"):
                            sig_e["force"] = False
                            s_e["client_software_update"] = sig_e
                            save_db(db_e)
                    except Exception:
                        pass

                def _done():
                    _refresh_status()
                    if err is not None:
                        try:
                            overlay.grab_release()
                        except Exception:
                            pass
                        try:
                            overlay.destroy()
                        except Exception:
                            pass
                        messagebox.showerror("خطای انتشار", str(err), parent=tab)
                        return
                    # فقط وقتی نتیجهٔ واقعی ثبت شده دکمه نشان بده — نه «موفقیت» جعلی
                    if not prog.get("done_ok"):
                        prog["outcome"] = prog.get("outcome") or "timeout"
                        prog["done_ok"] = True
                try:
                    tab.after(0, _done)
                except Exception:
                    _done()

            threading.Thread(target=_work, daemon=True).start()

        def _clear_sticky_force_on_tab_open():
            """تب ساخته شد: force چسبنده از session قبلی را خاموش کن (انتشار جدید دوباره True می‌گذارد)."""
            try:
                import time as _t
                db = load_db()
                s = db.setdefault("settings", {})
                sig = dict(s.get("client_software_update") or {})
                if not sig.get("force"):
                    return
                from server.update_service import progress_path
                # جلسهٔ خیلی تازه (<۳د) با waiting/updating → شاید انتشار همین الان؛ دست نزن
                try:
                    p = progress_path()
                    if p.is_file():
                        import json as _j
                        pr = _j.loads(p.read_text(encoding="utf-8"))
                        started = float((pr or {}).get("started_at") or 0)
                        fresh = started > 0 and (_t.time() - started) < 200.0
                        if fresh:
                            roles = (pr or {}).get("roles") or {}
                            for st in roles.values():
                                if isinstance(st, dict) and str(st.get("status") or "") in (
                                    "waiting", "updating", "pending",
                                ):
                                    return
                except Exception:
                    pass
                sig["force"] = False
                s["client_software_update"] = sig
                save_db(db)
            except Exception:
                pass

        btn_row = tk.Frame(inner, bg=C["card"])
        btn_row.pack(fill="x", pady=(8, 0))
        styled_btn(btn_row, "⬆  انتشار", do_publish, color=C["btn_success"]).pack(
            side="right", padx=4
        )
        styled_btn(btn_row, "🔄  تازه‌سازی وضعیت", _refresh_status, color=C["btn_ghost"]).pack(
            side="right", padx=4
        )
        _clear_sticky_force_on_tab_open()
        _refresh_status()

    def _build_backup_tab(self, tab):
        """تب بک‌آپ بدون رمز داخل نرم‌افزار — قفل روی پوشه ویندوز است."""
        _ensure_backup_dirs_protected()
        self._build_backup_tab_unlocked(tab)

    def _build_backup_tab_unlocked(self, tab):
        BG = C["panel"]
        _ensure_backup_dirs_protected()

        # ── اسکرول ──
        canvas = tk.Canvas(tab, bg=BG, highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a",
                           activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        sf = tk.Frame(canvas, bg=BG)
        _win = canvas.create_window((0,0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_win, width=e.width))
        register_scroll_canvas(canvas, sf)

        tk.Label(sf, text="💾  مدیریت بک‌آپ  (پوشه با رمز ادمین ویندوز قفل است — بدون رمز داخل نرم‌افزار)",
                 bg=BG, fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=(12,4))

        # ══════════════════════════════════════════════
        #  بخش ۱ — مسیرهای بک‌آپ
        # ══════════════════════════════════════════════
        sec1 = tk.LabelFrame(sf, text="  📁  مسیرهای بک‌آپ  ",
                              bg=BG, fg=C["accent"], font=FONT_NORM,
                              bd=1, relief="groove", labelanchor="e")
        sec1.pack(fill="x", padx=16, pady=(0,10))

        tk.Label(sec1,
                 text="می‌توانید تا ۳ مسیر مختلف تعریف کنید. بک‌آپ خودکار به همه مسیرها کپی می‌شود.",
                 bg=BG, fg=C["text_dim"], font=FONT_SMALL, justify="right").pack(anchor="e", padx=12, pady=(6,2))

        # خواندن مسیرها از settings
        def _get_backup_paths():
            _db = load_db()
            return _db.get("settings", {}).get("backup_paths", [
                BACKUP_DIR,
                BACKUP_DIR2,
                ""
            ])

        def _save_backup_paths(paths):
            _db = load_db()
            _db.setdefault("settings", {})["backup_paths"] = paths
            save_db(_db)
            # آپدیت متغیر global تا save_db از مسیرهای جدید استفاده کنه
            global BACKUP_DIR2
            valid = [p for p in paths if p and p.strip()]
            if len(valid) >= 2:
                BACKUP_DIR2 = valid[1]

        path_vars = []
        path_entries = []

        def _build_path_rows():
            for w in path_frame.winfo_children():
                w.destroy()
            paths = _get_backup_paths()
            while len(paths) < 3:
                paths.append("")
            path_vars.clear(); path_entries.clear()
            labels = ["مسیر اول (پیش‌فرض):", "مسیر دوم (فلش/هارد خارجی):", "مسیر سوم (شبکه/درایو دیگر):"]
            icons  = ["💻", "🔌", "🌐"]
            for i in range(3):
                row = tk.Frame(path_frame, bg=BG)
                row.pack(fill="x", padx=10, pady=4)
                tk.Label(row, text=f"{icons[i]}  {labels[i]}", bg=BG,
                         fg=C["text"], font=FONT_SMALL).pack(anchor="e")
                ent_row = tk.Frame(row, bg=BG)
                ent_row.pack(fill="x")
                v = tk.StringVar(value=paths[i])
                path_vars.append(v)
                ent = tk.Entry(ent_row, textvariable=v, font=FONT_SMALL,
                               bg=C["entry_bg"], fg=C["text"],
                               insertbackground=C["accent"], bd=0, relief="flat",
                               highlightthickness=1, highlightbackground=C["border"],
                               highlightcolor=C["accent"], justify="left", width=48)
                ent.pack(side="right", fill="x", expand=True, padx=(0,6))
                path_entries.append(ent)

                def _browse(var=v):
                    chosen = filedialog.askdirectory(title="انتخاب مسیر بک‌آپ", parent=self)
                    if chosen:
                        var.set(chosen)

                tk.Button(ent_row, text="📂 انتخاب", command=_browse,
                          bg=C["btn_ghost"], fg=C["text"], font=FONT_SMALL,
                          bd=0, relief="flat", cursor="hand2", padx=8, pady=4).pack(side="right")

        path_frame = tk.Frame(sec1, bg=BG)
        path_frame.pack(fill="x")
        _build_path_rows()

        status_path = tk.Label(sec1, text="", bg=BG, fg=C["success"], font=FONT_SMALL)
        status_path.pack(anchor="e", padx=12, pady=2)

        def save_paths():
            paths = [v.get().strip() for v in path_vars]
            _save_backup_paths(paths)
            # آپدیت مسیرهای global
            global BACKUP_DIR, BACKUP_DIR2
            valid = [p for p in paths if p]
            if valid:
                BACKUP_DIR = valid[0]
            if len(valid) >= 2:
                BACKUP_DIR2 = valid[1]
            status_path.config(text="✔  مسیرها ذخیره شدند", fg=C["success"])
            tab.after(3000, lambda: status_path.config(text=""))

        styled_btn(sec1, "💾  ذخیره مسیرها", save_paths,
                   color=C["btn_success"]).pack(anchor="w", padx=12, pady=(4,10))

        # ══════════════════════════════════════════════
        #  بخش ۲ — وضعیت و لیست بک‌آپ‌ها
        # ══════════════════════════════════════════════
        sec2 = tk.LabelFrame(sf, text="  📋  بک‌آپ‌های موجود  ",
                              bg=BG, fg=C["accent"], font=FONT_NORM,
                              bd=1, relief="groove", labelanchor="e")
        sec2.pack(fill="x", padx=16, pady=(0,10))

        info_lbl = tk.Label(sec2, text="", bg=BG, fg=C["text_dim"], font=FONT_SMALL)
        info_lbl.pack(anchor="e", padx=12, pady=4)

        cols_b = ("path","fn","sz","dt")
        heads_b = ("مسیر","نام فایل","حجم","تاریخ")
        tf_b, tree_b = scrolled_tree(sec2, cols_b, heads_b, height=10)
        tree_b.column("path", width=160, anchor="e")
        tree_b.column("fn",   width=220, anchor="e")
        tree_b.column("sz",   width=80,  anchor="center")
        tree_b.column("dt",   width=160, anchor="center")
        tf_b.pack(fill="both", expand=True, padx=12, pady=6)

        def refresh_b():
            tree_b.delete(*tree_b.get_children())
            paths = _get_backup_paths()
            total = 0
            for bdir in paths:
                if not bdir:
                    continue
                data_dir = _backup_data_dir(bdir)
                if not os.path.exists(data_dir):
                    continue
                try:
                    files = sorted([
                        f for f in os.listdir(data_dir)
                        if f.startswith("slab_db_") and (
                            f.endswith(".json") or f.endswith(".zip")
                        )
                    ], reverse=True)
                except Exception:
                    continue
                for fn in files:
                    fp = os.path.join(data_dir, fn)
                    try:
                        sz = os.path.getsize(fp)
                    except Exception:
                        continue
                    sz_s = f"{sz/1024:.1f} KB" if sz < 1048576 else f"{sz/1048576:.2f} MB"
                    mt = to_shamsi(datetime.datetime.fromtimestamp(os.path.getmtime(fp)))
                    short_dir = os.path.basename(bdir) or bdir
                    tree_b.insert("", "end", values=(short_dir, fn, sz_s, mt),
                                  tags=("file",))
                    total += 1
            info_lbl.config(text=f"📦 مجموع {total} بک‌آپ در {len([p for p in paths if p])} مسیر")

        refresh_b()

        # ══════════════════════════════════════════════
        #  بخش ۳ — دکمه‌های عملیات
        # ══════════════════════════════════════════════
        ctrl_b = tk.Frame(sf, bg=BG)
        ctrl_b.pack(fill="x", padx=16, pady=6)

        def manual_backup():
            paths = [p for p in _get_backup_paths() if p and p.strip()]
            if not paths:
                messagebox.showwarning("خطا", "ابتدا حداقل یک مسیر تعریف کنید.", parent=self)
                return
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            success = []
            try:
                if r"D:\SteelFactory2-v2" not in sys.path:
                    sys.path.insert(0, r"D:\SteelFactory2-v2")
                from shared.backup_vault import encrypt_bytes_to_zip, set_vault_password
                from pathlib import Path as _P
                import json as _json
                if is_backup_vault_password_required():
                    set_vault_password(get_backup_vault_password())
                else:
                    set_vault_password("")
                raw = _json.dumps(load_db(), ensure_ascii=False, indent=2).encode("utf-8")
            except Exception as ex:
                messagebox.showerror("خطا", str(ex), parent=self)
                return
            for bdir in paths:
                try:
                    data_dir = _backup_data_dir(bdir)
                    os.makedirs(data_dir, exist_ok=True)
                    dst = _P(data_dir) / f"slab_db_MANUAL_{ts}.json"
                    encrypt_bytes_to_zip(raw, dst, arcname=f"slab_db_MANUAL_{ts}.json")
                    success.append(bdir)
                except Exception as ex:
                    messagebox.showwarning("خطا", f"مسیر {bdir}:\n{ex}", parent=self)
            if success:
                messagebox.showinfo("موفق",
                    f"بک‌آپ در {len(success)} مسیر ذخیره شد:\n" +
                    "\n".join(success), parent=self)
                refresh_b()

        def restore_b():
            sel = tree_b.selection()
            if not sel:
                messagebox.showwarning("خطا", "یک فایل انتخاب کنید.", parent=self); return
            row_vals = tree_b.item(sel[0], "values")
            short_dir, fn = row_vals[0], row_vals[1]
            # پیدا کردن مسیر کامل (داخل .secure)
            full_path = None
            for bdir in _get_backup_paths():
                if not bdir:
                    continue
                data_dir = _backup_data_dir(bdir)
                candidate = os.path.join(data_dir, fn)
                if os.path.exists(candidate):
                    full_path = candidate
                    break
                # سازگاری: فایل هنوز در خود دروازه باشد
                candidate2 = os.path.join(bdir, fn)
                if os.path.exists(candidate2):
                    full_path = candidate2
                    break
            if not full_path:
                messagebox.showerror("خطا", "فایل پیدا نشد.", parent=self); return
            if not messagebox.askyesno("⚠️  بازیابی",
                    f"«{fn}»\nبازیابی شود؟\nداده فعلی جایگزین می‌شود!",
                    parent=self): return
            try:
                restore_path = full_path
                if fn.endswith(".protected.zip") or fn.endswith(".zip"):
                    if r"D:\SteelFactory2-v2" not in sys.path:
                        sys.path.insert(0, r"D:\SteelFactory2-v2")
                    from shared.backup_vault import decrypt_zip_to_bytes
                    from pathlib import Path as _P
                    import tempfile
                    # بک‌آپ‌های قدیمی ZIP — رمز اختیاری برای سازگاری
                    pw = get_backup_vault_password() or "Reza9063"
                    raw = decrypt_zip_to_bytes(_P(full_path), pw)
                    td = tempfile.mkdtemp(prefix="stf_restore_")
                    restore_path = os.path.join(td, "restore.json")
                    with open(restore_path, "wb") as f:
                        f.write(raw)
                if restore_db_from_file(restore_path):
                    messagebox.showinfo("موفق",
                        "بازیابی انجام شد.\nلطفاً برنامه را مجدد راه‌اندازی کنید.", parent=self)
                else:
                    messagebox.showerror("خطا", "بازیابی ناموفق بود.", parent=self)
            except Exception as ex:
                messagebox.showerror("خطا", f"بازیابی ناموفق:\n{ex}", parent=self)

        def export_db():
            path = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON", "*.json")],
                initialfile=f"slab_db_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                parent=self)
            if not path: return
            if export_current_db(path):
                messagebox.showinfo("موفق", f"دیتابیس صادر شد:\n{path}", parent=self)
            else:
                messagebox.showerror("خطا", "صادر کردن دیتابیس ناموفق بود.", parent=self)

        styled_btn(ctrl_b, "💾  بک‌آپ دستی", manual_backup,
                   color=C["btn_success"]).pack(side="right", padx=4)
        styled_btn(ctrl_b, "🔄  بازیابی انتخابی", restore_b,
                   color=C["warning"]).pack(side="right", padx=4)
        styled_btn(ctrl_b, "📤  صادر کردن دیتابیس", export_db,
                   color=C["accent2"]).pack(side="right", padx=4)
        styled_btn(ctrl_b, "🔃  بارگذاری مجدد لیست", refresh_b,
                   color=C["btn_ghost"]).pack(side="left", padx=4)



    # ════════════════════════════════════════════════════════════
    #  گزارش‌گیری هوشمند — تیک بزن، دریافت کن
    # ════════════════════════════════════════════════════════════
    def _build_report_smart(self, tab):
        """
        یه صفحه گزارش‌گیری با تیک‌باکس:
        کاربر هر گزارشی می‌خواد تیک میزنه → Excel میگیره.
        هر اسلب تاریخچه کامل دارد.
        """
        # ── اسکرول ──
        canvas = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a", activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)
        sf = tk.Frame(canvas, bg=C["panel"])
        _win = canvas.create_window((0,0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_win, width=e.width))
        register_scroll_canvas(canvas, sf)

        # ── عنوان ──
        tk.Label(sf, text="🎯  گزارش‌گیری هوشمند",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=(12,4))
        tk.Label(sf, text="گزارش‌های مورد نظر را تیک بزنید، بازه زمانی تعیین کنید و Excel دریافت کنید.",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16, pady=(0,10))

        # ── بازه زمانی ──
        tf_card = card_frame(sf)
        tf_card.pack(fill="x", padx=16, pady=(0,8))
        tk.Frame(tf_card, bg=C["accent"], height=2).pack(fill="x")
        tf_in = tk.Frame(tf_card, bg=C["card"])
        tf_in.pack(fill="x", padx=16, pady=12)
        tk.Label(tf_in, text="📅  بازه زمانی گزارش:", bg=C["card"],
                 fg=C["accent"], font=(_MAIN_FONT,10,"bold")).pack(anchor="e", pady=(0,8))

        _now_date = to_shamsi(datetime.datetime.now()).split("  ")[0]

        def _mk(parent, var, w):
            return tk.Entry(parent, textvariable=var,
                     bg=C["entry_bg"], fg=C["text"],
                     insertbackground=C["accent"], font=FONT_MONO,
                     justify="right", bd=0, relief="flat",
                     highlightthickness=1, highlightbackground=C["border"],
                     highlightcolor=C["accent"], width=w)

        def _mk_p(parent, lbl, var, w, fg=None):
            f = tk.Frame(parent, bg=C["card"])
            f.pack(side="right", padx=(0, 10))
            tk.Label(f, text=lbl, bg=C["card"],
                     fg=fg or C["accent"],
                     font=(_MAIN_FONT,11,"bold")).pack(side="right")
            _mk(f, var, w).pack(side="right", padx=(4, 0))

        tr1 = tk.Frame(tf_in, bg=C["card"]); tr1.pack(fill="x", pady=3)
        _fd = tk.StringVar(value=get_first_report_date_sh())
        _ft = tk.StringVar(value="00:00:00")
        _mk_p(tr1, "از تاریخ", _fd, 13)
        _mk_p(tr1, "از ساعت",  _ft,  9, fg=C["text_dim"])

        tr2 = tk.Frame(tf_in, bg=C["card"]); tr2.pack(fill="x", pady=3)
        _td = tk.StringVar(value=_now_date)
        _tt = tk.StringVar(value="23:59:59")
        _mk_p(tr2, "تا تاریخ", _td, 13)
        _mk_p(tr2, "تا ساعت",  _tt,  9, fg=C["text_dim"])

        from_var = tk.StringVar()
        to_var   = tk.StringVar()
        def _sync2(*_):
            from_var.set(f"{_fd.get().strip()}  {_ft.get().strip()}")
            to_var.set(f"{_td.get().strip()}  {_tt.get().strip()}")
        for _v in (_fd, _ft, _td, _tt): _v.trace_add("write", _sync2)
        _sync2()

        # ── فیلتر پرسنل (برای گزارش تکی/کلی پرسنل) ──
        tr3 = tk.Frame(tf_in, bg=C["card"]); tr3.pack(fill="x", pady=3)
        tk.Label(tr3, text="پرسنل:", bg=C["card"], fg=C["text"], font=FONT_NORM,
                 width=6, anchor="e").pack(side="right")
        db = load_db()
        persons = ["همه"] + [f"{ud.get('display',un)}  [{un}]"
                             for un,ud in db["users"].items() if un!="admin"]
        pers_cb = make_combo(tr3, persons, width=26)
        pers_cb.set("همه")
        pers_cb.pack(side="right", padx=6)
        tk.Label(tr3, text="(فقط برای گزارش پرسنل)",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(side="right")

        # ── انتخاب گزارش‌ها ──
        sel_card = card_frame(sf)
        sel_card.pack(fill="x", padx=16, pady=(0,8))
        tk.Frame(sel_card, bg=C["gold"], height=2).pack(fill="x")
        sel_in = tk.Frame(sel_card, bg=C["card"])
        sel_in.pack(fill="x", padx=16, pady=12)
        tk.Label(sel_in, text="☑  انتخاب گزارش‌ها:", bg=C["card"],
                 fg=C["gold"], font=(_MAIN_FONT,10,"bold")).pack(anchor="e", pady=(0,10))

        # لیست گزارش‌ها با توضیح
        REPORTS = [
            ("rep_all_slabs",    "📋  تاریخچه کامل همه اسلب‌ها",
             "هر اسلب: ثبت ذوب / QC / اسکارف / برش / باومن / انتقال / برگشت / خروج / محل فعلی"),
            ("rep_single_pers",  "👤  گزارش تکی پرسنل",
             "همه فعالیت‌های یک پرسنل خاص (انتخاب از فیلد پرسنل بالا)"),
            ("rep_all_pers",     "👥  گزارش کلی پرسنل",
             "یک شیت برای هر پرسنل — همه کارهایی که انجام داده"),
            ("rep_qc",           "✅  گزارش کنترل کیفی شده‌ها",
             "اسلب‌هایی که QC شده‌اند — توسط چه کسی، چه زمانی، الان کجاست"),
            ("rep_scarf",        "⚙  گزارش اسکارف‌ها",
             "اسلب‌هایی که اسکارف شده‌اند — نوع اسکارف / توسط چه کسی / زمان"),
            ("rep_cut",          "✂  گزارش برش‌ها",
             "اسلب‌هایی که برش خورده‌اند — نوع برش / باومن داشته یا نه / زمان"),
            ("rep_exit",         "🚪  گزارش خروجی‌ها (نوبت‌کار)",
             "اسلب‌هایی که خروج ثبت شده — توسط چه کسی / چه زمانی"),
            ("rep_outside1",     "🏭  موجودی انبار روباز ۱",
             "اسلب‌هایی که الان در انبار روباز ۱ هستند"),
            ("rep_outside2",     "🏭  موجودی انبار روباز ۲",
             "اسلب‌هایی که الان در انبار روباز ۲ هستند"),
            ("rep_inside",       "🏬  موجودی انبار داخلی",
             "اسلب‌هایی که الان در انبار داخلی هستند"),
            ("rep_returns",      "↩  گزارش برگشتی‌ها",
             "اسلب‌هایی که برگشت خورده‌اند — چند بار / از کجا / دلیل / تاریخ هر برگشت"),
            ("rep_bauman",       "🔬  گزارش تست باومن",
             "اسلب‌هایی که تست باومن داشته‌اند — وضعیت تحویل آزمایشگاه"),
        ]

        check_vars = {}
        # ۲ ستون
        grid = tk.Frame(sel_in, bg=C["card"])
        grid.pack(fill="x")
        grid.grid_columnconfigure(0, weight=1, uniform="col")
        grid.grid_columnconfigure(1, weight=1, uniform="col")

        for i, (key, title, desc) in enumerate(REPORTS):
            var = tk.BooleanVar(value=True)
            check_vars[key] = var
            col = i % 2; row = i // 2
            item_frame = tk.Frame(grid, bg="#0a1218", cursor="hand2",
                                  highlightthickness=1, highlightbackground=C["border"])
            item_frame.grid(row=row, column=col, padx=4, pady=3, sticky="nsew")
            inner_f = tk.Frame(item_frame, bg="#0a1218")
            inner_f.pack(fill="x", padx=10, pady=8)

            # چک‌مارک
            chk_lbl = tk.Label(inner_f, text="◉", bg="#0a1218", fg=C["gold"],
                                font=("Segoe UI Symbol",14, "bold"), cursor="hand2")
            chk_lbl.pack(side="left", padx=(0,8))

            # متن
            txt_f = tk.Frame(inner_f, bg="#0a1218")
            txt_f.pack(side="right", fill="x", expand=True)
            tk.Label(txt_f, text=title, bg="#0a1218", fg=C["text_bright"],
                     font=(_MAIN_FONT,10,"bold"), anchor="e").pack(anchor="e")
            tk.Label(txt_f, text=desc, bg="#0a1218", fg=C["text_dim"],
                     font=(_MAIN_FONT,8, "bold"), anchor="e", wraplength=280).pack(anchor="e")

            def mk_toggle(v=var, c=chk_lbl, f=item_frame, inf=inner_f):
                def toggle(e=None):
                    v.set(not v.get())
                    on = v.get()
                    c.config(text="◉" if on else "○",
                             fg=C["gold"] if on else C["text_dim"])
                    f.config(bg="#0a1218" if on else "#050a0e",
                             highlightbackground=C["gold"] if on else C["border"])
                    inf.config(bg="#0a1218" if on else "#050a0e")
                    for w in inf.winfo_children():
                        try: w.config(bg="#0a1218" if on else "#050a0e")
                        except: pass
                        try:
                            for ww in w.winfo_children():
                                ww.config(bg="#0a1218" if on else "#050a0e")
                        except: pass
                return toggle
            tog = mk_toggle()
            for w in [item_frame, inner_f, chk_lbl]:
                w.bind("<Button-1>", tog)
            for child in inner_f.winfo_children():
                child.bind("<Button-1>", tog)
                for cc in child.winfo_children():
                    cc.bind("<Button-1>", tog)

        # ── انتخاب همه / هیچ‌کدام ──
        quick = tk.Frame(sel_in, bg=C["card"])
        quick.pack(fill="x", pady=(10,0))
        def select_all():
            for v in check_vars.values(): v.set(True)
            # آپدیت ظاهر
            for w in grid.winfo_children():
                w.config(bg="#0a1218", highlightbackground=C["gold"])
        def select_none():
            for v in check_vars.values(): v.set(False)
        tk.Button(quick, text="✔ انتخاب همه", command=select_all,
                  bg=C["btn_ghost"], fg=C["gold"], font=(_MAIN_FONT,9, "bold"),
                  bd=0, relief="flat", cursor="hand2", padx=10, pady=4
                  ).pack(side="right", padx=4)
        tk.Button(quick, text="✕ هیچ‌کدام", command=select_none,
                  bg=C["btn_ghost"], fg=C["text_dim"], font=(_MAIN_FONT,9, "bold"),
                  bd=0, relief="flat", cursor="hand2", padx=10, pady=4
                  ).pack(side="right", padx=4)

        # ── وضعیت + دکمه دریافت ──
        status_lbl = tk.Label(sf, text="", bg=C["panel"], fg=C["success"], font=FONT_NORM)
        status_lbl.pack(anchor="e", padx=20, pady=4)

        def parse_sh(s):
            s = s.strip().replace("  "," ")
            import re as re2
            m = re2.match(r"(\d{4})/(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{2}):(\d{2})", s)
            if not m: return s
            jy,jm,jd,hh,mm,ss = [int(x) for x in m.groups()]
            return f"{jy:04d}/{jm:02d}/{jd:02d}  {hh:02d}:{mm:02d}:{ss:02d}"

        def flt_time(at, fsh, tsh):
            if not fsh or not tsh: return True
            at2 = at.replace("  "," ") if at else ""
            return not at2 or (fsh.replace("  "," ") <= at2 <= tsh.replace("  "," "))

        def slab_history_row(db, sid):
            """یک سطر کامل تاریخچه یک اسلب — تاریخ و ساعت جدا"""
            info = get_slab_full_info(db, sid)
            rets = info.get("returns_detail", [])
            ret_detail = ""
            for r in rets:
                ret_detail += f"{r.get('label','—')}: {r.get('returned_at','—')} | دلیل: {r.get('reason','—')} | "
            reg_d, reg_t = split_dt(info["registered_at"])
            qc_d,  qc_t  = split_dt(info["qc_at"])
            bm_d,  bm_t  = split_dt(info["bauman_at"])
            lb_d,  lb_t  = split_dt(info["lab_delivered_at"])
            tr_d,  tr_t  = split_dt(info["transfer_at"])
            ex_d,  ex_t  = split_dt(info["exit_at"])
            return (
                info["slab_id"],
                info["registered_by"], reg_d, reg_t,
                info["qc_status"],
                info["qc_by"], qc_d, qc_t,
                info["scarf"],
                info["cut"],
                info["bauman_done"], bm_d, bm_t, info["bauman_by"],
                info["lab_delivered"], lb_d, lb_t, info["lab_delivered_by"],
                info["transfer_dest"], tr_d, tr_t, info["transfer_by"],
                str(info["ret_to_internal"]) + " بار",
                str(info["ret_to_outside"]) + " بار",
                ret_detail.rstrip(" | ") or "ندارد",
                info["scrap"],
                info["exit_status"],
                info["exit_by"], ex_d, ex_t,
                info["current_location"],
                info.get("note",""),
            )

        SLAB_HEADS = [
            "شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت",
            "وضعیت QC","تأییدکننده QC","تاریخ QC","ساعت QC",
            "اسکارف","برش",
            "تست باومن","تاریخ باومن","ساعت باومن","برش‌کار باومن",
            "تحویل آزمایشگاه","تاریخ تحویل","ساعت تحویل","تحویل‌دهنده",
            "مقصد انتقال","تاریخ انتقال","ساعت انتقال","انتقال‌دهنده",
            "برگشت به داخلی","برگشت به روباز","جزئیات برگشت‌ها",
            "قراضه",
            "وضعیت خروج","خروج توسط","تاریخ خروج","ساعت خروج",
            "محل فعلی","توضیحات"
        ]

        def export_reports():
            selected = [k for k,v in check_vars.items() if v.get()]
            if not selected:
                messagebox.showwarning("خطا","حداقل یک گزارش انتخاب کنید.",parent=self)
                return
            if not XLSX:
                messagebox.showerror("خطا","openpyxl نصب نیست.\npip install openpyxl",parent=self)
                return

            fsh = parse_sh(from_var.get())
            tsh = parse_sh(to_var.get())
            db = load_db()

            # پرسنل انتخابی
            pers_sel = pers_cb.get()
            if pers_sel == "همه":
                sel_uname = None
            else:
                m_p = pers_sel.split("[")
                sel_uname = m_p[-1].rstrip("]").strip() if len(m_p)>1 else pers_sel
                sel_display = pers_sel.split("[")[0].strip()

            fname = f"گزارش_هوشمند_{shamsi_date_for_filename()}.xlsx"
            path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")], initialfile=fname, parent=self)
            if not path: return

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # شیت پیش‌فرض رو حذف کن

            CW_SLAB = [14,14,22,14,14,22,28,28,10,22,14,16,22,14,16,22,14,10,10,40,18,14,14,22,22,20]

            def add_sheet(title, heads, rows, col_widths=None):
                ws = wb.create_sheet(title[:30])
                self._write_sheet(ws, heads, rows, col_widths)

            # ── ۱. تاریخچه کامل همه اسلب‌ها ──
            if check_vars.get("rep_all_slabs",tk.BooleanVar()).get():
                rows = []
                for rec in db["melts"]:
                    at = rec.get("registered_at","")
                    if not flt_time(at, fsh, tsh): continue
                    rows.append(slab_history_row(db, rec["slab_id"]))
                add_sheet("تاریخچه کامل اسلب‌ها", SLAB_HEADS, rows, CW_SLAB)

            # ── ۲. گزارش تکی پرسنل ──
            if check_vars.get("rep_single_pers",tk.BooleanVar()).get() and sel_uname:
                rows = _person_rows(db, sel_uname, fsh, tsh)
                add_sheet(f"پرسنل_{sel_display[:20]}", 
                          ["بخش","شماره اسلب","عملیات","تاریخ","ساعت","جزئیات"], rows)

            # ── ۳. گزارش کلی پرسنل (یه شیت برای هر نفر) ──
            if check_vars.get("rep_all_pers",tk.BooleanVar()).get():
                for un, ud in db["users"].items():
                    if un == "admin": continue
                    rows = _person_rows(db, un, fsh, tsh)
                    if rows:
                        add_sheet(f"{ud.get('display',un)[:25]}",
                                  ["بخش","شماره اسلب","عملیات","تاریخ","ساعت","جزئیات"], rows)

            # ── ۴. کنترل کیفی شده‌ها ──
            if check_vars.get("rep_qc",tk.BooleanVar()).get():
                rows = []
                for rec in db["melts"]:
                    if rec.get("qc_status")!="کنترل کیفی شده": continue
                    if not flt_time(rec.get("qc_at",""), fsh, tsh): continue
                    info = get_slab_full_info(db, rec["slab_id"])
                    qc_d, qc_t = split_dt(rec.get("qc_at","—"))
                    reg_d, reg_t = split_dt(rec.get("registered_at","—"))
                    rows.append((
                        rec["slab_id"],
                        get_display_name(rec.get("qc_by","—"), db),
                        qc_d, qc_t,
                        get_display_name(rec.get("registered_by","—"), db),
                        reg_d, reg_t,
                        info["current_location"],
                        info["scarf"],
                        info["cut"],
                        info["bauman_done"],
                        info["exit_status"],
                    ))
                add_sheet("کنترل کیفی شده‌ها",
                          ["شماره اسلب","تأییدکننده QC","تاریخ QC","ساعت QC",
                           "ثبت‌کننده","تاریخ ثبت","ساعت ثبت",
                           "محل فعلی","اسکارف","برش","باومن","وضعیت خروج"],
                          rows)

            # ── ۵. اسکارف‌ها ──
            if check_vars.get("rep_scarf",tk.BooleanVar()).get():
                rows = []
                for rec in db["scarf_cut"]:
                    if rec.get("operation")!="اسکارفی": continue
                    if not flt_time(rec.get("registered_at",""), fsh, tsh): continue
                    _d, _t = split_dt(rec.get("registered_at","—"))
                    rows.append((
                        rec["slab_id"],
                        get_display_name(rec.get("registered_by","—"), db),
                        _d, _t,
                        rec.get("reason","—"),
                        rec.get("note","—"),
                        "✔" if rec.get("bauman_done") else "ندارد",
                    ))
                add_sheet("اسکارف‌ها",
                          ["شماره اسلب","ثبت‌کننده","تاریخ","ساعت","دلایل اسکارف","توضیحات","تست باومن"],
                          rows)

            # ── ۶. برش‌ها ──
            if check_vars.get("rep_cut",tk.BooleanVar()).get():
                rows = []
                for rec in db["scarf_cut"]:
                    if rec.get("operation")!="برشی": continue
                    if not flt_time(rec.get("registered_at",""), fsh, tsh): continue
                    bm = rec if rec.get("bauman_done") else None
                    lb = next((r for r in db["lab_deliveries"] if r["slab_id"]==rec["slab_id"]), None)
                    _d, _t = split_dt(rec.get("registered_at","—"))
                    _ld, _lt = split_dt(lb.get("delivered_at","—") if lb else "—")
                    rows.append((
                        rec["slab_id"],
                        get_display_name(rec.get("registered_by","—"), db),
                        _d, _t,
                        rec.get("reason","—"),
                        rec.get("note","—"),
                        "✔ دارد" if rec.get("bauman_done") else "ندارد",
                        ("تحویل داده شده" if lb else "در انتظار") if bm else "ندارد",
                        _ld, _lt,
                        get_display_name(lb.get("delivered_by","—"), db) if lb else "—",
                    ))
                add_sheet("برش‌ها",
                          ["شماره اسلب","ثبت‌کننده","تاریخ","ساعت","دلایل برش","توضیحات",
                           "تست باومن","وضعیت آزمایشگاه","تاریخ تحویل","ساعت تحویل","تحویل‌دهنده"],
                          rows)

            # ── ۷. خروجی‌ها ──
            if check_vars.get("rep_exit",tk.BooleanVar()).get():
                rows = []
                for rec in db["melts"]:
                    if rec.get("exit_status")!="خروج زده شده": continue
                    if not flt_time(rec.get("exit_at",""), fsh, tsh): continue
                    info = get_slab_full_info(db, rec["slab_id"])
                    ex_d, ex_t = split_dt(rec.get("exit_at","—"))
                    qc_d, qc_t = split_dt(rec.get("qc_at","—"))
                    rows.append((
                        rec["slab_id"],
                        get_display_name(rec.get("exit_by","—"), db),
                        ex_d, ex_t,
                        get_display_name(rec.get("qc_by","—"), db),
                        qc_d, qc_t,
                        info["transfer_dest"],
                        info["current_location"],
                    ))
                add_sheet("خروجی‌ها (نوبت‌کار)",
                          ["شماره اسلب","خروج توسط","تاریخ خروج","ساعت خروج",
                           "تأییدکننده QC","تاریخ QC","ساعت QC","آخرین مقصد","محل نهایی"],
                          rows)

            # ── ۸. موجودی انبار روباز ۱ ──
            if check_vars.get("rep_outside1",tk.BooleanVar()).get():
                rows = _warehouse_rows(db, "روباز ۱", fsh, tsh)
                add_sheet("انبار روباز ۱",
                          ["شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","انتقال‌دهنده","تاریخ انتقال","ساعت انتقال",
                           "تأییدکننده QC","اسکارف","برش","تعداد برگشت"],
                          rows)

            # ── ۹. موجودی انبار روباز ۲ ──
            if check_vars.get("rep_outside2",tk.BooleanVar()).get():
                rows = _warehouse_rows(db, "روباز ۲", fsh, tsh)
                add_sheet("انبار روباز ۲",
                          ["شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","انتقال‌دهنده","تاریخ انتقال","ساعت انتقال",
                           "تأییدکننده QC","اسکارف","برش","تعداد برگشت"],
                          rows)

            # ── ۱۰. موجودی انبار داخلی ──
            if check_vars.get("rep_inside",tk.BooleanVar()).get():
                rows = []
                for rec in db["melts"]:
                    info = get_slab_full_info(db, rec["slab_id"])
                    loc = info["current_location"]
                    if "داخلی" not in loc: continue
                    reg_d, reg_t = split_dt(rec.get("registered_at","—"))
                    rows.append((
                        rec["slab_id"],
                        get_display_name(rec.get("registered_by","—"), db),
                        reg_d, reg_t,
                        rec.get("qc_status","—"),
                        get_display_name(rec.get("qc_by","—"), db),
                        info["scarf"],
                        info["cut"],
                        info["bauman_done"],
                        str(info["ret_to_internal"]) + " بار",
                        loc,
                    ))
                add_sheet("انبار داخلی",
                          ["شماره اسلب","ثبت‌کننده","تاریخ ثبت","ساعت ثبت","وضعیت QC","تأییدکننده",
                           "اسکارف","برش","باومن","تعداد برگشت","محل فعلی"],
                          rows)

            # ── ۱۱. برگشتی‌ها ──
            if check_vars.get("rep_returns",tk.BooleanVar()).get():
                rows = []
                from collections import defaultdict
                by_slab = defaultdict(list)
                for r in db.get("return_log",[]):
                    if flt_time(r.get("returned_at",""), fsh, tsh):
                        by_slab[r.get("slab_id","")].append(r)
                for sid, rets in by_slab.items():
                    info = get_slab_full_info(db, sid)
                    for r in sorted(rets, key=lambda x: x.get("return_number",1)):
                        ret_d, ret_t = split_dt(r.get("returned_at","—"))
                        rows.append((
                            sid,
                            r.get("return_number","—"),
                            r.get("label","—"),
                            get_display_name(r.get("returned_by","—"), db),
                            ret_d, ret_t,
                            r.get("source","—"),
                            r.get("reason","—"),
                            info["current_location"],
                            info["qc_status"],
                        ))
                add_sheet("برگشتی‌ها",
                          ["شماره اسلب","شماره برگشت","نوع برگشت","ثبت‌کننده","تاریخ","ساعت",
                           "از انبار","دلیل","محل فعلی","وضعیت QC"],
                          rows)

            # ── ۱۲. تست باومن ──
            if check_vars.get("rep_bauman",tk.BooleanVar()).get():
                rows = []
                for rec in db["scarf_cut"]:
                    if rec.get("operation")!="برشی" or not rec.get("bauman_done"): continue
                    if not flt_time(rec.get("registered_at",""), fsh, tsh): continue
                    lb = next((r for r in db["lab_deliveries"] if r["slab_id"]==rec["slab_id"]), None)
                    melt = next((r for r in db["melts"] if r["slab_id"]==rec["slab_id"]), {})
                    cut_d, cut_t = split_dt(rec.get("registered_at","—"))
                    lb_d, lb_t = split_dt(lb.get("delivered_at","—") if lb else "—")
                    rows.append((
                        rec["slab_id"],
                        get_display_name(rec.get("registered_by","—"), db),
                        cut_d, cut_t,
                        "تحویل داده شده" if lb else "در انتظار",
                        get_display_name(lb.get("delivered_by","—"), db) if lb else "تحویل داده نشده",
                        lb_d, lb_t,
                        melt.get("qc_status","—"),
                        get_display_name(melt.get("qc_by","—"), db),
                    ))
                add_sheet("تست باومن",
                          ["شماره اسلب","برش‌کار","تاریخ برش","ساعت برش","وضعیت آزمایشگاه",
                           "تحویل‌دهنده","تاریخ تحویل","ساعت تحویل","وضعیت QC","تأییدکننده QC"],
                          rows)

            if not wb.worksheets:
                messagebox.showinfo("خطا","هیچ داده‌ای برای ذخیره یافت نشد.",parent=self)
                return

            wb.save(path)
            n_sheets = len(wb.worksheets)
            status_lbl.config(text=f"✔  Excel با {n_sheets} شیت ذخیره شد: {os.path.basename(path)}")
            messagebox.showinfo("✔  موفق",
                f"فایل Excel با {n_sheets} شیت ذخیره شد:\n{path}",parent=self)

        def _person_rows(db, uname, fsh, tsh):
            rows = []
            def flt(at): return flt_time(at, fsh, tsh)
            for r in db["melts"]:
                if r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    _d,_t = split_dt(r.get("registered_at","—"))
                    rows.append(("ثبت ذوب",r["slab_id"],"ثبت ذوب",_d,_t,r.get("qc_status","—")))
                if r.get("qc_by")==uname and flt(r.get("qc_at","")):
                    _d,_t = split_dt(r.get("qc_at","—"))
                    rows.append(("کنترل کیفی",r["slab_id"],"تأیید QC",_d,_t,r.get("location","—")))
                if r.get("exit_by")==uname and flt(r.get("exit_at","")):
                    _d,_t = split_dt(r.get("exit_at","—"))
                    rows.append(("خروج",r["slab_id"],"خروج",_d,_t,"خارج شده"))
            for r in db["scarf_cut"]:
                if r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    _d,_t = split_dt(r.get("registered_at","—"))
                    rows.append((r.get("operation","عملیات"),r["slab_id"],"ثبت",_d,_t,r.get("reason","—")))
                if r.get("operation")=="برشی" and r.get("bauman_done") and r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    _d,_t = split_dt(r.get("registered_at","—"))
                    rows.append(("باومن",r["slab_id"],"تست باومن",_d,_t,"آماده تحویل"))
            for r in db.get("lab_deliveries", []):
                if r.get("delivered_by")==uname and flt(r.get("delivered_at","")):
                    _d,_t = split_dt(r.get("delivered_at","—"))
                    rows.append(("آزمایشگاه",r["slab_id"],"تحویل آزمایشگاه",_d,_t,"تحویل شد"))
            for r in db["transfers_out"]:
                if r.get("transferred_by")==uname and flt(r.get("transferred_at","")):
                    _d,_t = split_dt(r.get("transferred_at","—"))
                    rows.append(("انتقال",r["slab_id"],"انتقال",_d,_t,r.get("destination","—")))
            for r in db.get("return_log",[]):
                if r.get("returned_by")==uname and flt(r.get("returned_at","")):
                    _d,_t = split_dt(r.get("returned_at","—"))
                    rows.append(("برگشت",r.get("slab_id",""),r.get("label","برگشت"),_d,_t,r.get("reason","—")))
            for r in db["lab_deliveries"]:
                if r.get("delivered_by")==uname and flt(r.get("delivered_at","")):
                    _d,_t = split_dt(r.get("delivered_at","—"))
                    rows.append(("آزمایشگاه",r["slab_id"],"تحویل",_d,_t,""))
            for r in db["scrap"]:
                if r.get("registered_by")==uname and flt(r.get("registered_at","")):
                    _d,_t = split_dt(r.get("registered_at","—"))
                    rows.append(("قراضه",r["slab_id"],"ثبت قراضه",_d,_t,r.get("reason","—")))
            rows.sort(key=lambda x: x[3])
            return rows

        def _warehouse_rows(db, which, fsh, tsh):
            rows = []
            for rec in db["melts"]:
                info = get_slab_full_info(db, rec["slab_id"])
                if which not in info["current_location"]: continue
                tr = next((r for r in db["transfers_out"] if r["slab_id"]==rec["slab_id"]), {})
                reg_d, reg_t = split_dt(rec.get("registered_at","—"))
                tr_d, tr_t = split_dt(tr.get("transferred_at","—"))
                rows.append((
                    rec["slab_id"],
                    get_display_name(rec.get("registered_by","—"), db),
                    reg_d, reg_t,
                    get_display_name(tr.get("transferred_by","—"), db),
                    tr_d, tr_t,
                    get_display_name(rec.get("qc_by","—"), db),
                    info["scarf"],
                    info["cut"],
                    str(info["ret_to_internal"]) + " بار",
                ))
            return rows

        # ── دکمه دریافت ──
        btn_f = tk.Frame(sf, bg=C["panel"])
        btn_f.pack(fill="x", padx=16, pady=14)
        big_btn = tk.Frame(btn_f, bg=C["accent"], cursor="hand2")
        big_btn.pack(side="right")
        big_lbl = tk.Label(big_btn,
                           text="  📥  دریافت Excel گزارش‌های انتخابی  ",
                           bg=C["accent"], fg="#000000",
                           font=(_MAIN_FONT,13,"bold"), padx=20, pady=14, cursor="hand2")
        big_lbl.pack()
        big_btn.bind("<Button-1>", lambda e: export_reports())
        big_lbl.bind("<Button-1>", lambda e: export_reports())
        big_btn.bind("<Enter>", lambda e: [big_btn.config(bg=C["accent_glow"]), big_lbl.config(bg=C["accent_glow"])])
        big_btn.bind("<Leave>", lambda e: [big_btn.config(bg=C["accent"]), big_lbl.config(bg=C["accent"])])

    def _admin_login_log(self, tab):
        """لاگ ورود کاربران — چه کسی چه زمانی وارد سیستم شده"""
        tab.configure(bg=C["panel"])
        tk.Label(tab, text="🔐  تاریخچه ورود به سیستم",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=8)
        tk.Label(tab, text="هر بار که یک کاربر وارد سیستم شود ثبت می‌شود",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        # فیلتر کاربر
        filter_f = card_frame(tab)
        filter_f.pack(fill="x", padx=16, pady=8)
        fi = tk.Frame(filter_f, bg=C["card"])
        fi.pack(padx=14, pady=10, fill="x")
        tk.Label(fi, text="فیلتر کاربر:", bg=C["card"], fg=C["text"],
                 font=FONT_NORM).pack(side="right", padx=(0,6))
        db0 = load_db()
        user_list = ["همه"] + [f"{ud.get('display',un)}  [{un}]"
                                for un,ud in db0["users"].items()]
        user_cb = make_combo(fi, user_list, width=26)
        user_cb.set("همه")
        user_cb.pack(side="right", padx=4)

        cols  = ("idx","display","username","role","login_date","login_time","logout_date","logout_time")
        heads = ("ردیف","نام کاربر","نام کاربری","نقش","تاریخ ورود","ساعت ورود","تاریخ خروج","ساعت خروج")
        tf, tree = scrolled_tree(tab, cols, heads, height=18)
        tf.pack(fill="both", expand=True, padx=16, pady=6)
        tree.column("idx",         width=50,  anchor="center")
        tree.column("display",     width=140, anchor="center")
        tree.column("username",    width=110, anchor="center")
        tree.column("role",        width=90,  anchor="center")
        tree.column("login_date",  width=110, anchor="center")
        tree.column("login_time",  width=90,  anchor="center")
        tree.column("logout_date", width=110, anchor="center")
        tree.column("logout_time", width=90,  anchor="center")

        cnt_lbl = tk.Label(tab, "", bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL)
        cnt_lbl.pack(anchor="e", padx=16)
        search_bar(tab, tree, col_indices=[1,2]).pack(anchor="e", padx=16, pady=2)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        def refresh(*_):
            tree.delete(*tree.get_children())
            db = load_db()
            logs = db.get("login_log", [])
            # فیلتر کاربر
            sel = user_cb.get()
            if sel != "همه":
                un = sel.split("[")[-1].rstrip("]").strip()
                logs = [l for l in logs if l.get("username")==un]
            # جدیدترین اول
            logs = list(reversed(logs))
            for i, entry in enumerate(logs, 1):
                role_fa = {
                    "admin":"مدیریت","shift":"شیفت",
                    "scarf":"برش‌کار","shift_n":"نوبت‌کار"
                }.get(entry.get("role",""),"—")
                in_d, in_t = split_dt(entry.get("at", "—"))
                if entry.get("logout_at"):
                    out_d, out_t = split_dt(entry.get("logout_at", "—"))
                else:
                    out_d, out_t = "—", "—"
                tree.insert("","end", values=(
                    i,
                    entry.get("display","—"),
                    entry.get("username","—"),
                    role_fa,
                    in_d, in_t, out_d, out_t,
                ))
            cnt_lbl.config(text=f"📊 {len(logs)} ورود ثبت شده")

        user_cb.bind("<<ComboboxSelected>>", refresh)
        refresh()

        def export_xl():
            if not XLSX:
                messagebox.showerror("خطا","openpyxl نصب نیست.",parent=self); return
            db = load_db()
            logs = list(reversed(db.get("login_log",[])))
            if not logs:
                messagebox.showinfo("خطا","هیچ لاگی ثبت نشده.",parent=self); return
            sel = user_cb.get()
            if sel != "همه":
                un = sel.split("[")[-1].rstrip("]").strip()
                logs = [l for l in logs if l.get("username")==un]
            rows = [(i, l.get("display","—"), l.get("username","—"),
                     l.get("role","—"),
                     *split_dt(l.get("at","—")),
                     *split_dt(l.get("logout_at","—") if l.get("logout_at") else "—"))
                    for i,l in enumerate(logs,1)]
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=f"login_log_{now_str().split()[0].replace('/','')}.xlsx",
                parent=self)
            if not path: return
            wb = openpyxl.Workbook(); ws = wb.active; ws.title="لاگ ورود"
            self._write_sheet(ws,
                ["ردیف","نام کاربر","نام کاربری","نقش","تاریخ ورود","ساعت ورود","تاریخ خروج","ساعت خروج"],
                rows, col_widths=[8,18,14,12,14,12,14,12])
            wb.save(path)
            messagebox.showinfo("موفق",f"Excel ذخیره شد:\n{path}",parent=self)

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=16, pady=4)
        styled_btn(ctrl,"📥  خروجی Excel",export_xl,color=C["btn_success"]).pack(side="right",padx=4)
        styled_btn(ctrl,"🔄 بروزرسانی",refresh,color=C["card"]).pack(side="right",padx=4)

    def _admin_movement_log(self, nb):
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="📜  تاریخچه جابجایی‌ها")
        tk.Label(tab, text="📜  تاریخچه کامل جابجایی اسلب‌ها بین انبارها",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=10)
        tk.Label(tab, text="تمام عملیات انتقال و برگشت اسلب‌ها — هیچ رکوردی حذف نمی‌شود.",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        ff = tk.Frame(tab, bg=C["panel"])
        ff.pack(fill="x", padx=16, pady=6)
        tk.Label(ff, text="فیلتر شماره اسلب:", bg=C["panel"],
                 fg=C["text"], font=FONT_NORM).pack(side="right", padx=(0,6))
        filter_var = tk.StringVar()
        tk.Entry(ff, textvariable=filter_var, bg=C["entry_bg"], fg=C["text"],
                 insertbackground=C["accent"], font=FONT_MONO, justify="right", bd=0, relief="flat",
                 highlightthickness=1, highlightbackground=C["border"],
                 highlightcolor=C["accent"], width=16).pack(side="right", padx=4)

        cols = ("slab_id","operation","from_loc","to_loc","reason","by","log_date","log_time")
        heads = ("شماره اسلب","نوع عملیات","از انبار","به انبار","دلیل","کاربر","تاریخ","ساعت")
        tf, tree = scrolled_tree(tab, cols, heads, height=16)
        tf.pack(fill="both", expand=True, padx=16, pady=4)
        for col in cols:
            tree.column(col, width=120, anchor="center")
        tree.column("slab_id", width=140)
        tree.column("reason", width=180)

        sb = search_bar(tab, tree)
        sb.pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        stats_lbl = tk.Label(tab, text="", bg=C["panel"], fg=C["warning"], font=FONT_SMALL)
        stats_lbl.pack(anchor="e", padx=16)

        def refresh(*_):
            tree.delete(*tree.get_children())
            db = load_db()
            logs = db.get("movement_log", [])
            filt = filter_var.get().strip()
            count_transfer = 0; count_return = 0
            for rec in reversed(logs):
                sid = rec.get("slab_id","—")
                if filt and filt not in sid: continue
                op = rec.get("operation","—")
                if "انتقال" in op: count_transfer += 1
                elif "برگشت" in op: count_return += 1
                clr = "transfer" if "انتقال" in op else "return_"
                tree.insert("", "end", values=(
                    sid, op, rec.get("from","—"), rec.get("to","—"),
                    rec.get("reason","—"), rec.get("by","—"), *split_dt(rec.get("at","—"))
                ), tags=(clr,))
            tree.tag_configure("transfer", background="#001a30", foreground=C["accent"])
            tree.tag_configure("return_",  background="#1a1000", foreground=C["warning"])
            total = count_transfer + count_return
            stats_lbl.config(text=f"📊  کل: {total}  |  انتقال: {count_transfer}  |  برگشت: {count_return}")

        filter_var.trace_add("write", refresh)
        refresh()

        ctrl = tk.Frame(tab, bg=C["panel"])
        ctrl.pack(fill="x", padx=16, pady=4)

        def export_log():
            if not XLSX:
                messagebox.showerror("خطا", "openpyxl نصب نیست.", parent=self)
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
                initialfile=f"movement_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", parent=self)
            if not path: return
            db = load_db()
            logs = db.get("movement_log", [])
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "تاریخچه جابجایی"; ws.sheet_view.rightToLeft = True
            headers = ["شماره اسلب","نوع عملیات","از انبار","به انبار","دلیل","کاربر","تاریخ","ساعت"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = Font(name="B Nazanin", bold=True, color="00D4FF")
                cell.fill = PatternFill("solid", fgColor="0D1520")
                cell.alignment = Alignment(horizontal="center")
            for r, rec in enumerate(reversed(logs), 2):
                ws.cell(r,1,rec.get("slab_id",""))
                ws.cell(r,2,rec.get("operation",""))
                ws.cell(r,3,rec.get("from",""))
                ws.cell(r,4,rec.get("to",""))
                ws.cell(r,5,rec.get("reason",""))
                ws.cell(r,6,rec.get("by",""))
                _d, _t = split_dt(rec.get("at",""))
                ws.cell(r,7,_d)
                ws.cell(r,8,_t)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22
            wb.save(path)
            messagebox.showinfo("موفق", f"فایل Excel ذخیره شد:\n{path}", parent=self)

        styled_btn(ctrl, "📥  خروجی Excel", export_log, color=C["btn_success"]).pack(side="right", padx=4)

    def _admin_return_log(self, nb):
        """تاریخچه کامل برگشت‌ها برای ادمین"""
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="↩️  تاریخچه برگشت‌ها")
        tk.Label(tab, text="↩️  تاریخچه کامل برگشت اسلب‌ها به انبار داخلی",
                 bg=C["panel"], fg=C["warning"], font=FONT_HEAD).pack(anchor="e", padx=16, pady=10)
        tk.Label(tab, text="این اطلاعات دائمی هستند و هیچ‌گاه حذف نمی‌شوند.",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=16)

        cols = ("slab_id","label","source","reason","returned_by","ret_date","ret_time")
        heads = ("شماره اسلب","نوع برگشت","از انبار","دلیل برگشت","ثبت‌کننده","تاریخ برگشت","ساعت برگشت")
        tf, tree = scrolled_tree(tab, cols, heads, height=16)
        tf.pack(fill="both", expand=True, padx=16, pady=8)
        tree.column("label", width=180, anchor="center")
        tree.column("reason", width=220, anchor="center")

        sb = search_bar(tab, tree, col_indices=[0])
        sb.pack(anchor="e", padx=16, pady=4)
        sort_toolbar(tab, tree, bg=C["panel"]).pack(anchor="e", padx=16, pady=2)

        stats = tk.Label(tab, text="", bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL)
        stats.pack(anchor="e", padx=16)

        def refresh():
            tree.delete(*tree.get_children())
            db = load_db()
            logs = db.get("return_log", [])
            for rec in reversed(logs):
                ret_num = rec.get("return_number", 1)
                label = rec.get("label", f"برگشت {return_ordinal_fa(ret_num)}")
                clr = "first_ret" if ret_num==1 else ("second_ret" if ret_num==2 else "multi_ret")
                tree.insert("", "end", values=(
                    rec.get("slab_id","—"), label, rec.get("source","—"),
                    rec.get("reason","—"), get_display_name(rec.get("returned_by","—"), db),
                    *split_dt(rec.get("returned_at","—"))
                ), tags=(clr,))
            tree.tag_configure("first_ret",  background="#1a1400", foreground=C["warning"])
            tree.tag_configure("second_ret", background="#1a0a00", foreground="#ff8c00")
            tree.tag_configure("multi_ret",  background="#200000", foreground=C["danger"])
            total = len(logs)
            unique = len(set(r.get("slab_id") for r in logs))
            stats.config(text=f"📊  کل برگشت‌ها: {total}  |  اسلب‌های منحصربه‌فرد: {unique}")

        refresh()

    def _admin_settings_inline(self, tab):
        """تنظیمات تم — مستقیم درون frame داده‌شده (بدون nb.add)"""
        tab.configure(bg=C["panel"])

        # ── اسکرول‌پذیر ──
        canvas = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb = tk.Scrollbar(tab, orient="vertical", command=canvas.yview,
                           bg="#707070", troughcolor="#1a1a1a", activebackground=C["accent"], width=16)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="left", fill="y")
        canvas.pack(fill="both", expand=True)
        inner_frame = tk.Frame(canvas, bg=C["panel"])
        win_id = canvas.create_window((0, 0), window=inner_frame, anchor="nw")

        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_configure(e):
            canvas.itemconfig(win_id, width=e.width)
        inner_frame.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        register_scroll_canvas(canvas, inner_frame)

        # ── تم‌های آماده ──
        PRESETS = {
            "🌑  پیش‌فرض (فولادی)": {
                "bg":"#3a3f45","panel":"#40464d","card":"#484e56","card2":"#444a52",
                "card_hover":"#525860","accent":"#d4a043","accent2":"#b8882e",
                "accent_glow":"#f0c060","text":"#f0f2f4","text_dim":"#a0a8b0",
                "text_bright":"#ffffff","entry_bg":"#353b41","header_bg":"#2e3338",
                "border":"#555c64","btn_primary":"#2a6090","btn_success":"#2a7850",
                "btn_danger":"#a03030",
            },
            "🌊  آبی نیلی": {
                "bg":"#0d1b2a","panel":"#112233","card":"#1a3050","card2":"#162840",
                "card_hover":"#1e3860","accent":"#4db8ff","accent2":"#2a90d0",
                "accent_glow":"#80d4ff","text":"#e0f0ff","text_dim":"#7090b0",
                "text_bright":"#ffffff","entry_bg":"#0d2040","header_bg":"#091525",
                "border":"#204060","btn_primary":"#1a5080","btn_success":"#1a6840",
                "btn_danger":"#882020",
            },
            "🌿  سبز جنگلی": {
                "bg":"#0f1f15","panel":"#152a1c","card":"#1c3825","card2":"#182f20",
                "card_hover":"#224530","accent":"#4caf70","accent2":"#3a8a55",
                "accent_glow":"#70d090","text":"#d0f0d8","text_dim":"#70a880",
                "text_bright":"#ffffff","entry_bg":"#0f2018","header_bg":"#0a1810",
                "border":"#2a5035","btn_primary":"#1a6040","btn_success":"#1a7030",
                "btn_danger":"#882020",
            },
            "🔴  قرمز تیره": {
                "bg":"#1f0f0f","panel":"#2a1515","card":"#381c1c","card2":"#301818",
                "card_hover":"#452222","accent":"#e05050","accent2":"#b83030",
                "accent_glow":"#f07070","text":"#f0d8d8","text_dim":"#a07070",
                "text_bright":"#ffffff","entry_bg":"#200f0f","header_bg":"#180a0a",
                "border":"#503030","btn_primary":"#6a2020","btn_success":"#2a6030",
                "btn_danger":"#a01010",
            },
            "🌸  بنفش شاهانه": {
                "bg":"#1a0f2e","panel":"#221440","card":"#2e1c58","card2":"#281850",
                "card_hover":"#382068","accent":"#c070ff","accent2":"#9040d0",
                "accent_glow":"#d898ff","text":"#ecdeff","text_dim":"#9070b8",
                "text_bright":"#ffffff","entry_bg":"#180d28","header_bg":"#120a20",
                "border":"#402860","btn_primary":"#502880","btn_success":"#286040",
                "btn_danger":"#882020",
            },
            "🌞  روشن (روز)": {
                "bg":"#e8edf2","panel":"#dde3ea","card":"#f0f4f8","card2":"#e8edf3",
                "card_hover":"#cdd5de","accent":"#2060a0","accent2":"#174080",
                "accent_glow":"#3080c0","text":"#1a2535","text_dim":"#5a6878",
                "text_bright":"#ffffff","entry_bg":"#ffffff","header_bg":"#c8d4e0",
                "border":"#b0bcc8","btn_primary":"#2060a0","btn_success":"#206840",
                "btn_danger":"#a02020",
            },
        }

        # ── آیکون‌های تب ──
        TAB_ICON_OPTIONS = {
            "home":     ["🏠","🏡","📊","🏗","🔷"],
            "melts":    ["🔥","⚡","🌡","🔶","➕"],
            "qc":       ["✅","☑","🔍","🏷","💎"],
            "rejected": ["⛔","❌","🚫","⚠️","🔴"],
            "transfer": ["🏭","📦","🚚","🗄","🏪"],
            "lab":      ["🧪","🔬","⚗","🧬","🏥"],
            "scrap":    ["♻️","🗑","♺","🔄","💀"],
            "pdf":      ["📄","📋","🖨","📑","📃"],
            "scarf":    ["⚙","🔧","🔩","⚒","🛠"],
            "cut":      ["✂","🔪","⚔","🪚","🗡"],
            "nobat":    ["🔄","🚪","🏃","📤","↗"],
            "admin":    ["👑","🛡","⚙️","🔐","🌟"],
        }
        TAB_LABELS = {
            "home":"نمای کلی","melts":"ثبت ذوب جدید","qc":"کنترل کیفی",
            "rejected":"تایید نشده","transfer":"موجودی انبار",
            "lab":"تحویل باومن","scrap":"قراضه","pdf":"گزارش PDF",
            "scarf":"اسکارف","cut":"برش","nobat":"خروج اسلب","admin":"مدیریت سیستم",
        }

        # بارگذاری تنظیمات ذخیره شده
        db_now = load_db()
        saved = db_now.get("settings", {})
        saved_icons = saved.get("tab_icons", {})

        # ────────────────────────────────────────
        #  بخش ۱: تم‌های آماده
        # ────────────────────────────────────────
        sec1 = card_frame(inner_frame)
        sec1.pack(fill="x", padx=16, pady=(16,6))
        s1in = tk.Frame(sec1, bg=C["card"])
        s1in.pack(padx=14, pady=12, fill="x")
        tk.Label(s1in, text="🎨  تم‌های آماده",
                 bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0,8))
        tk.Label(s1in, text="یک تم را انتخاب کنید — بلافاصله پیش‌نمایش نشان داده می‌شود",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(0,10))

        preset_var = tk.StringVar()
        preset_frame = tk.Frame(s1in, bg=C["card"])
        preset_frame.pack(fill="x")

        preview_swatch = tk.Frame(s1in, bg=C["card"], height=40)
        preview_swatch.pack(fill="x", pady=(8,4))

        def apply_preset_preview(name):
            p = PRESETS[name]
            for w in preview_swatch.winfo_children():
                w.destroy()
            colors = [p["accent"], p["bg"], p["panel"], p["card"],
                      p["entry_bg"], p["header_bg"], p["btn_success"], p["btn_danger"]]
            labels = ["تأکید","پس‌زمینه","پنل","کارت","ورودی","هدر","موفق","خطر"]
            for c, lbl in zip(colors, labels):
                f = tk.Frame(preview_swatch, bg=c, width=50, height=40)
                f.pack(side="right", padx=2)
                tk.Label(f, text=lbl, bg=c, fg=p["text"],
                         font=(_MAIN_FONT, 8, "bold")).pack(expand=True)

        def on_preset_select(name):
            preset_var.set(name)
            apply_preset_preview(name)
            for btn in preset_btns:
                n = btn.cget("text")
                btn.config(bg=C["accent"] if n == name else C["card2"],
                           fg="#ffffff" if n == name else C["text"])

        preset_btns = []
        row_f = None
        for i, name in enumerate(PRESETS):
            if i % 3 == 0:
                row_f = tk.Frame(preset_frame, bg=C["card"])
                row_f.pack(fill="x", pady=3)
            btn = tk.Button(row_f, text=name,
                            bg=C["card2"], fg=C["text"],
                            font=(_MAIN_FONT, 10, "bold"), bd=0, relief="flat",
                            cursor="hand2", padx=10, pady=8,
                            command=lambda n=name: on_preset_select(n))
            btn.pack(side="right", padx=4)
            preset_btns.append(btn)

        def apply_preset_full():
            name = preset_var.get()
            if not name:
                messagebox.showwarning("انتخاب نشده","ابتدا یک تم انتخاب کنید.", parent=tab); return
            p = PRESETS[name]
            C.update(p)
            _apply_theme_to_all()
            db2 = load_db()
            db2.setdefault("settings", {})["theme"] = p
            save_db(db2)
            messagebox.showinfo("✅  تم اعمال شد",
                f"تم «{name}» با موفقیت اعمال شد.\nبرای نمایش کامل، برنامه را مجدداً اجرا کنید.",
                parent=tab)

        styled_btn(s1in, "✔  اعمال تم انتخابی", apply_preset_full,
                   color=C["accent2"]).pack(anchor="e", pady=8)

        # ────────────────────────────────────────
        #  بخش ۲: ویرایش دستی رنگ‌ها
        # ────────────────────────────────────────
        sec2 = card_frame(inner_frame)
        sec2.pack(fill="x", padx=16, pady=6)
        s2in = tk.Frame(sec2, bg=C["card"])
        s2in.pack(padx=14, pady=12, fill="x")
        tk.Label(s2in, text="🖌  ویرایش دستی رنگ‌ها",
                 bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0,4))
        tk.Label(s2in, text="کد رنگ HEX وارد کنید — پیش‌نمایش فوری نمایش داده می‌شود",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(0,10))

        COLOR_FIELDS = [
            ("bg",          "پس‌زمینه اصلی"),
            ("panel",       "پنل/نوار کناری"),
            ("card",        "کارت‌ها"),
            ("card2",       "کارت ثانوی"),
            ("accent",      "رنگ تأکیدی (طلایی)"),
            ("accent2",     "رنگ تأکیدی ثانوی"),
            ("text",        "متن اصلی"),
            ("text_dim",    "متن کم‌رنگ"),
            ("text_bright", "متن درخشان"),
            ("entry_bg",    "پس‌زمینه ورودی"),
            ("header_bg",   "پس‌زمینه هدر"),
            ("border",      "حاشیه"),
            ("btn_primary", "دکمه اصلی"),
            ("btn_success", "دکمه موفقیت"),
            ("btn_danger",  "دکمه خطر"),
        ]

        color_vars = {}
        grid_f = tk.Frame(s2in, bg=C["card"])
        grid_f.pack(fill="x")

        def make_color_row(parent, key, label):
            row = tk.Frame(parent, bg=C["card"])
            row.pack(fill="x", pady=3)
            tk.Label(row, text=f"{label}:", bg=C["card"], fg=C["text_dim"],
                     font=FONT_SMALL, width=18, anchor="e").pack(side="right")
            var = tk.StringVar(value=C.get(key, "#ffffff"))
            color_vars[key] = var
            # swatch رنگی (بزرگ‌تر — کد HEX مخفی)
            swatch = tk.Label(row, text="        ", bg=C.get(key, "#ffffff"), width=6,
                              height=1, relief="flat",
                              highlightthickness=1, highlightbackground=C["border"],
                              cursor="hand2")
            swatch.pack(side="right", padx=6)
            # ورودی مخفی (برای نگه‌داری مقدار — نمایش داده نمی‌شود)
            ent = tk.Entry(row, textvariable=var, width=0)
            # دکمه انتخاب رنگ گرافیکی
            def open_picker(v=var, s=swatch):
                current = v.get()
                try:
                    result = colorchooser.askcolor(color=current, title="انتخاب رنگ")
                    if result and result[1]:
                        v.set(result[1])
                        s.config(bg=result[1])
                except Exception:
                    pass
            tk.Button(row, text="🎨  انتخاب", bg=C["card2"], fg=C["text"],
                      font=FONT_SMALL,
                      bd=0, relief="flat", cursor="hand2", padx=8, pady=3,
                      command=open_picker).pack(side="right", padx=(0,4))
            def upd(*_, s=swatch, v=var):
                try: s.config(bg=v.get())
                except: pass
            var.trace_add("write", upd)
            return var

        for key, lbl in COLOR_FIELDS:
            make_color_row(grid_f, key, lbl)

        def apply_custom_colors():
            changed = {}
            errors = []
            for key, var in color_vars.items():
                val = var.get().strip()
                if not val.startswith("#") or len(val) not in (4, 7):
                    errors.append(f"{key}: {val}")
                    continue
                changed[key] = val
            if errors:
                messagebox.showwarning("رنگ نامعتبر",
                    "رنگ‌های زیر نامعتبرند (باید #RGB یا #RRGGBB):\n" + "\n".join(errors),
                    parent=tab)
                return
            C.update(changed)
            _apply_theme_to_all()
            db2 = load_db()
            db2.setdefault("settings", {})["theme"] = dict(changed)
            save_db(db2)
            messagebox.showinfo("✅  رنگ‌ها اعمال شدند",
                "رنگ‌های سفارشی ذخیره شدند.\nبرای نمایش کامل، برنامه را مجدداً اجرا کنید.",
                parent=tab)

        styled_btn(s2in, "🎨  اعمال رنگ‌های سفارشی", apply_custom_colors,
                   color=C["accent2"]).pack(anchor="e", pady=8)

        # ────────────────────────────────────────
        #  بخش ۳: پس‌زمینه صفحه ورود
        # ────────────────────────────────────────
        sec_bg = card_frame(inner_frame)
        sec_bg.pack(fill="x", padx=16, pady=6)
        s_bg = tk.Frame(sec_bg, bg=C["card"])
        s_bg.pack(padx=14, pady=12, fill="x")
        tk.Label(s_bg, text="🖼  پس‌زمینه صفحه ورود",
                 bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0,4))
        tk.Label(s_bg, text="تصویر دلخواه برای پس‌زمینه صفحه لاگین انتخاب کنید (JPG / PNG)",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", pady=(0,10))

        # نمایش مسیر فعلی
        db_now2 = load_db()
        cur_img = db_now2.get("settings", {}).get("background_image", "")
        img_path_var = tk.StringVar(value=cur_img if cur_img else "انتخاب نشده")

        path_row = tk.Frame(s_bg, bg=C["card"])
        path_row.pack(fill="x", pady=4)

        # پیش‌نمایش کوچک
        preview_lbl = tk.Label(s_bg, bg=C["card2"], width=30, height=6,
                                text="پیش‌نمایش" if not cur_img else "",
                                fg=C["text_dim"], font=FONT_SMALL,
                                relief="flat", highlightthickness=1,
                                highlightbackground=C["border"])
        preview_lbl.pack(anchor="e", pady=(4,8))
        self._login_bg_preview = preview_lbl
        self._login_bg_img_ref = None

        def _show_preview(path):
            try:
                from PIL import Image, ImageTk
                img = Image.open(path)
                img.thumbnail((240, 120))
                ph = ImageTk.PhotoImage(img)
                preview_lbl.config(image=ph, text="")
                self._login_bg_img_ref = ph
            except Exception:
                preview_lbl.config(text="پیش‌نمایش در دسترس نیست", image="")

        if cur_img and os.path.exists(cur_img):
            _show_preview(cur_img)

        path_ent = tk.Entry(path_row, textvariable=img_path_var,
                            bg=C["entry_bg"], fg=C["text"],
                            insertbackground=C["accent"], font=FONT_SMALL,
                            justify="right", bd=0, relief="flat",
                            highlightthickness=1, highlightbackground=C["border"],
                            highlightcolor=C["accent"], state="readonly", width=36)
        path_ent.pack(side="right", padx=(0,6))

        def pick_image():
            path = filedialog.askopenfilename(
                title="انتخاب تصویر پس‌زمینه",
                filetypes=[("تصویر", "*.jpg *.jpeg *.png *.bmp *.gif"), ("همه فایل‌ها", "*.*")]
            )
            if not path:
                return
            img_path_var.set(path)
            _show_preview(path)

        def clear_image():
            img_path_var.set("انتخاب نشده")
            preview_lbl.config(image="", text="پیش‌نمایش")
            self._login_bg_img_ref = None

        def save_login_bg():
            path = img_path_var.get().strip()
            db2 = load_db()
            db2.setdefault("settings", {})
            if path == "انتخاب نشده" or not path:
                db2["settings"]["background_image"] = ""
                db2["settings"]["background_image_sha256"] = ""
                db2["settings"]["background_image_ext"] = ""
                _clear_local_background_files()
                try:
                    _delete_background_on_server()
                except Exception as ex:
                    # روی ادمین اگر API هنوز بالا نیامده، فقط محلی پاک می‌شود
                    pass
            else:
                if not os.path.exists(path):
                    messagebox.showwarning("فایل یافت نشد", "مسیر تصویر معتبر نیست.", parent=tab)
                    return
                try:
                    os.makedirs(ASSETS_DIR, exist_ok=True)
                    ext = os.path.splitext(path)[1].lower() or ".jpg"
                    _clear_local_background_files()
                    dest = os.path.join(ASSETS_DIR, "login_background" + ext)
                    if os.path.abspath(path) != os.path.abspath(dest):
                        shutil.copyfile(path, dest)
                    path = dest
                    sha = _sha256_file(path)
                    # انتشار روی سرور تا همه کلاینت‌ها همان تصویر را بگیرند
                    try:
                        meta = _publish_background_to_server(path)
                        sha = meta.get("sha256") or sha
                        ext = meta.get("ext") or ext
                    except Exception as ex:
                        # اگر API در دسترس نبود، حداقل فایل محلی و متادیتا ذخیره می‌شود
                        # (روی خود ادمین فایل در app_assets هست و API بعداً سرو می‌کند)
                        pass
                    db2["settings"]["background_image"] = f"login_background{ext}"
                    db2["settings"]["background_image_sha256"] = sha
                    db2["settings"]["background_image_ext"] = ext
                except Exception as ex:
                    messagebox.showwarning("خطا در کپی تصویر",
                        f"تصویر در پوشه‌ی برنامه ذخیره نشد:\n{ex}",
                        parent=tab)
                    return
            save_db(db2)
            # اعمال فوری روی LoginWindow اگر زنده باشد
            try:
                global _ACTIVE_LOGIN_WIN
                lw = _ACTIVE_LOGIN_WIN
                if lw is not None and lw.winfo_exists():
                    local_path = ensure_login_background_local(db2)
                    if local_path and os.path.exists(local_path):
                        lw.db = db2
                        lw._refresh_login_background(local_path)
                    else:
                        if getattr(lw, "_login_bg_lbl", None):
                            lw._login_bg_lbl.place_forget()
                        for w in (getattr(lw, "_scroll_container", None),
                                  getattr(lw, "_login_canvas", None),
                                  getattr(lw, "_users_frame", None)):
                            if w is not None and getattr(w, "_bg_layer_lbl", None):
                                w._bg_layer_lbl.place_forget()
            except Exception:
                pass
            messagebox.showinfo("✅  ذخیره شد",
                "تصویر پس‌زمینه ذخیره شد و برای همه کلاینت‌ها همگام می‌شود.\nدر صفحه ورود اعمال می‌شود.", parent=tab)

        btn_row = tk.Frame(s_bg, bg=C["card"])
        btn_row.pack(fill="x", pady=6)
        styled_btn(btn_row, "🗂  انتخاب تصویر", pick_image, color=C["btn_primary"]).pack(side="right", padx=4)
        styled_btn(btn_row, "🗑  حذف پس‌زمینه", clear_image, color=C["btn_danger"]).pack(side="right", padx=4)
        styled_btn(btn_row, "💾  ذخیره", save_login_bg, color=C["btn_success"]).pack(side="right", padx=4)

        # ────────────────────────────────────────
        #  بخش ۴: بایگانی دادهٔ حجیم
        # ────────────────────────────────────────
        sec_arc = card_frame(inner_frame)
        sec_arc.pack(fill="x", padx=16, pady=6)
        s_arc = tk.Frame(sec_arc, bg=C["card"])
        s_arc.pack(padx=14, pady=12, fill="x")
        tk.Label(s_arc, text="🗄  بایگانی داده‌های قدیمی (مقیاس‌پذیری)",
                 bg=C["card"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", pady=(0,4))
        tk.Label(s_arc,
                 text="داده‌های قدیمی‌تر از چند ماه اخیر به پوشه آرشیو منتقل می‌شوند تا دیتابیس سبک بماند و هنگ نکند",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL, wraplength=640, justify="right").pack(anchor="e", pady=(0,8))
        arc_info = tk.Label(s_arc, text="", bg=C["card"], fg=C["text"], font=FONT_SMALL, justify="right")
        arc_info.pack(anchor="e", pady=4)

        def refresh_arc_info():
            try:
                _ensure_stf_shared_on_path()
                from shared.archive_service import hot_db_stats, list_archives, HOT_KEEP_MONTHS
                dbx = load_db()
                st = hot_db_stats(dbx) or {}
                arcs = list_archives()
                meta = (dbx.get("settings") or {}).get("archive_meta") or {}
                total_rows = int(st.get("total_rows") or st.get("total") or 0)
                last_run = meta.get("last_run") or "—"
                arc_info.config(
                    text=(
                        f"ردیف‌های فعال: {total_rows:,}  |  فایل آرشیو: {len(arcs)}  |  "
                        f"نگهداری داغ: {HOT_KEEP_MONTHS} ماه اخیر\n"
                        f"آخرین بایگانی: {last_run}"
                    )
                )
            except Exception:
                arc_info.config(text="وضعیت آرشیو در دسترس نیست")

        def do_archive_now():
            if not messagebox.askyesno(
                    "بایگانی",
                    "داده‌های قدیمی از دیتابیس فعال به آرشیو منتقل شوند؟\n"
                    "داده حذف نمی‌شود — فقط از حافظهٔ فعال خارج می‌شود.",
                    parent=tab):
                return
            try:
                _ensure_stf_shared_on_path()
                from shared.archive_service import archive_cold_data
                db2 = load_db()
                stats = archive_cold_data(db2)
                save_db(db2)
                # اگر API در دسترس است، روی سرور هم اجرا شود
                try:
                    import urllib.request
                    base = _resolve_server_base_url()
                    req = urllib.request.Request(
                        f"{base}/api/v1/archives/run", data=b"{}", method="POST",
                        headers={"Content-Type": "application/json"})
                    urllib.request.urlopen(req, timeout=120).read()
                except Exception:
                    pass
                messagebox.showinfo(
                    "بایگانی انجام شد",
                    f"ردیف آرشیو شده: {stats.get('archived_rows', 0)}\n"
                    f"ردیف باقی‌ماندهٔ فعال: {stats.get('kept_rows', 0)}\n"
                    f"دوره‌ها: {', '.join(stats.get('periods', {}).keys()) or '—'}",
                    parent=tab)
                refresh_arc_info()
            except Exception as ex:
                messagebox.showerror("خطا", str(ex), parent=tab)

        styled_btn(s_arc, "🗄  بایگانی اکنون", do_archive_now,
                   color=C["warning"]).pack(side="right", padx=4, pady=8)
        styled_btn(s_arc, "🔃  بروزرسانی وضعیت", refresh_arc_info,
                   color=C["btn_ghost"]).pack(side="right", padx=4, pady=8)
        refresh_arc_info()

        # ── بارگذاری تم ذخیره شده برای پر کردن فیلدها ──
        saved_theme = saved.get("theme", {})
        if saved_theme:
            for key, var in color_vars.items():
                if key in saved_theme:
                    var.set(saved_theme[key])


    def _admin_settings(self, nb):
        tab = tk.Frame(nb, bg=C["panel"])
        nb.add(tab, text="🎨  تنظیمات تم")
        self._admin_settings_inline(tab)

    def _apply_theme_to_all(self):
        """اعمال رنگ‌های C به تمام ویجت‌های فعال"""

        # ── آپدیت ttk style ──
        style = ttk.Style()
        try:
            style.configure("Dark.Treeview",
                background=C["card2"], foreground=C["text"],
                fieldbackground=C["card2"], bordercolor=C["border"])
            style.configure("Dark.Treeview.Heading",
                background=C["header_bg"], foreground=C["accent"])
            style.map("Dark.Treeview",
                background=[("selected", C["accent"])],
                foreground=[("selected", C["header_bg"])])
            style.configure("Dark.TNotebook", background=C["bg"], bordercolor=C["border"])
            style.configure("Dark.TNotebook.Tab",
                background=C["tab_inactive"], foreground=C["text"])
            style.map("Dark.TNotebook.Tab",
                background=[("selected", C["accent"])],
                foreground=[("selected", "#ffffff")])
            style.configure("Sub.TNotebook", background=C["bg"], bordercolor=C["border"])
            style.configure("Sub.TNotebook.Tab",
                background=C["tab_inactive"], foreground=C["text"])
            style.map("Sub.TNotebook.Tab",
                background=[("selected", C["accent"])],
                foreground=[("selected", "#ffffff")])
        except Exception:
            pass

        # ── traverse کل ویجت‌ها ──
        def _recurse(widget):
            wtype = widget.winfo_class()
            try:
                if wtype == "Frame":
                    widget.configure(bg=C["panel"])
                elif wtype == "Label":
                    cur_bg = widget.cget("bg")
                    # هدر و کارت رو حفظ کن
                    widget.configure(fg=C["text"])
                    widget.configure(bg=C["panel"])
                elif wtype == "Canvas":
                    widget.configure(bg=C["bg"])
                elif wtype == "Button":
                    widget.configure(bg=C["card"], fg=C["text"],
                                     activebackground=C["card_hover"],
                                     activeforeground=C["text"])
                elif wtype == "Entry":
                    widget.configure(bg=C["entry_bg"], fg=C["text"],
                                     insertbackground=C["accent"],
                                     highlightbackground=C["border"],
                                     highlightcolor=C["accent"])
                elif wtype == "Scrollbar":
                    widget.configure(bg="#707070", troughcolor="#1a1a1a",
                                     activebackground=C["accent"])
            except Exception:
                pass
            for child in widget.winfo_children():
                _recurse(child)

        _recurse(self)

        # ── sidebar و tab buttons دقیق‌تر ──
        try:
            self._sidebar.configure(bg=C["panel"])
            for key, btn in self._tab_buttons.items():
                if key == self._current_tab:
                    btn.configure(bg=C["accent"], fg="#ffffff",
                                  activebackground=C["accent2"])
                else:
                    btn.configure(bg=C["panel"], fg=C["text"],
                                  activebackground=C["card_hover"])
        except Exception:
            pass

        # ── header ──
        try:
            self._header_frame.configure(bg=C["header_bg"])
        except Exception:
            pass

        # ── force redraw ──
        try:
            self.update_idletasks()
        except Exception:
            pass

    # ═══════════════════════════════════════════════════════════
    #  تب گزارش گیری (PDF)
    # ═══════════════════════════════════════════════════════════
    def _build_pdf_tab(self, tab):
        """گزارش‌گیری PDF — دو گزارش: QC شده + تحویلی به آزمایشگاه"""
        tab.configure(bg=C["panel"])

        # ── اسکرول ──
        canvas_outer = tk.Canvas(tab, bg=C["panel"], highlightthickness=0)
        vsb_outer = tk.Scrollbar(tab, orient="vertical", command=canvas_outer.yview,
                                  bg="#707070", troughcolor="#1a1a1a",
                                  activebackground=C["accent"], width=16)
        canvas_outer.configure(yscrollcommand=vsb_outer.set)
        vsb_outer.pack(side="right", fill="y")
        canvas_outer.pack(fill="both", expand=True)
        sf = tk.Frame(canvas_outer, bg=C["panel"])
        _win_outer = canvas_outer.create_window((0, 0), window=sf, anchor="nw")
        sf.bind("<Configure>", lambda e: canvas_outer.configure(scrollregion=canvas_outer.bbox("all")))
        canvas_outer.bind("<Configure>", lambda e: canvas_outer.itemconfig(_win_outer, width=e.width))
        register_scroll_canvas(canvas_outer, sf)

        # ── عنوان ──
        tk.Label(sf, text="📊  گزارش‌گیری",
                 bg=C["panel"], fg=C["accent"], font=FONT_HEAD).pack(anchor="e", padx=20, pady=(14,4))
        tk.Label(sf, text="برای کاربران شیفت — گزارش PDF اسلب‌های کنترل کیفی شده و تحویلی به آزمایشگاه",
                 bg=C["panel"], fg=C["text_dim"], font=FONT_SMALL).pack(anchor="e", padx=20)

        separator(sf)

        # ── فیلتر بازه زمانی ──
        dt_card = card_frame(sf)
        dt_card.pack(fill="x", padx=20, pady=8)
        dt_in = tk.Frame(dt_card, bg=C["card"])
        dt_in.pack(padx=16, pady=12, fill="x")
        tk.Label(dt_in, text="📅  بازه تاریخ و ساعت گزارش‌گیری",
                 bg=C["card"], fg=C["accent"], font=(_MAIN_FONT,10,"bold")).pack(anchor="e", pady=(0,8))

        _now_sh  = to_shamsi(datetime.datetime.now())
        _now_date = _now_sh.split("  ")[0]

        def _mk_entry(parent, var, w=13):
            return tk.Entry(parent, textvariable=var,
                     bg=C["entry_bg"], fg=C["text"],
                     insertbackground=C["accent"], font=FONT_MONO,
                     justify="right", bd=0, relief="flat",
                     highlightthickness=1, highlightbackground=C["border"],
                     highlightcolor=C["accent"], width=w)

        # ── ردیف تاریخ ──
        row_dates = tk.Frame(dt_in, bg=C["card"])
        row_dates.pack(fill="x", pady=3)

        def _mk_pdf_pair(parent, lbl, var, w):
            f = tk.Frame(parent, bg=C["card"])
            f.pack(side="right", padx=(0, 10))
            tk.Label(f, text=lbl, bg=C["card"], fg=C["text_dim"],
                     font=FONT_NORM).pack(side="right")
            _mk_entry(f, var, w).pack(side="right", padx=(4, 0))

        from_date_v = tk.StringVar(value=get_first_report_date_sh())
        to_date_v   = tk.StringVar(value=_now_date)
        _mk_pdf_pair(row_dates, "از تاریخ", from_date_v, 13)
        _mk_pdf_pair(row_dates, "تا تاریخ", to_date_v,   13)

        # ── ردیف ساعت ──
        row_times = tk.Frame(dt_in, bg=C["card"])
        row_times.pack(fill="x", pady=3)

        from_time_v = tk.StringVar(value="00:00:00")
        to_time_v   = tk.StringVar(value="23:59:59")
        _mk_pdf_pair(row_times, "از ساعت", from_time_v, 9)
        _mk_pdf_pair(row_times, "تا ساعت", to_time_v,   9)

        from_var = tk.StringVar()
        to_var   = tk.StringVar()
        def _sync_dt(*_):
            from_var.set(f"{from_date_v.get().strip()}  {from_time_v.get().strip()}")
            to_var.set(f"{to_date_v.get().strip()}  {to_time_v.get().strip()}")
        for v in (from_date_v, from_time_v, to_date_v, to_time_v):
            v.trace_add("write", _sync_dt)
        _sync_dt()

        # ── درخواست‌دهنده ──
        row_req = tk.Frame(dt_in, bg=C["card"])
        row_req.pack(fill="x", pady=6)
        tk.Label(row_req, text="نام درخواست‌دهنده:", bg=C["card"],
                 fg=C["text"], font=FONT_NORM, width=18, anchor="e").pack(side="right")
        req_name_var = tk.StringVar(value=self.udata.get("display",""))
        _mk_entry(row_req, req_name_var, 28).pack(side="right", padx=6)

        # ── وضعیت/پیش‌نمایش ──
        preview_lbl = tk.Label(sf, text="", bg=C["panel"], fg=C["warning"], font=FONT_NORM)
        preview_lbl.pack(anchor="e", padx=20, pady=4)

        def parse_sh(s):
            s = s.strip().replace("  "," ")
            m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{2}):(\d{2})", s)
            if not m: raise ValueError(f"فرمت اشتباه: {s}")
            jy,jm,jd,hh,mm,ss = [int(x) for x in m.groups()]
            return f"{jy:04d}/{jm:02d}/{jd:02d}  {hh:02d}:{mm:02d}:{ss:02d}"

        def get_in_range(at_field_val):
            try:
                from_sh = parse_sh(from_var.get())
                to_sh   = parse_sh(to_var.get())
                return from_sh <= at_field_val <= to_sh
            except:
                return True

        def update_preview(*_):
            try:
                from_sh = parse_sh(from_var.get())
                to_sh   = parse_sh(to_var.get())
                db = load_db()
                n_qc  = sum(1 for r in db["melts"]
                            if r.get("qc_status")=="کنترل کیفی شده"
                            and r.get("qc_at","")
                            and from_sh <= r.get("qc_at","") <= to_sh)
                n_lab = sum(1 for r in db.get("lab_deliveries",[])
                            if r.get("delivered_at","")
                            and from_sh <= r.get("delivered_at","") <= to_sh)
                preview_lbl.config(text=f"🔢  اسلب QC شده: {n_qc}  |  تحویلی آزمایشگاه: {n_lab}")
            except:
                preview_lbl.config(text="")
        from_var.trace_add("write", update_preview)
        to_var.trace_add("write", update_preview)
        update_preview()

        separator(sf)

        # ══════════════════════════════════════════════
        #  دو باکس کنار هم
        # ══════════════════════════════════════════════
        boxes_frame = tk.Frame(sf, bg=C["panel"])
        boxes_frame.pack(fill="x", padx=20, pady=8)

        # ── تابع مشترک ساخت PDF با 6 ستون / 20 ردیف هر ستون ──
        def _make_pdf_font_tools(parent_win):
            """تابع کمکی: بارگذاری font و reshape/bidi"""
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors as rl_colors
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                 Paragraph, Spacer, HRFlowable)
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.units import cm
            except ImportError:
                messagebox.showerror("خطا","reportlab نصب نیست.\npip install reportlab",parent=parent_win)
                return None

            try:
                import arabic_reshaper
                from bidi.algorithm import get_display as bidi_disp
                _reshaper = arabic_reshaper.ArabicReshaper(dict(
                    delete_harakat=False, support_ligatures=True))
                def rt(s):
                    return bidi_disp(_reshaper.reshape(str(s)))
            except ImportError:
                messagebox.showerror("خطا",
                    "arabic_reshaper یا python-bidi نصب نیست.\n"
                    "pip install arabic_reshaper python-bidi", parent=parent_win)
                return None

            FONT = "BNazanin"
            font_ok = False
            font_files = ["BNazanin.ttf","B Nazanin.ttf","b_nazanin.ttf","b nazanin.ttf",
                          "BNazanin Bold.ttf","B Nazanin Bold.ttf"]
            font_dirs  = [os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "Fonts"),
                          r"C:\Windows\Fonts",
                          r"C:\Users\{}\AppData\Local\Microsoft\Windows\Fonts".format(
                              os.environ.get("USERNAME","")),
                          _app_base_dir(),
                          os.getcwd()]
            for d in font_dirs:
                for ff in font_files:
                    fp = os.path.join(d, ff)
                    if os.path.exists(fp):
                        try:
                            pdfmetrics.registerFont(TTFont(FONT, fp))
                            font_ok = True
                            break
                        except: pass
                if font_ok: break

            if not font_ok:
                for fb_file in ["tahoma.ttf","Tahoma.ttf","arial.ttf","Arial.ttf","arialuni.ttf"]:
                    for d in [os.path.join(os.environ.get("SystemRoot", r"C:\Windows"), "Fonts"),
                              r"C:\Windows\Fonts", _app_base_dir(), os.getcwd()]:
                        fp = os.path.join(d, fb_file)
                        if os.path.exists(fp):
                            try:
                                pdfmetrics.registerFont(TTFont("PersianFB", fp))
                                FONT = "PersianFB"
                                font_ok = True
                                break
                            except: pass
                    if font_ok: break

            if not font_ok:
                messagebox.showerror("خطا","فونت فارسی پیدا نشد. BNazanin.ttf را کنار برنامه قرار دهید.",
                                     parent=parent_win)
                return None

            return dict(
                A4=A4, rl_colors=rl_colors,
                SimpleDocTemplate=SimpleDocTemplate, Table=Table, TableStyle=TableStyle,
                Paragraph=Paragraph, Spacer=Spacer, HRFlowable=HRFlowable,
                ParagraphStyle=ParagraphStyle, cm=cm,
                FONT=FONT, rt=rt
            )

        def _build_6col_pdf(path, title, data_rows, req_name, from_sh, to_sh,
                             col3_header, sig1_label, sig2_label, rl):
            """
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            جدول ۶ ستونی فارسی (RTL) — حداکثر ۲۰ سطر در هر صفحه

            ترتیب ستون‌ها از راست به چپ (دید کاربر فارسی):
              ستون ۱ (راست): ردیف        → شماره‌های فرد:  ۱، ۳، ۵، ...
              ستون ۲:         شماره اسلب
              ستون ۳:         اپراتور
              ─────── خط جدا ───────
              ستون ۴:         ردیف        → شماره‌های زوج:  ۲، ۴، ۶، ...
              ستون ۵:         شماره اسلب
              ستون ۶ (چپ):   اپراتور

            سطر ۱ جدول ← اسلب ۱ (راست) + اسلب ۲ (چپ)
            سطر ۲ جدول ← اسلب ۳ (راست) + اسلب ۴ (چپ)
            ...
            سطر ۲۰ جدول ← اسلب ۳۹ (راست) + اسلب ۴۰ (چپ)

            نکته ReportLab: ستون ایندکس ۰ = چپ‌ترین ستون فیزیکی.
            پس برای RTL باید ستون‌های چپ (زوج) را اول و ستون‌های راست (فرد) را آخر بچینیم:
            آرایه هر سطر: [who_even, slab_even, no_even, | who_odd, slab_odd, no_odd]
                            ←── ستون چپ (فیزیکی ۰,۱,۲) ──→   ←── ستون راست (فیزیکی ۳,۴,۵) ──→
            ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            """
            A4    = rl["A4"]
            cm    = rl["cm"]
            FONT  = rl["FONT"]
            rt    = rl["rt"]
            rl_colors = rl["rl_colors"]
            SimpleDocTemplate = rl["SimpleDocTemplate"]
            Table = rl["Table"]
            TableStyle = rl["TableStyle"]
            _RLPara = rl["Paragraph"]
            Spacer = rl["Spacer"]
            HRFlowable = rl["HRFlowable"]
            ParagraphStyle = rl["ParagraphStyle"]

            def P(text, style):
                return _RLPara(rt(str(text)), style)

            def ps(name, sz, al=1, c="#000000"):
                from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
                align = TA_CENTER if al == 1 else (TA_RIGHT if al == 2 else TA_LEFT)
                return ParagraphStyle(name, fontName=FONT, fontSize=sz,
                                      alignment=align, leading=sz * 1.6,
                                      textColor=rl_colors.HexColor(c),
                                      wordWrap="RTL")

            doc = SimpleDocTemplate(path, pagesize=A4,
                                    rightMargin=1.2*cm, leftMargin=1.2*cm,
                                    topMargin=1.2*cm, bottomMargin=1.2*cm)
            W = A4[0] - 2.4*cm   # عرض قابل استفاده

            ROWS_PER_PAGE = 20   # حداکثر سطر در هر صفحه = ۲۰ اسلب × ۲ ستون = ۴۰ اسلب

            # استایل‌ها
            hs   = ps("hs",  9, 1, "#ffffff")   # هدر جدول — وسط‌چین سفید
            cs_c = ps("csc", 8, 1, "#000000")   # ردیف شماره — وسط‌چین
            cs_r = ps("csr", 8, 2, "#000000")   # متن فارسی — راست‌چین

            # عرض ستون‌ها (۶ ستون)
            num_w  = 1.0 * cm
            slab_w = 3.8 * cm
            who_w  = (W - 2 * (num_w + slab_w)) / 2
            # ترتیب فیزیکی (چپ→راست):
            # [who_even | slab_even | no_even | who_odd | slab_odd | no_odd]
            col_ws = [who_w, slab_w, num_w, who_w, slab_w, num_w]

            def make_page(page_slabs, page_num, total_pages):
                story = []

                # ── هدر صفحه ──
                story.append(P("بسمه تعالی",
                               ps("bism", 9, 1, "#444444")))
                story.append(Spacer(1, .15*cm))
                story.append(HRFlowable(width="100%", thickness=2,
                                        color=rl_colors.HexColor("#003366")))
                story.append(Spacer(1, .1*cm))
                story.append(P("شرکت سازه پیشگام مدیسه — فولاد سفید دشت",
                               ps("co", 11, 1, "#003366")))
                story.append(Spacer(1, .05*cm))
                story.append(P(title, ps("ttl", 14, 1, "#000000")))
                story.append(Spacer(1, .05*cm))
                story.append(HRFlowable(width="100%", thickness=1,
                                        color=rl_colors.HexColor("#003366")))
                story.append(Spacer(1, .12*cm))

                # ── اطلاعات گزارش ──
                _now_d = to_shamsi(datetime.datetime.now())
                info_data = [[
                    P(f"درخواست‌دهنده: {req_name}", ps("i1", 9, 2, "#333333")),
                    P(f"تعداد اسلب: {len(data_rows)} عدد", ps("i2", 9, 1, "#003366")),
                    P(f"تاریخ چاپ: {_now_d}", ps("i3", 9, 0, "#333333")),
                ]]
                info_tbl = Table(info_data, colWidths=[W/3, W/3, W/3])
                info_tbl.setStyle(TableStyle([
                    ("FONTNAME",       (0,0), (-1,-1), FONT),
                    ("ALIGN",          (0,0), (-1,-1), "CENTER"),
                    ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
                    ("BACKGROUND",     (0,0), (-1,-1), rl_colors.HexColor("#eef4fb")),
                    ("BOX",            (0,0), (-1,-1), .5, rl_colors.HexColor("#aaa")),
                    ("TOPPADDING",     (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING",  (0,0), (-1,-1), 4),
                ]))
                story.append(info_tbl)
                story.append(Spacer(1, .08*cm))
                story.append(P(f"بازه زمانی: از  {from_sh}  تا  {to_sh}",
                               ps("rng", 9, 1, "#555555")))
                if total_pages > 1:
                    story.append(P(f"صفحه {page_num} از {total_pages}",
                                   ps("pg", 8, 1, "#888888")))
                story.append(Spacer(1, .18*cm))

                # ── ساخت جدول ──
                # هدر: ترتیب فیزیکی چپ→راست:
                # [col3_even | slab_even | no_even | col3_odd | slab_odd | no_odd]
                # که کاربر فارسی می‌بیند (راست→چپ):
                # [no_odd | slab_odd | col3_odd | no_even | slab_even | col3_even]
                #   ردیف    شماره اسلب  اپراتور    ردیف    شماره اسلب  اپراتور
                header = [
                    P(col3_header,   hs),   # ستون فیزیکی ۰ = چپ‌ترین = col3 زوج
                    P("شماره اسلب", hs),   # ستون فیزیکی ۱
                    P("ردیف",        hs),   # ستون فیزیکی ۲
                    P(col3_header,   hs),   # ستون فیزیکی ۳
                    P("شماره اسلب", hs),   # ستون فیزیکی ۴
                    P("ردیف",        hs),   # ستون فیزیکی ۵ = راست‌ترین = ردیف فرد
                ]
                rows_tbl = [header]

                base = (page_num - 1) * (ROWS_PER_PAGE * 2)  # offset شماره ردیف مطلق

                # هر سطر جدول = اسلب فرد (ستون راست) + اسلب زوج (ستون چپ)
                # اسلب فرد: ایندکس ۰، ۲، ۴، ... → شماره ردیف ۱، ۳، ۵، ...
                # اسلب زوج: ایندکس ۱، ۳، ۵، ... → شماره ردیف ۲، ۴، ۶، ...
                n = len(page_slabs)
                num_rows = (n + 1) // 2   # تعداد سطرهای واقعی (بدون سطر خالی)

                for i in range(num_rows):
                    odd_i  = i * 2        # ایندکس اسلب فرد  → ستون راست
                    even_i = i * 2 + 1    # ایندکس اسلب زوج  → ستون چپ

                    # اسلب فرد (ستون راست صفحه = فیزیکی ۳،۴،۵)
                    if odd_i < n:
                        s = page_slabs[odd_i]
                        odd_no  = P(str(base + odd_i + 1), cs_c)
                        odd_sid = P(s["slab_id"],           cs_r)
                        odd_who = P(s.get("_col3", "—"),    cs_r)
                    else:
                        odd_no = odd_sid = odd_who = P("", cs_c)

                    # اسلب زوج (ستون چپ صفحه = فیزیکی ۰،۱،۲)
                    if even_i < n:
                        s = page_slabs[even_i]
                        even_no  = P(str(base + even_i + 1), cs_c)
                        even_sid = P(s["slab_id"],            cs_r)
                        even_who = P(s.get("_col3", "—"),     cs_r)
                    else:
                        even_no = even_sid = even_who = P("", cs_c)

                    # ترتیب فیزیکی سطر:
                    # [who_even, slab_even, no_even, who_odd, slab_odd, no_odd]
                    # کاربر فارسی (RTL) می‌بیند:
                    # no_odd | slab_odd | who_odd || no_even | slab_even | who_even
                    #  ردیف     اسلب     اپراتور      ردیف      اسلب      اپراتور
                    rows_tbl.append([
                        even_who, even_sid, even_no,
                        odd_who,  odd_sid,  odd_no,
                    ])

                tbl = Table(rows_tbl, colWidths=col_ws, repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,0),  rl_colors.HexColor("#0d1e3c")),
                    ("TEXTCOLOR",     (0,0), (-1,0),  rl_colors.white),
                    ("ROWBACKGROUNDS",(0,1), (-1,-1),
                     [rl_colors.white, rl_colors.HexColor("#eef4fb")]),
                    ("TEXTCOLOR",     (0,1), (-1,-1), rl_colors.black),
                    ("FONTNAME",      (0,0), (-1,-1), FONT),
                    ("FONTSIZE",      (0,0), (-1,-1), 8),
                    ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                    ("GRID",          (0,0), (-1,-1), .4, rl_colors.HexColor("#bbbbbb")),
                    # خط جداکننده بین دو نیمه (بعد از ستون فیزیکی ۲)
                    ("LINEAFTER",     (2,0), (2,-1),  1.5, rl_colors.HexColor("#003366")),
                    ("TOPPADDING",    (0,0), (-1,-1), 5),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ]))
                story.append(tbl)
                story.append(Spacer(1, .8*cm))

                # ── جای امضا (فقط صفحه آخر) ──
                if page_num == total_pages:
                    sig_col = W / 2
                    sig_data = [[
                        P(sig2_label, ps("s2", 9, 1, "#333333")),
                        P(sig1_label, ps("s1", 9, 1, "#333333")),
                    ]]
                    sig_tbl = Table(sig_data, colWidths=[sig_col, sig_col])
                    sig_tbl.setStyle(TableStyle([
                        ("FONTNAME",      (0,0), (-1,-1), FONT),
                        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                        ("LINEABOVE",     (0,0), (0,0),  .8, rl_colors.black),
                        ("LINEABOVE",     (1,0), (1,0),  .8, rl_colors.black),
                        ("TOPPADDING",    (0,0), (-1,-1), 28),
                        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                    ]))
                    story.append(sig_tbl)

                story.append(Spacer(1, .3*cm))
                story.append(HRFlowable(width="100%", thickness=1,
                                        color=rl_colors.HexColor("#003366")))
                story.append(P(
                    "این گزارش به صورت سیستمی تولید شده است — سامانه مدیریت تختال، شرکت سازه پیشگام مدیسه",
                    ps("ft", 6, 1, "#999999")))
                return story

            # ── ساخت همه صفحات ──
            from reportlab.platypus import PageBreak
            all_story  = []
            slabs_per_page = ROWS_PER_PAGE * 2   # ۴۰ اسلب در هر صفحه
            total_pages = max(1, (len(data_rows) + slabs_per_page - 1) // slabs_per_page)

            for page_num in range(1, total_pages + 1):
                start = (page_num - 1) * slabs_per_page
                end   = start + slabs_per_page
                all_story += make_page(data_rows[start:end], page_num, total_pages)
                if page_num < total_pages:
                    all_story.append(PageBreak())

            doc.build(all_story)

        # ══════════════════
        #  باکس ۱: QC شده‌ها
        # ══════════════════
        box_qc = tk.Frame(boxes_frame, bg=C["card"],
                           highlightthickness=1, highlightbackground=C["accent"])
        box_qc.pack(side="right", fill="both", expand=True, padx=(0,6))
        tk.Frame(box_qc, bg=C["accent"], height=3).pack(fill="x")
        box_qc_in = tk.Frame(box_qc, bg=C["card"])
        box_qc_in.pack(fill="both", expand=True, padx=16, pady=12)

        # آیکون/تصویر نوشته
        tk.Label(box_qc_in, text="📋", bg=C["card"],
                 font=("Segoe UI Emoji", 36, "bold")).pack(pady=(8,4))
        tk.Label(box_qc_in, text="گزارش PDF",
                 bg=C["card"], fg=C["accent"], font=(_MAIN_FONT,13,"bold")).pack()
        tk.Label(box_qc_in, text="اسلب‌های کنترل کیفی شده",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL,
                 wraplength=220, justify="center").pack(pady=(4,12))

        status_qc = tk.Label(box_qc_in, text="", bg=C["card"],
                              fg=C["success"], font=FONT_SMALL)
        status_qc.pack()

        def gen_pdf_qc():
            req_name = req_name_var.get().strip() or self.udata.get("display","—")
            try:
                from_sh = parse_sh(from_var.get())
                to_sh   = parse_sh(to_var.get())
            except ValueError:
                messagebox.showerror("خطا","فرمت تاریخ اشتباه.\nمثال: 1404/03/15  07:00:00",parent=self)
                return
            db = load_db()
            filtered = [
                r for r in db["melts"]
                if r.get("qc_status")=="کنترل کیفی شده"
                and r.get("qc_at","")
                and from_sh <= r.get("qc_at","") <= to_sh
            ]
            if not filtered:
                messagebox.showinfo("اطلاع","هیچ اسلب QC‌شده‌ای در این بازه یافت نشد.",parent=self)
                return

            # افزودن ستون سوم: نام اپراتور ثبت‌کننده
            for r in filtered:
                r["_col3"] = get_display_name(r.get("registered_by","—"), db)

            rl = _make_pdf_font_tools(self)
            if not rl: return

            fname = f"گزارش_کنترل_کیفی_{shamsi_date_for_filename()}.pdf"
            path = self._resolve_report_save_path(
                "qc_pdf", fname, [("PDF","*.pdf")], ".pdf")
            if not path: return
            try:
                _build_6col_pdf(
                    path=path,
                    title="اسلب‌های کنترل کیفی شده",
                    data_rows=filtered,
                    req_name=req_name,
                    from_sh=from_sh,
                    to_sh=to_sh,
                    col3_header="اپراتور ثبت‌کننده",
                    sig1_label="امضای اپراتور ثبت‌کننده",
                    sig2_label="امضای مسئول کنترل کیفی",
                    rl=rl,
                )
                status_qc.config(text=f"✔ ذخیره شد: {os.path.basename(path)}")
                self._remember_last_report("qc_pdf", path)
                messagebox.showinfo("موفق",f"PDF با موفقیت ذخیره شد:\n{path}",parent=self)
            except Exception as ex:
                messagebox.showerror("خطا",f"خطا در ساخت PDF:\n{ex}",parent=self)

        styled_btn(box_qc_in, "📄  دریافت گزارش PDF",
                   gen_pdf_qc, color=C["accent2"]).pack(pady=8, fill="x")
        styled_btn(box_qc_in, "🖨️  نمایش آخرین گزارش",
                   lambda: self._open_last_report("qc_pdf", "کنترل کیفی شده"),
                   color=C["text_dim"]).pack(pady=(0,8), fill="x")

        # ═══════════════════════════
        #  باکس ۲: تحویلی آزمایشگاه
        # ═══════════════════════════
        box_lab = tk.Frame(boxes_frame, bg=C["card"],
                            highlightthickness=1, highlightbackground=C["btn_success"])
        box_lab.pack(side="right", fill="both", expand=True, padx=(6,0))
        tk.Frame(box_lab, bg=C["btn_success"], height=3).pack(fill="x")
        box_lab_in = tk.Frame(box_lab, bg=C["card"])
        box_lab_in.pack(fill="both", expand=True, padx=16, pady=12)

        tk.Label(box_lab_in, text="🧪", bg=C["card"],
                 font=("Segoe UI Emoji", 36, "bold")).pack(pady=(8,4))
        tk.Label(box_lab_in, text="گزارش PDF",
                 bg=C["card"], fg=C["btn_success"], font=(_MAIN_FONT,13,"bold")).pack()
        tk.Label(box_lab_in, text="گزارش اسلب‌های تحویلی به آزمایشگاه",
                 bg=C["card"], fg=C["text_dim"], font=FONT_SMALL,
                 wraplength=220, justify="center").pack(pady=(4,12))

        status_lab = tk.Label(box_lab_in, text="", bg=C["card"],
                               fg=C["success"], font=FONT_SMALL)
        status_lab.pack()

        def gen_pdf_lab():
            req_name = req_name_var.get().strip() or self.udata.get("display","—")
            try:
                from_sh = parse_sh(from_var.get())
                to_sh   = parse_sh(to_var.get())
            except ValueError:
                messagebox.showerror("خطا","فرمت تاریخ اشتباه.\nمثال: 1404/03/15  07:00:00",parent=self)
                return
            db = load_db()
            filtered = [
                r for r in db.get("lab_deliveries",[])
                if r.get("delivered_at","")
                and from_sh <= r.get("delivered_at","") <= to_sh
            ]
            if not filtered:
                messagebox.showinfo("اطلاع","هیچ اسلب تحویلی به آزمایشگاه در این بازه یافت نشد.",parent=self)
                return

            # افزودن ستون سوم: نام تحویل‌دهنده
            for r in filtered:
                r["_col3"] = get_display_name(r.get("delivered_by","—"), db)

            rl = _make_pdf_font_tools(self)
            if not rl: return

            fname = f"گزارش_آزمایشگاه_{shamsi_date_for_filename()}.pdf"
            path = self._resolve_report_save_path(
                "lab_pdf", fname, [("PDF","*.pdf")], ".pdf")
            if not path: return
            try:
                _build_6col_pdf(
                    path=path,
                    title="لیست اسلب‌های تحویلی به آزمایشگاه",
                    data_rows=filtered,
                    req_name=req_name,
                    from_sh=from_sh,
                    to_sh=to_sh,
                    col3_header="تحویل دهنده",
                    sig1_label="امضای تحویل دهنده",
                    sig2_label="امضای مسئول آزمایشگاه",
                    rl=rl,
                )
                status_lab.config(text=f"✔ ذخیره شد: {os.path.basename(path)}")
                self._remember_last_report("lab_pdf", path)
                messagebox.showinfo("موفق",f"PDF با موفقیت ذخیره شد:\n{path}",parent=self)
            except Exception as ex:
                messagebox.showerror("خطا",f"خطا در ساخت PDF:\n{ex}",parent=self)

        styled_btn(box_lab_in, "📄  دریافت گزارش PDF",
                   gen_pdf_lab, color=C["btn_success"]).pack(pady=8, fill="x")
        styled_btn(box_lab_in, "🖨️  نمایش آخرین گزارش",
                   lambda: self._open_last_report("lab_pdf", "آزمایشگاه"),
                   color=C["text_dim"]).pack(pady=(0,8), fill="x")



# ═══════════════════════════════════════════════════════════
#  نقطه ورود
# ═══════════════════════════════════════════════════════════
def hide_console_window():
    """پنجرهٔ سیاه CMD را مخفی می‌کند (فقط ویندوز)."""
    if os.name != "nt":
        return
    try:
        import ctypes
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 0)
    except Exception:
        pass


if __name__ == "__main__":
    # ── اجرای پروسه‌ی نگهبان تسک‌بار (safety net) ──
    if _is_watchdog_argv():
        pid = _watchdog_parent_pid()
        if pid:
            try:
                _run_taskbar_watchdog(pid)
            except Exception:
                pass
        sys.exit(0)

    hide_console_window()

    missing = []
    if not XLSX:
        missing.append("openpyxl  (pip install openpyxl)")
    if not SHAMSI_OK:
        missing.append("jdatetime  (pip install jdatetime)  — بدون این کتابخانه از الگوریتم داخلی استفاده می‌شود")
    if missing:
        print("⚠️  پیش‌نیازهای زیر نصب نیستند:\n" + "\n".join(missing))
    ensure_report_dirs()
    app = LoginWindow()
    app.mainloop()
